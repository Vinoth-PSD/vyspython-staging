from django.contrib.auth import authenticate
from rest_framework.views import APIView
from rest_framework.response import Response 
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from .serializers import ChangePasswordSerializer, ProfileEduDetailsSerializer, ProfileFamilyDetailsSerializer, ProfileHoroscopeSerializer, ProfilePartnerPrefSerializer 
from rest_framework import viewsets
from .models import Country, ProfileEduDetails, ProfileFamilyDetails, ProfilePartnerPref, State, District, ProfileHolder, MaritalStatus, Height, Complexion, ParentsOccupation, HighestEducation, UgDegree, AnnualIncome, BirthStar, Rasi, Lagnam, DasaBalance, FamilyType, FamilyStatus, FamilyValue, LoginDetailsTemp ,Get_profiledata , Mode , Property , Gothram , EducationLevel , Profession , Match , MasterStatePref , AdminUser , Role , City , Express_interests , Profile_visitors, Profile_wishlists , Photo_request , PlanDetails , Image_Upload  ,ProfileStatus , MatchingStarPartner, Image_Upload, Profile_personal_notes, Registration1 , Get_profiledata_Matching , Profespref , Profile_vysassist , Homepage,ProfileLoginLogs,ProfileSendFromAdmin , ProfileSubStatus , Profile_PlanFeatureLimit , ProfileVysAssistFollowup , VysAssistcomment ,ProfileSuggestedPref , Profile_callogs , ProfileHoroscope , MasterhighestEducation ,PlanSubscription , ProfileVisibility ,Addonpackages ,Roles ,CallManagement,CallLog,ActionLog,AssignLog ,CallTypeMaster,ParticularsMaster,CallStatusMaster,ActionPointMaster , AssignLog_New , ActionLog_New , CallLog_New , CallManagement_New
from collections import defaultdict
from .serializers import CountrySerializer, StateSerializer, DistrictSerializer,ProfileHolderSerializer, MaritalStatusSerializer, HeightSerializer, ComplexionSerializer, ParentsOccupationSerializer, HighestEducationSerializer, UgDegreeSerializer, AnnualIncomeSerializer,BirthStarSerializer, RasiSerializer, LagnamSerializer, DasaBalanceSerializer, FamilyTypeSerializer, FamilyStatusSerializer, FamilyValueSerializer, LoginDetailsTempSerializer,Getnewprofiledata , ModeSerializer, PropertySerializer , GothramSerializer , EducationLevelSerializer ,ProfessionSerializer , MatchSerializer ,MasterStatePrefSerializer , CitySerializer , Getnewprofiledata_new , QuickUploadSerializer , ProfileStatusSerializer , LoginEditSerializer , GetproflistSerializer , ImageGetSerializer , MatchingscoreSerializer , HomepageSerializer, Profile_idValidationSerializer , UpdateAdminComments_Serializer , ProfileSubStatusSerializer , PlandetailsSerializer ,ProfileplanSerializer , ProfileVysAssistFollowupSerializer , VysassistSerializer , ProfileSuggestedPrefSerializer  , AdminUserDropdownSerializer , ProfileVisibilitySerializer ,LoginSerializer ,RoleSerializers ,RoleDropdownSerializer
from rest_framework.decorators import action
from rest_framework import generics, filters
from rest_framework.pagination import PageNumberPagination
from django.db.models import Q
import openpyxl
from django.http import HttpResponse
from .models import Get_profiledata
from datetime import datetime
from django.http import JsonResponse
from django.conf import settings
from django.shortcuts import get_object_or_404
from datetime import datetime, timedelta , date
from django.utils import timezone
import calendar
import csv
from django.utils.encoding import smart_str
from django.core.mail import send_mail
from django.core.mail import EmailMessage
from email.mime.image import MIMEImage
from django.http import StreamingHttpResponse
from rest_framework.parsers import JSONParser, MultiPartParser
import json
from django.forms.models import model_to_dict
from django.core.exceptions import ObjectDoesNotExist
from PyPDF2 import PdfMerger
import tempfile
from . import models
from authentication.views import My_horoscope_generate,WithoutAddressSendEmailAPI,WithoutAddressPrintPDF,generate_pdf_without_address
from rest_framework.views import APIView
from django.http import JsonResponse
from rest_framework.test import APIRequestFactory  # Helps to create a request object
from rest_framework.request import Request
import io
from rest_framework.viewsets import ModelViewSet
from .models import SentFullProfileEmailLog
from .models import SentShortProfileEmailLog
from .models import SentFullProfilePrintPDFLog
from .models import SentShortProfilePrintPDFLog
from .models import SentFullProfilePrintwpLog
from .models import SentShortProfilePrintwpLog
from .models import CallType
from .serializers import CallTypeSerializer
from .models import CallStatus
from .serializers import CallStatusSerializer,PersonalnotesSerializer
from .models import CallAction
from .serializers import CallActionSerializer
from .models import ProfileCallManagement
from .serializers import ProfileCallManagementSerializer
from .models import MarriageSettleDetails
from .serializers import MarriageSettleDetailsSerializer
from .models import PaymentTransaction
from .serializers import PaymentTransactionSerializer
from .serializers import InvoiceSerializer
from .serializers import LoginLogSerializer,Renewalprofiledata,PaymentTransactionSerializer
from django.db.models.functions import Cast
from django.db.models import DateTimeField
from .models import Invoice,AdminPrintLogs
import tempfile
from xhtml2pdf import pisa
from io import BytesIO,StringIO
import base64
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from authentication.models import Horoscope
import requests
import re
from dateutil import parser
from datetime import datetime, time
from .serializers import CommonProfileSearchSerializer
from azure.storage.blob import BlobServiceClient, ContentSettings
from PIL import Image, ImageFilter
from io import BytesIO
import os
import logging
from collections import OrderedDict
from authentication.views import fetch_porutham_details,get_dasa_name,dasa_format_date,format_date_of_birth
from authentication.models import Get_profiledata as gpt
from .serializers import PlanSubscriptionSerializer,PlanSubscriptionListSerializer,PaymentTransactionListSerializer
from authentication.helpers.matching import preload_matching_scores , get_matching_score_util
from django.core.cache import caches
from django.core.cache import cache
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

from authentication.views import get_profile_image_azure_optimized
from .models import DataHistory
from django.db.models import QuerySet
from django.shortcuts import render

from rest_framework import viewsets, permissions

from accounts.utils.round_robin_assign import assign_user_for_state
from django.core import signing





from rest_framework.permissions import IsAuthenticated, AllowAny
from django.contrib.auth import get_user_model
from django.contrib.auth import authenticate
from rest_framework.authtoken.models import Token

from .models import Role, RolePermission , User
from .serializers import UserSerializer,DashboardSerializer

from django.db.models import F, Value
from django.db.models.functions import Concat

import pandas as pd
import numpy as np
from django.db.models import OuterRef, Subquery
from django.db.models import Exists
from openpyxl.utils import get_column_letter
# User = get_user_model()


# from authentication.models import ProfileVisibility
# from authentication.serializers import ProfileVisibilityListSerializer

#
# class ModeViewSet(viewsets.ModelViewSet):
#     queryset = Mode.objects.filter(is_deleted=False)  # Only show non-deleted records
#     serializer_class = ModeSerializer

#     def destroy(self, request, args, *kwargs):
#         # Override the destroy method to implement soft delete
#         instance = self.get_object()
#         instance.is_deleted = True
#         instance.save()
#         return Response({"status": "deleted"})


class DashboardcountView(APIView):
      def get(self, request):

        try:
            # Counts based on assumptions
            new_profiles = LoginDetails.objects.filter(status=0).count()
            approved_profiles = LoginDetails.objects.filter(status=1).count()
            pending_profiles = LoginDetails.objects.filter(status=2).count()
            hidden_profiles = LoginDetails.objects.filter(status=3).count()
            photo_request_count = Image_Upload.objects.filter(image_approved=0,is_deleted=0).count()
            quick_upload_count = LoginDetails.objects.filter(quick_registration=1).count()
            paidprofiles_count = LoginDetails.objects.filter(~Q(Plan_id__in=[6, 7, 8, 9, 11, 12, 13])).count()
            prospect_profiles = LoginDetails.objects.filter(Plan_id=8).exclude(status__in=[0, 3, 4]).count()
            featured_profiles = LoginDetails.objects.filter(Plan_id__in=[3, 4]).exclude(status__in=[0, 3, 4]).count()
            deletedprofiles = LoginDetails.objects.filter(status=4).count()

            return Response({
                "new_profiles": new_profiles,
                "approved_profiles": approved_profiles,
                "pending_profiles":pending_profiles,
                "photo_request_count": photo_request_count,
                "hidden_profiles_count": hidden_profiles,
                "quick_upload_count": quick_upload_count,
                "paidprofiles_count":paidprofiles_count,
                "prospect_profiles":prospect_profiles,
                "featured_profiles":featured_profiles,
                "deletedprofiles":deletedprofiles
            })

        except Exception as e:
            return Response({"error": str(e)}, status=500)  


class GetMasterStatus(APIView):
    def get(self, request):
        statuses = ProfileStatus.objects.all()
        serializer = ProfileStatusSerializer(statuses, many=True)
        return Response({
            'status': 'success',
            'data': serializer.data
        })
        
class GetSubMasterStatus(APIView):
    def post(self, request):
        primary_status=request.data.get('primary_status')
        owner_id = request.data.get('admin_user_id')
        try:
            owner_id = int(owner_id)
            user = User.objects.get(id=owner_id)
        except Exception:
            user = None

        if user:
            role = user.role
            permissions = RolePermission.objects.filter(role=role).select_related('action')
            data = permissions.values('action__code', 'value')
            edit_permission = data.filter(action__code='membership_activation').first()
            perm = edit_permission['value'] if edit_permission else None
        else:
            perm = None

        if not primary_status:
            return Response({
                'status': 'error',
                'message': 'primary_status is required'
            }, status=status.HTTP_400_BAD_REQUEST)

        statuses = ProfileSubStatus.objects.filter(status_code=primary_status)
        serializer = ProfileSubStatusSerializer(statuses, many=True)
        response_data = serializer.data
        if str(primary_status) == "1" and perm is not None:

            for item in response_data:
                if perm == 0:
                    item["value"] = 0
                elif perm == 1:
                    item["value"] = 1
                elif perm in [2, 3]:
                    item["value"] = 1 if item["id"] in [1, 2, 3, 4] else 0

        return Response({
            'status': 'success',
            'data': response_data
        })
    

class GetPlanbyStatus(APIView):
    def post(self, request):
        secondary_status=request.data.get('secondary_status')

        if not secondary_status:
            return Response({
                'status': 'error',
                'message': 'secondary_status is required'
            }, status=status.HTTP_400_BAD_REQUEST)

        statuses = PlanDetails.objects.filter(master_substatus =secondary_status)
        serializer = PlandetailsSerializer(statuses, many=True)
                
        return Response({
            'status': 'success',
            'data': serializer.data
        })

class GetallPlans(APIView):
    def post(self, request):


        statuses = PlanDetails.objects.filter()
        serializer = PlandetailsSerializer(statuses, many=True)
                
        return Response({
            'status': 'success',
            'data': serializer.data
        })



class ModeViewSet(viewsets.ModelViewSet):
    serializer_class = ModeSerializer

    # Only show non-deleted records
    queryset = Mode.objects.filter(is_deleted=False)

    # Retrieve an object by 'mode' (primary key)
    def get_object(self):
        try:
            return Mode.objects.get(mode=self.kwargs.get('pk'), is_deleted=False)
        except Mode.DoesNotExist:
            raise Http404("Mode not found")

    # Override the destroy method to implement soft delete by mode ID
    # def destroy(self, request, args, *kwargs):
    #     instance = self.get_object()
    #     instance.is_deleted = True
    #     instance.save()
    #     return Response({"status": "deleted"}, status=status.HTTP_204_NO_CONTENT)
    def destroy(self, request, *args, **kwargs):
        # 'pk' is already part of kwargs, no need to pass it explicitly
        instance = self.get_object()  # Will use the 'pk' from kwargs
        self.perform_destroy(instance)
        return Response(status=status.HTTP_204_NO_CONTENT)

    # Update method to edit a Mode object based on the mode ID (pk)
    def update(self, request, *args, **kwargs):
        instance = self.get_object()  # This will fetch the object using pk from kwargs
        serializer = self.get_serializer(instance, data=request.data, partial=True)  # Allow partial updates
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return Response(serializer.data)


    # Optionally, override the list method to customize behavior for retrieving objects
    def list(self, request):
        queryset = Mode.objects.filter(is_deleted=False)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
    
from accounts.models import ProfileHoroscope


class ProfileHoroscopeViewSet(viewsets.ModelViewSet):
    queryset = ProfileHoroscope.objects.all()
    serializer_class = ProfileHoroscopeSerializer


class PropertyViewSet(viewsets.ModelViewSet):
    queryset = Property.objects.filter(is_deleted=False)  # Filter out soft-deleted entries
    serializer_class = PropertySerializer

    def perform_destroy(self, instance):
        """Override delete behavior to implement soft delete."""
        instance.is_deleted = True
        instance.save()


class GothramViewSet(viewsets.ModelViewSet):
    queryset = Gothram.objects.filter(is_deleted=False)  # Filter out soft-deleted entries
    serializer_class = GothramSerializer
    
    def list(self, request, *args, **kwargs):
        # Custom flattened response
        flattened_data = []
        queryset = self.get_queryset()
        for gothram in queryset:
            sankethas = [s.strip() for s in gothram.sanketha_namam.split('-')]
            for sanketha in sankethas:
                flattened_data.append({
                    "id": gothram.id,
                    "gothram_name": gothram.gothram_name,
                    "rishi": gothram.rishi,
                    "sanketha_namam": sanketha
                })
        return Response(flattened_data)
    
    def perform_destroy(self, instance):
        """Override delete behavior to implement soft delete."""
        instance.is_deleted = True
        instance.save()

class EducationLevelViewSet(viewsets.ModelViewSet):
    queryset = EducationLevel.objects.filter(is_deleted=False)  # Exclude soft-deleted entries
    serializer_class = EducationLevelSerializer

    def perform_destroy(self, instance):
        """Override delete to implement soft delete."""
        instance.is_deleted = True
        instance.save()

class ProfessionViewSet(viewsets.ModelViewSet):
    queryset = Profession.objects.filter(is_deleted=False)  # Only show non-deleted records
    serializer_class = ProfessionSerializer

    def destroy(self, request, args, *kwargs):
        profession = self.get_object()
        profession.is_deleted = True
        profession.save()
        return Response(status=status.HTTP_204_NO_CONTENT)


class MatchViewSet(viewsets.ModelViewSet):
    queryset = Match.objects.filter(is_deleted=False)  # Filter out soft-deleted records
    serializer_class = MatchSerializer

    def destroy(self, request, args, *kwargs):
        match = self.get_object()
        match.is_deleted = True  # Perform soft delete by setting is_deleted to True
        match.save()
        return Response(status=status.HTTP_204_NO_CONTENT)


class MasterStatePrefViewSet(viewsets.ModelViewSet):
    queryset = MasterStatePref.objects.filter(is_deleted=False)  # Exclude soft deleted items
    serializer_class = MasterStatePrefSerializer

    @action(detail=True, methods=['patch'])
    def soft_delete(self, request, pk=None):
        try:
            state = MasterStatePref.objects.get(pk=pk)
            state.is_deleted = True
            state.save()
            return Response({"message": "Deleted successfully."}, status=status.HTTP_200_OK)
        except MasterStatePref.DoesNotExist:
            return Response({"error": "State not found."}, status=status.HTTP_404_NOT_FOUND)


# class SignInView(APIView):
#     def post(self, request):
#         email = request.data.get('email')
#         password = request.data.get('password')
#         user = authenticate(request, username=email, password=password)
#         if user is not None:
#             return Response({'message': 'Success'}, status=status.HTTP_200_OK)
#         return Response({'message': 'Failed'}, status=status.HTTP_401_UNAUTHORIZED)



# class SignInView(APIView):
#     def post(self, request):
#         email = request.data.get('email')
#         password = request.data.get('password')
        
#         #user = authenticate(request, username=email, password=password)
        
#         user=AdminUser.objects.get(email=email,password=password)
        
#         if user is not None:

            


#             return Response({'message': 'Success'}, status=status.HTTP_200_OK)
#         return Response({'message': 'Failed'}, status=status.HTTP_401_UNAUTHORIZED)


# class SignInView(APIView):
#     def post(self, request):
#         email = request.data.get('email')
#         password = request.data.get('password')
        
#         # Assuming you are directly querying the user with email and password.
#         try:
#             user = AdminUser.objects.get(email=email, password=password)
#         except AdminUser.DoesNotExist:
#             return Response({'message': 'Failed'}, status=status.HTTP_401_UNAUTHORIZED)

#         # If user is found, include user data in the response.
#         user_data = {
#             'id': user.id,
#             'email': user.email,
#             'name': user.first_name,  # Assuming 'name' is a field in the AdminUser model
#             'role':user.role_id
#         }

#         return Response({'message': 'Success', 'user': user_data}, status=status.HTTP_200_OK)

class SignInView(APIView):

    def post(self, request):
        username = request.data.get('email')
        password = request.data.get('password')

        if not username or not password:
            return Response({
                "status": "error",
                "message": "Eamil and password are required"
            }, status=400)

        try:
            user = AdminUser.objects.get(email=username)

            
            if user.password == password:
                role = user.role_id  

                
                permissions = {
                    # "view_permission": "1" if role.view_only else "0",
                    # "add_permission": "1" if role.sales else "0",
                    # "edit_permission": "1" if role.support else "0",
                    # "delete_permission": "1" if role.biz_dev else "0"

                    "view_permission": "1" ,
                    "add_permission": "1",
                    "edit_permission": "1",
                    "delete_permission": "1"

                }

                full_name=user.first_name + user.last_name

                return Response({
                    "status": "success",
                    "message": "Login successful",
                    "data": {
                        "user": {
                            "id": user.id,
                            "username": user.username,
                            "email": user.email,
                            "first_name": user.first_name,
                            "last_name": user.last_name,
                            "full_name":full_name,
                            "role_id": role.id,
                            "role":role.role_name,
                            "permissions": permissions  
                        }
                    }
                })
            else:
                return Response({
                    "status": "error",
                    "message": "Invalid password"
                }, status=401)

        except AdminUser.DoesNotExist:
            return Response({
                "status": "error",
                "message": "User not found"
            }, status=404)





class ChangePasswordView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = ChangePasswordSerializer(data=request.data)
        if serializer.is_valid():
            user = request.user
            if not user.check_password(serializer.validated_data['old_password']):
                return Response({'old_password': 'Old password is incorrect'}, status=status.HTTP_400_BAD_REQUEST)

            user.set_password(serializer.validated_data['new_password'])
            user.save()
            return Response({'message': 'Password successfully changed'}, status=status.HTTP_200_OK)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class CountryViewSet(viewsets.ModelViewSet):
    queryset = Country.objects.filter(is_deleted=False)  # Filter out deleted countries
    serializer_class = CountrySerializer

    def destroy(self, request, *args, **kwargs):
        country = self.get_object()
        country.is_deleted = True
        country.save()
        return Response(status=status.HTTP_204_NO_CONTENT)

# class StateViewSet(viewsets.ModelViewSet):
#     queryset = State.objects.filter(is_deleted=False)  # Only fetch non-deleted states
#     serializer_class = StateSerializer

#     def destroy(self, request, *args, **kwargs):
#         state = self.get_object()
#         state.is_deleted = True
#         state.save()
#         return Response(status=status.HTTP_204_NO_CONTENT)

class StateViewSet(viewsets.ModelViewSet):
    queryset = State.objects.filter(is_deleted=False)  # Only fetch non-deleted states
    serializer_class = StateSerializer

    # def destroy(self, request, args, *kwargs):
    #     state = self.get_object()
    #     state.is_deleted = True
    #     state.save()
    #     return Response(status=status.HTTP_204_NO_CONTENT)

    def destroy(self, request, *args, **kwargs):
        # 'pk' is already part of kwargs, no need to pass it explicitly
        instance = self.get_object()  # Will use the 'pk' from kwargs
        self.perform_destroy(instance)
        return Response(status=status.HTTP_204_NO_CONTENT)

# class DistrictViewSet(viewsets.ModelViewSet):
#     queryset = District.objects.filter(is_deleted=False)  # Only fetch non-deleted districts
#     serializer_class = DistrictSerializer

#     def destroy(self, request, *args, **kwargs):
#         district = self.get_object()
#         district.is_deleted = True
#         district.save()
#         return Response(status=status.HTTP_204_NO_CONTENT)

class DistrictViewSet(viewsets.ModelViewSet):
    queryset = District.objects.all()  # Provide a default queryset for the router
    serializer_class = DistrictSerializer

    def get_queryset(self):
        queryset = District.objects.filter(is_deleted=False)  # Filter non-deleted districts
        state_id = self.request.query_params.get('state_id', None)
        if state_id is not None:
            return queryset.filter(state_id=state_id)
        return queryset

# class CityViewSet(viewsets.ModelViewSet):
#     queryset = City.objects.filter(is_deleted=False)  # Only fetch non-deleted cities
#     serializer_class = CitySerializer

#     def get_queryset(self):
#         district_id = self.request.query_params.get('district_id', None)
#         if district_id is not None:
#             return self.queryset.filter(district_id=district_id)
#         return self.queryset

#     # Soft delete the city by updating is_deleted to True
#     def destroy(self, request, args, *kwargs):
#         city = self.get_object()
#         city.is_deleted = True
#         city.save(update_fields=['is_deleted'])  # Only update the is_deleted field
#         return Response(status=status.HTTP_204_NO_CONTENT)

class CityViewSet(viewsets.ModelViewSet):
    queryset = City.objects.filter(is_deleted=False)  # Only fetch non-deleted cities
    serializer_class = CitySerializer

    def get_queryset(self):
        district_id = self.request.query_params.get('district_id', None)
        if district_id is not None:
            return self.queryset.filter(district_id=district_id, is_deleted=False)
        return self.queryset.filter(is_deleted=False)

    # Soft delete the city by updating is_deleted to True
    def destroy(self, request, *args, **kwargs):  # Corrected signature
        city = self.get_object()
        city.is_deleted = True
        city.save(update_fields=['is_deleted'])  # Only update the is_deleted field
        return Response(status=status.HTTP_204_NO_CONTENT)

class ProfileHolderViewSet(viewsets.ModelViewSet):
    queryset = ProfileHolder.objects.all()
    serializer_class = ProfileHolderSerializer


class MaritalStatusViewSet(viewsets.ModelViewSet):
    queryset = MaritalStatus.objects.filter(is_deleted=False)  # Only fetch non-deleted statuses
    serializer_class = MaritalStatusSerializer
    lookup_field = 'StatusId'  # Ensure lookups use StatusId

    def destroy(self, request, *args, **kwargs):
        marital_status = self.get_object()
        marital_status.is_deleted = True
        marital_status.save()
        return Response(status=status.HTTP_204_NO_CONTENT)


from rest_framework import status
from rest_framework.response import Response

# class HeightViewSet(viewsets.ModelViewSet):
#     queryset = Height.objects.filter(is_deleted=False)  # Only fetch non-deleted heights
#     serializer_class = HeightSerializer  # Assuming the correct serializer class is HeightSerializer

#     def destroy(self, request, *args, **kwargs):
#         height = self.get_object()
#         height.is_deleted = True
#         height.save()
#         return Response(status=status.HTTP_204_NO_CONTENT)

class HeightViewSet(viewsets.ModelViewSet):
    queryset = Height.objects.filter(is_deleted=False)  # Fetch only non-deleted heights
    serializer_class = HeightSerializer

    # def destroy(self, request, args, *kwargs):
    #     height = self.get_object()
    #     height.is_deleted = True  # Perform soft delete
    #     height.save()
    #     return Response(status=status.HTTP_204_NO_CONTENT)
    def destroy(self, request, *args, **kwargs):
        # 'pk' is already part of kwargs, no need to pass it explicitly
        instance = self.get_object()  # Will use the 'pk' from kwargs
        self.perform_destroy(instance)
        return Response(status=status.HTTP_204_NO_CONTENT)

    def get_queryset(self):
        return Height.objects.filter(is_deleted=False) 


class ComplexionViewSet(viewsets.ModelViewSet):
    queryset = Complexion.objects.filter(is_deleted=False)  # Only fetch non-deleted complexions
    serializer_class = ComplexionSerializer

    def destroy(self, request, *args, **kwargs):
        complexion = self.get_object()
        complexion.is_deleted = True
        complexion.save()
        return Response(status=status.HTTP_204_NO_CONTENT)




class ParentsOccupationViewSet(viewsets.ModelViewSet):
    queryset = ParentsOccupation.objects.filter(is_deleted=False)  # Only fetch non-deleted occupations
    serializer_class = ParentsOccupationSerializer

    def destroy(self, request, *args, **kwargs):
        occupation = self.get_object()
        occupation.is_deleted = True
        occupation.save()
        return Response(status=status.HTTP_204_NO_CONTENT)


class HighestEducationViewSet(viewsets.ModelViewSet):
    queryset = HighestEducation.objects.all()
    serializer_class = HighestEducationSerializer

class UgDegreeViewSet(viewsets.ModelViewSet):
    queryset = UgDegree.objects.filter(is_deleted=False)  # Only fetch non-deleted degrees
    serializer_class = UgDegreeSerializer

    def destroy(self, request, *args, **kwargs):
        degree = self.get_object()
        degree.is_deleted = True
        degree.save()
        return Response(status=status.HTTP_204_NO_CONTENT)


class AnnualIncomeViewSet(viewsets.ModelViewSet):
    queryset = AnnualIncome.objects.filter(is_deleted=False)  # Only fetch non-deleted annual incomes
    serializer_class = AnnualIncomeSerializer

    def destroy(self, request, *args, **kwargs):
        annual_income = self.get_object()
        annual_income.is_deleted = True
        annual_income.save()
        return Response(status=status.HTTP_204_NO_CONTENT)

    
class BirthStarViewSet(viewsets.ModelViewSet):
    queryset = BirthStar.objects.exclude(is_deleted=1)
    serializer_class = BirthStarSerializer



class RasiViewSet(viewsets.ModelViewSet):
    queryset = Rasi.objects.filter(is_deleted=False)  # Only fetch non-deleted rasies
    serializer_class = RasiSerializer

    def destroy(self, request, *args, **kwargs):
        rasi = self.get_object()
        rasi.is_deleted = True
        rasi.save()
        return Response(status=status.HTTP_204_NO_CONTENT)



class LagnamViewSet(viewsets.ModelViewSet):
    queryset = Lagnam.objects.filter(is_deleted=False)  # Only fetch non-deleted lagnams
    serializer_class = LagnamSerializer

    def destroy(self, request, *args, **kwargs):
        lagnam = self.get_object()
        lagnam.is_deleted = True
        lagnam.save()
        return Response(status=status.HTTP_204_NO_CONTENT)


class DasaBalanceViewSet(viewsets.ModelViewSet):
    queryset = DasaBalance.objects.all()
    serializer_class = DasaBalanceSerializer

class FamilyTypeViewSet(viewsets.ModelViewSet):
    queryset = FamilyType.objects.filter(is_deleted=False)  # Only fetch non-deleted family types
    serializer_class = FamilyTypeSerializer

    def destroy(self, request, *args, **kwargs):
        family_type = self.get_object()
        family_type.is_deleted = True
        family_type.save()
        return Response(status=status.HTTP_204_NO_CONTENT)



class FamilyStatusViewSet(viewsets.ModelViewSet):
    queryset = FamilyStatus.objects.filter(is_deleted=False)  # Only fetch non-deleted statuses
    serializer_class = FamilyStatusSerializer

    def destroy(self, request, *args, **kwargs):
        family_status = self.get_object()
        family_status.is_deleted = True
        family_status.save()
        return Response(status=status.HTTP_204_NO_CONTENT)


class FamilyValueViewSet(viewsets.ModelViewSet):
    queryset = FamilyValue.objects.filter(is_deleted=False)  # Only fetch non-deleted family values
    serializer_class = FamilyValueSerializer

    def destroy(self, request, *args, **kwargs):
        family_value = self.get_object()
        family_value.is_deleted = True
        family_value.save()
        return Response(status=status.HTTP_204_NO_CONTENT)

# views.py
from rest_framework import viewsets
from .models import LoginDetailsTemp
from .serializers import LoginDetailsTempSerializer

class LoginDetailsTempViewSet(viewsets.ModelViewSet):
    queryset = LoginDetailsTemp.objects.all()
    serializer_class = LoginDetailsTempSerializer

    @action(detail=True, methods=['patch'])
    def approve(self, request, pk=None):
        login_detail = self.get_object()
        login_detail.status = 1
        last_profile = LoginDetailsTemp.objects.filter(ProfileId__regex=r'^vy\d{3}$').order_by('ProfileId').last()
        if last_profile:
            last_serial_number = int(last_profile.ProfileId[2:])
            new_serial_number = last_serial_number + 1
        else:
            new_serial_number = 1
        login_detail.ProfileId = f'vy{new_serial_number:03}'
        login_detail.save()
        return Response({'status': 'accepted', 'ProfileId': login_detail.ProfileId})

    @action(detail=True, methods=['patch'])
    def disapprove(self, request, pk=None):
        login_detail = self.get_object()
        login_detail.status = 0
        login_detail.save()
        return Response({'status': 'disapproved'})
    
from rest_framework import status
from rest_framework.decorators import api_view
from rest_framework.response import Response
from .models import Profile
from .serializers import ProfileSerializer

@api_view(['PUT'])
def Update_AdminComments(request, profile_id):
    try:
        # Fetch the record to update
        instance = LoginDetails.objects.get(ProfileId=profile_id)
    except LoginDetails.DoesNotExist:
        return Response({"error": "Record not found"}, status=status.HTTP_404_NOT_FOUND)
    instance.Admin_comment_date=datetime.today()
    # Deserialize the input data and validate
    serializer = UpdateAdminComments_Serializer(instance, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()

        response_data = {
        "message": "Admin Comments updated successfully",
    }

        return Response(response_data,status=status.HTTP_200_OK)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
def basic_details(request):
    if request.method == 'POST':
        serializer = ProfileSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)  # Correct status code
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)  # Correct status code
from rest_framework import viewsets, status
from rest_framework.response import Response
from .models import LoginDetails
from .serializers import LoginDetailsSerializer
from django.db import transaction

class LoginDetailsViewSet(viewsets.ModelViewSet):
    queryset = LoginDetails.objects.all()
    serializer_class = LoginDetailsSerializer

    def generate_unique_profile_id(self):
        try:
            last_profile = LoginDetails.objects.latest('ContentId')

            if last_profile:
                # Assuming ContentId is an integer or a string that can be converted to an integer
                last_content_id = int(last_profile.ContentId)
                numeric_part = str(last_content_id + 1).zfill(3)
                new_profile_id = f"VY240{numeric_part}"
            else:
                # Handle the case when there is no previous profile
                new_profile_id = "VY240001"
        except LoginDetails.DoesNotExist:
            # Handle the case when there are no records in the table
            new_profile_id = "VY240001"
        
        return new_profile_id
    
    @transaction.atomic
    def perform_create(self, serializer):
        # Check if ProfileId exists in the incoming data
        profile_id = self.request.data.get('ProfileId', None)

        if not profile_id:
            # Generate a new ProfileId if not provided
            profile_id = self.generate_unique_profile_id()
        
        # Save the instance with the generated or provided ProfileId
        serializer.save(ProfileId=profile_id)

    def destroy(self, request, *args, **kwargs):
        return Response(
            {"detail": "Delete operation is not allowed for this resource."},
            status=status.HTTP_405_METHOD_NOT_ALLOWED
        )


class ProfileFamilyDetailsViewSet(viewsets.ModelViewSet):
    queryset = ProfileFamilyDetails.objects.all()
    serializer_class = ProfileFamilyDetailsSerializer

class ProfileEduDetailsViewSet(viewsets.ModelViewSet):
    queryset = ProfileEduDetails.objects.all()
    serializer_class = ProfileEduDetailsSerializer

class ProfilePartnerPrefViewSet(viewsets.ModelViewSet):
    queryset = ProfilePartnerPref.objects.all()
    serializer_class = ProfilePartnerPrefSerializer




# data table server side responses  #
class StandardResultsPaging(PageNumberPagination):
    page_size = 5
    page_size_query_param = 'page_size'
    max_page_size = 100



def dictfetchall(cursor):
    "Return all rows from a cursor as a dict"
    columns = [col[0] for col in cursor.description]
    return [
        dict(zip(columns, row))
        for row in cursor.fetchall()
    ]

EXPORT_HEADER_CONFIG = {
    ("0", ""): [
        ("ProfileId", "Profile ID"),
        ("DateOfJoin", "Date of Registration"),
        ("Profile_name", "Name"),
        ("years", "Age"),
        ("MaritalStatus", "Marital Status"),
        ("Gender", "Gender"),
        ("Profile_city", "City"),
        ("state_name", "State"),
        ("ModeName", "Created By"),
        ("EducationLevel", "Education Details"),
        ("family_status_name", "Family Status"),
        ("income", "Annual Income"),
        ("profession", "Profession"),
        ("username", "Profile Owner"),
        ("Last_login_date", "Last Login"),
    ],

    ("1", ""): [
        ("ProfileId", "Profile ID"),
        ("DateOfJoin", "Date of Registration"),
        ("Profile_name", "Name"),
        ("years", "Age"),
        ("MaritalStatus", "Marital Status"),
        ("Gender", "Gender"),
        ("Profile_city", "City"),
        ("state_name", "State"),
        ("birthstar_name", "Birth Star"),
        ("profile_status", "Profile Status"),
        ("plan_name", "Profile Mode"),
        ("ModeName", "Created By"),
        ("EducationLevel", "Education Details"),
        ("family_status_name", "Family Status"),
        ("income", "Annual Income"),
        ("profession", "Profession"),
        ("username", "Profile Owner"),
        ("Last_login_date", "Last Login"),
    ],

    ("2", ""): [
        ("ProfileId", "Profile ID"),
        ("DateOfJoin", "Date of Registration"),
        ("Profile_name", "Name"),
        ("years", "Age"),
        ("MaritalStatus", "Marital Status"),
        ("Gender", "Gender"),
        ("Profile_city", "City"),
        ("state_name", "State"),
        ("profile_status", "Profile Status"),
        ("plan_name", "Profile Mode"),
        ("ModeName", "Created By"),
        ("EducationLevel", "Education Details"),
        ("family_status_name", "Family Status"),
        ("income", "Annual Income"),
        ("profession", "Profession"),
        ("username", "Profile Owner"),
        ("Last_login_date", "Last Login"),
    ],

    ("1", "paid"): [
        ("ProfileId", "Profile ID"),
        ("DateOfJoin", "Date of Registration"),
        ("Profile_name", "Name"),
        ("years", "Age"),
        ("MaritalStatus", "Marital Status"),
        ("Gender", "Gender"),
        ("Profile_city", "City"),
        ("state_name", "State"),
        ("membership_startdate", "Membership Date"),
        ("profile_status", "Profile Status"),
        ("plan_name", "Profile Mode"),
        ("ModeName", "Created By"),
        ("EducationLevel", "Education Details"),
        ("family_status_name", "Family Status"),
        ("income", "Annual Income"),
        ("profession", "Profession"),
        ("username", "Profile Owner"),
        ("Last_login_date", "Last Login"),
        ("has_horo", "Horoscope"),
        ("has_photo", "Photo"),
    ],

    ("1", "prospect"): [
        ("ProfileId", "Profile ID"),
        ("DateOfJoin", "Date of Registration"),
        ("Profile_name", "Name"),
        ("years", "Age"),
        ("MaritalStatus", "Marital Status"),
        ("Gender", "Gender"),
        ("Profile_city", "City"),
        ("state_name", "State"),
        ("profile_status", "Profile Status"),
        ("plan_name", "Profile Mode"),
        ("ModeName", "Created By"),
        ("EducationLevel", "Education Details"),
        ("family_status_name", "Family Status"),
        ("income", "Annual Income"),
        ("profession", "Profession"),
        ("username", "Profile Owner"),
        ("Last_login_date", "Last Login"),
        ("has_horo", "Horoscope"),
        ("has_photo", "Photo"),
    ],

    ("3", ""): [
        ("ProfileId", "Profile ID"),
        ("DateOfJoin", "Date of Registration"),
        ("Profile_name", "Name"),
        ("years", "Age"),
        ("MaritalStatus", "Marital Status"),
        ("Gender", "Gender"),
        ("Profile_city", "City"),
        ("state_name", "State"),
        ("birthstar_name", "Birth Star"),
        ("ModeName", "Created By"),
        ("plan_name", "Profile Mode"),
        ("profile_status", "Profile Status"),
        ("EducationLevel", "Education Details"),
        ("family_status_name", "Family Status"),
        ("income", "Annual Income"),
        ("profession", "Profession"),
        ("username", "Profile Owner"),
        ("Last_login_date", "Last Login"),
    ],

    ("4", ""): [
        ("ProfileId", "Profile ID"),
        ("DateOfJoin", "Date of Registration"),
        ("Profile_name", "Name"),
        ("years", "Age"),
        ("MaritalStatus", "Marital Status"),
        ("Gender", "Gender"),
        ("Profile_city", "City"),
        ("state_name", "State"),
        ("birthstar_name", "Birth Star"),
        ("ModeName", "Created By"),
        ("plan_name", "Profile Mode"),
        ("profile_status", "Profile Status"),
        ("EducationLevel", "Education Details"),
        ("family_status_name", "Family Status"),
        ("income", "Annual Income"),
        ("profession", "Profession"),
        ("username", "Profile Owner"),
        ("Last_login_date", "Last Login"),
    ],
}


class Newprofile_get(generics.ListAPIView):
    serializer_class = Getnewprofiledata_new
    pagination_class = StandardResultsPaging
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['ProfileId', 'Gender', 'EmailId', 'Profile_dob', 'Profile_city']


    def get_queryset(self):
        search_query = self.request.query_params.get('search', None)
        page_id=self.request.query_params.get('page_name', None)
        plan_ids=self.request.query_params.get('plan_ids', None)
        plan_id_list_new = [] 
        if plan_ids: 
            plan_id_list_new = [int(pid.strip()) for pid in plan_ids.split(',') if pid.strip().isdigit()]
        if page_id is None:
            status_id = 0  # Default status to 0 when page_id is '1'
        else:
            status_id = page_id  # Otherwise, set the status as the page_id
        
        # Base SQL query with JOINs
        # sql = """
        #     SELECT ld.ContentId, ld.ProfileId, ld.Profile_name, ld.Gender, ld.Mobile_no, ld.EmailId, 
        #            ld.Profile_dob, ld.Profile_city, ld.Profile_whatsapp, ld.Profile_alternate_mobile, ld.Plan_id, ld.status, 
        #            ld.DateOfJoin, ld.Last_login_date, ld.Profile_for, ms.MaritalStatus, cm.complexion_desc, s.name AS state_name, 
        #            cy.city_name, c.name AS country_name, d.name AS district_name,
        #            pfd.family_status, ped.highest_education, ped.profession, ped.anual_income, ph.birthstar_name
        #     FROM logindetails ld
        #     LEFT JOIN maritalstatusmaster ms ON ld.Profile_marital_status = ms.StatusId
        #     LEFT JOIN complexionmaster cm ON ld.Profile_complexion = cm.complexion_id
        #     LEFT JOIN masterstate s ON ld.Profile_state = s.id
        #     LEFT JOIN mastercity cy ON ld.Profile_city = cy.id
        #     LEFT JOIN mastercountry c ON ld.Profile_country = c.id
        #     LEFT JOIN masterdistrict d ON ld.Profile_district = d.name
        #     LEFT JOIN profile_familydetails pfd ON ld.ProfileId = pfd.profile_id
        #     LEFT JOIN profile_edudetails ped ON ld.ProfileId = ped.profile_id
        #     LEFT JOIN profile_horoscope ph ON ld.ProfileId = ph.profile_id  
        #     """

        sql = """
            SELECT ld.ContentId, ld.ProfileId, ld.Profile_name, ld.Gender, ld.Mobile_no, ld.EmailId, 
                   ld.Profile_dob,  ld.Profile_whatsapp, ld.Profile_alternate_mobile, ld.Plan_id, ld.status, 
                   ld.DateOfJoin, ld.Last_login_date, ld.Profile_for, ld.membership_startdate, ld.membership_enddate, ms.MaritalStatus, cm.complexion_desc, s.name AS state_name, 
                   ld.Profile_city AS Profile_city, cy.city_name , c.name AS country_name, d.name AS district_name,
                   pfd.family_status, ped.highest_education, ped.anual_income, ph.birthstar_name , mp.profession AS profession ,pl.plan_name ,ld.Owner_id,
                   md.ModeName,ou.username,IF(pi.profile_id IS NULL, 0, 1) AS has_photo,IF(hi.profile_id IS NULL, 0, 1) AS has_horo,
                   dh.deleted_date
            FROM logindetails ld
            LEFT JOIN maritalstatusmaster ms ON ld.Profile_marital_status = ms.StatusId
            LEFT JOIN complexionmaster cm ON ld.Profile_complexion = cm.complexion_id
            LEFT JOIN masterstate s ON ld.Profile_state = s.id
            LEFT JOIN mastercity cy ON ld.Profile_city = cy.id
            LEFT JOIN mastercountry c ON ld.Profile_country = c.id
            LEFT JOIN masterdistrict d ON ld.Profile_district = d.name
            LEFT JOIN profile_familydetails pfd ON ld.ProfileId = pfd.profile_id
            LEFT JOIN profile_edudetails ped ON ld.ProfileId = ped.profile_id
            LEFT JOIN profile_horoscope ph ON ld.ProfileId = ph.profile_id 
            LEFT JOIN plan_master pl ON ld.Plan_id = pl.id 
            LEFT JOIN masterprofession mp ON ped.profession = mp.RowId 
            LEFT JOIN mastermode md ON md.Mode = ld.Profile_for
            LEFT JOIN users ou ON ou.id = ld.Owner_id
            LEFT JOIN (SELECT DISTINCT profile_id FROM profile_images
               WHERE is_deleted=0 AND image<>'' AND image IS NOT NULL) pi
                ON pi.profile_id=ld.ProfileId
                
            LEFT JOIN ( SELECT DISTINCT profile_id FROM profile_horoscope 
            WHERE (horoscope_file IS NOT NULL AND horoscope_file <> '') OR (horoscope_file_admin IS NOT NULL AND horoscope_file_admin <> '') ) hi 
            ON hi.profile_id = ld.ProfileId
            """
        if int(page_id) == 4 or int(page_id) == 3 or int(page_id) == 2 or (int(page_id) == 1 and plan_ids in [None,'']):
            print("Inside page_id 4 block")
            sql += """
            LEFT JOIN (
                SELECT profile_id, MAX(date_time) AS deleted_date
                FROM datahistory
                GROUP BY profile_id
            ) dh ON dh.profile_id = ld.ProfileId
            """
        else:
            sql += """
            LEFT JOIN (
                SELECT NULL AS profile_id, NULL AS deleted_date
            ) dh ON 1 = 0
            """

            

        # Add the search query conditions if provided
        if search_query:
            sql += """
            WHERE (
                ld.ProfileId LIKE %s OR
                ld.temp_profileid LIKE %s OR
                ld.Gender LIKE %s OR
                ld.Mobile_no LIKE %s OR
                ld.EmailId LIKE %s OR
                ms.MaritalStatus LIKE %s OR
                ld.Profile_dob LIKE %s OR
                cm.complexion_desc LIKE %s OR
                ld.Profile_address LIKE %s OR
                ld.Profile_country LIKE %s OR
                s.name LIKE %s OR
                cy.city_name LIKE %s OR
                ld.Profile_pincode LIKE %s
            ) AND ld.status= %s
            """
            search_pattern = f'%{search_query}%'
            params = [search_pattern] * 13 +  [status_id]  # Same pattern for all fields
        else:
            sql += "WHERE ld.status = %s"
            params = [status_id]

        if plan_ids is not None:
            plan_id_list = [pid.strip() for pid in plan_ids.split(',') if pid.strip()]
            if plan_id_list:
                placeholders = ','.join(['%s'] * len(plan_id_list))
                sql += f" AND ld.Plan_id IN ({placeholders})"
                params.extend(plan_id_list)
        if int(page_id) == 4 or int(page_id) == 3 or int(page_id) == 2 or (int(page_id) == 1 and plan_ids in [None,'']):
            sql += " ORDER BY dh.deleted_date DESC"
        elif int(page_id) == 1 and 8 in plan_id_list_new:
            sql += " ORDER BY ld.DateOfJoin DESC"
        elif int(page_id) == 1:
            sql += " ORDER BY ld.membership_startdate DESC"
        elif int(page_id) == 0:
            sql += " ORDER BY ld.DateOfJoin DESC"
        else:
            sql += " ORDER BY ld.ContentId DESC"
        

        with connection.cursor() as cursor:
            cursor.execute(sql, params)
            rows = dictfetchall(cursor)  # Fetch rows as a dictionary

        # Return the rows to the serializer
        return rows

class NewProfileExportAPI(APIView):
    PAID_PLANS = {"1", "2", "3", "11", "12", "13", "14", "15"}

    def get(self, request):
        export_type = request.query_params.get("export")
        page_name = request.query_params.get("page_name", "")
        plan_ids = request.query_params.get("plan_ids", "")
        search_query = request.query_params.get("search")

        if export_type not in ("csv", "xlsx"):
            return Response({"error": "Invalid export type"}, status=400)

        plan_set = set(plan_ids.split(",")) if plan_ids else set()
        export_key = ""

        if page_name == "1":
            if plan_set & self.PAID_PLANS:
                export_key = "paid"
            elif "8" in plan_set:
                export_key = "prospect"

        headers = EXPORT_HEADER_CONFIG.get((page_name, export_key)) \
                  or EXPORT_HEADER_CONFIG.get((page_name, ""))

        rows = self.get_export_queryset(page_name, plan_ids, search_query)

        if export_type == "csv":
            return export_csv_from_dict(rows, headers)

        return export_xlsx_from_dict(rows, headers)

    def get_export_queryset(self, page_id, plan_ids, search_query):
        status_id = int(page_id or 0)
        params = []

        sql = """
        SELECT
            ld.ProfileId,
            ld.Profile_name,
            ld.Gender,
            ld.DateOfJoin,
            ld.Last_login_date,
            ld.membership_startdate,
            ld.Profile_city,
            ms.MaritalStatus,
            s.name AS state_name,
            mps.status_name as profile_status,
            pl.plan_name,
            ld.status,
            ou.username,
            ped.highest_education,
            pfd.family_status,
            ped.anual_income,
            mai.income,
            mp.profession,
            ld.Owner_id,
            ph.birthstar_name,
            mph.ModeName,
            me.EducationLevel,
            mfs.status as family_status_name,
            TIMESTAMPDIFF(
                YEAR,
                STR_TO_DATE(ld.Profile_dob, '%%Y-%%m-%%d'),
                CURDATE()) AS years,
            IF(pi.profile_id IS NULL, 'No', 'Yes') AS has_photo,
            IF(hi.profile_id IS NULL, 'No', 'Yes') AS has_horo
        FROM logindetails ld
        LEFT JOIN maritalstatusmaster ms ON ld.Profile_marital_status = ms.StatusId
        LEFT JOIN mastercity cy ON ld.Profile_city = cy.id
        LEFT JOIN masterstate s ON ld.Profile_state = s.id
        LEFT JOIN plan_master pl ON ld.Plan_id = pl.id
        LEFT JOIN users ou ON ou.id = ld.Owner_id
        LEFT JOIN profile_edudetails ped ON ld.ProfileId = ped.profile_id
        LEFT JOIN profile_familydetails pfd ON ld.ProfileId = pfd.profile_id
        LEFT JOIN masterprofession mp ON ped.profession = mp.RowId
        LEFT JOIN profile_horoscope ph ON ld.ProfileId = ph.profile_id
        LEFT JOIN (
            SELECT DISTINCT profile_id FROM profile_images
            WHERE is_deleted=0 AND image IS NOT NULL AND image <> ''
        ) pi ON pi.profile_id = ld.ProfileId
        LEFT JOIN (
            SELECT DISTINCT profile_id FROM profile_horoscope
            WHERE horoscope_file IS NOT NULL OR horoscope_file_admin IS NOT NULL
        ) hi ON hi.profile_id = ld.ProfileId
        LEFT JOIN masterprofilestatus mps ON ld.status = mps.status_code
        LEFT JOIN mastermode mph ON ld.Profile_for = mph.Mode
        LEFT JOIN mastereducation me ON ped.highest_education = me.RowId
        LEFT JOIN masterfamilystatus mfs ON pfd.family_status = mfs.id
        LEFT JOIN masterannualincome mai ON ped.anual_income = mai.id
        WHERE ld.status = %s
        """

        params.append(status_id)

        if search_query:
            sql += """
            AND (
                ld.ProfileId LIKE %s OR
                ld.Profile_name LIKE %s OR
                ld.EmailId LIKE %s OR
                cy.city_name LIKE %s
            )
            """
            pattern = f"%{search_query}%"
            params.extend([pattern] * 4)

        if plan_ids:
            plan_list = [p for p in plan_ids.split(",") if p]
            placeholders = ",".join(["%s"] * len(plan_list))
            sql += f" AND ld.Plan_id IN ({placeholders})"
            params.extend(plan_list)

        sql += " ORDER BY ld.DateOfJoin DESC"

        with connection.cursor() as cursor:
            cursor.execute(sql, params)
            return dictfetchall(cursor)

def export_csv_from_dict(rows, headers, filename="profiles.csv"):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)

    field_keys = [h[0] for h in headers]
    header_titles = [h[1] for h in headers]

    writer.writerow(header_titles)

    for row in rows:
        writer.writerow([row.get(k, "") for k in field_keys])

    return response



def export_xlsx_from_dict(rows, headers, filename="profiles.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Profiles"

    field_keys = [h[0] for h in headers]
    header_titles = [h[1] for h in headers]

    ws.append(header_titles)

    for row in rows:
        excel_row = []
        for key in field_keys:
            value = row.get(key, "")

            if isinstance(value, datetime):
                value = value.date()
            elif value is None:
                value = ""
            elif isinstance(value, bool):
                value = "Yes" if value else "No"

            excel_row.append(value)

        ws.append(excel_row)

    # Auto width
    for i, title in enumerate(header_titles, 1):
        ws.column_dimensions[get_column_letter(i)].width = min(len(title) + 15, 35)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response

# class Newprofile_get(generics.ListAPIView):
#     queryset = LoginDetails.objects.all()
#     serializer_class = Getnewprofiledata
#     pagination_class = StandardResultsPaging
#     filter_backends = [filters.OrderingFilter]
#     ordering_fields = ['ProfileId', 'Gender', 'EmailId', 'Profile_dob', 'Profile_city']

#     def get_queryset(self):
#         queryset = LoginDetails.objects.all()
#         search_query = self.request.query_params.get('search', None)
#         if search_query:
#             queryset = queryset.filter(
#                 Q(ProfileId__icontains=search_query) |
#                 Q(temp_profileid__icontains=search_query) |
#                 Q(Gender__icontains=search_query) |
#                 Q(Mobile_no__icontains=search_query) |
#                 Q(EmailId__icontains=search_query) |
#                 Q(Profile_marital_status__icontains=search_query) |
#                 Q(Profile_dob__icontains=search_query) |
#                 Q(Profile_complexion__icontains=search_query) |
#                 Q(Profile_address__icontains=search_query) |
#                 Q(Profile_country__icontains=search_query) |
#                 Q(Profile_state__icontains=search_query) |
#                 Q(Profile_city__icontains=search_query) |
#                 Q(Profile_pincode__icontains=search_query)
#             )
#         return queryset
    

# class Get_Profile_data(APIView):

#     def post(self, request):
#             profile_id='VY240013'
            
#             data = Get_profiledata.get_edit_profile(profile_id)
#             # output_serializer = serializers.MatchingStarSerializer(data, many=True)

#             # Construct the response structure
#             response = data

#             return Response(response, status=status.HTTP_200_OK, safe=False)
#         #return Response(input_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class GetProfileDataView(APIView):

    def post(self, request):
        #profile_id = 'VY240013'
        profile_id = request.data.get('profile_id')

        try:
            data = Get_profiledata.get_edit_profile(profile_id)
            # Uncomment and modify the following line if you have a serializer
            # output_serializer = serializers.MatchingStarSerializer(data, many=True)

            # Construct the response structure
            response = data

            return Response(response, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
class ProfileViewSet(viewsets.ModelViewSet):
    queryset = LoginDetails.objects.all()
    serializer_class = Getnewprofiledata

    def retrieve(self, request, *args, **kwargs):
        # print("Retrieving profile with ID:", kwargs.get('pk'))
        return super().retrieve(request, *args, **kwargs)
    
    
from rest_framework import generics, status
from rest_framework.response import Response
from rest_framework.views import APIView
from .models import Profile
from .serializers import ProfileSerializer

class GetProfileDataView(generics.ListAPIView):
    queryset = Profile.objects.all()
    serializer_class = ProfileSerializer

class ProfileDetailView(APIView):

    def get(self, request, pk, format=None):
        try:
            profile = Profile.objects.get(pk=pk)
        except Profile.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)
        
        serializer = ProfileSerializer(profile)
        return Response(serializer.data)
    
    def put(self, request, pk, format=None):
        try:
            profile = Profile.objects.get(pk=pk)
        except Profile.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)
        
        serializer = ProfileSerializer(profile, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    def delete(self, request, pk, format=None):
        try:
            profile = Profile.objects.get(pk=pk)
        except Profile.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)
        
        profile.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)



from rest_framework import generics
from .models import LoginDetailsTemp
from .serializers import LoginDetailsTempSerializer

class LoginDetailsListCreateView(generics.ListCreateAPIView):
    queryset = LoginDetailsTemp.objects.all()
    serializer_class = LoginDetailsTempSerializer

class LoginDetailsDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = LoginDetailsTemp.objects.all()
    serializer_class = LoginDetailsTempSerializer



import logging




logger = logging.getLogger(__name__)

class LoginDetailsDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = LoginDetailsTemp.objects.all()
    serializer_class = LoginDetailsTempSerializer

    def delete(self, request, *args, **kwargs):
        logger.info(f"Delete request received for ID: {kwargs.get('pk')}")
        return super().delete(request, *args, **kwargs)




# class Get_all_profiles(generics.ListAPIView):
#     serializer_class = Getnewprofiledata
#     pagination_class = StandardResultsPaging
#     filter_backends = [filters.OrderingFilter]
#     ordering_fields = ['ProfileId', 'Gender', 'EmailId', 'Profile_dob', 'Profile_city']

#     def get_queryset(self, profile_status=None):
#         search_query = self.request.query_params.get('search', None)
#         query = '''
#             SELECT l.*, pe.*, pf.*, ph.*, pi.*, pp.*
#             FROM logindetails l
#             LEFT JOIN profile_edudetails pe ON pe.profile_id = l.ProfileId
#             LEFT JOIN profile_familydetails pf ON pf.profile_id = l.ProfileId
#             LEFT JOIN profile_horoscope ph ON ph.profile_id = l.ProfileId
#             LEFT JOIN profile_images pi ON pi.profile_id = l.ProfileId
#             LEFT JOIN profile_partner_pref pp ON pp.profile_id = l.ProfileId
#             WHERE l.status=%s
#         '''

#         if search_query:
#             query += '''
#                 AND (
#                     l.ProfileId LIKE %s
#                     OR l.Gender LIKE %s
#                     OR l.Mobile_no LIKE %s
#                     OR l.EmailId LIKE %s
#                     OR l.Profile_marital_status LIKE %s
#                     OR l.Profile_dob LIKE %s
#                     OR l.Profile_complexion LIKE %s
#                     OR l.Profile_address LIKE %s
#                     OR l.Profile_country LIKE %s
#                     OR l.Profile_state LIKE %s
#                     OR l.Profile_city LIKE %s
#                     OR l.Profile_pincode LIKE %s
#                 )
#             '''
#             search_query = f"%{search_query}%"

#         with connection.cursor() as cursor:
#             if search_query:
#                 cursor.execute(query, [profile_status] + [search_query] * 12)
#             else:
#                 cursor.execute(query, [profile_status])
#             columns = [col[0] for col in cursor.description]
#             rows = cursor.fetchall()
#             result = [dict(zip(columns, row)) for row in rows]
        
#         return result

#     def get(self, request, *args, **kwargs):
#         # profile_status = request.data.get('profile_status')

#         # if profile_status is None:
#         #     return Response({"detail": "profile_status is required."}, status=400)
        
#         profile_status=2

#         queryset = self.get_queryset(profile_status)
#         page = self.paginate_queryset(queryset)
#         if page is not None:
#             serializer = self.get_serializer(page, many=True)
#             return self.get_paginated_response(serializer.data)
#         serializer = self.get_serializer(queryset, many=True)
#         return Response(serializer.data)





from rest_framework import generics, filters
from rest_framework.response import Response
from django.db import connection
from django.urls import path
from .serializers import Getnewprofiledata
from .pagination import StandardResultsPaging

class Get_all_profiles(generics.ListAPIView):
    serializer_class = Getnewprofiledata
    pagination_class = StandardResultsPaging
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['ProfileId', 'Gender', 'EmailId', 'Profile_dob', 'Profile_city']

    def get_queryset(self, profile_status=None):
        search_query = self.request.query_params.get('search', None)
        query = '''
            SELECT l.*, pe.*, pf.*, ph.*, pi.*, pp.*
            FROM logindetails l
            LEFT JOIN profile_edudetails pe ON pe.profile_id = l.ProfileId
            LEFT JOIN profile_familydetails pf ON pf.profile_id = l.ProfileId
            LEFT JOIN profile_horoscope ph ON ph.profile_id = l.ProfileId
            LEFT JOIN profile_images pi ON pi.profile_id = l.ProfileId
            LEFT JOIN profile_partner_pref pp ON pp.profile_id = l.ProfileId
            WHERE l.status=%s
        '''

        if search_query:
            query += '''
                AND (
                    l.ProfileId LIKE %s
                    OR l.Gender LIKE %s
                    OR l.Mobile_no LIKE %s
                    OR l.EmailId LIKE %s
                    OR l.Profile_marital_status LIKE %s
                    OR l.Profile_dob LIKE %s
                    OR l.Profile_complexion LIKE %s
                    OR l.Profile_address LIKE %s
                    OR l.Profile_country LIKE %s
                    OR l.Profile_state LIKE %s
                    OR l.Profile_city LIKE %s
                    OR l.Profile_pincode LIKE %s
                )
            '''
            search_query = f"%{search_query}%"

        with connection.cursor() as cursor:
            if search_query:
                cursor.execute(query, [profile_status] + [search_query] * 12)
            else:
                cursor.execute(query, [profile_status])
            columns = [col[0] for col in cursor.description]
            rows = cursor.fetchall()
            result = [dict(zip(columns, row)) for row in rows]
        
        return result

    def get(self, request, *args, **kwargs):
        profile_status = kwargs.get('profile_status', None)
        if profile_status is None:
            return Response({"detail": "profile_status is required."}, status=400)
        
        queryset = self.get_queryset(profile_status)
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
    
#CSM Page
from rest_framework import generics, status
from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser  # Make sure to import these
from rest_framework.views import APIView
from .models import Page
from .models import AdminSettings
from .serializers import AdminSettingsSerializer
from .serializers import PageSerializer, PageListSerializer
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
import os

class PageViewSet(viewsets.ModelViewSet):
    queryset = Page.objects.all()  # Default manager, no filter
    serializer_class = PageSerializer

class PageListViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Page.objects.filter(deleted=False)  # Custom manager filters out deleted pages and pages with status not equal to 'active'
    serializer_class = PageListSerializer

class PageEditView(APIView):
    def put(self, request, pk):
        try:
            page = Page.objects.get(pk=pk, deleted=False)
        except Page.DoesNotExist:
            return Response({'error': 'Page not found'}, status=status.HTTP_404_NOT_FOUND)
        
        serializer = PageSerializer(page, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class PageDeleteView(APIView):
    def delete(self, request, pk):
        try:
            page = Page.objects.get(pk=pk, deleted=False)
            page.deleted = True
            page.save()
            return Response({'status': 'Page deleted successfully'})
        except Page.DoesNotExist:
            return Response({'error': 'Page not found'}, status=status.HTTP_404_NOT_FOUND)
        
class ImageUploadView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request, format=None):
        if 'upload' not in request.FILES:
            return Response({"error": "No file was submitted."}, status=status.HTTP_400_BAD_REQUEST)

        file = request.FILES['upload']
        original_file_name = file.name
        save_path = os.path.join('ckeditor/images', original_file_name)

        file_name = default_storage.save(save_path, ContentFile(file.read()))
        file_url = default_storage.url(file_name)

        #base_url = 'http://103.214.132.20:8000'  
        file_url = file_url

        return Response({"uploaded": True, "url": file_url}, status=status.HTTP_200_OK)
    
from django.db import transaction

#Adminsettings Page
class AdminSettingsView(APIView):
    def get(self, request):
        settings = AdminSettings.objects.first()
        if settings:
            serializer = AdminSettingsSerializer(settings)
            return Response(serializer.data)
        return Response({'error': 'Settings not found'}, status=status.HTTP_404_NOT_FOUND)

class AdminSettingsUpdateView(generics.UpdateAPIView):
    queryset = AdminSettings.objects.all()
    serializer_class = AdminSettingsSerializer

    def get_object(self):
        # Fetch the specific row to update, assume there's only one
        return AdminSettings.objects.first()

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        if not instance:
            return Response({'error': 'Settings not found'}, status=status.HTTP_404_NOT_FOUND)
        
        # If site_name is being updated, handle it carefully to avoid creating a new row
        old_site_name = instance.site_name
        new_site_name = request.data.get('site_name', old_site_name)

        # Check if the site_name is actually changing
        if old_site_name != new_site_name:
            # You need to handle the update carefully to avoid new row creation
            with transaction.atomic():
                # Delete the old instance and create a new one with the updated site_name
                instance.delete()
                instance.site_name = new_site_name
                serializer = self.get_serializer(instance, data=request.data, partial=partial)
                serializer.is_valid(raise_exception=True)
                serializer.save()
        else:
            serializer = self.get_serializer(instance, data=request.data, partial=partial)
            serializer.is_valid(raise_exception=True)
            self.perform_update(serializer)

        return Response(serializer.data)

    def perform_update(self, serializer):
        # Save the updated instance without creating a new row
        serializer.save()


# #Admin users 
# from rest_framework import viewsets
# from .models import AdminUser
# from .serializers import AdminUserSerializer,AdminUserListSerializer


# class AdminUserViewSet(viewsets.ModelViewSet):
#     queryset = AdminUser.objects.all()
#     serializer_class = AdminUserSerializer

# class AdminUserListViewSet(viewsets.ReadOnlyModelViewSet):
#     queryset = AdminUser.objects.filter(deleted=False)
#     serializer_class = AdminUserListSerializer

    
# class AdminEditView(APIView):
#     def put(self, request, pk):
#         try:
#             user = AdminUser.objects.get(pk=pk, deleted=False)
#         except AdminUser.DoesNotExist:
#             return Response({'error': 'AdminUser not found'}, status=status.HTTP_404_NOT_FOUND)
        
#         serializer = AdminUserSerializer(user, data=request.data)
#         if serializer.is_valid():
#             serializer.save()
#             return Response(serializer.data)
#         return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
# class AdminDeleteView(APIView):
#     def delete(self, request, pk):
#         try:
#             user = AdminUser.objects.get(pk=pk, deleted=False)
#             user.deleted = True  # Mark the user as deleted
#             user.save()
#             return Response({'status': 'AdminUser deleted successfully'})
#         except AdminUser.DoesNotExist:
#             return Response({'error': 'AdminUser not found'}, status=status.HTTP_404_NOT_FOUND)


#AdminUser Roles and permissions  
from rest_framework import viewsets
from rest_framework.views import APIView
from rest_framework.response import Response
from .models import AdminUser , Role
from .serializers import AdminUserSerializer,AdminUserListSerializer
from rest_framework.decorators import api_view

@api_view(['POST'])
def add_admin_user(request):
    serializer = AdminUserSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response({
            "message": "Admin user added successfully"
        }, status=status.HTTP_201_CREATED)
    

    formatted_errors = {field: errors[0] for field, errors in serializer.errors.items()}
    return Response(formatted_errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
def list_roles(request):
    roles = Role.objects.all().values('id', 'role_name') 
    
    # print('roles',roles) 
    
    return Response(roles, status=status.HTTP_200_OK)

@api_view(['GET'])
def list_admin_users(request, pk=None):
    queryset = AdminUser.objects.filter(status=False)  
    serializer_class = AdminUserSerializer 

    if pk:
        queryset = queryset.filter(pk=pk)

    serializer = serializer_class(queryset, many=True)
    return Response(serializer.data, status=status.HTTP_200_OK)


@api_view(['PUT'])
def edit_admin_user(request, pk):
    try:
        user = AdminUser.objects.get(pk=pk)
    except AdminUser.DoesNotExist:
        return Response({
            "message": "Admin user not found"
        }, status=status.HTTP_404_NOT_FOUND)
    
    # Check if the user is marked as deleted (status=1)
    if user.status == 1:
        return Response({
            "message": "Cannot edit a deleted admin user"
        }, status=status.HTTP_400_BAD_REQUEST)

    serializer = AdminUserSerializer(user, data=request.data)  # Remove partial=True

    if serializer.is_valid():
        serializer.save()
        return Response({
            "message": "Admin user updated successfully"
        }, status=status.HTTP_200_OK)

    # Directly return the validation errors for required fields
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)



@api_view(['DELETE'])
def delete_admin_user(request, pk):
    try:
        user = AdminUser.objects.get(pk=pk)
    except AdminUser.DoesNotExist:
        return Response({
            "message": "Admin user not found"
        }, status=status.HTTP_404_NOT_FOUND)
    
    user.status = True  
    user.deleted = True # Set status to True to indicate deleted
    user.save()
    return Response({
        "message": "Admin user deleted successfully"
    }, status=status.HTTP_200_OK)

class AdminUserDetailView(APIView):

    def post(self, request):
        username = request.data.get('username')
        password = request.data.get('password')

        if not username or not password:
            return Response({
                "status": "error",
                "message": "Username and password are required"
            }, status=400)

        try:
            user = AdminUser.objects.get(email=username,deleted=False)

            
            if user.password == password:
                role = user.role_id  

                
                permissions = {
                    "view_users": "1" ,
                    "add_users": "1" ,
                    "edit_users": "1" ,
                    "delete_users": "1",
                    "view_orders": "1",
                    "edit_orders": "1",
                    
                }

                permissions_1 = {
                    "search_profile": "1" if role.search_profile else "0",
                    "add_profile": "1" if role.add_profile else "0",
                    "edit_profile_all_fields": "1" if role.edit_profile_all_fields else "0",
                    "edit_profile_admin_comments_and_partner_settings": "1" if role.edit_profile_admin_comments_and_partner_settings else "0",
                    "membership_activation": "1" if role.membership_activation else "0",
                    "new_photo_update": "1" if role.new_photo_update else "0",
                    "edit_horo_photo": "1" if role.edit_horo_photo else "0",
                    "add_users": "1" if role.add_users else "0"
                }

                full_name=user.first_name + user.last_name

                return Response({
                    "status": "success",
                    "message": "Login successful",
                    "data": {
                        "user": {
                            "id": user.id,
                            "username": user.username,
                            "email": user.email,
                            "first_name": user.first_name,
                            "last_name": user.last_name,
                            "full_name":full_name,
                            "password": user.password,
                            "role_id": role.id,
                            "permissions": permissions ,
                            "permissions_1": permissions_1  
                        }
                    }
                })
            else:
                return Response({
                    "status": "error",
                    "message": "Invalid password"
                }, status=401)

        except AdminUser.DoesNotExist:
            return Response({
                "status": "error",
                "message": "User not found"
            }, status=404)



from .models import SuccessStory, Award , Testimonial
from .serializers import SuccessStorySerializer, SuccessStoryListSerializer, AwardSerializer, AwardListSerializer ,TestimonialSerializer,TestimonialListSerializer , HomepageSerializer , Homepage
class SuccessStoryViewSet(viewsets.ModelViewSet):
    queryset = SuccessStory.objects.all()
    serializer_class = SuccessStorySerializer

    def create(self, request, *args, **kwargs):
        owner_id = request.data.get('admin_user_id')

        try:
            owner_id = int(owner_id)
            user = User.objects.get(id=owner_id)
        except Exception:
            user = None
        
        if user:
            role = user.role
            permissions = RolePermission.objects.filter(role=role).select_related('action')
            data = permissions.values('action__code', 'value')
            edit_permission = data.filter(action__code='marriage_photo_upload').first()
            edit = edit_permission['value'] if edit_permission else None
        else:
            edit = None

        if user and edit != 1:
            return Response({
                "status": "error",
                "message": "Permission Error"
            }, status=status.HTTP_403_FORBIDDEN)

        return super().create(request, *args, **kwargs)


class SuccessStoryListViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = SuccessStory.objects.filter(deleted=False)
    serializer_class = SuccessStoryListSerializer

class SuccessStoryEditView(APIView):
    def put(self, request, pk):
        owner_id = request.data.get('admin_user_id')
        try:
            owner_id = int(owner_id)
            user = User.objects.get(id=owner_id)
        except Exception:
            user = None
             
        if user:
            role = user.role
            permissions = RolePermission.objects.filter(role=role).select_related('action')
            data = permissions.values('action__code', 'value')
            edit_permission = data.filter(action__code='marriage_photo_upload').first()
            edit=edit_permission['value'] if edit_permission else None
        else:
            edit =None
          
        if user:  
            if edit ==1:
                pass
            else:
                return Response({
                    "status": "error",
                    "message": "Permission Error"
                }, status=status.HTTP_403_FORBIDDEN)
        try:
            success_story = SuccessStory.objects.get(pk=pk, deleted=False)
        except SuccessStory.DoesNotExist:
            return Response({'error': 'Success Story not found'}, status=status.HTTP_404_NOT_FOUND)
        
        serializer = SuccessStorySerializer(success_story, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class SuccessStoryDeleteView(APIView):
    def delete(self, request, pk):
        try:
            success_story = SuccessStory.objects.get(pk=pk, deleted=False)
            success_story.deleted = True
            success_story.save()
            return Response({'status': 'Success Story deleted successfully'})
        except SuccessStory.DoesNotExist:
            return Response({'error': 'Success Story not found'}, status=status.HTTP_404_NOT_FOUND)


class AwardViewSet(viewsets.ModelViewSet):
    queryset = Award.objects.filter(deleted=False)
    serializer_class = AwardSerializer

class AwardListViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Award.objects.filter(deleted=False)
    serializer_class = AwardListSerializer

class AwardEditView(APIView):
    def put(self, request, pk):
        try:
            award = Award.objects.get(pk=pk, deleted=False)
        except Award.DoesNotExist:
            return Response({'error': 'Award not found'}, status=status.HTTP_404_NOT_FOUND)

        serializer = AwardSerializer(award, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class AwardDeleteView(APIView):
    def delete(self, request, pk):
        try:
            award = Award.objects.get(pk=pk, deleted=False)
            award.deleted = True
            award.save()
            return Response({'status': 'Award deleted successfully'})
        except Award.DoesNotExist:
            return Response({'error': 'Award not found'}, status=status.HTTP_404_NOT_FOUND)


class TestimonialViewSet(viewsets.ModelViewSet):
    queryset = Testimonial.objects.filter(deleted=False)
    serializer_class = TestimonialSerializer

class TestimonialListViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Testimonial.objects.filter(deleted=False)
    serializer_class = TestimonialSerializer  

class VysycommentsListViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = VysAssistcomment.objects.filter()
    serializer_class = VysassistSerializer  


class TestimonialEditView(APIView):
    def put(self, request, pk):
        try:
            testimonial = Testimonial.objects.get(pk=pk, deleted=False)
        except Testimonial.DoesNotExist:
            return Response({'error': 'Testimonial not found'}, status=status.HTTP_404_NOT_FOUND)

        serializer = TestimonialListSerializer(testimonial, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class TestimonialDeleteView(APIView):
    def delete(self, request, pk):
        try:
            testimonial = Testimonial.objects.get(pk=pk, deleted=False)
            testimonial.deleted = True
            testimonial.save()
            return Response({'status': 'Testimonial deleted successfully'}, status=status.HTTP_204_NO_CONTENT)
        except Testimonial.DoesNotExist:
            return Response({'error': 'Testimonial not found'}, status=status.HTTP_404_NOT_FOUND)
        


# class HomepageViewSet(viewsets.ModelViewSet):
#     queryset = Homepage.objects.all()
#     serializer_class = HomepageSerializer

# class HomepageListViewSet(viewsets.ReadOnlyModelViewSet):
#     queryset = Homepage.objects.filter(deleted=False)
#     serializer_class = HomepageSerializer

# class HomepageEditView(APIView):
#     def put(self, request, pk):
#         try:
#             homepage = Homepage.objects.get(pk=pk, deleted=False)
#         except Homepage.DoesNotExist:
#             return Response({'error': 'Homepage not found'}, status=status.HTTP_404_NOT_FOUND)

#         serializer = HomepageSerializer(homepage, data=request.data)
#         if serializer.is_valid():
#             serializer.save()
#             return Response(serializer.data)
#         return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# class HomepageDeleteView(APIView):
#     def delete(self, request, pk):
#         try:
#             homepage = Homepage.objects.get(pk=pk, deleted=False)
#             homepage.deleted = True
#             homepage.save()
#             return Response({'status': 'Homepage deleted successfully'})
#         except Homepage.DoesNotExist:
#             return Response({'error': 'Homepage not found'}, status=status.HTTP_404_NOT_FOUND)


class HomepageListView(APIView):
    def get(self, request):
        # Fetching all homepage entries
        homepages = Homepage.objects.filter(deleted=False)

        if not homepages.exists():
            return JsonResponse({'status': 'error', 'message': 'No homepage entries found.'}, status=status.HTTP_404_NOT_FOUND)

        # Serializing the data
        serializer = HomepageSerializer(homepages, many=True)

        # Return a structured response
        return JsonResponse({
            'status': 'success',
            'message': 'Homepage fetched successfully',
            'data': serializer.data
        }, status=status.HTTP_200_OK)

    def put(self, request):
        # Fetching the homepage entry (assuming there should be only one active homepage entry)
        try:
            homepage = Homepage.objects.get(deleted=False)
        except Homepage.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Homepage entry not found.'}, status=status.HTTP_404_NOT_FOUND)

        # Update the existing entry with the new data
        serializer = HomepageSerializer(homepage, data=request.data, partial=False)
        if serializer.is_valid():
            serializer.save()
            return JsonResponse({
                'status': 'success',
                'message': 'Homepage updated successfully',
                'data': serializer.data
            }, status=status.HTTP_200_OK)
        else:
            return JsonResponse({'status': 'error', 'message': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


# from rest_framework import status
# from rest_framework.response import Response
# from rest_framework.views import APIView
# from rest_framework import serializers
# from django.db import transaction
# from django.core.exceptions import ValidationError

# from .models import (
#     LoginDetails,
#     ProfileFamilyDetails,
#     ProfileEduDetails,
#     ProfileHoroscope,
#     ProfilePartnerPref
# )

# from .serializers import (
#     LoginDetailsSerializer,
#     ProfileFamilyDetailsSerializer,
#     ProfileEduDetailsSerializer,
#     ProfileHoroscopeSerializer,
#     ProfilePartnerPrefSerializer
# )
# class SubmitProfileAPIView(APIView):
#     """
#     This API view will accept data and save it to all 5 models.
#     """

#     @transaction.atomic
#     def post(self, request, *args, **kwargs):
#         # Extract the respective data from the request payload
#         login_data = request.data.get('login_details', {})
#         family_data = request.data.get('family_details', {})
#         edu_data = request.data.get('education_details', {})
#         horoscope_data = request.data.get('horoscope_details', {})
#         partner_pref_data = request.data.get('partner_pref_details', {})

#         # Initialize error tracking
#         errors = {}

#         # Step 1: Validate if mobile_no or email_id already exists
#         Mobile_no = login_data.get('Mobile_no')
#         EmailId = login_data.get('EmailId')

#         if Mobile_no and LoginDetails.objects.filter(Mobile_no=Mobile_no).exists():
#             errors['Mobile_no'] = ['This mobile number is already registered.']
#         if EmailId and LoginDetails.objects.filter(EmailId=EmailId).exists():
#             errors['EmailId'] = ['This email address is already registered.']

#         # If there are any errors in the validation, return early
#         if errors:
#             return Response(errors, status=status.HTTP_400_BAD_REQUEST)

#         # Proceed to save LoginDetails
#         login_serializer = LoginDetailsSerializer(data=login_data)
#         if login_serializer.is_valid():
#             login_detail = login_serializer.save()

#             # Generate the ProfileId based on ContentId
#             content_id = login_detail.ContentId
#             profile_id = f'VY{content_id:04}'  # Zero-pad ContentId to 4 digits

#             # Set and save the ProfileId
#             login_detail.ProfileId = profile_id
#             login_detail.save()
#         else:
#             errors['login_details'] = login_serializer.errors

#         # Check for errors again after trying to save LoginDetails
#         if errors:
#             return Response(errors, status=status.HTTP_400_BAD_REQUEST)

#         # Step 2: Process and save ProfileFamilyDetails
#         if family_data:
#             family_data['profile_id'] = profile_id  # Set the ProfileId
#             family_serializer = ProfileFamilyDetailsSerializer(data=family_data)
#             if family_serializer.is_valid():
#                 family_serializer.save()
#             else:
#                 errors['family_details'] = family_serializer.errors

#         # Step 3: Process and save ProfileEduDetails
#         if edu_data:
#             edu_data['profile_id'] = profile_id  # Set the ProfileId
#             edu_serializer = ProfileEduDetailsSerializer(data=edu_data)
#             if edu_serializer.is_valid():
#                 edu_serializer.save()
#             else:
#                 errors['education_details'] = edu_serializer.errors

#         # Step 4: Process and save ProfileHoroscope
#         if horoscope_data:
#             horoscope_data['profile_id'] = profile_id  # Set the ProfileId
#             horoscope_serializer = ProfileHoroscopeSerializer(data=horoscope_data)
#             if horoscope_serializer.is_valid():
#                 horoscope_serializer.save()
#             else:
#                 errors['horoscope_details'] = horoscope_serializer.errors

#         # Step 5: Process and save ProfilePartnerPref
#         if partner_pref_data:
#             partner_pref_data['profile_id'] = profile_id  # Set the ProfileId
#             partner_pref_serializer = ProfilePartnerPrefSerializer(data=partner_pref_data)
#             if partner_pref_serializer.is_valid():
#                 partner_pref_serializer.save()
#             else:
#                 errors['partner_pref_details'] = partner_pref_serializer.errors

#         # If there are any errors, rollback and return error response
#         if errors:
#             return Response(errors, status=status.HTTP_400_BAD_REQUEST)

#         # Success response
#         return Response({"status": "success", "ProfileId": profile_id}, status=status.HTTP_201_CREATED)


from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework import serializers
from django.db import transaction
from django.core.exceptions import ValidationError

from .models import (
    LoginDetails,
    ProfileFamilyDetails,
    ProfileEduDetails,
    ProfileHoroscope,
    ProfilePartnerPref,
    PlanFeatureLimit
)

from .serializers import (
    LoginDetailsSerializer,
    ProfileFamilyDetailsSerializer,
    ProfileEduDetailsSerializer,
    ProfileHoroscopeSerializer,
    ProfilePartnerPrefSerializer
)


class SubmitProfileAPIView(APIView):
    """
    This API view will accept data and save it to all 5 models.
    """
    parser_classes = [JSONParser, MultiPartParser]


    @transaction.atomic
    def post(self, request):
        # Extract the respective data from the request payload

        owner_id = request.data.get('admin_user_id')
        try:
            owner_id = int(owner_id)
            user = User.objects.get(id=owner_id)
        except Exception:
            user = None
             
        if user:
            role = user.role
            permissions = RolePermission.objects.filter(role=role).select_related('action')
            data = permissions.values('action__code', 'value')
            edit_permission = data.filter(action__code='add_profile').first()
            edit=edit_permission['value'] if edit_permission else None
        else:
            edit =None
           
        if user: 
            if edit ==1:
                pass
            else:
                return Response({
                    "status": "error",
                    "message": "Permission Error"
                }, status=status.HTTP_403_FORBIDDEN)
                
        def parse_json_field(field):
          if isinstance(field, str):
              try:
                  return json.loads(field)
              except json.JSONDecodeError:
                  return {}
          return field
        login_data = parse_json_field(request.data.get('login_details', {}))
        family_data = parse_json_field(request.data.get('family_details', {}))
        edu_data = parse_json_field(request.data.get('education_details', {}))
        horoscope_data = parse_json_field(request.data.get('horoscope_details', {}))
        partner_pref_data = parse_json_field(request.data.get('partner_pref_details', {}))
        suggested_pref_data = parse_json_field(request.data.get('suggested_pref_details', {}))
        owner = request.data.get('owner_id')
        
        horoscope_file = request.FILES.get('horoscope_file')
        if horoscope_file:
              horoscope_data['horoscope_file'] = horoscope_file


        errors = {}

     
        Mobile_no = login_data.get('Mobile_no')
        EmailId = login_data.get('EmailId')

       
        if errors:
            return Response(errors, status=status.HTTP_400_BAD_REQUEST)

        # Step 2: Validate all serializers before saving anything
        login_serializer = LoginDetailsSerializer(data=login_data)
        family_serializer = ProfileFamilyDetailsSerializer(data=family_data) if family_data else None
        edu_serializer = ProfileEduDetailsSerializer(data=edu_data) if edu_data else None
        horoscope_serializer = ProfileHoroscopeSerializer(data=horoscope_data) if horoscope_data else None
        partner_pref_serializer = ProfilePartnerPrefSerializer(data=partner_pref_data) if partner_pref_data else None
        suggested_pref_serializer = ProfileSuggestedPrefSerializer(data=suggested_pref_data) if suggested_pref_data else None

        # Collect validation errors if any
        if not login_serializer.is_valid():
            errors['login_details'] = login_serializer.errors
        if family_serializer and not family_serializer.is_valid():
            errors['family_details'] = family_serializer.errors
        if edu_serializer and not edu_serializer.is_valid():
            errors['education_details'] = edu_serializer.errors
        if horoscope_serializer and not horoscope_serializer.is_valid():
            errors['horoscope_details'] = horoscope_serializer.errors
        if partner_pref_serializer and not partner_pref_serializer.is_valid():
            errors['partner_pref_details'] = partner_pref_serializer.errors
        if suggested_pref_serializer and not suggested_pref_serializer.is_valid():
            errors['suggested_pref_serializer'] = suggested_pref_serializer.errors

        # If any errors exist, return all errors at once
        if errors:
            return Response(errors, status=status.HTTP_400_BAD_REQUEST)

        # Step 3: Save all validated data inside an atomic transaction
        # Save LoginDetails and generate ProfileId
        login_detail = login_serializer.save()
        content_id = login_detail.ContentId
        # numeric_part = f'{content_id:04}'  # Zero-pad ContentId to 4 digits
        numeric_part = content_id
        profile_id = f'VM{numeric_part}' if login_detail.Gender.lower() == 'male' else f'VF{numeric_part}'
        login_detail.ProfileId = profile_id
        login_detail.status = 0 
        login_detail.Plan_id=7
        login_detail.primary_status=0
        login_detail.secondary_status=26
        login_detail.plan_status=7
        login_detail.Profile_for = 8
        login_detail.save()

        profile_idproof_file = request.FILES.get('Profile_idproof')
        if profile_idproof_file:
            login_detail.Profile_idproof = profile_idproof_file

        # Handle Profile_divorceproof file
        profile_divorceproof_file = request.FILES.get('Profile_divorceproof')
        if profile_divorceproof_file:
             login_detail.Profile_divorceproof = profile_divorceproof_file

        # login_detail['status']=0
        # login_detail.save()
        family_data['profile_id'] = profile_id
        edu_data['profile_id'] = profile_id 
        horoscope_data['profile_id'] = profile_id
        partner_pref_data['profile_id'] = profile_id
        suggested_pref_data['profile_id'] = profile_id
        # Save ProfileFamilyDetails
        if family_serializer:
            family_serializer = ProfileFamilyDetailsSerializer(data=family_data)
            if family_serializer.is_valid():
                updated_instance = family_serializer.save()
                uncle_gothram = family_data.get('uncle_gothram')
                if uncle_gothram:
                    updated_instance.madulamn = uncle_gothram
                    updated_instance.save()
            else:
                errors['family_details'] = family_serializer.errors

        # Save ProfileEduDetails
        if edu_serializer:
            edu_serializer = ProfileEduDetailsSerializer(data=edu_data)
            if edu_serializer.is_valid():
                edu_serializer.save()
            else:
                errors['education_details'] = edu_serializer.errors

        # Save ProfileHoroscope
        if horoscope_serializer:
            horoscope_serializer = ProfileHoroscopeSerializer(data=horoscope_data)
            if horoscope_serializer.is_valid():
                horoscope_detail = horoscope_serializer.save()


                rasi_input_text = horoscope_data.get("rasi_kattam")

            if rasi_input_text:  # Check if rasi_kattam exists and has value
            # Run dosham calculation
                mars_dosham, rahu_kethu_dosham = GetMarsRahuKethuDoshamDetails(rasi_input_text)
                
                # Update calculated fields directly on model
                horoscope_detail.calc_chevvai_dhosham = "True" if mars_dosham else "False"
                horoscope_detail.calc_raguketu_dhosham = "True" if rahu_kethu_dosham else "False"

                # Save only calculated fields
                horoscope_detail.save(update_fields=['calc_chevvai_dhosham', 'calc_raguketu_dhosham'])
                
            else:
                errors['horoscope_details'] = horoscope_serializer.errors


    
    
        # Save ProfilePartnerPref
        if partner_pref_serializer:
            partner_pref_serializer = ProfilePartnerPrefSerializer(data=partner_pref_data)
            if partner_pref_serializer.is_valid():
                partner_pref_serializer.save()
            else:
                errors['partner_pref_details'] = partner_pref_serializer.errors
        
        if suggested_pref_serializer:
            suggested_pref_serializer = ProfileSuggestedPrefSerializer(data=suggested_pref_data)
            if suggested_pref_serializer.is_valid():
                suggested_pref_serializer.save()
            else:
                errors['suggested_pref_details'] = suggested_pref_serializer.errors

        plan_features = PlanFeatureLimit.objects.filter(plan_id=7)

        membership_fromdate = date.today()
        membership_todate = membership_fromdate + timedelta(days=365)

        profile_feature_objects = [
                            Profile_PlanFeatureLimit(
                                **{k: v for k, v in model_to_dict(feature).items() if k != 'id'},  # Exclude 'id'
                                profile_id=profile_id,
                                # plan_id=7,
                                membership_fromdate=membership_fromdate,
                                membership_todate=membership_todate,
                                status=1
                            )
                            for feature in plan_features
                        ]

        Profile_PlanFeatureLimit.objects.bulk_create(profile_feature_objects)

        # Step to handle multiple image uploads
        images = request.FILES.getlist('images')  # Get the list of uploaded images
        if len(images) > 10:
            return Response({"error": "You can upload a maximum of 10 images."}, status=status.HTTP_400_BAD_REQUEST)

        for image in images:
            Image_Upload.objects.create(profile_id=profile_id, image=image)

        # Return errors if any exist
        if errors:
            return Response(errors, status=status.HTTP_400_BAD_REQUEST)

        try:
            if owner:
                owner_id =int(owner)
            else:
                owner_id = None
            DataHistory.objects.create(
                                profile_id=profile_id,
                                profile_status=0,
                                owner_id=owner_id
                            )
        except Exception as e:
            pass
        # Success response

        state_id=login_detail.Profile_state

        assigned_user = assign_user_for_state(state_id)

        if assigned_user:
            login_detail.Owner_id = str(assigned_user.id)
            login_detail.save()
        return Response({"status": "success", "ProfileId": profile_id}, status=status.HTTP_201_CREATED)

def parse_membership_date(date_str):
    if not date_str:
        return None
    try:
        # Parse the incoming date or datetime
        dt = parser.isoparse(date_str)
        # If input had only date (time = 00:00:00), set to end of day
        if dt.time() == time(0, 0, 0):
            dt = datetime.combine(dt.date(), time(23, 59, 59))
        # Make timezone aware if naive
        if timezone.is_naive(dt):
            dt = timezone.make_aware(dt)
        return dt
    except Exception as e:
        raise ValueError(f"Invalid date format: {date_str} ({e})")
    

class EditProfileAPIView(APIView):
    """
    This API view will allow users to edit and update their profile details based on ProfileId.
    """

    @transaction.atomic
    def put(self, request, profile_id, *args, **kwargs):
        # Extract the data from the request payload
        login_data = request.data.get('login_details', {})
        family_data = request.data.get('family_details', {})
        edu_data = request.data.get('education_details', {})
        horoscope_data = request.data.get('horoscope_details', {})
        partner_pref_data = request.data.get('partner_pref_details', {})
        suggested_pref_data = request.data.get('suggested_pref_details', {})
        profile_common_data = request.data.get('profile_common_details', {})
        profile_visibility_data=request.data.get('profile_visibility_details', {})


        # print(profile_visibility_data,'123456')

        # Initialize error tracking
        errors = {}

        # Step 1: Retrieve and update LoginDetails based on ProfileId
        try:
            login_detail = LoginDetails.objects.get(ProfileId=profile_id)
        except LoginDetails.DoesNotExist:
            return Response({'error': 'Profile not found.'}, status=status.HTTP_404_NOT_FOUND)
        if login_data:
            login_serializer = LoginEditSerializer(instance=login_detail, data=login_data, partial=True)
            if login_serializer.is_valid():
                login_serializer.save()
            else:
                errors['login_details'] = login_serializer.errors 

        # Step 2: Retrieve and update ProfileFamilyDetails
        if family_data:
            try:
                family_detail = ProfileFamilyDetails.objects.get(profile_id=profile_id)
            except ProfileFamilyDetails.DoesNotExist:
                return Response({'error': 'Family details not found.'}, status=status.HTTP_404_NOT_FOUND)

            family_serializer = ProfileFamilyDetailsSerializer(instance=family_detail, data=family_data, partial=True)
            if family_serializer.is_valid():
                updated_instance = family_serializer.save()
                uncle_gothram = family_data.get('uncle_gothram')
                if uncle_gothram:
                    updated_instance.madulamn = uncle_gothram
                    updated_instance.save()
            else:
                errors['family_details'] = family_serializer.errors

        # Step 3: Retrieve and update ProfileEduDetails
        if edu_data:
            try:
                edu_detail = ProfileEduDetails.objects.get(profile_id=profile_id)
            except ProfileEduDetails.DoesNotExist:
                return Response({'error': 'Education details not found.'}, status=status.HTTP_404_NOT_FOUND)

            edu_serializer = ProfileEduDetailsSerializer(instance=edu_detail, data=edu_data, partial=True)
            if edu_serializer.is_valid():
                edu_serializer.save()
            else:
                errors['education_details'] = edu_serializer.errors

        # Step 4: Retrieve and update ProfileHoroscope
        if horoscope_data:
            # print('1234567890')
            try:
                horoscope_detail = ProfileHoroscope.objects.get(profile_id=profile_id)
            except ProfileHoroscope.DoesNotExist:
                return Response({'error': 'Horoscope details not found.'}, status=status.HTTP_404_NOT_FOUND)
            
            # Get input text
            rasi_input_text = horoscope_data.get("rasi_kattam")
            # print(rasi_input_text,'123456')
            if rasi_input_text:
                # Update input field
                horoscope_detail.rasi_kattam = rasi_input_text
        
                # Run dosham logic
                mars_dosham, rahu_kethu_dosham = GetMarsRahuKethuDoshamDetails(rasi_input_text)
                # print(mars_dosham)
                # print(rahu_kethu_dosham)
                # Save dosham results directly to model fields
                horoscope_detail.calc_chevvai_dhosham = "True" if mars_dosham else "False"
                horoscope_detail.calc_raguketu_dhosham = "True" if rahu_kethu_dosham else "False"
        
            # Update other fields in horoscope_data using serializer (excluding the calculated fields)
            # horoscope_data.pop("calc_chevvai_dhosham", None)
            # horoscope_data.pop("calc_raguketu_dhosham", None)
        
            horoscope_serializer = ProfileHoroscopeSerializer(
                instance=horoscope_detail,
                data=horoscope_data,
                partial=True
            )

            horoscope_serializer = ProfileHoroscopeSerializer(instance=horoscope_detail, data=horoscope_data, partial=True)
            if horoscope_serializer.is_valid():
                horoscope_serializer.save()
            else:
                errors['horoscope_details'] = horoscope_serializer.errors

        # Step 5: Retrieve and update ProfilePartnerPref
        if partner_pref_data:
            try:
                partner_pref_detail = ProfilePartnerPref.objects.get(profile_id=profile_id)
            except ProfilePartnerPref.DoesNotExist:
                return Response({'error': 'Partner preference details not found.'}, status=status.HTTP_404_NOT_FOUND)
            
        #prefered porutham rasi-stat value storing in the database mythili code 25-06-25

                    # Make a proper mutable copy of the input dict
            if isinstance(partner_pref_data, dict):
                partner_pref_payload = partner_pref_data.copy()
            else:
                # If it's a QueryDict (e.g., from request.data), convert to normal dict
                partner_pref_payload = dict(partner_pref_data.lists())
                # flatten single-item lists: {'key': ['value']} -> {'key': 'value'}
                for key in partner_pref_payload:
                    if isinstance(partner_pref_payload[key], list) and len(partner_pref_payload[key]) == 1:
                        partner_pref_payload[key] = partner_pref_payload[key][0]
       
            # Extract and process 'pref_porutham_star'
            pref_star_ids = partner_pref_payload.get('pref_porutham_star')
            if pref_star_ids:
                try:
                    id_list = [int(i.strip()) for i in str(pref_star_ids).split(',') if i.strip().isdigit()]
                    matches = MatchingStarPartner.objects.filter(id__in=id_list)
       
                    star_rasi_pairs = [f"{m.dest_star_id}-{m.dest_rasi_id}" for m in matches]
       
                    # Save both cleaned values
                    partner_pref_payload['pref_porutham_star'] = ",".join(map(str, id_list))
                    partner_pref_payload['pref_porutham_star_rasi'] = ",".join(star_rasi_pairs)
       
                except Exception as e:
                    errors['partner_pref_details'] = {
                        'pref_porutham_star': [f"Invalid input or failed to process star IDs: {str(e)}"]
                    }

            partner_pref_serializer = ProfilePartnerPrefSerializer(instance=partner_pref_detail, data=partner_pref_data, partial=True)
            if partner_pref_serializer.is_valid():
                partner_pref_serializer.save()
            else:
                errors['partner_pref_details'] = partner_pref_serializer.errors
        

        # Step 6: RetriSuggestedeve and update ProfilePartnerPref
        if suggested_pref_data:
            try:
                suggested_pref_detail = ProfileSuggestedPref.objects.get(profile_id=profile_id)
            except ProfileSuggestedPref.DoesNotExist:
                #return Response({'error': 'suggested pref not found.'}, status=status.HTTP_404_NOT_FOUND)
                suggested_pref_detail = ProfileSuggestedPref.objects.create(
                    profile_id=profile_id
                )

            suggested_pref_serializer = ProfileSuggestedPrefSerializer(instance=suggested_pref_detail, data=suggested_pref_data, partial=True)
            if suggested_pref_serializer.is_valid():
                suggested_pref_serializer.save()
            else:
                errors['suggested_pref_details'] = suggested_pref_serializer.errors
         
         

        # Step 7: Retrieve and update ProfileEduDetails
        if profile_visibility_data:
            # print('inside profile visibility')
            try:
                print('update the existing record')
                profvis_detail = ProfileVisibility.objects.get(profile_id=profile_id)
                provis_serializer = ProfileVisibilitySerializer(instance=profvis_detail, data=profile_visibility_data, partial=True)

            except ProfileVisibility.DoesNotExist:
                print('insert the new record')
                # return Response({'error': 'Profile Visibility details not found.'}, status=status.HTTP_404_NOT_FOUND)
                profile_visibility_data['profile_id'] = profile_id
                provis_serializer = ProfileVisibilitySerializer(data=profile_visibility_data)
                
                #Insert if data not exists
        
            if provis_serializer.is_valid():
                provis_serializer.save()
            else:
                errors['profile_visibility_details'] = provis_serializer.errors


        #common data to be update code is below


        if profile_common_data:
            owner = profile_common_data.get("owner_id")
            # print('inside profile common data update',profile_common_data.get("primary_status"))
            # Only include the common data keys that are available in the request
            login_common_data = clean_none_fields({
                "Addon_package": profile_common_data.get("Addon_package"),
                "Notifcation_enabled": profile_common_data.get("Notifcation_enabled"),
                "PaymentExpire": profile_common_data.get("PaymentExpire"),
                "Package_name": profile_common_data.get("Package_name"),
                "status": profile_common_data.get("status"),
                "DateOfJoin": profile_common_data.get("DateOfJoin"),
                "Profile_name": profile_common_data.get("Profile_name"),
                "Gender": profile_common_data.get("Gender"),
                "Mobile_no": profile_common_data.get("Mobile_no"),
                "membership_startdate": parse_membership_date(profile_common_data.get("membership_fromdate")),
                "membership_enddate": parse_membership_date(profile_common_data.get("membership_todate")),
                "Profile_for": profile_common_data.get("Profile_for"),
                "primary_status":profile_common_data.get("status"),
                "secondary_status":profile_common_data.get("primary_status"),
                "plan_status":profile_common_data.get("secondary_status"),
                "Plan_id": str(profile_common_data.get("secondary_status")),
                "Otp_verify":profile_common_data.get("mobile_otp_verify"),
            })
            family_common_data=clean_none_fields({
                "family_status":profile_common_data.get("family_status")
            })
            horos_common_data=clean_none_fields({
                "calc_chevvai_dhosham":profile_common_data.get("calc_chevvai_dhosham"),
                "calc_raguketu_dhosham":profile_common_data.get("calc_raguketu_dhosham"),
                "horoscope_hints":profile_common_data.get("horoscope_hints")
            })
            profileplan_common_data=clean_none_fields({
                "exp_int_lock":profile_common_data.get("exp_int_lock"),
                "express_int_count":profile_common_data.get("exp_int_count"),
                "profile_permision_toview":profile_common_data.get("visit_count"),
                "plan_id":profile_common_data.get("secondary_status"),
                # "membership_fromdate":profile_common_data.get("membership_fromdate"),
                # "membership_todate":profile_common_data.get("membership_todate")
                "membership_fromdate": parse_membership_date(profile_common_data.get("membership_fromdate")),
                "membership_todate": parse_membership_date(profile_common_data.get("membership_todate")),

            })
            
            try:
                if owner:
                    owner_id =int(owner)
                else:
                    owner_id = None
                old_status = getattr(login_detail, 'status', None)
                new_status = profile_common_data.get("status") or old_status

                old_plan_id = getattr(login_detail, 'Plan_id', None)
                new_plan_id = profile_common_data.get("secondary_status") or old_plan_id
                others_id = profile_common_data.get("primary_status")
                try:
                    others = ProfileSubStatus.objects.get(id=others_id)
                except:    
                    others = None
                
                if old_status is not None and int(old_status) != int(new_status) and int(old_plan_id) != int(new_plan_id):
                    try:
                        DataHistory.objects.create(
                            profile_id=profile_id,
                            profile_status=new_status,
                            plan_id=new_plan_id,
                            owner_id = owner_id
                        )
                    except Exception as e:
                        pass
                elif old_status is not None and int(old_status) != int(new_status):
                    try:
                        DataHistory.objects.create(
                            profile_id=profile_id,
                            profile_status=new_status,
                            owner_id = owner_id,
                            others=others
                        )
                    except Exception as e:
                        pass
                    
                elif int(old_plan_id) != int(new_plan_id):
                    try:
                        DataHistory.objects.create(
                            profile_id=profile_id,
                            profile_status=new_status, 
                            plan_id=new_plan_id,
                            owner_id = owner_id
                        )
                    except Exception as e:
                        pass
            except Exception as e:
                pass
            # print('login_common_data', login_common_data)
            # Update Login Details
            login_detail = LoginDetails.objects.get(ProfileId=profile_id)
            login_serializer = LoginEditSerializer(instance=login_detail, data=login_common_data, partial=True)
            if login_serializer.is_valid():
                login_serializer.save()
            else:
                return Response({'error': login_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

            # Update Family Details
            family_detail = ProfileFamilyDetails.objects.get(profile_id=profile_id)
            family_serializer = ProfileFamilyDetailsSerializer(instance=family_detail, data=family_common_data, partial=True)
            if family_serializer.is_valid():
                family_serializer.save()
            else:
                return Response({'error': family_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

            # Update Horoscope Details
            horoscope_detail = ProfileHoroscope.objects.get(profile_id=profile_id)
            horoscope_serializer = ProfileHoroscopeSerializer(instance=horoscope_detail, data=horos_common_data, partial=True)
            if horoscope_serializer.is_valid():
                horoscope_serializer.save()
            else:
                return Response({'error': horoscope_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

            plan_id = profile_common_data.get("secondary_status")
            plan_features = models.PlanFeatureLimit.objects.filter(plan_id=plan_id).values().first()
            

            if plan_features:
                # Remove the 'id' field if present
                plan_features.pop('id', None)
                plan_features.pop('plan_id', None)  # optional, if you don't want to override plan_id

                # Add membership dates
                plan_features.update({
                    'plan_id': plan_id,
                    'membership_fromdate': parse_membership_date(profile_common_data.get("membership_fromdate")),
                    'membership_todate': parse_membership_date(profile_common_data.get("membership_todate")),
                    'status':1
                })
                # print(plan_features,'plan features updated')
                # Update the profile_plan_features row for profile_id
                models.Profile_PlanFeatureLimit.objects.filter(profile_id=profile_id).update(**plan_features)
                # print(pro_plan,'profile plan feature updated')
                
            # # Update profileplan Details
            profileplan_detail = Profile_PlanFeatureLimit.objects.get(profile_id=profile_id,status=1)
            profileplan_serializer = ProfileplanSerializer(instance=profileplan_detail, data=profileplan_common_data, partial=True)
            if profileplan_serializer.is_valid():
                # print('profile plan serializer is valid', profileplan_serializer.validated_data)
                profileplan_serializer.save()
            else:
                return Response({'error': profileplan_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

            addon_package_ids = profile_common_data.get("Addon_package", "")

            if addon_package_ids:
                # Split comma-separated string into list of ints
                addon_package_id_list = [int(pk.strip()) for pk in addon_package_ids.split(",") if pk.strip().isdigit()]

                # Check if ID 1 is in the list
                if 1 in addon_package_id_list:
                    # print("Addon Package ID 1 found. Updating Profile_plan_feature...")

                    # Example: update all rows (or filter if needed)
                    Profile_PlanFeatureLimit.objects.filter(profile_id=profile_id).update(vys_assist=1,vys_assist_count=5)
     
        # If there are any validation errors, return them
        if errors:
            return Response(errors, status=status.HTTP_400_BAD_REQUEST)

        # Success response
        return Response({"status": "success", "message": "Profile updated successfully."}, status=status.HTTP_200_OK)


def clean_none_fields(data_dict):
    return {k: v for k, v in data_dict.items() if v is not None}


# class GetProfileDetailsAPIView(APIView):
#     """
#     This API view will fetch all profile-related details to populate the edit page based on ProfileId.
#     """

#     def get(self, request, profile_id, *args, **kwargs):
#         # Initialize a dictionary to hold the response data
#         response_data = {}

#         # Step 1: Fetch LoginDetails based on ProfileId
#         try:
#             login_detail = LoginDetails.objects.get(ProfileId=profile_id)
#             response_data['login_details'] = LoginDetailsSerializer(login_detail).data
#         except LoginDetails.DoesNotExist:
#             return Response({'error': 'Profile not found.'}, status=status.HTTP_404_NOT_FOUND)

#         # Step 2: Fetch ProfileFamilyDetails
#         try:
#             family_detail = ProfileFamilyDetails.objects.get(profile_id=profile_id)
#             response_data['family_details'] = ProfileFamilyDetailsSerializer(family_detail).data
#         except ProfileFamilyDetails.DoesNotExist:
#             response_data['family_details'] = {}  # Return an empty object if not found

#         # Step 3: Fetch ProfileEduDetails
#         try:
#             edu_detail = ProfileEduDetails.objects.get(profile_id=profile_id)
#             response_data['education_details'] = ProfileEduDetailsSerializer(edu_detail).data
#         except ProfileEduDetails.DoesNotExist:
#             response_data['education_details'] = {}  # Return an empty object if not found

#         # Step 4: Fetch ProfileHoroscope
#         try:
#             horoscope_detail = ProfileHoroscope.objects.get(profile_id=profile_id)
#             response_data['horoscope_details'] = ProfileHoroscopeSerializer(horoscope_detail).data
#         except ProfileHoroscope.DoesNotExist:
#             response_data['horoscope_details'] = {}  # Return an empty object if not found

#         # Step 5: Fetch ProfilePartnerPref
#         try:
#             partner_pref_detail = ProfilePartnerPref.objects.get(profile_id=profile_id)
#             response_data['partner_pref_details'] = ProfilePartnerPrefSerializer(partner_pref_detail).data
#         except ProfilePartnerPref.DoesNotExist:
#             response_data['partner_pref_details'] = {}  # Return an empty object if not found

#         # Return all the gathered data
#         return Response(response_data, status=status.HTTP_200_OK)
def format_time_am_pm(time_str):
    if not time_str:  # Handles None or empty strings
        return "N/A"
    try:
        time_obj = datetime.strptime(str(time_str), "%H:%M:%S")
        return time_obj.strftime("%I:%M %p")  # 12-hour format with AM/PM
    except ValueError:
        return str(time_str)

def calculate_idle_days(last_login_date):
    if last_login_date:
        try:
            # Parse if it's a string
            if isinstance(last_login_date, str):
                last_login_date = datetime.strptime(last_login_date, "%Y-%m-%d %H:%M:%S")

            idle_duration = datetime.now() - last_login_date  # returns timedelta
            return f"{idle_duration.days} day(s)"
        except Exception as e:
            return "N/A"
    return "N/A"

def get_profile_status(status_id, primary_status,sub_status):
    status_list=[]
    # print(status_id,primary_status,sub_status,'status details')
    if status_id or status_id==0:
        try:
            status = ProfileStatus.objects.get(status_code=status_id)
            status_list.append(status.status_name)
        except Exception as e:
            pass
    if primary_status or primary_status==0:
        try:
            prim_status=None
            pri = ProfileSubStatus.objects.filter(status_code=status_id)
            for p in pri:
                if str(p.id) == str(primary_status):
                    prim_status = p
                    break
            if prim_status:
                status_list.append(prim_status.sub_status_name)
        except Exception as e:
            print("Error fetching primary status:", e)
            pass
    if sub_status or sub_status==0:
        try:
            sub_status_obj=None
            sub = PlanDetails.objects.filter(master_substatus=primary_status)
            for s in sub:
                if str(s.id) == str(sub_status):
                    sub_status_obj = s
                    break
            if sub_status_obj:
                status_list.append(sub_status_obj.plan_name)
        except Exception as e:
            print("Error fetching sub status:", e)
            pass
    return " / ".join(status_list) if status_list else "N/A"

def get_owner_name(user_id):
    try:
        if user_id:
            user = User.objects.get(id=user_id)
            return user.username
        else:
            return None
    except Exception as e:
        print(f"Profile_owner:{str(e)}")
        return None

def get_mem_status(profile_id):
    try:
        try:
            login_detail = LoginDetails.objects.get(ProfileId=profile_id)
        except Exception:
            return None
        
        if not login_detail.membership_enddate:
            return None
        print(login_detail.membership_enddate,login_detail.Plan_id)
        if login_detail.membership_enddate.date() < timezone.now().date() and int(login_detail.Plan_id) in [1,2,3]:
            return "Renew"
        return None
    except Exception as e:
        print(f"error:{str(e)}")
        return None
    
def get_profile_relation(user_id):
    try:
        if user_id:
            user = Mode.objects.get(mode=user_id)
            return user.mode_name
        else:
            return None
    except Exception as e:
        print(f"Profile_owner:{str(e)}")
        return None

def get_preferred_gothram(primary_id, fallback_value):
    try:
        if primary_id and str(primary_id).isdigit():
            gothram = Gothram.objects.filter(id=primary_id, is_deleted=False).first()
            if gothram:
                return gothram.gothram_name
            
        return fallback_value if fallback_value not in [None, "", "0", "N/A", "~"] else "N/A"
    except Exception:
        return None
    
def get_others(id,status):
    try:
        others=None
        dh = DataHistory.objects.filter(profile_id=id,profile_status=status).order_by('-date_time').first()
        if status == 4:
            others = dh.delete_others
        elif status == 3:
            others = dh.hide_others
        elif status == 2:
            others = dh.pending_others
        else:
            others = None
        return others
    except Exception:
        return None

def map_eye_wear(value): 
    try:
        if str(value) == "1": 
            return "Yes" 
        elif str(value) == "0": 
            return "No" 
        return str(value)
    except Exception:
        return "No"

class GetProfEditDetailsAPIView(APIView):
    """
    This API view will fetch all profile-related details to populate the edit page based on ProfileId.
    """

    def get(self, request, profile_id, *args, **kwargs):
        # Initialize a dictionary to hold the response data
        response_data = {}

        # Step 1: Fetch LoginDetails based on ProfileId
        try:
            login_detail = LoginDetails.objects.get(ProfileId=profile_id)
            response_data['login_details'] = LoginDetailsSerializer(login_detail).data           
        except LoginDetails.DoesNotExist:
            return Response({'error': 'Profile not found.'}, status=status.HTTP_404_NOT_FOUND)

        # Step 2: Fetch ProfileFamilyDetails
        try:
            family_detail = ProfileFamilyDetails.objects.get(profile_id=profile_id)
            family_data = ProfileFamilyDetailsSerializer(family_detail).data
            family_data['uncle_gothram'] = family_data.get('madulamn')
            family_data['eye_wear'] = map_eye_wear(family_data.get('eye_wear'))
            response_data['family_details']= family_data
        except ProfileFamilyDetails.DoesNotExist:
            response_data['family_details'] = {}  # Return an empty object if not found

        # Step 3: Fetch ProfileEduDetails
        try:
            edu_detail = ProfileEduDetails.objects.get(profile_id=profile_id)
            response_data['education_details'] = ProfileEduDetailsSerializer(edu_detail).data
        except ProfileEduDetails.DoesNotExist:
            response_data['education_details'] = {}  # Return an empty object if not found

        # Step 4: Fetch ProfileHoroscope
        try:
            horoscope_detail = ProfileHoroscope.objects.get(profile_id=profile_id)
            serialized_data = ProfileHoroscopeSerializer(horoscope_detail).data
            raw_time = serialized_data.get('time_of_birth')
            formatted_time = format_time_am_pm(raw_time)

            # Inject the formatted time back into the response
            serialized_data['time_of_birth'] = formatted_time
            response_data['horoscope_details'] = serialized_data
            
        except ProfileHoroscope.DoesNotExist:
            response_data['horoscope_details'] = {}  # Return an empty object if not found

        # Step 5: Fetch ProfilePartnerPref
        try:
            partner_pref_detail = ProfilePartnerPref.objects.get(profile_id=profile_id)
            response_data['partner_pref_details'] = ProfilePartnerPrefSerializer(partner_pref_detail).data
        except ProfilePartnerPref.DoesNotExist:
            response_data['partner_pref_details'] = {}  # Return an empty object if not found

        payment_detail = PlanSubscription.objects.filter(profile_id=profile_id).order_by('-payment_date').first()
        if payment_detail:
            payment_date = payment_detail.payment_date if payment_detail.payment_date else None
            payment_mode = payment_detail.payment_mode if payment_detail.payment_mode else ''
        else:
            payment_date = None
            payment_mode = ''

        # print('payment_detail:', payment_detail)
        # print('payment_date:', payment_date)
        # print('payment_mode:', payment_mode)

        try:
            profile_plan_features = Profile_PlanFeatureLimit.objects.get(profile_id=profile_id)
            # print(profile_plan_features.plan_id,'profile_plan_features')
            if isinstance(profile_plan_features.membership_fromdate, str):
                profile_plan_features.membership_fromdate = datetime.strptime(
                    profile_plan_features.membership_fromdate, "%Y-%m-%d %H:%M:%S"
                )

            if isinstance(profile_plan_features.membership_todate, str):
                profile_plan_features.membership_todate = datetime.strptime(
                    profile_plan_features.membership_todate, "%Y-%m-%d %H:%M:%S"
                )


            response_data['profile_plan_features'] = ProfileplanSerializer(profile_plan_features).data
        except Profile_PlanFeatureLimit.DoesNotExist:

            profile_plan_features = Profile_PlanFeatureLimit.objects.create(
                    profile_id=profile_id,
                    status=1
                ) 
            response_data['profile_plan_features'] = ProfileplanSerializer(profile_plan_features).data
            # response_data['profile_plan_features'] = {}  # Return an empty object if not found



        try:
            plan_details = PlanDetails.objects.get(id=profile_plan_features.plan_id)

            plan_name= plan_details.plan_name
           
        except PlanDetails.DoesNotExist:
            
            plan_name= ""


        result_percen=calculate_points_and_get_empty_fields(profile_id)

        gender=login_detail.Gender

        def dosham_value_formatter(value):
                if isinstance(value, str):
                    return {"0": "Unknown", "1": "Yes", "2": "No","False":"No","True":"Yes"}.get(value, value)
                elif isinstance(value, int):
                    return {0: "Unknown", 1: "Yes", 2: "No"}.get(value, value)
                
                return value
        
        profession_name = safe_get_by_id(Profession, edu_detail.profession, 'profession')
        try:
            if edu_detail:
                if getattr(edu_detail, "degree", None) and edu_detail.degree not in [86,'86']:
                    qualification_name = safe_get_by_id(MasterhighestEducation, edu_detail.degree, 'degeree_name')
                else:
                    qualification_name=edu_detail.other_degree
            else:
                qualification_name=""
        except Exception as e:
            qualification_name=""
        
        
        country_name=safe_get_by_id(Country,edu_detail.work_country,'name')
        state_name=safe_get_by_id(State,edu_detail.work_state,'name')
        district_name=safe_get_by_id(District,edu_detail.work_district,'name')
        city_name = safe_get_by_id(City, edu_detail.work_city, 'city_name')
        

        location_name = next((name for name in [city_name, district_name, state_name, country_name] if name),None)

        # print('12345',profession_name,qualification_name , city_name )
        
        # print("Profession to match:", repr(edu_detail.profession))
        # print("Degree to match:", repr(edu_detail.highest_education))
        # print("City to match:", repr(login_detail.Profile_city))

        about_self = response_data['family_details'].get('about_self')
        if not about_self:  # Checks for None, '', or missing

            profile = {
                "name": login_detail.Profile_name,
                "profession": profession_name,
                "company": getattr(edu_detail, "company_name", ""),
                "designation": getattr(edu_detail, "designation", ""),
                "business": getattr(edu_detail, "business_name", ""),
                "qualification": qualification_name,
                "location": location_name,
                "profile_type": getattr(edu_detail, "profession", None),
                "nature_of_business": getattr(edu_detail, "nature_of_business", "")
            }
            myself = generate_about_myself_summary(profile)
            response_data['family_details']['about_self'] = myself
        

        response_data['profile_common_details']={
                "Addon_package": login_detail.Addon_package,
                "Notifcation_enabled":  login_detail.Notifcation_enabled,
                "PaymentExpire": login_detail.PaymentExpire,
                "Package_name": plan_name, #login_detail.Package_name,
                "status":login_detail.status,
                "DateOfJoin":login_detail.DateOfJoin,
                "ProfileId":login_detail.ProfileId,
                "Profile_name":login_detail.Profile_name,
                "Gender":login_detail.Gender,
                "Mobile_no":login_detail.Mobile_no,
                "Profile_for":login_detail.Profile_for,
                "calc_chevvai_dhosham": dosham_value_formatter(horoscope_detail.calc_chevvai_dhosham),
                "calc_raguketu_dhosham": dosham_value_formatter(horoscope_detail.calc_raguketu_dhosham),
                "horoscope_hints": horoscope_detail.horoscope_hints,
                "family_status":family_detail.family_status,
                "Admin_comments":login_detail.Admin_comments,
                "suya_gothram": get_preferred_gothram(family_detail.suya_gothram_admin,family_detail.suya_gothram),
                "profile_completion":int(result_percen['completion_percentage']),
                "exp_int_lock": getattr(profile_plan_features, "exp_int_lock", None),
                "exp_int_count": getattr(profile_plan_features, "express_int_count", None),
                "visit_count": getattr(profile_plan_features, "profile_permision_toview", None),
                "primary_status":login_detail.primary_status,
                "secondary_status":login_detail.secondary_status,
                "plan_status":login_detail.plan_status,
                "profile_image":Get_profile_image(profile_id,gender,1,0,is_admin=True),
                #"valid_till":getattr(profile_plan_features, "membership_todate", None),
                "valid_till":membership_todate.strftime("%d-%m-%Y") if (membership_todate := getattr(profile_plan_features, "membership_todate", None)) else None,
                "created_date":login_detail.DateOfJoin,
                "idle_days":calculate_idle_days(login_detail.Last_login_date),
                "membership_fromdate":getattr(profile_plan_features, "membership_fromdate", None),
                "membership_todate":getattr(profile_plan_features, "membership_todate", None),
                # "membership_fromdate": format(profile_plan_features.membership_fromdate, '0000-0-0') if profile_plan_features.membership_fromdate else '0000-0-0',
                # "membership_todate": format(profile_plan_features.membership_todate, '0000-0-0') if profile_plan_features.membership_todate else '0000-0-0',
                "age":calculate_age(login_detail.Profile_dob),
                # "payment_date":payment_date,
                "payment_date": payment_date.strftime("%d-%m-%Y") if payment_date else None ,
                "payment_mode":payment_mode,
                "profile_status": get_profile_status(login_detail.status,login_detail.secondary_status,login_detail.Plan_id),
                "add_on_pack_name":", ".join(
    Addonpackages.objects.filter(
        package_id__in=[package_id.strip() for package_id in login_detail.Addon_package.split(",")] if login_detail.Addon_package else []
    ).values_list("name", flat=True)
),
                "mobile_otp_verify":login_detail.Otp_verify,
                "profile_owner_id":login_detail.Owner_id,
                "profile_owner":get_owner_name(login_detail.Owner_id),
                "membership_status":get_mem_status(login_detail.ProfileId),
                "profile_relation":get_profile_relation(login_detail.Profile_for),
                "others":get_others(login_detail.ProfileId,login_detail.status)
                #"myself":myself
                }
    
                
        

        # print('profile_common_details',response_data['profile_common_details'])
        
       
        suggest_profile_details=Get_profiledata_Matching.get_unique_suggested_match_count(gender,profile_id)
        # print('suggest_profile_details:', suggest_profile_details)
        matching_profile_count = (
            Get_profiledata_Matching.get_profile_match_count(gender, profile_id)
            or 0
        ) 

        pstatus = login_detail.plan_status

        if pstatus in (3,17,'3', '17'):
            visibility_count = (
                Get_profiledata_Matching.get_visibility_match_count(gender, profile_id) or 0
            )
        else:
            visibility_count = 0


        suggest_profile_count = suggest_profile_details  # This will not cause an error
        # print('suggest_profile_count:', suggest_profile_count)

        mutual_condition = Q(status=2) & (Q(profile_from=profile_id) | Q(profile_to=profile_id))
        # personal_notes_condition={'status': 1,'profile_id':profile_id}
        # wishlist_condition = {'status': 1,'profile_from':profile_id}
        received_intrests_count = {'status': 1,'profile_to':profile_id}
        sent_intrest_count = {'status': 1,'profile_from':profile_id}
        viewed_profile_count = {'status': 1,'profile_id':profile_id}
        my_vistor_count = {'status': 1,'viewed_profile':profile_id}
        photo_int_count = {'status': 1,'profile_to':profile_id}
        vys_assist_count = {'status': 1,'profile_from':profile_id}

        call_sent_count = {'status': 1,'profile_from':profile_id}
        call_rec_count = {'status': 1,'profile_to':profile_id}

        mutual_int_count = count_records_forQ(Express_interests, mutual_condition)
        # personal_notes_count = count_records(Profile_personal_notes, personal_notes_condition)
        # wishlist_count = count_records(Profile_wishlists, wishlist_condition)
        received_int_count = count_records(Express_interests, received_intrests_count)
        sent_int_count = count_records(Express_interests, sent_intrest_count)
        myvisitor_count = count_records(Profile_visitors, my_vistor_count)
        viewed_prof_count = count_records(Profile_visitors, viewed_profile_count)

        photo_int_count = count_records(Photo_request, photo_int_count)
        vys_prof_count = count_records(Profile_vysassist, vys_assist_count)
        call_sent_act_count = count_records(Profile_callogs, call_sent_count)

        call_recev_act_count = count_records(Profile_callogs, call_rec_count)

        

        
        response_data['profile_matching_counts'] = {
                "matchingprofile_count": matching_profile_count,
                "suggestedprofile_count":suggest_profile_count,
                "viewedprofile_count":viewed_prof_count,
                "visitorprofile_count":myvisitor_count,
                "ctocsend_count":call_sent_act_count,
                "ctocreceived_count":call_recev_act_count,
                "exp_int_sentcount":sent_int_count,
                "exp_int_reccount": received_int_count,
                "mutual_int_count":mutual_int_count,
                "shortlisted_count":0,
                "prsent_count":photo_int_count,
                "varequest_count":vys_prof_count,
                "visibility_count": visibility_count 
        }

        try:
            suggests_pref_detail = ProfileSuggestedPref.objects.get(profile_id=profile_id)
        except ProfileSuggestedPref.DoesNotExist:
            suggests_pref_detail = ProfileSuggestedPref.objects.create(profile_id=profile_id)
        response_data['suggests_pref_details'] = ProfileSuggestedPrefSerializer(suggests_pref_detail).data

        # Profile Visibility
        try:
            profile_visibility_qs = ProfileVisibility.objects.filter(profile_id=profile_id).first()
            if profile_visibility_qs:
                visibility_serializer = ProfileVisibilitySerializer(profile_visibility_qs)
                response_data['profile_visibility'] = visibility_serializer.data
            else:
                response_data['profile_visibility'] = {}
        except Exception as e:
            response_data['profile_visibility'] = {}
            response_data['profile_visibility_error'] = str(e)

        return Response(response_data, status=status.HTTP_200_OK)

# def safe_get_by_id(model, pk_value, return_field):
#     if not pk_value:
#         return ""
#     try:
#         obj = model.objects.get(pk=pk_value)
#         return getattr(obj, return_field)
#     except model.DoesNotExist:
#         return ""
#     except Exception as e:
#         print(f"[ERROR] {e}")
#         return ""


def safe_get_by_id(model, pk_value, return_field):
    """
    Safely fetch a field value from a model using the primary key.
    If pk_value is a name (e.g., "Chennai"), return it as is.
    """

    if not pk_value:
        return ""

    # Convert comma-separated values → take first ID
    if isinstance(pk_value, str) and "," in pk_value:
        pk_value = pk_value.split(",")[0].strip()

    # If pk_value is a pure number → treat as ID
    if isinstance(pk_value, int) or (isinstance(pk_value, str) and pk_value.isdigit()):
        try:
            pk_int = int(pk_value)
            obj = model.objects.filter(pk=pk_int).first()
            if not obj:
                return ""   # ID not found
            return getattr(obj, return_field, "")
        except Exception as e:
            print(f"[ERROR] Failed DB fetch for {model.__name__} with id={pk_value}: {e}")
            return ""

    # Otherwise → pk_value is not an ID, it's a name → return directly
    return pk_value





# def generate_about_myself_summary(profile):
#     name = profile.get("name", "Name")
#     profession = profile.get("profession", "your profession")
#     designation= profile.get("designation", "your designation")
#     business = profile.get("business", "your business")
#     company = profile.get("company", "your company")
#     qualification = profile.get("qualification", "your qualification")
#     institution = profile.get("institution", None)
#     location = profile.get("location", "your location")
#     profile_type = profile.get("profile_type")  # 'employee', 'business', or 'not_working'

#     if profile_type == "1":
#         summary = (
#             f"I am {name}, currently working as a {designation} at {company}. "
#             f"I hold a degree in {qualification}"
#         )
#         #if institution:
#             #summary += f" and have completed my education from {institution}."
#         if location not in (None, "", "null", "NULL"):
#             summary += f" I live in {location}."

#     elif profile_type == "2":
#         summary = (
#             f"I am {name}, a business professional engaged in {business}. "
#             f"I hold a degree in {qualification}"
#         )
#         #if institution:
#             #summary += f" from {institution}."
#         summary += f" I operate my business from {location}."

#     else:  # Not working or student, etc.
#         summary = (
#             f"I am {name}, "
#             f"I have completed my education in {qualification}"
#         )
#         #if institution:
#             #summary += f" from {institution}."
            
#     if location not in (None, "", "null", "NULL"):
#         summary += f" I live in {location}."

#     return summary

# def generate_about_myself_summary(profile):
#     name = profile.get("name", "Name")
#     profession = profile.get("profession", "your profession")
#     business = profile.get("business", "your business")
#     company = profile.get("company", "your company")
#     designation =  profile.get("designation", "your designation")
#     qualification = profile.get("qualification", "your qualification")
#     institution = profile.get("institution", None)
#     location = profile.get("location", "your location")
#     profile_type = profile.get("profile_type")  # 'employee', 'business', or 'not_working'
#     print(location)
#     if profile_type == "1":
#         summary = (
#             f"I am {name}, currently working as a {designation} at {company}. "
#             f"I hold a degree in {qualification}"
#         )
#         #if institution:
#             #summary += f" and have completed my education from {institution}."
#         if location not in (None, "", "null", "NULL"):
#             summary += f" I live in {location}."

#     elif profile_type == "2":
#         summary = (
#             f"I am {name}, a business professional engaged in {business}. "
#             f"I hold a degree in {qualification}"
#         )
#         #if institution:
#             #summary += f" from {institution}."
#         if location not in (None, "", "null", "NULL"):
#             summary += f" I operate my business from {location}."

#     else:  # Not working or student, etc.
#         summary = (
#             f"I am {name}, "
#             f"I have completed my education in {qualification}"
#         )
#         #if institution:
#             #summary += f" from {institution}."
#         # summary += f" I live in {location}."
#     if location not in (None, "", "null", "NULL"):
#         summary += f" I live in {location}."

#     return summary


def generate_about_myself_summary(profile):
    name = (profile.get("name") or "").strip()
    qualification = (profile.get("qualification") or "").strip()
    designation = (profile.get("designation") or "").strip()
    company = (profile.get("company") or "").strip()
    business_name = (profile.get("business") or "").strip()
    nature_of_business = (profile.get("nature_of_business") or "").strip()
    location = (profile.get("location") or "").strip()
    profile_type = str(profile.get("profile_type")) if profile.get("profile_type") else None

    summary_parts = []

    if name:
        summary_parts.append(f"My name is {name}.")

    if qualification:
        summary_parts.append(f"I have completed my {qualification}.")

    if profile_type == "1":
        if designation and company:
            summary_parts.append(f"I am currently employed as {designation} at {company}.")
        elif designation:
            summary_parts.append(f"I am currently employed as {designation}.")
        elif company:
            summary_parts.append(f"I am currently employed at {company}.")
        else:
            summary_parts.append("I am currently employed.")

    elif profile_type == "2": 
        if nature_of_business and business_name:
            summary_parts.append(f"I am managing my own business in {nature_of_business}, named {business_name}.")
        elif nature_of_business:
            summary_parts.append(f"I am managing my own business in {nature_of_business}.")
        elif business_name:
            summary_parts.append(f"I am managing my own business named {business_name}.")
        else:
            summary_parts.append("I am managing my own business.")

    elif profile_type == "6": 
        emp_bus_parts = []

        if designation and company:
            emp_bus_parts.append(f"I am employed as {designation} at {company}")
        elif designation:
            emp_bus_parts.append(f"I am employed as {designation}")
        elif company:
            emp_bus_parts.append(f"I am employed at {company}")
        else:
            emp_bus_parts.append("I am employed")

        if nature_of_business and business_name:
            emp_bus_parts.append(f"and also run a business in {nature_of_business}, named {business_name}.")
        elif nature_of_business:
            emp_bus_parts.append(f"and also run a business in {nature_of_business}.")
        elif business_name:
            emp_bus_parts.append(f"and also run a business named {business_name}.")
        else:
            emp_bus_parts.append("and also run a business.")

        summary_parts.append(" ".join(emp_bus_parts))

    elif profile_type == "7":
        if designation and company:
            summary_parts.append(f"I work in a government job as {designation} at {company}.")
        else:
            summary_parts.append("I work in a government job.")

    elif profile_type == "3": 
        summary_parts.append("I am currently a student.")

    if location:
        summary_parts.append(f"I live in {location}.")
    if not any(summary_parts):
        return "" 

    return " ".join(summary_parts)

def calculate_points_and_get_empty_fields(profile_id):
    total_points = 0
    completed_points = 0
    empty_fields = []  # List to store empty fields

    # Define field weights
    field_weights = {
        'logindetails': {'Profile_idproof': 15},  # ID Proof Upload - 15%
        'profile_images': {'image': 15},  # Photo Upload - 15%
        'profile_horoscope': {'horoscope_file': 15},  # Horoscope Upload - 15%
        'logindetails_additional': {'EmailId': 5},  # Email Verification - 5%
        'profile_familydetails': {'property_worth': 5},  # Property Worth - 5%
        'about_myself': {'about_self': 10},  # About Myself - 10%
        'about_my_family': {'about_family': 10},  # About My Family - 10%
        'profile_edudetails': {'career_plans': 10, 'anual_income': 5},  # Career Plan (10%), Annual Income (5%)
        'profile_videos': {'Video_url': 10},  # Videos - 10%
    }

    # 1. ID Proof Upload
    logindetails = LoginDetails.objects.filter(ProfileId=profile_id).first()
    if logindetails:
        for field, weight in field_weights['logindetails'].items():
            total_points += weight
            if getattr(logindetails, field):
                completed_points += weight
            else:
                #empty_fields.append(field)
                empty_fields.append({'tab': 'Personal_info', 'field': field})

    # 2. Photo Upload
    profile_images = Image_Upload.objects.filter(profile_id=profile_id).first()
    if profile_images:
        for field, weight in field_weights['profile_images'].items():
            total_points += weight
            if getattr(profile_images, field):
                completed_points += weight
            else:
                #empty_fields.append(field)
                empty_fields.append({'tab': 'profile_images', 'field': field})
    else:
        for field, weight in field_weights['profile_images'].items():
            total_points += weight
        empty_fields.append({'tab': 'profile_images', 'field': field})

                
    # 3. Horoscope Upload
    profile_horoscope = ProfileHoroscope.objects.filter(profile_id=profile_id).first()
    if profile_horoscope:
        for field, weight in field_weights['profile_horoscope'].items():
            total_points += weight
            if getattr(profile_horoscope, field):
                completed_points += weight
            else:
                # empty_fields.append(field)
                empty_fields.append({'tab': 'profile_horoscope', 'field': field})


    # 4. Email Verification
    if logindetails:
        for field, weight in field_weights['logindetails_additional'].items():
            total_points += weight
            if getattr(logindetails, field):
                completed_points += weight
            else:
                #empty_fields.append(field)
                empty_fields.append({'tab': 'Personal_info', 'field': field})


    # 5. Property Worth
    profile_familydetails = ProfileFamilyDetails.objects.filter(profile_id=profile_id).first()
    if profile_familydetails:
        for field, weight in field_weights['profile_familydetails'].items():
            total_points += weight
            if getattr(profile_familydetails, field):
                completed_points += weight
            else:
                #empty_fields.append(field)
                empty_fields.append({'tab': 'profile_familydetails', 'field': field})

    # 6. About Myself
    if profile_familydetails:  # Assuming "about_self" is in logindetails
        for field, weight in field_weights['about_myself'].items():
            total_points += weight
            if getattr(profile_familydetails, field):
                completed_points += weight
            else:
                #empty_fields.append(field)
                empty_fields.append({'tab': 'Personal_info', 'field': field})

    # 7. About My Family
    if profile_familydetails:
        for field, weight in field_weights['about_my_family'].items():
            total_points += weight
            if getattr(profile_familydetails, field):
                completed_points += weight
            else:
                #empty_fields.append(field)
                empty_fields.append({'tab': 'profile_familydetails', 'field': field})

    # 8. Career Plan and Annual Income
    profile_edudetails = ProfileEduDetails.objects.filter(profile_id=profile_id).first()
    if profile_edudetails:
        for field, weight in field_weights['profile_edudetails'].items():
            total_points += weight
            if getattr(profile_edudetails, field):
                completed_points += weight
            else:
                #empty_fields.append(field)
                empty_fields.append({'tab': 'profile_edudetails', 'field': field})

    # 9. Videos
    profile_videos = LoginDetails.objects.filter(ProfileId=profile_id).first()
    if profile_videos:
        for field, weight in field_weights['profile_videos'].items():
            total_points += weight
            if getattr(profile_videos, field):
                completed_points += weight
            else:
                #empty_fields.append(field)
                empty_fields.append({'tab': 'Personal_info', 'field': field})

    # Calculate completion percentage
    completion_percentage = (completed_points / total_points) * 100 if total_points else 0

    return {
        'total_points': total_points,
        'completed_points': completed_points,
        'completion_percentage': completion_percentage,
        'empty_fields': empty_fields
    }


# class SubmitProfileAPIView(APIView):
#     """
#     This API view will accept data and save it to all 5 models.
#     """

#     @transaction.atomic
#     def post(self, request, *args, **kwargs):
#         # Extract the respective data from the request payload
#         login_data = request.data.get('login_details', {})
#         family_data = request.data.get('family_details', {})
#         edu_data = request.data.get('education_details', {})
#         horoscope_data = request.data.get('horoscope_details', {})
#         partner_pref_data = request.data.get('partner_pref_details', {})

#         # Initialize error tracking
#         errors = {}

#         # Step 1: Validate if mobile_no or email_id already exists
#         Mobile_no = login_data.get('Mobile_no')
#         EmailId = login_data.get('EmailId')

#         if Mobile_no and LoginDetails.objects.filter(Mobile_no=Mobile_no).exists():
#             errors['Mobile_no'] = ['This mobile number is already registered.']
#         if EmailId and LoginDetails.objects.filter(EmailId=EmailId).exists():
#             errors['EmailId'] = ['This email address is already registered.']

#         # If there are any errors in the validation, return early
#         if errors:
#             return Response(errors, status=status.HTTP_400_BAD_REQUEST)

#         # Proceed to save LoginDetails
#         login_serializer = LoginDetailsSerializer(data=login_data)
#         if login_serializer.is_valid():
#             login_detail = login_serializer.save()

#             # Generate the ProfileId based on ContentId and Gender
#             content_id = login_detail.ContentId
#             numeric_part = f'{content_id:04}'  # Zero-pad ContentId to 4 digits

#             # Gender-based profile ID generation
#             if login_detail.Gender.lower() == 'male':
#                 profile_id = f'VM240{numeric_part}'
#             else:
#                 profile_id = f'VF240{numeric_part}'

#             # Set and save the ProfileId
#             login_detail.ProfileId = profile_id
#             login_detail.save()
#         else:
#             errors['login_details'] = login_serializer.errors

#         # Check for errors again after trying to save LoginDetails
#         if errors:
#             return Response(errors, status=status.HTTP_400_BAD_REQUEST)

#         # Step 2: Process and save ProfileFamilyDetails
#         if family_data:
#             family_data['profile_id'] = profile_id  # Set the ProfileId
#             family_serializer = ProfileFamilyDetailsSerializer(data=family_data)
#             if family_serializer.is_valid():
#                 family_serializer.save()
#             else:
#                 errors['family_details'] = family_serializer.errors

#         # Step 3: Process and save ProfileEduDetails
#         if edu_data:
#             edu_data['profile_id'] = profile_id  # Set the ProfileId
#             edu_serializer = ProfileEduDetailsSerializer(data=edu_data)
#             if edu_serializer.is_valid():
#                 edu_serializer.save()
#             else:
#                 errors['education_details'] = edu_serializer.errors

#         # Step 4: Process and save ProfileHoroscope
#         if horoscope_data:
#             horoscope_data['profile_id'] = profile_id  # Set the ProfileId
#             horoscope_serializer = ProfileHoroscopeSerializer(data=horoscope_data)
#             if horoscope_serializer.is_valid():
#                 horoscope_serializer.save()
#             else:
#                 errors['horoscope_details'] = horoscope_serializer.errors

#         # Step 5: Process and save ProfilePartnerPref
#         if partner_pref_data:
#             partner_pref_data['profile_id'] = profile_id  # Set the ProfileId
#             partner_pref_serializer = ProfilePartnerPrefSerializer(data=partner_pref_data)
#             if partner_pref_serializer.is_valid():
#                 partner_pref_serializer.save()
#             else:
#                 errors['partner_pref_details'] = partner_pref_serializer.errors

#         # If there are any errors, rollback and return error response
#         if errors:
#             return Response(errors, status=status.HTTP_400_BAD_REQUEST)

#         # Success response
#         return Response({"status": "success", "ProfileId": profile_id}, status=status.HTTP_201_CREATED)


# def export_excel(request, profile_id):
#     profile_data = Get_profiledata.get_edit_profile(profile_id)

#     if not profile_data:
#         return HttpResponse("No data found for the provided profile ID.", status=404)

    
#     if len(profile_data) > 1:
#         profile_data = [profile_data[0]]

#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = 'Profile Data'

#     header = profile_data[0].keys()  
#     ws.append(list(header))

#     for row in profile_data:
#         ws.append(list(row.values()))

#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = f'attachment; filename="profile_{profile_id}_data.xlsx"'

#     wb.save(response)

#     return response


def export_excel(request):
    # Fetch only profiles where status is 1
    profile_data = Get_profiledata.get_all_profiles(status=1)

    if not profile_data:
        return HttpResponse("No data found.", status=404)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Profile Data'

    # Extract the header from the first record
    header = profile_data[0].keys()  
    ws.append(list(header))

    # Add all profile data rows
    for row in profile_data:
        ws.append(list(row.values()))

    # Set the response as an Excel file download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="vysya_profiles.xlsx"'

    # Save workbook to the response
    wb.save(response)

    return response


def export_csv(data, filename="call_management_export.csv"):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'


    if not data:
        return response


    writer = csv.DictWriter(response, fieldnames=data[0].keys())
    writer.writeheader()
    writer.writerows(data)


    return response

def export_excel_call(data, filename="call_management_export.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Call Management"


    if data:
        ws.append(list(data[0].keys()))
        for row in data:
            ws.append(list(row.values()))


    response = HttpResponse(
    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

QUICK_UPLOAD_EXPORT_HEADERS = [
    ("ProfileId", "Profile ID"),
    ("Profile_name", "Name"),
    ("Gender", "Gender"),
    ("Mobile_no", "Mobile"),
    ("EmailId", "Email"),
    ("DateOfJoin", "Date Of Join"),
]

def export_csv_queryset(queryset, headers, filename="quick_upload.csv"):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)
    fields = [h[0] for h in headers]
    titles = [h[1] for h in headers]

    writer.writerow(titles)

    for obj in queryset:
        row = []
        for field in fields:
            value = getattr(obj, field, "")
            if isinstance(value, datetime):
                value = value.date()
            row.append(value)
        writer.writerow(row)

    return response

def export_xlsx_queryset(queryset, headers, filename="quick_upload.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quick Upload"

    fields = [h[0] for h in headers]
    titles = [h[1] for h in headers]

    ws.append(titles)

    for obj in queryset:
        excel_row = []
        for field in fields:
            value = getattr(obj, field, "")
            if isinstance(value, datetime):
                value = value.date()
            excel_row.append(value)
        ws.append(excel_row)

    for i, title in enumerate(titles, 1):
        ws.column_dimensions[get_column_letter(i)].width = min(len(title) + 15, 35)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response

class QuickUploadAPIView(generics.ListAPIView):
    serializer_class = QuickUploadSerializer
    pagination_class = StandardResultsPaging

    def get_queryset(self):
        quick_upload_data = LoginDetails.objects.filter(quick_registration='1').order_by('-DateOfJoin')
        return quick_upload_data

    def list(self, request, *args, **kwargs):
        export_type = request.query_params.get("export")

        queryset = self.get_queryset()

        if export_type == "csv":
            return export_csv_queryset(
                queryset,
                QUICK_UPLOAD_EXPORT_HEADERS,
                filename="quick_upload.csv"
            )

        if export_type == "xlsx":
            return export_xlsx_queryset(
                queryset,
                QUICK_UPLOAD_EXPORT_HEADERS,
                filename="quick_upload.xlsx"
            )
        return super().list(request, *args, **kwargs)

EXPRESS_INTEREST_EXPORT_HEADERS = [
    ("profile_from_id", "From Profile ID"),
    ("profile_from_name", "From Name"),
    ("profile_from_mobile", "From Mobile"),
    ("from_state", "From State"),
    ("from_plan", "From Plan"),

    ("profile_to_id", "To Profile ID"),
    ("profile_to_name", "To Name"),
    ("profile_to_mobile", "To Mobile"),
    ("to_state", "To State"),
    ("to_plan", "To Plan"),

    ("to_express_message", "Message"),
    ("req_datetime", "Request Date"),
    ("response_datetime", "Response Date"),
    ("status", "Status"),
]


class ExpressInterestView(APIView):
    pagination_class = StandardResultsPaging
    STATUS_MAP = {
        "0": "Removed",
        "1": "Request Sent",
        "2": "Accepted",
        "3": "Rejected"
    }

    def get(self, request):
        export_type = request.query_params.get("export")
        from_date = request.query_params.get('from_date')
        to_date = request.query_params.get('to_date')
        profile_state = request.query_params.get('profile_state')
        status_param = request.query_params.get('status')
        
        if not from_date or not to_date:
            return Response(
                {"error": "Please provide both from_date and to_date."},
                status=400
            )

        try:
            from_date = datetime.strptime(from_date, '%Y-%m-%d')
            to_date = datetime.strptime(to_date, '%Y-%m-%d')
        except ValueError:
            return Response(
                {"error": "Invalid date format. Use YYYY-MM-DD."},
                status=400
            )

        if not profile_state:
            return Response(
                {"error": "Please provide profile_state."},
                status=400
            )

        profile_state_list = profile_state.split(',')
        
        status_list = None
        if status_param:
            status_list = status_param.split(',')

        express_interests = Express_interests.objects.filter(
            req_datetime__range=(from_date, to_date)
        )
        response_statuses = {"2", "3"}
        if status_list:
            express_interests = express_interests.filter(status__in=status_list)
            if any(s in response_statuses for s in status_list):
                express_interests = express_interests.order_by('-response_datetime')
            else:
                express_interests = express_interests.order_by('-req_datetime')
        else:
            express_interests = express_interests.order_by('-req_datetime')
        
        from_ids = set(express_interests.values_list('profile_from', flat=True))
        to_ids = set(express_interests.values_list('profile_to', flat=True))
        all_profile_ids = from_ids | to_ids

        profiles = LoginDetails.objects.filter(
            ProfileId__in=all_profile_ids
        ).values(
            'ProfileId',
            'Profile_name',
            'Mobile_no',
            'Profile_state',
            'Plan_id'
        )

        profile_map = {p['ProfileId']: p for p in profiles}
        state_ids = set()
        plan_ids = set()

        for p in profiles:
            state_id = safe_int(p.get('Profile_state'))
            plan_id = safe_int(p.get('Plan_id'))
            if p['Profile_state']:
                state_ids.add(state_id)
            if p.get('Plan_id'):
                plan_ids.add(plan_id)
                
        state_map = {
            s.id: s.name
            for s in State.objects.filter(
                id__in=state_ids,
                is_active=True,
                is_deleted=False
            )
        }
        plan_map = {
            p.id: p.plan_name
            for p in PlanDetails.objects.filter(id__in=plan_ids)
        }


        result = []
        for interest in express_interests:
            profile_from = profile_map.get(interest.profile_from)
            profile_to = profile_map.get(interest.profile_to)

            if not profile_from or not profile_to:
                continue

            from_state_id = safe_int(profile_from.get('Profile_state'))
            to_state_id = safe_int(profile_to.get('Profile_state'))
            from_plan_id = safe_int(profile_from.get('Plan_id'))
            to_plan_id = safe_int(profile_to.get('Plan_id'))


            result.append({
                'profile_from_id': profile_from['ProfileId'],
                'profile_from_name': profile_from['Profile_name'],
                'profile_from_mobile': profile_from['Mobile_no'],
                'from_state': state_map.get(int(from_state_id)) if from_state_id else None,
                'from_plan': plan_map.get(int(from_plan_id)) if from_plan_id else None,

                'profile_to_id': profile_to['ProfileId'],
                'profile_to_name': profile_to['Profile_name'],
                'profile_to_mobile': profile_to['Mobile_no'],
                'to_state': state_map.get(int(to_state_id)) if to_state_id else None,
                'to_plan': plan_map.get(int(to_plan_id)) if to_plan_id else None,

                'to_express_message': interest.to_express_message,
                'req_datetime': interest.req_datetime,
                'response_datetime': interest.response_datetime,
                'status': self.STATUS_MAP.get(str(interest.status), "Sent")
            })

        if export_type == "csv":
            return export_csv_from_dict(result, EXPRESS_INTEREST_EXPORT_HEADERS)

        if export_type == "xlsx":
            return export_xlsx_from_dict(result, EXPRESS_INTEREST_EXPORT_HEADERS)

        paginator = self.pagination_class()
        paginated_result = paginator.paginate_queryset(result, request)

        if paginated_result is not None:
            return paginator.get_paginated_response(paginated_result)

        return Response(result, status=200)


def export_login_csv(filename, rows):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    rows = list(rows)
    if not rows:
        return response

    writer = csv.writer(response)

    # headers from dict keys
    headers = rows[0].keys()
    writer.writerow(headers)

    for row in rows:
        writer.writerow(row.values())

    return response

def export_login_xlsx(filename, rows):
    rows = list(rows)

    wb = openpyxl.Workbook()
    ws = wb.active

    if rows:
        ws.append(list(rows[0].keys()))
        for row in rows:
            ws.append(list(row.values()))

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def export_view_csv(filename, headers, rows):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    return response


def export_view_xlsx(filename, headers, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Viewed Profiles"

    ws.append(headers)

    for row in rows:
        ws.append(row)

    # Auto column width
    for i, col in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(len(col) + 2, 18)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


class ViewedProfileByDateRangeView(APIView):
    pagination_class = StandardResultsPaging

    def get(self, request):
        from_date = request.query_params.get('from_date')
        to_date = request.query_params.get('to_date')
        export_type = request.query_params.get('export')  
        mutual_only = request.query_params.get('mutual_only') == '1'

        if not from_date or not to_date:
            return Response({"error": "Please provide both from_date and to_date."}, status=400)

        try:
            from_date = datetime.strptime(from_date, '%Y-%m-%d')
            to_date = datetime.strptime(to_date, '%Y-%m-%d')
        except ValueError:
            return Response({"error": "Invalid date format. Use YYYY-MM-DD."}, status=400)

        visitors = Profile_visitors.objects.filter(
            datetime__range=(from_date, to_date)
        ).order_by('-datetime')

        if not visitors.exists():
            return Response({"message": "No profile visitors found in the given date range."}, status=404)


        viewer_ids = set(visitors.values_list('profile_id', flat=True))
        viewed_ids = set(visitors.values_list('viewed_profile', flat=True))
        all_profile_ids = viewer_ids | viewed_ids

        profiles = LoginDetails.objects.filter(
            ProfileId__in=all_profile_ids
        )

        profile_map = {p.ProfileId: p for p in profiles}


        plan_map = {
            p.id: p.plan_name
            for p in PlanDetails.objects.all()
        }

        mode_map = {
            m.mode: m.mode_name
            for m in Mode.objects.all()
        }
        state_map = {
            s.id: s.name
            for s in State.objects.filter(is_deleted=False)
        }


        view_pairs = set(
            visitors.values_list('profile_id', 'viewed_profile')
        )

        result = []

        for v in visitors:
            viewer = profile_map.get(v.profile_id)
            viewed = profile_map.get(v.viewed_profile)
            is_mutual = (v.viewed_profile, v.profile_id) in view_pairs
            if not viewer or not viewed:
                continue
            
            if mutual_only and not is_mutual:
                continue

            def resolve_plan(plan_id):
                try:
                    return plan_map.get(int(plan_id))
                except (TypeError, ValueError):
                    return None


            def resolve_mode(profile_for):
                try:
                    return mode_map.get(int(profile_for))
                except (TypeError, ValueError):
                    return None


            def resolve_state(profile):
                state = profile.Profile_state
                if isinstance(state, str) and not state.isdigit():
                    return state  # already a name
                try:
                    return state_map.get(int(state))
                except (TypeError, ValueError):
                    return None
            result.append({
                'profile_viewer_contentId': viewer.ContentId,
                'profile_viewer_profileId': viewer.ProfileId,
                'profile_viewer_name': viewer.Profile_name,
                'profile_viewer_dob': viewer.Profile_dob.isoformat() if viewer.Profile_dob else None,
                'profile_viewer_city': viewer.Profile_city if viewer.Profile_city else None,
                'profile_viewer_gender': viewer.Gender,
                'profile_viewer_planid': resolve_plan(viewer.Plan_id),
                'profile_viewer_created_by': resolve_mode(viewer.Profile_for),
                'profile_viewer_state': resolve_state(viewer),

                'viewed_profile_contentId': viewed.ContentId,
                'viewed_profile_profileId': viewed.ProfileId,
                'viewed_profile_name': viewed.Profile_name,
                'viewed_profile_dob': viewed.Profile_dob.isoformat() if viewed.Profile_dob else None,
                'viewed_profile_city': viewed.Profile_city if viewed.Profile_city else None,
                'viewed_profile_gender': viewed.Gender,
                'viewed_profile_planid': resolve_plan(viewed.Plan_id),
                'viewed_profile_created_by': resolve_mode(viewed.Profile_for),
                'viewed_profile_state': resolve_state(viewed),

                'datetime': v.datetime.isoformat(),
                'status': v.status,
                'is_mutual_viewed': is_mutual
            })

        if export_type in ['csv', 'xlsx']:
            headers = [
                "Viewer Profile ID", "Viewer Name", "Viewer Gender", "Viewer DOB",
                "Viewer City", "Viewer State", "Viewer Plan", "Viewer Created By",

                "Viewed Profile ID", "Viewed Name", "Viewed Gender", "Viewed DOB",
                "Viewed City", "Viewed State", "Viewed Plan", "Viewed Created By",

                "Viewed DateTime", "Status", "Mutual Viewed"
            ]

            rows = []
            for r in result:
                rows.append([
                    r['profile_viewer_profileId'],
                    r['profile_viewer_name'],
                    r['profile_viewer_gender'],
                    r['profile_viewer_dob'],
                    r['profile_viewer_city'],
                    r['profile_viewer_state'],
                    r['profile_viewer_planid'],
                    r['profile_viewer_created_by'],

                    r['viewed_profile_profileId'],
                    r['viewed_profile_name'],
                    r['viewed_profile_gender'],
                    r['viewed_profile_dob'],
                    r['viewed_profile_city'],
                    r['viewed_profile_state'],
                    r['viewed_profile_planid'],
                    r['viewed_profile_created_by'],

                    r['datetime'],
                    r['status'],
                    r['is_mutual_viewed']
                ])

            filename = f"viewed_profiles_{from_date.date()}_{to_date.date()}.{export_type}"

            if export_type == 'csv':
                return export_view_csv(filename, headers, rows)

            return export_view_xlsx(filename, headers, rows)

        paginator = self.pagination_class()
        page = paginator.paginate_queryset(result, request)

        if page is not None:
            return paginator.get_paginated_response(page)

        return Response(result, status=200)


class BookmarksView(APIView):
    pagination_class = StandardResultsPaging
    STATUS_MAP = {
        "0": "Bookmark Removed",
        "1": "Bookmarked"
    }
    def get(self, request):
        from_date = request.query_params.get('from_date')
        to_date = request.query_params.get('to_date')
        export_type = request.query_params.get('export')

        if not from_date or not to_date:
            return Response(
                {"error": "Please provide both from_date and to_date."},
                status=400
            )

        try:
            from_date = datetime.strptime(from_date, '%Y-%m-%d')
            to_date = datetime.strptime(to_date, '%Y-%m-%d')
        except ValueError:
            return Response(
                {"error": "Invalid date format. Use YYYY-MM-DD."},
                status=400
            )

        bookmarks = Profile_wishlists.objects.filter(
            marked_datetime__range=(from_date, to_date)
        ).order_by('-marked_datetime')

        if not bookmarks.exists():
            return Response({"message": "No bookmarks found."}, status=404)

        profile_ids = set()
        for b in bookmarks:
            profile_ids.add(b.profile_from)
            profile_ids.add(b.profile_to)

        profiles = LoginDetails.objects.filter(ProfileId__in=profile_ids)
        profile_map = {p.ProfileId: p for p in profiles}


        state_ids = set()
        plan_ids = set()
        profile_status_ids = set()
        
        for p in profiles:

            if p.Profile_state and str(p.Profile_state).isdigit():
                state_ids.add(int(p.Profile_state))
                
            if p.Plan_id and str(p.Plan_id).isdigit():
                plan_ids.add(int(p.Plan_id))
                
            if p.status and str(p.status).isdigit():
                profile_status_ids.add(int(p.status))


        state_map = {
            s.id: s.name
            for s in State.objects.filter(id__in=state_ids, is_deleted=False)
        }
        
        plan_map = {
            p.id: p.plan_name
            for p in PlanDetails.objects.filter(id__in=plan_ids)
        }
        
        profile_status_map = {
            ps.status_code: ps.status_name
            for ps in ProfileStatus.objects.filter(status_code__in=profile_status_ids)
        }


        def resolve_state(profile):
            state = profile.Profile_state
            if isinstance(state, str) and not state.isdigit():
                return state
            return state_map.get(int(state)) if state else None

        result = []

        for b in bookmarks:
            profile_from = profile_map.get(b.profile_from)
            profile_to = profile_map.get(b.profile_to)

            if not profile_from or not profile_to:
                continue

            result.append({
                    'profile_from_id': profile_from.ProfileId,
                    'profile_from_name': profile_from.Profile_name,
                    'profile_from_gender': profile_from.Gender,
                    'profile_from_city': profile_from.Profile_city,
                    'profile_from_state': resolve_state(profile_from),
                    'profile_from_plan': plan_map.get(int(profile_from.Plan_id)) if profile_from.Plan_id else None,
                    'profile_from_status': profile_status_map.get(int(profile_from.status)) if profile_from.status else None,

                    'profile_to_id': profile_to.ProfileId,
                    'profile_to_name': profile_to.Profile_name,
                    'profile_to_gender': profile_to.Gender,
                    'profile_to_city': profile_to.Profile_city,
                    'profile_to_state': resolve_state(profile_to),
                    'profile_to_plan': plan_map.get(int(profile_to.Plan_id)) if profile_to.Plan_id else None,
                    'profile_to_status': profile_status_map.get(int(profile_to.status)) if profile_to.status else None,

                    'marked_datetime': b.marked_datetime.isoformat(),
                    'status': self.STATUS_MAP.get(str(b.status), None)
                })

        if export_type in ['csv', 'xlsx']:
            headers = [
                "From Profile ID", "From Name", "From Gender", "From City",
                "From State", "From Plan", "From Status",

                "To Profile ID", "To Name", "To Gender", "To City",
                "To State", "To Plan", "To Status",

                "Marked DateTime", "Bookmark Status"
            ]

            rows = []
            for r in result:
                rows.append([
                    r['profile_from_id'],
                    r['profile_from_name'],
                    r['profile_from_gender'],
                    r['profile_from_city'],
                    r['profile_from_state'],
                    r['profile_from_plan'],
                    r['profile_from_status'],

                    r['profile_to_id'],
                    r['profile_to_name'],
                    r['profile_to_gender'],
                    r['profile_to_city'],
                    r['profile_to_state'],
                    r['profile_to_plan'],
                    r['profile_to_status'],

                    r['marked_datetime'],
                    r['status']
                ])

            filename = f"bookmarks_{from_date.date()}_{to_date.date()}.{export_type}"

            if export_type == 'csv':
                return export_renew_csv(filename, headers, rows)

            return export_renew_xlsx(filename, headers, rows)

        paginator = self.pagination_class()
        page = paginator.paginate_queryset(result, request)

        if page is not None:
            return paginator.get_paginated_response(page)

        return Response(result, status=200)


class PhotoRequestView(APIView):
    pagination_class = StandardResultsPaging
    STATUS_MAP = {
        "0": "Removed",
        "1": "Request Sent",
        "2": "Accepted",
        "3": "Rejected"
    }

    def get(self, request):
        from_date = request.query_params.get('from_date')
        to_date = request.query_params.get('to_date')
        export_type = request.query_params.get('export')

        if not from_date or not to_date:
            return Response(
                {"error": "Please provide both from_date and to_date."},
                status=400
            )

        try:
            from_date = datetime.strptime(from_date, '%Y-%m-%d')
            to_date = datetime.strptime(to_date, '%Y-%m-%d')
        except ValueError:
            return Response(
                {"error": "Invalid date format. Use YYYY-MM-DD."},
                status=400
            )

        photo_requests = Photo_request.objects.filter(
            req_datetime__range=(from_date, to_date)
        ).order_by('-req_datetime')

        if not photo_requests.exists():
            return Response(
                {"message": "No photo request records found."},
                status=404
            )

        profile_ids = set()
        for r in photo_requests:
            profile_ids.add(r.profile_from)
            profile_ids.add(r.profile_to)

        profiles = LoginDetails.objects.filter(ProfileId__in=profile_ids)
        profile_map = {p.ProfileId: p for p in profiles}

        state_ids = set()
        plan_ids = set()
        profile_status_ids = set()

        for p in profiles:
            if p.Profile_state and str(p.Profile_state).isdigit():
                state_ids.add(int(p.Profile_state))
                
            if p.Plan_id and str(p.Plan_id).isdigit():
                plan_ids.add(int(p.Plan_id))
                
            if p.status and str(p.status).isdigit():
                profile_status_ids.add(int(p.status))

        state_map = {
            s.id: s.name
            for s in State.objects.filter(id__in=state_ids, is_deleted=False)
        }

        plan_map = {
            p.id: p.plan_name
            for p in PlanDetails.objects.filter(id__in=plan_ids)
        }
        
        profile_status_map = {
            ps.status_code: ps.status_name
            for ps in ProfileStatus.objects.filter(status_code__in=profile_status_ids)
        }

        def resolve_state(profile):
            state = profile.Profile_state
            if isinstance(state, str) and not state.isdigit():
                return state
            return state_map.get(int(state)) if state else None

        result = []

        for r in photo_requests:
            profile_from = profile_map.get(r.profile_from)
            profile_to = profile_map.get(r.profile_to)

            if not profile_from or not profile_to:
                continue
            

            result.append({
                'profile_from_id': profile_from.ProfileId,
                'profile_from_name': profile_from.Profile_name,
                'profile_from_mobile': profile_from.Mobile_no,
                'profile_from_gender': profile_from.Gender,
                'profile_from_city': profile_from.Profile_city,
                'profile_from_state': resolve_state(profile_from),
                'profile_from_plan': plan_map.get(safe_int(profile_from.Plan_id)) if profile_from.Plan_id else None,
                'profile_from_status': profile_status_map.get(safe_int(profile_from.status)) if profile_from.status else None,

                'profile_to_id': profile_to.ProfileId,
                'profile_to_name': profile_to.Profile_name,
                'profile_to_mobile': profile_to.Mobile_no,
                'profile_to_gender': profile_to.Gender,
                'profile_to_city': profile_to.Profile_city,
                'profile_to_state': resolve_state(profile_to),
                'profile_to_plan': plan_map.get(safe_int(profile_to.Plan_id)) if profile_to.Plan_id else None,
                'profile_to_status': profile_status_map.get(safe_int(profile_to.status)) if profile_to.status else None,

                'req_datetime': r.req_datetime.isoformat() if r.req_datetime else None,
                'response_datetime': r.response_datetime.isoformat() if r.response_datetime else None,
                'response_message': r.response_message,
                'status': self.STATUS_MAP.get(str(r.status), "Sent")
            })
            
        if export_type == "csv":
            return export_renew_csv("photo_req_profiles", result)

        if export_type == "xlsx":
            return export_renew_xlsx("photo_req_profiles", result)
            
        paginator = self.pagination_class()
        page = paginator.paginate_queryset(result, request)

        if page is not None:
            return paginator.get_paginated_response(page)

        return Response(result, status=200)
    


#Get All profile_images

# class ProfileImages(APIView):
#     pagination_class = StandardResultsPaging

#     def get(self, request):
#         profile_id = request.query_params.get('profile_id')  # Get the profile_id from query params

#         # Fetch images based on whether profile_id is provided
#         if profile_id:
#             images = Image_Upload.objects.filter(profile_id=profile_id)
#         else:
#             images = Image_Upload.objects.all()

#         # Check if images exist
#         if not images.exists():
#             if profile_id:
#                 return Response({"message": "No images found for the provided profile_id."}, status=status.HTTP_404_NOT_FOUND)
#             return Response({"message": "No images found."}, status=status.HTTP_404_NOT_FOUND)

#         # Create a dictionary to group images by profile_id
#         profile_images = {}
#         for image in images:
#             if image.profile_id not in profile_images:
#                 profile_images[image.profile_id] = []

#             profile_images[image.profile_id].append(request.build_absolute_uri(image.image.url))

#         # Convert the dictionary to the desired list format
#         result = []
#         for profile_id, urls in profile_images.items():
#             result.append({
#                 'profile_id': profile_id,
#                 'image_url': urls  # List of image URLs
#             })

#         # Implement pagination
#         paginator = self.pagination_class()
#         paginated_result = paginator.paginate_queryset(result, request)

#         # If there are paginated results, return the paginated response
#         if paginated_result is not None:
#             return paginator.get_paginated_response(paginated_result)

#         # If no pagination is needed, return the full result set
#         return Response(result, status=status.HTTP_200_OK)


class ProfileImages(APIView):
    pagination_class = StandardResultsPaging

    def get(self, request):
        search = request.query_params.get('search')
        profile_id = request.query_params.get('profile_id')
        from_date = request.query_params.get('from_date')
        to_date = request.query_params.get('to_date')
        export_type = request.query_params.get('export')

        login_q = Q()

        if search:
            login_q |= Q(ProfileId__icontains=search)
            login_q |= Q(Profile_name__icontains=search)
            login_q |= Q(Gender__iexact=search)
            login_q |= Q(EmailId__icontains=search)

            try:
                login_q |= Q(Profile_dob=datetime.strptime(search, "%Y-%m-%d").date())
            except ValueError:
                pass

        if profile_id:
            login_q &= Q(ProfileId=profile_id)

        profile_ids_qs = (
                LoginDetails.objects
                .filter(login_q, status__in=[0, 1])
                .values_list('ProfileId', flat=True)
            )

        image_filter = Q(
            profile_id__in=profile_ids_qs,
            image_approved__isnull=True
        ) & (Q(is_deleted=False) | Q(is_deleted__isnull=True)) & ~Q(image='')

        if from_date:
            image_filter &= Q(uploaded_at__date__gte=from_date)
        if to_date:
            image_filter &= Q(uploaded_at__date__lte=to_date)

        latest_profiles_qs = (
            Image_Upload.objects
            .filter(image_filter)
            .values('profile_id')
            .annotate(latest_uploaded_at=Max('uploaded_at'))
            .order_by('-latest_uploaded_at')
        )

        paginator = self.pagination_class()
        page = paginator.paginate_queryset(latest_profiles_qs, request)

        if not page:
            return Response(
                {"message": "No images found."},
                status=status.HTTP_404_NOT_FOUND
            )

        profile_ids = [row['profile_id'] for row in page]
        latest_upload_map = {
            row['profile_id']: row['latest_uploaded_at']
            for row in page
        }


        images = (
            Image_Upload.objects
            .filter(image_filter, profile_id__in=profile_ids)
            .only('profile_id', 'image', 'image_approved', 'is_deleted', 'uploaded_at')
            .order_by('-uploaded_at')
        )


        profile_images = defaultdict(list)
        for img in images:
            profile_images[img.profile_id].append({
                "image_url": request.build_absolute_uri(img.image.url),
                "image_approved": img.image_approved,
                "is_deleted": img.is_deleted,
                "uploaded_at": img.uploaded_at.isoformat()
            })

        login_map = {
            ld.ProfileId: ld
            for ld in LoginDetails.objects
            .filter(ProfileId__in=profile_ids)
            .only(
                'ProfileId', 'Profile_name', 'Mobile_no',
                'Profile_dob', 'Gender', 'EmailId',
                'Profile_whatsapp', 'Plan_id', 'status'
            )
        }

        plan_map = dict(
            PlanDetails.objects.values_list('id', 'plan_name')
        )

        profile_status_map = dict(
            ProfileStatus.objects.values_list('status_code', 'status_name')
        )
        result = []
        for pid in profile_ids:
            ld = login_map.get(pid)
            result.append({
                "profile_id": pid,
                "Profile_name": ld.Profile_name if ld else None,
                "Profile_mobile_no": ld.Mobile_no if ld else None,
                "Profile_dob": ld.Profile_dob.isoformat() if ld and ld.Profile_dob else None,
                "profile_gender": ld.Gender if ld and ld.Gender else None,
                "Profile_email": ld.EmailId if ld else None,
                "profile_whats_app_no": ld.Profile_whatsapp if ld else None,
                "profile_plan": plan_map.get(int(ld.Plan_id)) if ld and ld.Plan_id else None,
                "profile_status": (
                    profile_status_map.get(int(ld.status))
                    if ld.status is not None
                    else None
                ),
                "latest_uploaded_at": latest_upload_map.get(pid).isoformat(),
                "images": profile_images.get(pid, [])
            })
        def remove_keys(data, keys):
            return [
                {k: v for k, v in row.items() if k not in keys}
                for row in data
            ]
        if export_type in ("csv", "xlsx"):
            export_data = remove_keys(result, ["images"])

            if export_type == "csv":
                return export_renew_csv("Approve_image_profiles", export_data)

            if export_type == "xlsx":
                return export_renew_xlsx("Approve_image_profiles", export_data)


        return paginator.get_paginated_response(result)


#Get Profile_imagesbyId with Personal details
#Get Profile_imagesbyId with Personal details
class ProfileImagesView(APIView):

    def get(self, request):
        profile_id = request.query_params.get('profile_id')  # Get the profile_id from query params

        # Fetch profile details from LoginDetails and ProfileHoroscope
        if profile_id:
            try:
                login_details = LoginDetails.objects.get(ProfileId=profile_id)
                horoscope = ProfileHoroscope.objects.get(profile_id=profile_id)
            except LoginDetails.DoesNotExist:
                return Response({"message": "Profile not found for the provided profile_id."}, status=status.HTTP_404_NOT_FOUND)
            except ProfileHoroscope.DoesNotExist:
                return Response({"message": "Horoscope not found for the provided profile_id."}, status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"message": "profile_id query param is required."}, status=status.HTTP_400_BAD_REQUEST)

        # Fetch city and state values from their respective models if stored as IDs
        # For Profile_city: check if it's an ID, fetch name from City model if so
        if login_details.Profile_city.isdigit():
            profile_city_obj = City.objects.filter(id=login_details.Profile_city).first()
            profile_city = profile_city_obj.name if profile_city_obj else "Unknown City"
        else:
            profile_city = login_details.Profile_city  # It's already a name, no need to query

        # For Profile_state: check if it's an ID, fetch name from State model if so
        if login_details.Profile_state.isdigit():
            profile_state_obj = State.objects.filter(id=login_details.Profile_state).first()
            profile_state = profile_state_obj.name if profile_state_obj else "Unknown State"
        else:
            profile_state = login_details.Profile_state  # It's already a name, no need to query

        # Fetch images associated with the profile
        images = Image_Upload.objects.filter(profile_id=profile_id)

        if not images.exists():
            return Response({"message": "No images found for the provided profile_id."}, status=status.HTTP_404_NOT_FOUND)

        # Prepare image URLs
        image_urls = [request.build_absolute_uri(image.image.url) for image in images]

        # Prepare response data
        result = {
            "profile_id": login_details.ProfileId,
            "Profile_name": login_details.Profile_name,
            "Gender": login_details.Gender,
            "Profile_dob": login_details.Profile_dob,
            "Profile_state": profile_state,  # Use the resolved state value
            "Profile_city": profile_city,    # Use the resolved city value
            "Profile_mobile_no": login_details.Profile_mobile_no,
            "horoscope_file": request.build_absolute_uri(horoscope.horoscope_file) if horoscope.horoscope_file else None,
            "image_url": image_urls
        }

        # Return full result set if no pagination is needed
        return Response(result, status=status.HTTP_200_OK)

def get_plan(plan_id):
    if not plan_id:
        return "No Plan"
    plan = PlanDetails.objects.filter(id=plan_id).first()
    return plan.plan_name if plan else "No Plan"

def get_family_status(family_status_id):
    if not family_status_id:
        return "N/A"
    family_status = FamilyStatus.objects.filter(id=family_status_id).first()
    return family_status.status if family_status else "N/A"

def get_annual_income(anual_income_id, actual_income):
    if actual_income and actual_income not in [None,"0", "N/A","~"]:
        return actual_income if actual_income else "N/A"
    income = AnnualIncome.objects.filter(id=anual_income_id).first()
    return income.income if income else (actual_income if actual_income else "N/A")

def get_location(city,state_id,country_id):
    location=[]
    if city:
        if isinstance(city, str) and not city.isdigit():
            # print(city,"city is string")
            location.append(city)
    if state_id:
        # print(state_id,"state_id")
        if isinstance(state_id, str) and not state_id.isdigit():
            location.append(state_id)
        else:
            state = State.objects.filter(id=state_id, is_deleted=False).values_list('name', flat=True).first()
            if state:
                # print(state,"state")
                location.append(state)
    else:
        if country_id:
            if isinstance(country_id, str) and not country_id.isdigit():
                location.append(country_id)
            else:
                country = Country.objects.filter(id=country_id, is_deleted=False).values_list('name', flat=True).first()
                if country:
                    location.append(country)
    return ", ".join(location) if location else "N/A"

def get_dhosham(dhosham_id):
    if not dhosham_id:
        return "N/A"
    if dhosham_id == 1:
        return "Yes"
    elif dhosham_id == 2:
        return "No"
    elif dhosham_id == "1":
        return "Yes"
    elif dhosham_id == "2":
        return "No"
    elif dhosham_id == True:
        return "Yes"
    elif dhosham_id == False:
        return "No"
    else:
        return "N/A"
    
def get_designation_or_nature(des,nature):
    if des and des not in [None,"0", "N/A","~"]:
        return des
    elif nature and nature not in [None,"0", "N/A","~"]:
        return nature
    else:
        return "N/A"
    
def get_company_or_business(des,nature):
    if des and des not in [None,"0", "N/A",""]:
        return des
    elif nature and nature not in [None,"0", "N/A",""]:
        return nature
    else:
        return "N/A"


dhosham_map = {
    1: "Yes",
    2: "No",
    "1": "Yes",
    "2": "No",
    True: "Yes",
    False: "No",
    "True": "Yes",
    "False": "No",
    None: "N/A"
}

def get_dhosham_new(dhosham_id):
    # Use the map and default to "N/A"
    return dhosham_map.get(dhosham_id, "N/A")


# {s.id: s.name for s in State.objects.all()}

# # Use the dictionary keys directly
# states_map = {s["id"]: s["name"] for s in State.objects.filter(is_deleted=False).values("id", "name")}
# countries_map = {c["id"]: c["name"] for c in Country.objects.filter(is_deleted=False).values("id", "name")}



def get_states_map():
    return {
        s["id"]: s["name"]
        for s in State.objects.filter(is_deleted=False).values("id", "name")
    }

def get_countries_map():
    return {
        c["id"]: c["name"]
        for c in Country.objects.filter(is_deleted=False).values("id", "name")
    }



def get_location_new(city, state_id, country_id):

    states_map = get_states_map()
    countries_map = get_countries_map()

    parts = []

    # City
    if city and city not in [None, "0", "N/A", "~"]:
        parts.append(city)

    # State
    if state_id:
        if str(state_id).isdigit():
            state_name = states_map.get(int(state_id))
        else:
            state_name = state_id

        if state_name:
            parts.append(state_name)

    # Country (only if state missing)
    elif country_id:
        if str(country_id).isdigit():
            country_name = countries_map.get(int(country_id))
        else:
            country_name = country_id

        if country_name:
            parts.append(country_name)

    return ", ".join(parts) if parts else "N/A"



def get_designation_or_nature_new(des, nature):
    return des if des and des not in [None, "0", "N/A", "~"] else \
           (nature if nature and nature not in [None, "0", "N/A", "~"] else "N/A")


def get_company_or_business_new(company, business):
    return company if company and company not in [None, "0", "N/A", ""] else \
           (business if business and business not in [None, "0", "N/A", ""] else "N/A")

def get_action_log(profile_id,to_id):
    try:
        logs = AdminPrintLogs.objects.filter(profile_id=profile_id, sentprofile_id=to_id)
        if not logs.exists():
            return "No logs"

        formatted_logs = ",".join(f"{log.action_type}({log.format_type})" for log in logs)
        return formatted_logs
    except Exception as e:
        return "No logs"

def get_bulk_action_logs(from_profile_id, to_profile_ids):
    format_short_map = {
        "match_full_profile": "fp",
        "match_full_profile_black": "fp",
        "match_compatability_color": "mch",
        "match_compatability_black": "mch",
        "match_compatability_without_horo": "mc",
        "match_compatability_without_horo_black": "mc"
    }
    action_short_map = {
        "print":"prnt",
        "whatsapp":"whts",
        "email":"email"
    }

    logs = AdminPrintLogs.objects.filter(
        profile_id=from_profile_id,
        sentprofile_id__in=to_profile_ids
    )

    logs_map = {}
    for log in logs:
        pid = log.sentprofile_id
        if pid not in logs_map:
            logs_map[pid] = []

        short_format = format_short_map.get(log.format_type, log.format_type)
        short_action = action_short_map.get(log.action_type.lower(), log.action_type)
        updated_at_str = log.updated_at.strftime("%Y-%m-%d")
        logs_map[pid].append(f"{short_action}-{short_format}({updated_at_str})")

    return {pid: ", ".join(actions) for pid, actions in logs_map.items()}

class Get_prof_list_match(APIView):

    async_capable = False
    def get_action_scores_bulk(self, profile_from, profile_ids):
            """
            Get action scores for multiple profiles at once.
            Returns a dict: { profile_to: {score: int, actions: list} }
            """
            scores = {pid: {"score": 0, "actions": []} for pid in profile_ids}

            # Express Interests (Sent / Received / Accepted / Rejected)
            interests = Express_interests.objects.filter(
                Q(profile_from=profile_from, profile_to__in=profile_ids) |
                Q(profile_to=profile_from, profile_from__in=profile_ids),
                status__in=[1, 2, 3]
            )

            for ei in interests:
                if ei.profile_from == profile_from:
                    target = ei.profile_to
                    if ei.status == 1:
                        scores[target]["score"] += 1
                        scores[target]["actions"].append({"action": "Express Interest Sent", "datetime": ei.req_datetime})
                    elif ei.status == 2:
                        scores[target]["score"] += 1
                        scores[target]["actions"].append({"action": "Express Interest Accepted", "datetime": ei.req_datetime})
                    elif ei.status == 3:
                        scores[target]["score"] += 1
                        scores[target]["actions"].append({"action": "Express Interest Rejected", "datetime": ei.req_datetime})
                else:
                    target = ei.profile_from
                    if ei.status == 1:
                        scores[target]["score"] += 1
                        scores[target]["actions"].append({"action": "Express Interest Received", "datetime": ei.req_datetime})

            # Wishlists
            wishlists = Profile_wishlists.objects.filter(
                Q(profile_from=profile_from, profile_to__in=profile_ids, status=1) |
                Q(profile_to=profile_from, profile_from__in=profile_ids, status=1)
            )
            for wl in wishlists:
                if wl.profile_from == profile_from:
                    scores[wl.profile_to]["score"] += 1
                    scores[wl.profile_to]["actions"].append({"action": "Bookmarked", "datetime": wl.marked_datetime})
                else:
                    scores[wl.profile_from]["score"] += 1
                    scores[wl.profile_from]["actions"].append({"action": "Bookmark Received", "datetime": wl.marked_datetime})

            # Photo Requests
            photo_requests = Photo_request.objects.filter(
                Q(profile_from=profile_from, profile_to__in=profile_ids, status=1) |
                Q(profile_to=profile_from, profile_from__in=profile_ids, status=1)
            )
            for pr in photo_requests:
                if pr.profile_from == profile_from:
                    scores[pr.profile_to]["score"] += 1
                    scores[pr.profile_to]["actions"].append({"action": "Photo Request Sent", "datetime": pr.req_datetime})
                else:
                    scores[pr.profile_from]["score"] += 1
                    scores[pr.profile_from]["actions"].append({"action": "Photo Request Received", "datetime": pr.req_datetime})

            # Visitors
            visitors = Profile_visitors.objects.filter(
                Q(profile_id=profile_from, viewed_profile__in=profile_ids, status=1) |
                Q(viewed_profile=profile_from, profile_id__in=profile_ids, status=1)
            )
            for v in visitors:
                if v.profile_id == profile_from:
                    if v.viewed_profile in scores:
                        scores[v.viewed_profile]["score"] += 1
                        scores[v.viewed_profile]["actions"].append({"action": "Visited", "datetime": v.datetime})
                else:
                    if v.profile_id in scores:
                        scores[v.profile_id]["score"] += 1
                        scores[v.profile_id]["actions"].append({"action": "Viewed", "datetime": v.datetime})

            return scores



    def post(self, request):
        serializer = GetproflistSerializer(data=request.data)
        if not serializer.is_valid():
            return JsonResponse(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        profile_id = serializer.validated_data['profile_id']
        try:
            profile_data = Registration1.objects.get(ProfileId=profile_id)
        except Registration1.DoesNotExist:
            return JsonResponse({"Status": 0, "message": "Profile not found"}, status=status.HTTP_404_NOT_FOUND)

        gender = profile_data.Gender

        # Pagination
        per_page = int(request.data.get("per_page", 10))
        page_number = int(request.data.get("page_number", 1))
        start = (page_number - 1) * per_page

        # Fetch profiles
        profile_details, total_count, profile_with_indices = Get_profiledata_Matching.get_profile_list(
            gender=gender,
            profile_id=profile_id,
            start=start,
            per_page=per_page,
            search_profile_id=request.data.get('search_profile_id'),
            order_by=request.data.get('order_by'),
            search_profession=request.data.get('search_profession'),
            search_age=request.data.get('search_age'),
            search_location=request.data.get('search_location'),
            complexion=request.data.get('complexion'),
            city=request.data.get('city'),
            state=request.data.get('state'),
            education=request.data.get('education'),
            foreign_intrest=request.data.get('foreign_intrest'),
            has_photos=request.data.get('has_photos'),
            height_from=request.data.get('height_from'),
            height_to=request.data.get('height_to'),
            matching_stars=request.data.get('matching_stars'),
            min_anual_income=request.data.get('min_anual_income'),
            max_anual_income=request.data.get('max_anual_income'),
            membership=request.data.get('membership'),
            ragu=request.data.get('ragu'),
            chev=request.data.get('chev'),
            father_alive=request.data.get('father_alive'),
            mother_alive=request.data.get('mother_alive'),
            marital_status=request.data.get('marital_status') ,
            family_status=request.data.get('family_status'),
            whatsapp_field=request.data.get('whatsapp_field'),
            field_of_study=request.data.get('pref_fieldof_study'),
            degree = request.data.get('degree'),
            from_date=request.data.get('from_dateofjoin'),
            to_date=request.data.get('to_dateofjoin'),
            action_type=request.data.get('action_type'),
            status=request.data.get('status'),
            search = request.data.get('search'),
            except_viewed = request.data.get('except_viewed'),
            except_visitor = request.data.get('except_visitor')
        )

        if not profile_details:
            return JsonResponse({"Status": 0, "message": "No matching records", "search_result": "1"}, status=status.HTTP_200_OK)

        my_profile_details = get_profile_details([profile_id])[0]
        my_star_id = str(my_profile_details['birthstar_name'])
        my_rasi_id = str(my_profile_details['birth_rasi_name'])

        profile_ids = [detail.get("ProfileId") for detail in profile_details]
        action_scores =0
        try:
            action_scores = self.get_action_scores_bulk(profile_id, profile_ids)
        except Exception:
            action_scores = 0

        preload_matching_scores()
        score_map = cache.get("matching_score_map", {})

        result_profiles = []
        base_url = settings.MEDIA_URL
        plans = {p.id: p.plan_name for p in PlanDetails.objects.all()}  # or raw query
        family_statuses = {f.id: f.status for f in FamilyStatus.objects.all()}
        professions = {p.row_id: p.profession for p in Profession.objects.all()}
        states = {s.id: s.name for s in State.objects.all()}
        anualincomes = {i.id: i.income for i in AnnualIncome.objects.all()}


        # print('plans',plans)
        # print('family_statuses',family_statuses)
        logs_map = get_bulk_action_logs(profile_id, profile_ids)
        for detail in profile_details:

            # print(detail)

            dest_star = str(detail.get("birthstar_name"))
            dest_rasi = str(detail.get("birth_rasi_name"))


            chevvai = get_dhosham_new(detail.get("calc_chevvai_dhosham"))
            raguketu = get_dhosham_new(detail.get("calc_raguketu_dhosham"))


            key = (my_star_id, my_rasi_id, dest_star, dest_rasi, gender.lower())
            
            match_count = score_map.get(key, 0)
            matching_score = 100 if match_count == 15 else match_count * 10



            # image_function = lambda detail: get_profile_image_azure_optimized(detail.get("ProfileId"), gender, 1,0)

            pid = detail.get("ProfileId")
            result_profiles.append({
                "profile_id": detail.get("ProfileId"),
                "profile_name": detail.get("Profile_name"),
                # "profile_img":image_function(detail),
                "profile_img": base_url + (detail.get("profile_image") or "default_img.png"),
                "profile_age": detail.get("profile_age"),
                # "plan": get_plan(detail.get("Plan_id")),
                "plan": plans.get(int(detail.get("Plan_id")) if detail.get("Plan_id") else "N/A"),
               "family_status": family_statuses.get(int(detail.get("family_status")) if detail.get("family_status") else "N/A"),
                # "degree": degree(detail.get("degree"),detail.get("other_degree")),
                # "anual_income":get_annual_income(detail.get("anual_income"),detail.get("actual_income")),
                "degree": degree(detail.get("degree"),detail.get("other_degree")),
                "anual_income": (detail.get("actual_income") if detail.get("actual_income") not in [None, "", "0"] else anualincomes.get(int(detail.get("anual_income"))) if detail.get("anual_income") else "N/A"),
                "star": detail.get("star"), 
                "profession": professions.get(int(detail.get("profession")) if detail.get("profession") else "N/A"),
                "city": detail.get("Profile_city") if detail.get("Profile_city") not in [None,"0", "N/A","~"] else "N/A",
                # "state":  get_state_name(detail.get("Profile_state")) if detail.get("Profile_state") not in [None,"0", "N/A","~"] else "N/A",
                "state":states.get(int(detail.get("Profile_state")) if detail.get("Profile_state") else "N/A"),
                # "work_place": get_location(detail.get("work_city"),detail.get("work_state"),detail.get("work_country")),
                # "designation": get_designation_or_nature(detail.get("designation"),detail.get("nature_of_business")),
                # "company_name": get_company_or_business(detail.get("company_name"),detail.get("business_name")),

                "work_place": get_location_new(detail.get("work_city"), detail.get("work_state"), detail.get("work_country")),
                "designation": get_designation_or_nature_new(detail.get("designation"), detail.get("nature_of_business")),
                "company_name": get_company_or_business_new(detail.get("company_name"), detail.get("business_name")),
                "father_occupation":detail.get("father_occupation") if detail.get("father_occupation") not in [None,"0", "N/A","~"] else "N/A",
                "suya_gothram": detail.get("suya_gothram") if detail.get("suya_gothram") not in [None,"0", "N/A","~"] else "N/A",
                #"matching_score": Get_matching_score(my_star_id, my_rasi_id, detail.get("birthstar_name"), detail.get("birth_rasi_name"), gender),
                "chevvai": chevvai,
                "raguketu": raguketu,
                "matching_score":matching_score,
                "action_score": action_scores[pid],
                "action_log":logs_map.get(pid, "No logs"),
                # get_action_log(profile_id,detail.get("ProfileId"))
                "dateofjoin": detail.get("DateOfJoin") if detail.get("DateOfJoin") else None,
                "profile_status":detail.get("Status"),
            })

        return JsonResponse({
            "Status": 1,
            "message": "Matching records fetched successfully",
            "profiles": result_profiles,
            "total_count": total_count,
            "received_per_page": per_page,
            "received_page_number": page_number,
            "all_profile_ids": profile_with_indices,
            "search_result": "1"
        }, status=status.HTTP_200_OK)
        

class Get_suggest_list_match(APIView):

    def get_action_scores_bulk(self, profile_from, profile_ids):
        scores = {pid: {"score": 0, "actions": []} for pid in profile_ids}

        interests = Express_interests.objects.filter(
            Q(profile_from=profile_from, profile_to__in=profile_ids) |
            Q(profile_to=profile_from, profile_from__in=profile_ids),
            status__in=[1, 2, 3]
        )

        for ei in interests:
            if ei.profile_from == profile_from:
                target = ei.profile_to
                action = "Express Interest Sent" if ei.status == 1 else (
                         "Express Interest Accepted" if ei.status == 2 else "Express Interest Rejected")
            else:
                target = ei.profile_from
                action = "Express Interest Received"

            if target in scores:
                scores[target]["score"] += 1
                scores[target]["actions"].append({"action": action, "datetime": ei.req_datetime})

        wishlists = Profile_wishlists.objects.filter(
            Q(profile_from=profile_from, profile_to__in=profile_ids, status=1) |
            Q(profile_to=profile_from, profile_from__in=profile_ids, status=1)
        )
        for wl in wishlists:
            if wl.profile_from == profile_from:
                scores[wl.profile_to]["score"] += 1
                scores[wl.profile_to]["actions"].append({"action": "Bookmarked", "datetime": wl.marked_datetime})
            else:
                scores[wl.profile_from]["score"] += 1
                scores[wl.profile_from]["actions"].append({"action": "Bookmark Received", "datetime": wl.marked_datetime})

        photo_requests = Photo_request.objects.filter(
            Q(profile_from=profile_from, profile_to__in=profile_ids, status=1) |
            Q(profile_to=profile_from, profile_from__in=profile_ids, status=1)
        )
        for pr in photo_requests:
            if pr.profile_from == profile_from:
                scores[pr.profile_to]["score"] += 1
                scores[pr.profile_to]["actions"].append({"action": "Photo Request Sent", "datetime": pr.req_datetime})
            else:
                scores[pr.profile_from]["score"] += 1
                scores[pr.profile_from]["actions"].append({"action": "Photo Request Received", "datetime": pr.req_datetime})

        visitors = Profile_visitors.objects.filter(
            Q(profile_id=profile_from, viewed_profile__in=profile_ids, status=1) |
            Q(viewed_profile=profile_from, profile_id__in=profile_ids, status=1)
        )
        for v in visitors:
            if v.profile_id == profile_from:
                scores[v.viewed_profile]["score"] += 1
                scores[v.viewed_profile]["actions"].append({"action": "Visited", "datetime": v.datetime})
            else:
                scores[v.profile_id]["score"] += 1
                scores[v.profile_id]["actions"].append({"action": "Viewed", "datetime": v.datetime})

        return scores

    def post(self, request):
        serializer = GetproflistSerializer(data=request.data)

        if not serializer.is_valid():
            return JsonResponse(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        profile_id = serializer.validated_data['profile_id']

        try:
            profile_data = Registration1.objects.get(ProfileId=profile_id)
        except Registration1.DoesNotExist:
            return JsonResponse({"Status": 0, "message": "Profile not found"}, status=status.HTTP_404_NOT_FOUND)

        gender = profile_data.Gender

        # Pagination setup
        try:
            per_page = int(request.data.get("per_page", 10))
        except (ValueError, TypeError):
            per_page = 10

        try:
            page_number = int(request.data.get("page_number", 1))
        except (ValueError, TypeError):
            page_number = 1

        per_page = max(1, per_page)
        page_number = max(1, page_number)

        start = (page_number - 1) * per_page

        # Exclude existing partner preferences
        partner_profiles = gpt.get_profile_list_for_pref_type(
            profile_id=profile_id,
            use_suggested=False
        )
        exclude_profile_ids = [r['ProfileId'] for r in partner_profiles]
        print("match",len(exclude_profile_ids))

        # Fetch matching profiles with filtering and pagination inside query
        profile_details, total_count, profile_with_indices = Get_profiledata_Matching.get_suggest_profile_list(
            gender=gender,
            profile_id=profile_id,
            start=start,
            per_page=per_page,
            search_profile_id=request.data.get('search_profile_id'),
            order_by=request.data.get('order_by'),
            search_profession=request.data.get('search_profession'),
            search_age=request.data.get('search_age'),
            search_location=request.data.get('search_location'),
            complexion=request.data.get('complexion'),
            city=request.data.get('city'),
            state=request.data.get('state'),
            education=request.data.get('education'),
            foreign_intrest=request.data.get('foreign_intrest'),
            has_photos=request.data.get('has_photos'),
            height_from=request.data.get('height_from'),
            height_to=request.data.get('height_to'),
            matching_stars=request.data.get('matching_stars'),
            min_anual_income=request.data.get('min_anual_income'),
            max_anual_income=request.data.get('max_anual_income'),
            membership=request.data.get('membership'),
            ragu=request.data.get('ragu'),
            chev=request.data.get('chev'),
            father_alive=request.data.get('father_alive'),
            mother_alive=request.data.get('mother_alive'),
            marital_status=request.data.get('marital_status'),
            family_status=request.data.get('family_status'),
            whatsapp_field=request.data.get('whatsapp_field'),
            field_of_study=request.data.get('pref_fieldof_study'),
            degree=request.data.get('degree'),
            exclude_profile_ids=exclude_profile_ids
        )
        print("tol",total_count)
        if not profile_details:
            return JsonResponse({
                "Status": 0,
                "message": "No matching records",
                "search_result": "1"
            }, status=status.HTTP_200_OK)

        profile_ids = [p["ProfileId"] for p in profile_details]

        my_profile_details = get_profile_details([profile_id])[0]
        my_star_id = str(my_profile_details['birthstar_name'])
        my_rasi_id = str(my_profile_details['birth_rasi_name'])

        preload_matching_scores()
        score_map = cache.get("matching_score_map", {})

        # Action logs/scores
        try:
            action_scores = self.get_action_scores_bulk(profile_id, profile_ids)
        except Exception:
            action_scores = {}

        logs_map = get_bulk_action_logs(profile_id, profile_ids)

        # Lookup dictionaries
        plans = {p.id: p.plan_name for p in PlanDetails.objects.all()}
        family_statuses = {f.id: f.status for f in FamilyStatus.objects.all()}
        professions = {p.row_id: p.profession for p in Profession.objects.all()}
        states = {s.id: s.name for s in State.objects.all()}
        anualincomes = {i.id: i.income for i in AnnualIncome.objects.all()}

        result_profiles = []

        for detail in profile_details:
            pid = detail["ProfileId"]
            dest_star = str(detail.get("birthstar_name"))
            dest_rasi = str(detail.get("birth_rasi_name"))
            key = (my_star_id, my_rasi_id, dest_star, dest_rasi, gender.lower())
            match_count = score_map.get(key, 0)
            matching_score = 100 if match_count == 15 else match_count * 10

            chevvai = get_dhosham_new(detail.get("calc_chevvai_dhosham"))
            raguketu = get_dhosham_new(detail.get("calc_raguketu_dhosham"))

            result_profiles.append({
                "profile_id": pid,
                "profile_name": detail.get("Profile_name"),
                "profile_img": settings.MEDIA_URL + (detail.get("profile_image") or "default_img.png"),
                "profile_age": detail.get("profile_age"),
                "plan": plans.get(int(detail.get("Plan_id"))) if detail.get("Plan_id") else "N/A",
                "family_status": family_statuses.get(int(detail.get("family_status"))) if detail.get("family_status") else "N/A",
                "degree": degree(detail.get("degree"), detail.get("other_degree")),
                "anual_income": detail.get("actual_income") if detail.get("actual_income") not in [None, "", "0"]
                                  else anualincomes.get(int(detail.get("anual_income"))) if detail.get("anual_income") else "N/A",
                "star": detail.get("star"),
                "profession": professions.get(int(detail.get("profession"))) if detail.get("profession") else "N/A",
                "city": detail.get("Profile_city") if detail.get("Profile_city") not in [None, "0", "N/A", "~"] else "N/A",
                "state": states.get(int(detail.get("Profile_state"))) if detail.get("Profile_state") else "N/A",
                "work_place": get_location_new(detail.get("work_city"), detail.get("work_state"), detail.get("work_country")),
                "designation": get_designation_or_nature_new(detail.get("designation"), detail.get("nature_of_business")),
                "company_name": get_company_or_business_new(detail.get("company_name"), detail.get("business_name")),
                "father_occupation": detail.get("father_occupation") if detail.get("father_occupation") not in [None, "0", "N/A", "~"] else "N/A",
                "suya_gothram": detail.get("suya_gothram") if detail.get("suya_gothram") not in [None, "0", "N/A", "~"] else "N/A",
                "chevvai": chevvai,
                "raguketu": raguketu,
                "matching_score": matching_score,
                "action_score": action_scores.get(pid, {}),
                "action_log": logs_map.get(pid, "No logs"),
                "dateofjoin": detail.get("DateOfJoin") if detail.get("DateOfJoin") else None,
                "verified": detail.get("Profile_verified"),
                "profile_status": detail.get("Status")
            })

        return JsonResponse({
            "Status": 1,
            "message": "Matching records fetched successfully",
            "profiles": result_profiles,
            "total_count": total_count,
            "received_per_page": per_page,
            "received_page_number": page_number,
            "all_profile_ids": profile_with_indices,
            "search_result": "1"
        }, status=status.HTTP_200_OK)
       
class Get_visibility_list_match(APIView):
    def post(self, request):
        serializer = GetproflistSerializer(data=request.data)

        if not serializer.is_valid():
            return JsonResponse(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        profile_id = serializer.validated_data['profile_id']

        # ✅ Safely get profile
        try:
            profile_data = Registration1.objects.get(ProfileId=profile_id)
        except Registration1.DoesNotExist:
            return JsonResponse({"Status": 0, "message": "Profile not found"}, status=status.HTTP_404_NOT_FOUND)

        gender = profile_data.Gender

        # ✅ Get filters safely
        search_profile_id = request.data.get('search_profile_id')
        search_profession = request.data.get('search_profession')
        search_age = request.data.get('search_age')
        search_location = request.data.get('search_location')
        order_by = request.data.get('order_by')

        # ✅ Pagination (defaults)
        try:
            per_page = int(request.data.get('per_page', 10))
        except (ValueError, TypeError):
            per_page = 10

        try:
            page_number = int(request.data.get('page_number', 1))
        except (ValueError, TypeError):
            page_number = 1

        per_page = max(1, per_page)
        page_number = max(1, page_number)

        # Fetch all matching profiles first
        # ✅ Fetch all matching profiles first (parameter names now match the function)
        profile_details, total_count, profile_with_indices = Get_profiledata_Matching.get_profile_visibility(
            gender,
            profile_id,
            start=0,
            per_page=100000,  # Get all then paginate
            order_by=order_by,
            profession=request.data.get('profession'),  # ✅ renamed
            age_from=request.data.get('from_age'),      # ✅ renamed
            age_to=request.data.get('to_age'),          # ✅ renamed
            education=request.data.get('education'),
            foreign_intrest=request.data.get('foreign_intrest'),
            height_from=request.data.get('height_from'),
            height_to=request.data.get('height_to'),
            min_anual_income=request.data.get('min_anual_income'),
            max_anual_income=request.data.get('max_anual_income'),
            ragu=request.data.get('ragu'),
            chev=request.data.get('chev'),
            marital_status=request.data.get('marital_status'),
            family_status=request.data.get('family_status'),
            field_of_study=request.data.get('pref_fieldof_study'),
            degree=request.data.get('degree')
        )


        # Paginate results
        paginator = Paginator(profile_details, per_page)
        try:
            page_obj = paginator.page(page_number)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = []

        my_profile_details = get_profile_details([profile_id])
        my_gender = my_profile_details[0]['Gender']
        my_star_id = my_profile_details[0]['birthstar_name']
        my_rasi_id = my_profile_details[0]['birth_rasi_name']

        restricted_profile_details = [
            {
                "profile_id": detail.get("ProfileId"),
                "profile_name": detail.get("Profile_name"),
                "profile_img": Get_profile_image(
                    detail.get("ProfileId"),
                    gender="female" if gender.lower() == "male" else "male",
                    no_of_image=1,
                    photo_protection=0,
                    is_admin=True
                ),
                "profile_age": calculate_age(detail.get("Profile_dob")),
                "plan": get_plan(detail.get("Plan_id")),
                "family_status": get_family_status(detail.get("family_status")),
                "degree": degree(detail.get("degree"), detail.get("other_degree")),
                "anual_income": get_annual_income(detail.get("anual_income"), detail.get("actual_income")),
                "star": detail.get("star"),
                "profession": getprofession(detail.get("profession")),
                "city": detail.get("Profile_city") if detail.get("Profile_city") not in [None, "0", "N/A", "~"] else "N/A",
                "state": get_state_name(detail.get("Profile_state")) if detail.get("Profile_state") not in [None, "0", "N/A", "~"] else "N/A",
                "work_place": get_location(detail.get("work_city"), detail.get("work_state"), detail.get("work_country")),
                "designation": get_designation_or_nature(detail.get("designation"), detail.get("nature_of_business")),
                "company_name": get_company_or_business(detail.get("company_name"), detail.get("business_name")),
                "father_occupation": detail.get("father_occupation") if detail.get("father_occupation") not in [None, "0", "N/A", "~"] else "N/A",
                "suya_gothram": detail.get("suya_gothram") if detail.get("suya_gothram") not in [None, "0", "N/A", "~"] else "N/A",
                "chevvai": get_dhosham(detail.get("calc_chevvai_dhosham")),
                "raguketu": get_dhosham(detail.get("calc_raguketu_dhosham")),
                "photo_protection": detail.get("Photo_protection"),
                "matching_score": Get_matching_score(my_star_id, my_rasi_id, detail.get("birthstar_name"), detail.get("birth_rasi_name"), my_gender),
                "dateofjoin": detail.get("DateOfJoin") if detail.get("DateOfJoin") else None,
            }
            for detail in page_obj
        ]

        return JsonResponse({
            "Status": 1 if restricted_profile_details else 0,
            "message": "Matching records fetched successfully" if restricted_profile_details else "No matching records",
            "profiles": restricted_profile_details,
            "total_count": total_count,
            "received_per_page": per_page,
            "received_page_number": page_number,
            "calculated_per_page": per_page,
            "calculated_page_number": page_number,
            "all_profile_ids": profile_with_indices,
            "search_result": "1"
        }, status=status.HTTP_200_OK)


def send_bulk_email(request):
 
 if request.method == "POST":
        subject = request.POST.get('subject', '')
        message = request.POST.get('message', '')
        profile_ids = request.POST.get('profile_id', '').strip().split(',')
        stars = request.POST.getlist('stars')
        gender = request.POST.get('gender', None)
        plan = request.POST.get('plan', None)
        from_age = request.POST.get('from_age', None)
        to_age = request.POST.get('to_age', None)
        from_date = request.POST.get('from_date', None)
        to_date = request.POST.get('to_date', None)

        # Handle file attachment from request (assuming form uses 'file' as input name)
        attached_file = request.FILES.get('file')

        # Convert age and date parameters
        try:
            from_age = int(from_age) if from_age else None
            to_age = int(to_age) if to_age else None
        except ValueError:
            return JsonResponse({
                "success": False,
                "message": "Invalid age range provided.",
                "total_emails_sent": 0
            })

        try:
            from_date = datetime.strptime(from_date, '%Y-%m-%d').date() if from_date else None
            to_date = datetime.strptime(to_date, '%Y-%m-%d').date() if to_date else None
        except ValueError:
            return JsonResponse({
                "success": False,
                "message": "Invalid date format. Use YYYY-MM-DD.",
                "total_emails_sent": 0
            })

        # Check if any filters are provided
        if not profile_ids and not stars and not gender and not plan and not from_age and not to_age and not from_date and not to_date:
            return JsonResponse({
                "success": False,
                "message": "No profile IDs, stars, gender, plan, age range, or date range provided to send emails.",
                "total_emails_sent": 0
            })

        # Initialize email list
        emails = []

        # Filter emails by profile IDs if provided
        if profile_ids:
            profile_emails = LoginDetails.objects.filter(ProfileId__in=profile_ids).values_list('EmailId', flat=True)
            emails.extend(profile_emails)

        # Filter emails by birth stars if provided
        if stars:
            star_profile_ids = ProfileHoroscope.objects.filter(birthstar_name__in=stars).values_list('profile_id', flat=True)
            star_emails = LoginDetails.objects.filter(ProfileId__in=star_profile_ids).values_list('EmailId', flat=True)
            emails.extend(star_emails)

        # Filter emails by gender and status=8 if provided
        if gender:
            gender_emails = LoginDetails.objects.filter(Gender=gender, status=8).values_list('EmailId', flat=True)
            emails.extend(gender_emails)

        # Filter emails by Plan_id if provided
        if plan:
            plan_emails = LoginDetails.objects.filter(Plan_id=plan).values_list('EmailId', flat=True)
            emails.extend(plan_emails)

        # Filter emails by age range if provided
        if from_age is not None or to_age is not None:
            profiles = LoginDetails.objects.all()
            for profile in profiles:
                age = calculate_age(profile.Profile_dob)
                if age is not None and (from_age is None or age >= from_age) and (to_age is None or age <= to_age):
                    emails.append(profile.EmailId)

        # Filter emails by registration date range if provided
        if from_date or to_date:
            date_filter = {}
            if from_date:
                date_filter['DateOfJoin__gte'] = from_date
            if to_date:
                date_filter['DateOfJoin__lte'] = to_date

            date_emails = LoginDetails.objects.filter(**date_filter).values_list('EmailId', flat=True)
            emails.extend(date_emails)

        # Remove duplicate emails
        emails = list(set(emails))

        # If no emails were found, return a message
        if not emails:
            return JsonResponse({
                "success": False,
                "message": "No emails found for the given filters.",
                "total_emails_sent": 0
            })

        # HTML content with embedded logo in a container
        # HTML content with embedded logo from URL and attached image
        html_content = f"""
        <html>
        <body>
            <table width="100%" cellpadding="0" cellspacing="0" border="0" bgcolor="#f4f4f4">
                <tr>
                    <td align="center">
                        <table width="600" cellpadding="0" cellspacing="0" border="0" bgcolor="#ffffff" style="margin: 20px auto; padding: 20px; border-radius: 10px;">
                            <tr>
                                <td align="center">
                                    <!-- Static Logo Image from URL -->
                                    <img src="https://vysyamat.blob.core.windows.net/vysyamala/newvysyamalalogo2.png" alt="Logo" style="max-width: 200px;"/>
                                </td>
                            </tr>
                            <tr>
                                <td align="center">
                                    <!-- Embedded Attached Image using CID -->
                                    <img src="cid:attached_image" alt="Attached Image" style="max-width: 600px;"/>
                                </td>
                            </tr>
                            <tr>
                                <td style="padding: 20px; text-align: left;">
                                    <h2 style="color: #333;">Dear Customer,</h2>
                                    <p style="color: #555; font-size: 16px;">
                                        {message}
                                    </p>
                                    <p style="color: #555; font-size: 16px;">
                                        Best regards,<br/>
                                        <strong>Vysyamala</strong>
                                    </p>
                                </td>
                            </tr>
                        </table>
                    </td>
                </tr>
            </table>
        </body>
        </html>
        """


        # Send emails in bulk with HTML content and file attachment
        for email in emails:
            email_message = EmailMessage(
                subject=subject,
                body=html_content,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=[email],
            )
            email_message.content_subtype = 'html'  # Set content to HTML

             # Embed the attached image using Content-ID (CID)
            if attached_file:
                # Attach the file and set its CID to use it in the email body
                image = MIMEImage(attached_file.read())
                image.add_header('Content-ID', '<attached_image>')
                email_message.attach(image)

            email_message.send(fail_silently=False)

        return JsonResponse({
            "success": True,
            "message": "Emails sent successfully.",
            "total_emails_sent": len(emails)
        })

 return JsonResponse({
        "success": False,
        "message": "Invalid request method. Only POST is allowed."
    }, status=400)




def Get_wishlist(profile_id,user_profile_id):
   
    if profile_id and user_profile_id:
        
        
         existing_entry=Profile_wishlists.objects.filter(profile_from=profile_id,profile_to=user_profile_id,status=1)

         if existing_entry:

            return 1
                  
         else:
              return 0
    return None


def Get_expressstatus(profile_id, user_profile_id):
    if profile_id and user_profile_id:
        print(f'profile_id: {profile_id}, user_profile_id: {user_profile_id}')

        # Get the first matching entry
        existing_entry = Express_interests.objects.filter(profile_from=profile_id, profile_to=user_profile_id).first()

        #print('existing_entry:', existing_entry)

        if existing_entry:
            # Serialize the single instance
            serializer = ExpressInterestsSerializer(existing_entry)
            # Return only the status
            return serializer.data['status']
        else:
            
            return 0

    return 0  # Return 0 if no entry exists or profile_id/user_profile_id are not provided



def Get_personalnotes_value(profile_id, user_profile_id):
    if profile_id and user_profile_id:
        print(f'profile_id: {profile_id}, user_profile_id: {user_profile_id}')

        # Get the first matching entry
        existing_entry = Profile_personal_notes.objects.filter(profile_id=profile_id, profile_to=user_profile_id).first()

        #print('existing_entry:', existing_entry)

        if existing_entry:
            # Serialize the single instance
            serializer = PersonalnotesSerializer(existing_entry)
            # Return only the status
            return serializer.data['notes']
        else:
            
            return ''

    return ''  # Return 0 if no entry exists or profile_id/user_profile_id are not provided



# def get_degree(degeree):

#     # print('degeree',degeree)

#     try:
        
#         Profile_ug_degree = UgDegree.objects.get(id=degeree).degree
    
#     except UgDegree.DoesNotExist:
#                 Profile_ug_degree = None 
    
#     return Profile_ug_degree
def get_degree(degeree):

    # print('degeree',degeree)
    if isinstance(degeree, str):
        return degeree
    
    try:
        
        Profile_ug_degree = models.UgDegree.objects.get(id=degeree).degree
    
    except models.UgDegree.DoesNotExist:
        Profile_ug_degree = None 
    
    return Profile_ug_degree



def getprofession(profession):

    if profession in [None,"0", "N/A","~"," ",'']:
        return "N/A"

    try:
        
        Profile_profession = Profespref.objects.get(RowId=profession).profession
    
    except Profespref.DoesNotExist:
                Profile_profession = None 
    
    return Profile_profession


def Get_matching_score(source_star_id, source_rasi_id,dest_star_id,dest_rasi_id,gender):
    
    # print('source_star_id : ',source_star_id,'source_rasi_id: ',source_rasi_id,'dest_star_id: ', dest_star_id , 'dest_rasi_id: ',dest_rasi_id,'gender',gender)

    if source_star_id and source_rasi_id and dest_star_id and dest_rasi_id:
        
       

        # Get the first matching entry
        existing_entry = MatchingStarPartner.objects.filter(source_star_id=source_star_id, source_rasi_id=source_rasi_id, dest_star_id=dest_star_id,dest_rasi_id=dest_rasi_id,gender=gender)


        if existing_entry:

            # print('sddgdfgfg')
            # Serialize the single instance
            serializer = MatchingscoreSerializer(existing_entry,many=True)

            match_count = serializer.data[0].get('match_count', 0)
            # Return only the status
            if(match_count==15):
                matching_score=100
            else:
                matching_score=match_count*10            

            return matching_score
        else:
            
            return 0

    return 0  # Return 0 if no entry exists or profile_id/user_profile_id are not provided




# def Get_profile_image(user_profile_id, is_admin=False):
#     base_url = settings.IMAGE_BASEURL

#     # Fetch all images for the given user profile ID
#     if user_profile_id:
#         # If admin, skip photo protection logic
#         if is_admin:
#             get_entry = Image_Upload.objects.filter(profile_id=user_profile_id)[:10]
#             if get_entry.exists():
#                 # Serialize the multiple instances for admin
#                 serializer = ImageGetSerializer(get_entry, many=True)
#                 images_dict = {
#                     str(index + 1): base_url + entry['image']
#                     for index, entry in enumerate(serializer.data)
#                 }
#                 return images_dict
#             else:
#                 # Return an empty array if no entry is found
#                 return []
#         else:
#             # Non-admin users: consider photo protection
#             get_entry = Image_Upload.objects.filter(profile_id=user_profile_id).first()
#             if get_entry:
#                 serializer = ImageGetSerializer(get_entry)
#                 # Check if the user has photo protection enabled
#                 if get_entry.photo_protection:
#                     # Return blurred image for regular users with photo protection
#                     img_base64 = get_blurred_image(serializer.data['image'])
#                     return {"1": img_base64}
#                 else:
#                     # Return the image directly if no photo protection
#                     return {"1": base_url + serializer.data['image']}
#             else:
#                 # Return an empty array if no entry exists
#                 return []
#     else:
#         # Return an empty array if no user_profile_id is provided
#         return []

def Get_profile_image(user_profile_id, gender, no_of_image, photo_protection, is_admin=False):
    base_url = settings.MEDIA_URL
    default_img_bride = 'default_bride.png'
    default_img_groom = 'default_groom.png'

    # Admin bypasses photo protection logic
    if is_admin:
        if user_profile_id:
            if no_of_image == 1:
                get_entry = Image_Upload.objects.filter(profile_id=user_profile_id).filter(
                    Q(is_deleted=False) | Q(is_deleted__isnull=True)
                ).first()
                if get_entry:
                    # Admin gets unblurred image
                    serializer = ImageGetSerializer(get_entry)
                    return serializer.data['image']
                else:
                    # Return default image based on gender if no image is found
                    return base_url + (default_img_groom if gender.lower() == 'male' else default_img_bride)
            else:
                # Fetch up to 10 images for admin without any photo protection
                get_entry = Image_Upload.objects.filter(profile_id=user_profile_id).filter(
                    Q(is_deleted=False) | Q(is_deleted__isnull=True)
                )[:10]
                if get_entry.exists():
                    serializer = ImageGetSerializer(get_entry, many=True)
                    # Return a dictionary of images
                    images_dict = {
                        str(index + 1):  entry['image']
                        for index, entry in enumerate(serializer.data)
                    }
                    return images_dict
                else:
                    # Return default images if none are found
                    default_img = default_img_groom if gender.lower() == 'male' else default_img_bride
                    return {"1": base_url + default_img, "2": base_url + default_img}

    # Non-admin logic
    if photo_protection != 1:        
        if user_profile_id:
            if no_of_image == 1:
                get_entry = Image_Upload.objects.filter(profile_id=user_profile_id).filter(
                    Q(is_deleted=False) | Q(is_deleted__isnull=True)
                ).first()
                if get_entry:
                    serializer = ImageGetSerializer(get_entry)
                    return serializer.data['image']
                else:
                    return base_url + (default_img_groom if gender.lower() == 'male' else default_img_bride)
            else:
                get_entry = Image_Upload.objects.filter(profile_id=user_profile_id).filter(
                    Q(is_deleted=False) | Q(is_deleted__isnull=True)
                )[:10]
                if get_entry.exists():
                    serializer = ImageGetSerializer(get_entry, many=True)
                    images_dict = {
                        str(index + 1): entry['image']
                        for index, entry in enumerate(serializer.data)
                    }
                    return images_dict
                else:
                    default_img = default_img_groom if gender.lower() == 'male' else default_img_bride
                    return {"1": base_url + default_img, "2": base_url + default_img}
    else:
        # Photo protection enabled
        if no_of_image == 1:
            get_entry = Image_Upload.objects.filter(profile_id=user_profile_id).filter(
                    Q(is_deleted=False) | Q(is_deleted__isnull=True)
                ).first()
            if get_entry:
                serializer = ImageGetSerializer(get_entry)
                img_base64 = get_blurred_image(serializer.data['image'])
                return img_base64
        else:
            get_entry = Image_Upload.objects.filter(profile_id=user_profile_id).filter(
                    Q(is_deleted=False) | Q(is_deleted__isnull=True)
                ).first()
            if get_entry:
                serializer = ImageGetSerializer(get_entry)
                img_base64 = get_blurred_image(serializer.data['image'])
                return {"1": img_base64}
            else:
                return base_url + (default_img_groom if gender.lower() == 'male' else default_img_bride)

             
def Get_image_profile(user_profile_id):
    base_url = settings.MEDIA_URL
    default_img_bride = 'default_bride.png'
    default_img_groom = 'default_groom.png'
    user_profile = Registration1.objects.get(ProfileId=user_profile_id)
    
    gender = user_profile.Gender
    photo_protection = user_profile.Photo_protection

    # Default to the appropriate image based on gender
    if not photo_protection:
        get_entry = Image_Upload.objects.filter(profile_id=user_profile_id).filter(
                    Q(is_deleted=False) | Q(is_deleted__isnull=True)
                ).first()
        if get_entry:
            serializer = ImageGetSerializer(get_entry)
            return base_url + serializer.data['image']
        
        return base_url + (default_img_groom if gender.lower() == 'male' else default_img_bride)
    
    get_entry = Image_Upload.objects.filter(profile_id=user_profile_id).filter(
                    Q(is_deleted=False) | Q(is_deleted__isnull=True)
                ).first()
    if get_entry:
        serializer = ImageGetSerializer(get_entry)
        img_base64 = get_blurred_image(serializer.data['image'])
        return img_base64  # Ensure this returns a string
    
    # Fallback to a default blurred image in case of no entry found
    return settings.MEDIA_URL + 'default_img.png'

def get_blurred_image(image_name):
    # Construct the image path
    #print('image_name',image_name)

    image_name = image_name[len(''):]
    
    image_path = os.path.join(settings.MEDIA_ROOT,image_name)

    # print('image_path',image_path)
    
    # Check if the file exists
    if not os.path.isfile(image_path):
        return settings.MEDIA_URL+'default_img.png'
    
    try:
        # Open the image using Pillow
        with Image.open(image_path) as img:
            # Apply blur effect
            blurred_image = img.filter(ImageFilter.GaussianBlur(10))  # Adjust the blur radius if needed
            
            # Save the blurred image to a BytesIO object
            buffered = BytesIO()
            blurred_image.save(buffered, format="JPEG")
            
            # Encode the image in base64
            img_base64 = base64.b64encode(buffered.getvalue()).decode('utf-8')
            
            # Return the base64 encoded image in a JSON response
            return 'data:image/jpeg;base64,'+img_base64
    
    except Exception as e:
        return settings.MEDIA_URL+'default_img.png'

def get_profile_details(profile_ids):
    print('profile_details')
    #profiles = models.Get_profiledata.get_profile_details.objects.filter(ProfileId__in=profile_ids)
    profiles = Get_profiledata_Matching.get_profile_details(profile_ids)
       
    
    return profiles


def calculate_age(dob):
    """
    Calculate age based on date of birth.
    
    Args:
    dob (datetime.date): The date of birth.
    
    Returns:
    int or None: The calculated age or None if dob is not provided.
    """
    if dob:
        today = datetime.today()
        age = today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))
        return age
    return None



#Profile Vysassist
class ProfileVysAssistView(APIView):
    pagination_class = StandardResultsPaging

    def get(self, request):
        profile_assists = Profile_vysassist.objects.all().order_by('-req_datetime')

        if not profile_assists.exists():
            return Response({"message": "No profile_vysassist records found."}, status=404)

        # Create a result list to include profile information
        result = []

        for assist in profile_assists:
            # Fetch profile_from data (without filtering by Profile_state)
            profile_from_data = LoginDetails.objects.filter(
                ProfileId=assist.profile_from
            ).first()

            # Fetch profile_to data (without filtering by Profile_state)
            profile_to_data = LoginDetails.objects.filter(
                ProfileId=assist.profile_to
            ).first()

            status_comments = {
                0: "Removed",
                1: "Request Sent",
                2: "Accepted",
                3: "Rejected"
            }
            # Only add to the result if both profile_from and profile_to exist
            if profile_from_data and profile_to_data:
                result.append({
                    'profile_vysasst_id': assist.id,
                    'profile_from_id': profile_from_data.ProfileId,
                    'profile_from_name': profile_from_data.Profile_name,
                    'profile_from_mobile': profile_from_data.Mobile_no,
                    'profile_to_id': profile_to_data.ProfileId,
                    'profile_to_name': profile_to_data.Profile_name,
                    'profile_to_mobile': profile_to_data.Mobile_no,
                    'to_message': assist.to_message,
                    'req_datetime': assist.req_datetime,
                    'response_datetime': assist.response_datetime,
                    'status': status_comments.get(assist.status)  # 1: request sent, 2: accepted, 3: rejected, 0: removed
                })

        # Implement pagination
        paginator = self.pagination_class()
        paginated_result = paginator.paginate_queryset(result, request)

        # If there are paginated results, return the paginated response
        if paginated_result is not None:
            return paginator.get_paginated_response(paginated_result)

        # If no pagination is needed, return the full result set
        return Response(result, status=200)
    
#Delete Profile
def delete_profile(request, profile_id):
    if request.method == 'DELETE':
        try:
            # Get the profile by ProfileId
            profile = get_object_or_404(LoginDetails, ProfileId=profile_id)
            
            # Update the status to 7 (soft delete)
            profile.status = '7'
            profile.save()

            # Return success response
            return JsonResponse({'message': 'Profile soft deleted successfully'}, status=200)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=400)
    
#Viewed Profiles    
class My_viewed_profiles(APIView):
    def post(self, request):
        serializer = Profile_idValidationSerializer(data=request.data)

        if serializer.is_valid():
            profile_id = serializer.validated_data.get('profile_id')
            page = int(request.data.get('page_number', 1))
            per_page = int(request.data.get('per_page', 10)) 

            # Extract from_date and to_date from the request
            from_date = request.data.get('from_date')
            to_date = request.data.get('to_date')

            try:
                # Initialize the base queryset to filter by profile_id
                all_profiles = Profile_visitors.objects.filter(profile_id=profile_id)

                # Apply date filters if from_date and to_date are provided
                if from_date:
                    from_date = datetime.strptime(from_date, '%Y-%m-%d').date()  # Convert to date
                    all_profiles = all_profiles.filter(datetime__date__gte=from_date)

                if to_date:
                    to_date = datetime.strptime(to_date, '%Y-%m-%d').date()  # Convert to date
                    all_profiles = all_profiles.filter(datetime__date__lte=to_date)

                # Get all profile IDs in the filtered queryset
                all_profile_ids = {str(index + 1): profile_id for index, profile_id in enumerate(all_profiles.values_list('viewed_profile', flat=True))}

                total_records = all_profiles.count()

                start = (page - 1) * per_page
                end = start + per_page

                # Fetch paginated data
                fetch_data = all_profiles[start:end]

                if fetch_data.exists():
                    profile_ids = fetch_data.values_list('viewed_profile', flat=True)
                    profile_details = get_profile_details(profile_ids)
                    profile_datetime_map = {
                        record.viewed_profile: (
                            record.datetime.strftime("%Y-%m-%d %H:%M:%S") if record.datetime else None
                        )
                        for record in fetch_data
                    }
                    profile_data = Registration1.objects.get(ProfileId=profile_id)
                    horo_data = ProfileHoroscope.objects.get(profile_id=profile_id)

                    my_star_id = horo_data.birthstar_name
                    my_rasi_id = horo_data.birth_rasi_name
                    my_gender = profile_data.Gender

                    restricted_profile_details = [
                        {
                            "visited_profileid": detail.get("ProfileId"),
                            "visited_profile_name": detail.get("Profile_name"),
                            "visited_Profile_img": Get_profile_image(detail.get("ProfileId"), my_gender, 1, 0),
                            "visited_profile_age": calculate_age(detail.get("Profile_dob")),
                            "visited_verified": detail.get("Profile_verified"),
                            "visited_height": detail.get("Profile_height"),
                            "visited_star": detail.get("star_name"),
                            "visited_profession": getprofession(detail.get("profession")),
                            "visited_city": detail.get("Profile_city"),
                            "visited_degree": " ",
                            "visited_match_score": Get_matching_score(my_star_id, my_rasi_id, detail.get("birthstar_name"), detail.get("birth_rasi_name"), my_gender),
                            "visited_views": count_records(Profile_visitors, {'status': 1, 'viewed_profile': detail.get("ProfileId")}),
                            "visited_lastvisit":  profile_datetime_map.get(detail.get("ProfileId")),
                            "visited_userstatus": get_user_statusandlastvisit(detail.get("Last_login_date"))[1],
                            "visited_horoscope": "Horoscope Available" if detail.get("horoscope_file") else "Horoscope Not Available",
                            "visited_profile_wishlist": Get_wishlist(profile_id, detail.get("ProfileId")),
                        }
                        for detail in profile_details
                    ]

                    combined_data = {
                        "profiles": restricted_profile_details,
                        "page": page,
                        "per_page": per_page,
                        "total_pages": (total_records + per_page - 1) // per_page,  # Calculate total pages
                        "total_records": total_records,
                        "all_profile_ids": all_profile_ids
                    }

                    return JsonResponse({"Status": 1, "message": "Fetched viewed profile lists successfully", "data": combined_data, "viewed_profile_count": total_records}, status=status.HTTP_200_OK)
                else:
                    return JsonResponse({"Status": 0, "message": "No viewed profiles found for the given profile ID"}, status=status.HTTP_200_OK)
            except Profile_visitors.DoesNotExist:
                return JsonResponse({"Status": 0, "message": "No viewed profiles found for the given profile ID"}, status=status.HTTP_200_OK)
        else:
            return JsonResponse(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
#Visitors Profiles    
# class My_profiles_vistors(APIView):
#     def post(self, request):
#         serializer = Profile_idValidationSerializer(data=request.data)

#         if serializer.is_valid():
#             profile_id = serializer.validated_data.get('profile_id')
#             page = int(request.data.get('page_number', 1))
#             per_page = int(request.data.get('per_page', 10)) 

#             # Extract from_date and to_date from the request
#             from_date = request.data.get('from_date')
#             to_date = request.data.get('to_date')

#             try:
#                 # Initialize the base queryset to filter by profile_id
#                 all_profiles = Profile_visitors.objects.filter(viewed_profile=profile_id)

#                 # Apply date filters if from_date and to_date are provided
#                 if from_date:
#                     from_date = datetime.strptime(from_date, '%Y-%m-%d').date()  # Convert to date
#                     all_profiles = all_profiles.filter(datetime__date__gte=from_date)

#                 if to_date:
#                     to_date = datetime.strptime(to_date, '%Y-%m-%d').date()  # Convert to date
#                     all_profiles = all_profiles.filter(datetime__date__lte=to_date)

#                 # Get all profile IDs in the filtered queryset
#                 all_profile_ids = {str(index + 1): profile_id for index, profile_id in enumerate(all_profiles.values_list('viewed_profile', flat=True))}

#                 total_records = all_profiles.count()

#                 start = (page - 1) * per_page
#                 end = start + per_page

#                 # Fetch paginated data
#                 fetch_data = all_profiles[start:end]

#                 if fetch_data.exists():
#                     profile_ids = fetch_data.values_list('viewed_profile', flat=True)
#                     profile_details = get_profile_details(profile_ids)

#                     profile_data = Registration1.objects.get(ProfileId=profile_id)
#                     horo_data = ProfileHoroscope.objects.get(profile_id=profile_id)

#                     my_star_id = horo_data.birthstar_name
#                     my_rasi_id = horo_data.birth_rasi_name
#                     my_gender = profile_data.Gender

#                     # profile_call_city_name=get_city_name(profile_call.Profile_city)
#                     # profile_call_state_name=get_state_name(profile_call.Profile_state)

#                     restricted_profile_details = [
#                         {
#                             "visited_profileid": detail.get("ProfileId"),
#                             "visited_profile_name": detail.get("Profile_name"),
#                             "visited_Profile_img": Get_profile_image(detail.get("ProfileId"), my_gender, 1, 0),
#                             "visited_profile_age": calculate_age(detail.get("Profile_dob")),
#                             "visited_verified": detail.get("Profile_verified"),
#                             "visited_height": detail.get("Profile_height"),
#                             "visited_star": detail.get("star_name"),
#                             "visited_profession": getprofession(detail.get("profession")),
#                             "visited_city": get_city_name(detail.get("Profile_city")),
#                             "visited_degree": get_degree(detail.get("ug_degeree")),
#                             "visited_match_score": Get_matching_score(my_star_id, my_rasi_id, detail.get("birthstar_name"), detail.get("birth_rasi_name"), my_gender),
#                             "visited_views": count_records(Profile_visitors, {'status': 1, 'viewed_profile': detail.get("ProfileId")}),
#                             "visited_lastvisit": get_user_statusandlastvisit(detail.get("Last_login_date"))[0],
#                             "visited_userstatus": get_user_statusandlastvisit(detail.get("Last_login_date"))[1],
#                             "visited_horoscope": "Horoscope Available" if detail.get("horoscope_file") else "Horoscope Not Available",
#                             "visited_profile_wishlist": Get_wishlist(profile_id, detail.get("ProfileId")),
#                         }
#                         for detail in profile_details
#                     ]

#                     combined_data = {
#                         "profiles": restricted_profile_details,
#                         "page": page,
#                         "per_page": per_page,
#                         "total_pages": (total_records + per_page - 1) // per_page,  # Calculate total pages
#                         "total_records": total_records,
#                         "all_profile_ids": all_profile_ids
#                     }

#                     return JsonResponse({"Status": 1, "message": "Fetched viewed profile lists successfully", "data": combined_data, "viewed_profile_count": total_records}, status=status.HTTP_200_OK)
#                 else:
#                     return JsonResponse({"Status": 0, "message": "No viewed profiles found for the given profile ID"}, status=status.HTTP_200_OK)
#             except Profile_visitors.DoesNotExist:
#                 return JsonResponse({"Status": 0, "message": "No viewed profiles found for the given profile ID"}, status=status.HTTP_200_OK)
#         else:
#             return JsonResponse(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class My_profiles_vistors(APIView):
    def post(self, request):
        serializer = Profile_idValidationSerializer(data=request.data)
        
        if serializer.is_valid():
            profile_id = serializer.validated_data.get('profile_id')
            page = int(request.data.get('page_number', 1))
            per_page = int(request.data.get('per_page', 10)) 

            # Extract from_date and to_date from the request
            from_date = request.data.get('from_date')
            to_date = request.data.get('to_date')

            try:
                # Initialize the base queryset to filter by profile_id
                all_profiles = Profile_visitors.objects.filter(profile_id=profile_id)

                # Apply date filters if from_date and to_date are provided
                if from_date:
                    from_date = datetime.strptime(from_date, '%Y-%m-%d').date()  # Convert to date
                    all_profiles = all_profiles.filter(datetime__date__gte=from_date)

                if to_date:
                    to_date = datetime.strptime(to_date, '%Y-%m-%d').date()  # Convert to date
                    all_profiles = all_profiles.filter(datetime__date__lte=to_date)

                # Get all profile IDs in the filtered queryset
                all_profile_ids = {str(index + 1): profile_to for index, profile_to in enumerate(all_profiles.values_list('viewed_profile', flat=True))}

                total_records = all_profiles.count()

                start = (page - 1) * per_page
                end = start + per_page

                # Fetch paginated data
                fetch_data = all_profiles[start:end]

                if fetch_data.exists():
                    profile_ids = fetch_data.values_list('viewed_profile', flat=True)
                    profile_details = get_profile_details(profile_ids)

                    profile_data = Registration1.objects.get(ProfileId=profile_id)
                    horo_data = ProfileHoroscope.objects.get(profile_id=profile_id)

                    my_star_id = horo_data.birthstar_name
                    my_rasi_id = horo_data.birth_rasi_name
                    my_gender = profile_data.Gender

                    restricted_profile_details = [
                        {
                            "visited_profileid": detail.get("ProfileId"),
                            "visited_profile_name": detail.get("Profile_name"),
                            "visited_Profile_img": Get_profile_image(detail.get("ProfileId"), my_gender, 1, 0),
                            "visited_profile_age": calculate_age(detail.get("Profile_dob")),
                            "visited_verified": detail.get("Profile_verified"),
                            "visited_height": detail.get("Profile_height"),
                            "visited_star": detail.get("star_name"),
                            "visited_profession": getprofession(detail.get("profession")),
                            "visited_city": detail.get("Profile_city"),
                            "visited_degree": " ",
                            "visited_match_score": Get_matching_score(my_star_id, my_rasi_id, detail.get("birthstar_name"), detail.get("birth_rasi_name"), my_gender),
                            "visited_views": count_records(Profile_visitors, {'status': 1, 'viewed_profile': detail.get("ProfileId")}),
                            "visited_lastvisit": get_user_statusandlastvisit(detail.get("Last_login_date"))[0],
                            "visited_userstatus": get_user_statusandlastvisit(detail.get("Last_login_date"))[1],
                            "visited_horoscope": "Horoscope Available" if detail.get("horoscope_file") else "Horoscope Not Available",
                            "visited_profile_wishlist": Get_wishlist(profile_id, detail.get("ProfileId")),
                        }
                        for detail in profile_details
                    ]

                    combined_data = {
                        "profiles": restricted_profile_details,
                        "page": page,
                        "per_page": per_page,
                        "total_pages": (total_records + per_page - 1) // per_page,  # Calculate total pages
                        "total_records": total_records,
                        "all_profile_ids": all_profile_ids
                    }

                    return JsonResponse({"Status": 1, "message": "Fetched viewed profile lists successfully", "data": combined_data, "viewed_profile_count": total_records}, status=status.HTTP_200_OK)
                else:
                    return JsonResponse({"Status": 0, "message": "No viewed profiles found for the given profile ID"}, status=status.HTTP_200_OK)
            except Profile_visitors.DoesNotExist:
                return JsonResponse({"Status": 0, "message": "No viewed profiles found for the given profile ID"}, status=status.HTTP_200_OK)
        else:
            return JsonResponse(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        

def get_user_statusandlastvisit(lastlogindate):

    now = timezone.now()
    # Convert now to a naive datetime
    now_naive = now.replace(tzinfo=None)
    one_month_ago = now_naive - timedelta(days=30)


    Profile_status_active = ''
    last_login_date=lastlogindate
    last_visit=''

    if last_login_date:
    # Check if the date is the default invalid value
        if last_login_date == '0000-00-00 00:00:00':
            last_login_date = None
            Profile_status_active = "Newly registered"
            # print(last_login_date,'last_login_date0000')
        else:
                print('Hai')
                # if isinstance(last_login_date, str):
                #     print(last_login_date,'last_login_date123')
                try:
                        # Convert string to datetime
                        # print(last_login_date,'last_login_date12345')                          

                        last_visit =lastlogindate.strftime("(%B %d, %Y)") 
                            

                        #last_login_date = datetime.strptime(last_login_date, "%Y-%m-%d %H:%M:%S")


                except ValueError:
                    # print(last_login_date,'8521478523')
                    last_login_date = None
                # elif not isinstance(last_login_date, datetime):
                # print(last_login_date,'878787878')
                last_login_date = None

            # Compare the last_login_date with one_month_ago
                if last_login_date and last_login_date < one_month_ago:
                        Profile_status_active = "In Active User"  # Mark as inactive if last login is older than one month
                else:
                        Profile_status_active = "Active User"
    else:
            Profile_status_active = "Newly registered"  # Handle case where Last_login_date is None or empty


    return last_visit , Profile_status_active


def count_records(model_n, filter_condition):
    """
    Counts records based on the given filter condition.
    
    :param model: The Django model to query.
    :param filter_condition: A dictionary of conditions to filter the records.
    :return: The count of records that match the filter condition.
    """
    # Filter the records based on the condition
    # queryset = model_n.objects.filter(**filter_condition)
    queryset = model_n.objects.filter(**filter_condition)
    # Get the count of the filtered records
    count = queryset.count()
    
    return count


def count_records_forQ(model_n, filter_condition):
    """
    Counts records based on the given filter condition.
    
    :param model: The Django model to query.
    :param filter_condition: A dictionary of conditions to filter the records.
    :return: The count of records that match the filter condition.
    """
    # Filter the records based on the condition
    # queryset = model_n.objects.filter(**filter_condition)
    queryset = model_n.objects.filter(filter_condition)
    # Get the count of the filtered records
    count = queryset.count()
    
    return count


#Photo Request
class Get_photo_request_list(APIView):

    def post(self, request):
        serializer = Profile_idValidationSerializer(data=request.data)

        if serializer.is_valid():
            profile_id = serializer.validated_data.get('profile_id')
            page = int(request.data.get('page_number', 1))
            per_page = int(request.data.get('per_page', 10))  
            
            from_date = request.data.get('from_date')  # Optional filter date
            to_date = request.data.get('to_date')      # Optional filter date
            
            try:
                # Filter based on the 'from_date' and 'to_date' if provided
                filter_conditions = {'profile_to': profile_id, 'status__in': [1, 2, 3]}
                
                # Apply date range filters if provided
                if from_date:
                    from_date = datetime.strptime(from_date, '%Y-%m-%d')  # Assuming 'from_date' is in 'YYYY-MM-DD' format
                    filter_conditions['req_datetime__gte'] = from_date
                
                if to_date:
                    to_date = datetime.strptime(to_date, '%Y-%m-%d')  # Assuming 'to_date' is in 'YYYY-MM-DD' format
                    filter_conditions['req_datetime__lte'] = to_date

                all_profiles = Photo_request.objects.filter(**filter_conditions)
                all_profile_ids = {str(index + 1): profile_id for index, profile_id in enumerate(all_profiles.values_list('profile_to', flat=True))}

                total_records = all_profiles.count()

                start = (page - 1) * per_page
                end = start + per_page
                
                fetch_data = all_profiles[start:end]
                if fetch_data.exists():
                    profile_ids = fetch_data.values_list('profile_from', flat=True)
                    profile_details = get_profile_details(profile_ids)

                    profile_data = Registration1.objects.get(ProfileId=profile_id)

                    horo_data = ProfileHoroscope.objects.get(profile_id=profile_id)
                    my_star_id = horo_data.birthstar_name
                    my_rasi_id = horo_data.birth_rasi_name
                    my_gender = profile_data.Gender

                    restricted_profile_details = [
                        {
                            "req_profileid": detail.get("ProfileId"),
                            "req_profile_name": detail.get("Profile_name"),
                            "req_Profile_img": Get_profile_image(detail.get("ProfileId"), my_gender, 1, 0),
                            "req_profile_age": calculate_age(detail.get("Profile_dob")),
                            "response_message": fetch_data[index].response_message,
                            "req_status": fetch_data[index].status,
                            "req_verified": detail.get('Profile_verified'),
                            "req_height": detail.get("Profile_height"),
                            "req_star": detail.get("star_name"),
                            "req_profession": getprofession(detail.get("profession")),
                            "req_city": detail.get("Profile_city"),
                            "req_degree": get_degree(detail.get("ug_degeree")),
                            "req_match_score": Get_matching_score(my_star_id, my_rasi_id, detail.get("birthstar_name"), detail.get("birth_rasi_name"), my_gender),
                            "req_views": count_records(Profile_visitors, {'status': 1, 'viewed_profile': detail.get("ProfileId")}),
                            "req_lastvisit": get_user_statusandlastvisit(detail.get("Last_login_date"))[0],
                            "req_userstatus": get_user_statusandlastvisit(detail.get("Last_login_date"))[1],
                            "req_horoscope": "Horoscope Available" if detail.get("horoscope_file") else "Horoscope Not Available",
                            "req_profile_wishlist": Get_wishlist(profile_id, detail.get("ProfileId")),
                        }
                        for index, detail in enumerate(profile_details)
                    ]

                    combined_data = {
                        "profiles": restricted_profile_details,
                        "page": page,
                        "per_page": per_page,
                        "total_pages": (total_records + per_page - 1) // per_page,  # Calculate total pages
                        "total_records": total_records,
                        "all_profile_ids": all_profile_ids
                    }

                    return JsonResponse({"Status": 1, "message": "Fetched Photo request and profile details successfully", "data": combined_data, "photoreq_count": total_records}, status=status.HTTP_200_OK)
                else:
                    return JsonResponse({"Status": 0, "message": "No photo request found for the given profile ID"}, status=status.HTTP_200_OK)
            except Photo_request.DoesNotExist:
                return JsonResponse({"Status": 0, "message": "No photo request found for the given profile ID"}, status=status.HTTP_200_OK)
        else:
            return JsonResponse(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


#Vysyassist
class My_vysassist_list(APIView):

    def post(self, request):
        serializer = Profile_idValidationSerializer(data=request.data)

        if serializer.is_valid():
            profile_id = serializer.validated_data.get('profile_id')
            page = int(request.data.get('page_number', 1))
            per_page = int(request.data.get('per_page', 10))  
            from_date = request.data.get('from_date')  
            to_date = request.data.get('to_date')  

            try:
                # Start building the query
                query = Profile_vysassist.objects.filter(profile_from=profile_id, status=1)
                
                # Apply date filtering if the dates are provided in the request
                if from_date:
                    from_date = datetime.strptime(from_date, '%Y-%m-%d')  
                    query = query.filter(req_datetime__gte=from_date)

                if to_date:
                    to_date = datetime.strptime(to_date, '%Y-%m-%d')  
                    query = query.filter(req_datetime__lte=to_date)

                # Get all profiles and filter them by date
                all_profiles = query
                all_profile_ids = {str(index + 1): profile_id for index, profile_id in enumerate(all_profiles.values_list('profile_to', flat=True))}

                total_records = all_profiles.count()

                start = (page - 1) * per_page
                end = start + per_page
                              
                fetch_data = query[start:end]  # Fetch paginated data
                if fetch_data.exists():
                    profile_ids = fetch_data.values_list('profile_to', flat=True)
                    profile_details = get_profile_details(profile_ids)

                    profile_data = Registration1.objects.get(ProfileId=profile_id)

                    horo_data = ProfileHoroscope.objects.get(profile_id=profile_id)

                    my_star_id = horo_data.birthstar_name
                    my_rasi_id = horo_data.birth_rasi_name
                    my_gender = profile_data.Gender

                    restricted_profile_details = [
                        {
                            "vys_assist_id": detail.get("id"),
                            "vys_profileid": detail.get("ProfileId"),
                            "vys_profile_name": detail.get("Profile_name"),
                            "vys_Profile_img": Get_profile_image(detail.get("ProfileId"), my_gender, 1, detail.get("Photo_protection")),
                            "vys_profile_age": calculate_age(detail.get("Profile_dob")),
                            "vys_verified": detail.get("Profile_verified"),
                            "vys_height": detail.get("Profile_height"),
                            "vys_star": detail.get("star_name"),
                            "vys_profession": getprofession(detail.get("profession")),
                            "vys_city": detail.get("Profile_city"),
                            "vys_degree": get_degree(detail.get("ug_degeree")),
                            "vys_match_score": Get_matching_score(my_star_id, my_rasi_id, detail.get("birthstar_name"), detail.get("birth_rasi_name"), my_gender),
                            "vys_views": count_records(Profile_visitors, {'status': 1, 'viewed_profile': detail.get("ProfileId")}),
                            "vys_lastvisit": get_user_statusandlastvisit(detail.get("Last_login_date"))[0],
                            "vys_userstatus": get_user_statusandlastvisit(detail.get("Last_login_date"))[1],
                            "vys_horoscope": "Horoscope Available" if detail.get("horoscope_file") else "Horoscope Not Available",
                            "vys_profile_wishlist": Get_wishlist(profile_id, detail.get("ProfileId")),
                        }
                        for detail in profile_details
                    ]
                    
                    combined_data = {
                        "profiles": restricted_profile_details,
                        "page": page,
                        "per_page": per_page,
                        "total_pages": (total_records + per_page - 1) // per_page,  # Calculate total pages
                        "total_records": total_records,
                        "all_profile_ids": all_profile_ids
                    }

                    return JsonResponse({"Status": 1, "message": "Fetched Vysassist and profile details successfully", "data": combined_data, "vysassist_count": total_records}, status=status.HTTP_200_OK)
                else:
                    return JsonResponse({"Status": 0, "message": "No Vysassist found for the given profile ID"}, status=status.HTTP_200_OK)
            except Profile_vysassist.DoesNotExist:
                return JsonResponse({"Status": 0, "message": "No Vysassist found for the given profile ID"}, status=status.HTTP_200_OK)
        else:
            return JsonResponse(serializer.errors, status=status.HTTP_200_OK)


#Personal Notes
class Get_personal_notes(APIView):

    def post(self, request):
        serializer = Profile_idValidationSerializer(data=request.data)

        if serializer.is_valid():
            profile_id = serializer.validated_data.get('profile_id')
            page = int(request.data.get('page_number', 1))
            per_page = int(request.data.get('per_page', 10))
            
            # Get the from_date and to_date from request (optional)
            from_date_str = request.data.get('from_date')
            to_date_str = request.data.get('to_date')

            # Parse the dates if provided
            try:
                if from_date_str:
                    from_date = datetime.strptime(from_date_str, "%Y-%m-%d")
                else:
                    from_date = None

                if to_date_str:
                    to_date = datetime.strptime(to_date_str, "%Y-%m-%d")
                else:
                    to_date = None
            except ValueError:
                return JsonResponse({"Status": 0, "message": "Invalid date format. Use YYYY-MM-DD."}, status=status.HTTP_400_BAD_REQUEST)

            try:
                # Filter profiles by profile_id and status=1
                all_profiles = Profile_personal_notes.objects.filter(profile_id=profile_id)

                # Apply additional date filters if provided
                if from_date:
                    all_profiles = all_profiles.filter(datetime__gte=from_date)
                if to_date:
                    all_profiles = all_profiles.filter(datetime__lte=to_date)

                all_profile_ids = {str(index + 1): profile_id for index, profile_id in enumerate(all_profiles.values_list('profile_id', flat=True))}

                total_records = all_profiles.count()

                # Pagination logic
                start = (page - 1) * per_page
                end = start + per_page
                fetch_data = all_profiles[start:end]

                if fetch_data.exists():
                    profile_ids = fetch_data.values_list('profile_id', flat=True)
                    profile_details = get_profile_details(profile_ids)

                    profile_data = Registration1.objects.get(ProfileId=profile_id)
                    horo_data = ProfileHoroscope.objects.get(profile_id=profile_id)

                    my_star_id = horo_data.birthstar_name
                    my_rasi_id = horo_data.birth_rasi_name
                    my_gender = profile_data.Gender

                    personal_notes = fetch_data.values_list('profile_id', 'notes', 'datetime')

                    notes_mapping = {profile_id: (notes, datetime) for profile_id, notes, datetime in personal_notes}

                    restricted_profile_details = [
                        {
                            "notes_profileid": detail.get("ProfileId"),
                            "notes_profile_name": detail.get("Profile_name"),
                            "notes_Profile_img": Get_profile_image(detail.get("ProfileId"), my_gender, 1, 0),
                            "notes_profile_age": calculate_age(detail.get("Profile_dob")),
                            "notes_details": notes_mapping.get(detail.get("ProfileId"), ('notes', ''))[0],
                            "notes_datetime": notes_mapping.get(detail.get("ProfileId"), ('datetime', ''))[1],
                            "notes_verified": detail.get("Profile_verified"),
                            "notes_height": detail.get("Profile_height"),
                            "notes_star": detail.get("star_name"),
                            "notes_profession": getprofession(detail.get("profession")),
                            "notes_city": detail.get("Profile_city"),
                            "notes_degree": get_degree(detail.get("ug_degeree")),
                            "notes_match_score": Get_matching_score(my_star_id, my_rasi_id, detail.get("birthstar_name"), detail.get("birth_rasi_name"), my_gender),
                            "notes_views": count_records(Profile_visitors, {'status': 1, 'viewed_profile': detail.get("ProfileId")}),
                            "notes_lastvisit": get_user_statusandlastvisit(detail.get("Last_login_date"))[0],
                            "notes_userstatus": get_user_statusandlastvisit(detail.get("Last_login_date"))[1],
                            "notes_horoscope": "Horoscope Available" if detail.get("horoscope_file") else "Horoscope Not Available",
                            "notes_profile_wishlist": Get_wishlist(profile_id, detail.get("ProfileId")),
                        }
                        for detail in profile_details
                    ]

                    combined_data = {
                        "profiles": restricted_profile_details,
                        "page": page,
                        "per_page": per_page,
                        "total_pages": (total_records + per_page - 1) // per_page,
                        "total_records": total_records,
                        "all_profile_ids": all_profile_ids
                    }

                    return JsonResponse({"Status": 1, "message": "Fetched Notes lists successfully", "data": combined_data, "personal_note_count": total_records}, status=status.HTTP_200_OK)
                else:
                    return JsonResponse({"Status": 0, "message": "No Notes found for the given profile ID"}, status=status.HTTP_404_NOT_FOUND)

            except Profile_personal_notes.DoesNotExist:
                return JsonResponse({"Status": 0, "message": "No Notes found for the given profile ID"}, status=status.HTTP_404_NOT_FOUND)
        else:
            return JsonResponse(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


#Express Interest sent
class Exp_intrests_list(APIView):

    def post(self, request):
        serializer = Profile_idValidationSerializer(data=request.data)

        if serializer.is_valid():
            profile_id = serializer.validated_data.get('profile_id')
            page = int(request.data.get('page_number', 1))
            per_page = int(request.data.get('per_page', 10))  

            try:
                # Base query for Express Interests
                query = Express_interests.objects.filter(profile_from=profile_id, status=1)

                # # Filter by date if provided
                # if from_date:
                #     from_date = datetime.strptime(from_date, "%Y-%m-%d")  # Convert string to date
                #     query = query.filter(req_datetime__gte=from_date)

                # if to_date:
                #     to_date = datetime.strptime(to_date, "%Y-%m-%d")  # Convert string to date
                #     query = query.filter(req_datetime__lte=to_date)

                # Now, create the dictionary of all profile IDs.
                all_profile_ids = {str(index + 1): profile_id for index, profile_id in enumerate(query.values_list('profile_to', flat=True))}

                # Get the total number of records.
                total_records = query.count()

                start = (page - 1) * per_page
                end = start + per_page
                
                fetch_data = query[start:end]
                if fetch_data.exists():
                    profile_ids = fetch_data.values_list('profile_to', flat=True)
                    profile_details = get_profile_details(profile_ids)

                    profile_data = Registration1.objects.get(ProfileId=profile_id)

                    horo_data = ProfileHoroscope.objects.get(profile_id=profile_id)

                    my_star_id = horo_data.birthstar_name
                    my_rasi_id = horo_data.birth_rasi_name

                    my_gender = profile_data.Gender

                    restricted_profile_details = [
                        {
                            "myint_profileid": detail.get("ProfileId"),
                            "myint_profile_name": detail.get("Profile_name"),
                            "myint_Profile_img": Get_profile_image(detail.get("ProfileId"), my_gender, 1, 0),
                            "myint_profile_age": calculate_age(detail.get("Profile_dob")),
                            "myint_verified": detail.get("Profile_verified"),
                            "myint_height": detail.get("Profile_height"),
                            "myint_star": detail.get("star_name"),
                            "myint_profession": getprofession(detail.get("profession")),
                            "myint_city": detail.get("Profile_city"),
                            "myint_degree": get_degree(detail.get("ug_degeree")),
                            "myint_match_score": Get_matching_score(my_star_id, my_rasi_id, detail.get("birthstar_name"), detail.get("birth_rasi_name"), my_gender),
                            "myint_views": count_records(Profile_visitors, {'status': 1, 'viewed_profile': detail.get("ProfileId")}),
                            "myint_lastvisit": get_user_statusandlastvisit(detail.get("Last_login_date"))[0],
                            "myint_userstatus": get_user_statusandlastvisit(detail.get("Last_login_date"))[1],
                            "myint_horoscope": "Horoscope Available" if detail.get("horoscope_file") else "Horoscope Not Available",
                            "myint_profile_wishlist": Get_wishlist(profile_id, detail.get("ProfileId")),
                        }
                        for detail in profile_details
                    ]

                    combined_data = {
                        "profiles": restricted_profile_details,
                        "page": page,
                        "per_page": per_page,
                        "total_pages": (total_records + per_page - 1) // per_page,  # Calculate total pages
                        "total_records": total_records,
                        "all_profile_ids": all_profile_ids
                    }

                    return JsonResponse({"Status": 1, "message": "Fetched interests and profile details successfully", "data": combined_data, "myint_count": total_records}, status=status.HTTP_200_OK)
                else:
                    return JsonResponse({"Status": 0, "message": "No interests found for the given profile ID"}, status=status.HTTP_200_OK)
            except Express_interests.DoesNotExist:
                return JsonResponse({"Status": 0, "message": "No interests found for the given profile ID"}, status=status.HTTP_200_OK)
        else:
            return JsonResponse(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

#Express Interest sent
class Exp_intrests_received(APIView):

    def post(self, request):
        serializer = Profile_idValidationSerializer(data=request.data)

        if serializer.is_valid():
            profile_id = serializer.validated_data.get('profile_id')
            page = int(request.data.get('page_number', 1))
            per_page = int(request.data.get('per_page', 10))  

            try:
                # Base query for Express Interests
                query = Express_interests.objects.filter(profile_to=profile_id, status=1)

                # Filter by date if provided
                # if from_date:
                #     from_date = datetime.strptime(from_date, "%Y-%m-%d")  # Convert string to date
                #     query = query.filter(req_datetime__gte=from_date)

                # if to_date:
                #     to_date = datetime.strptime(to_date, "%Y-%m-%d")  # Convert string to date
                #     query = query.filter(req_datetime__lte=to_date)

                # Now, create the dictionary of all profile IDs.
                all_profile_ids = {str(index + 1): profile_id for index, profile_id in enumerate(query.values_list('profile_from', flat=True))}

                # Get the total number of records.
                total_records = query.count()

                start = (page - 1) * per_page
                end = start + per_page
                
                fetch_data = query[start:end]
                if fetch_data.exists():
                    profile_ids = fetch_data.values_list('profile_from', flat=True)
                    profile_details = get_profile_details(profile_ids)

                    profile_data = Registration1.objects.get(ProfileId=profile_id)

                    horo_data = ProfileHoroscope.objects.get(profile_id=profile_id)

                    my_star_id = horo_data.birthstar_name
                    my_rasi_id = horo_data.birth_rasi_name

                    my_gender = profile_data.Gender

                    restricted_profile_details = [
                        {
                            "myint_profileid": detail.get("ProfileId"),
                            "myint_profile_name": detail.get("Profile_name"),
                            "myint_Profile_img": Get_profile_image(detail.get("ProfileId"), my_gender, 1, 0),
                            "myint_profile_age": calculate_age(detail.get("Profile_dob")),
                            "myint_verified": detail.get("Profile_verified"),
                            "myint_height": detail.get("Profile_height"),
                            "myint_star": detail.get("star_name"),
                            "myint_profession": getprofession(detail.get("profession")),
                            "myint_city": detail.get("Profile_city"),
                            "myint_degree": get_degree(detail.get("ug_degeree")),
                            "myint_match_score": Get_matching_score(my_star_id, my_rasi_id, detail.get("birthstar_name"), detail.get("birth_rasi_name"), my_gender),
                            "myint_views": count_records(Profile_visitors, {'status': 1, 'viewed_profile': detail.get("ProfileId")}),
                            "myint_lastvisit": get_user_statusandlastvisit(detail.get("Last_login_date"))[0],
                            "myint_userstatus": get_user_statusandlastvisit(detail.get("Last_login_date"))[1],
                            "myint_horoscope": "Horoscope Available" if detail.get("horoscope_file") else "Horoscope Not Available",
                            "myint_profile_wishlist": Get_wishlist(profile_id, detail.get("ProfileId")),
                        }
                        for detail in profile_details
                    ]

                    combined_data = {
                        "profiles": restricted_profile_details,
                        "page": page,
                        "per_page": per_page,
                        "total_pages": (total_records + per_page - 1) // per_page,  # Calculate total pages
                        "total_records": total_records,
                        "all_profile_ids": all_profile_ids
                    }

                    return JsonResponse({"Status": 1, "message": "Fetched interests and profile details successfully", "data": combined_data, "myint_count": total_records}, status=status.HTTP_200_OK)
                else:
                    return JsonResponse({"Status": 0, "message": "No interests found for the given profile ID"}, status=status.HTTP_200_OK)
            except Express_interests.DoesNotExist:
                return JsonResponse({"Status": 0, "message": "No interests found for the given profile ID"}, status=status.HTTP_200_OK)
        else:
            return JsonResponse(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        

class Exp_intrests_mutual(APIView):

    def post(self, request):
        serializer = Profile_idValidationSerializer(data=request.data)

        if serializer.is_valid():
            profile_id = serializer.validated_data.get('profile_id')
            page = int(request.data.get('page_number', 1))
            per_page = int(request.data.get('per_page', 10)) 

            try:
                all_profiles = Express_interests.objects.filter(
                    (Q(profile_from=profile_id) | Q(profile_to=profile_id)) & Q(status=2)
                )

                # Get both profile_from and profile_to IDs, and exclude the current profile_id
                profile_to_ids = all_profiles.values_list('profile_to', flat=True)
                profile_from_ids = all_profiles.values_list('profile_from', flat=True)

                # Combine and exclude the current profile_id
                all_profile_ids = set(profile_to_ids) | set(profile_from_ids)
                # all_profile_ids_1 = {str(index + 1): pid for index, pid in enumerate(all_profile_ids) if pid != profile_id}
                # all_profile_ids_1 = {str(i + 1): pid for i=0, pid in enumerate(all_profile_ids) if pid != profile_id}
                #all_profile_ids_1 = {str(i + 1): pid for i, pid in enumerate(all_profile_ids) if pid != profile_id}
                # all_profile_ids_1 = {str(i + 1): pid for i, pid in enumerate(all_profile_ids) if pid != profile_id}
                all_profile_ids_1 = {str(index + 1): pid for index, pid in enumerate([pid for pid in all_profile_ids if pid != profile_id])}



                total_records=len(all_profile_ids_1)
                start = (page - 1) * per_page
                end = start + per_page

                #fetch_data = models.Express_interests.objects.filter(profile_from=profile_id , profile_to=profile_id)
                fetch_data = Express_interests.objects.filter(
                    (Q(profile_from=profile_id) | Q(profile_to=profile_id)) &  Q(status=2))[start:end]

                if fetch_data.exists():
                    #profile_ids = fetch_data.values_list('profile_to', flat=True)
                    
                                                            
                    # Get profile_to IDs
                    profile_to_ids = fetch_data.values_list('profile_to', flat=True)

                    # Get profile_from IDs
                    profile_from_ids = fetch_data.values_list('profile_from', flat=True)

                    # Combine both sets of IDs
                    all_profile_ids = set(profile_to_ids) | set(profile_from_ids)

                    # Exclude the current profile_id
                    profile_ids = [pid for pid in all_profile_ids if pid != profile_id]
                                        
                    
                    
                    profile_details = get_profile_details(profile_ids)


                    profile_data =  Registration1.objects.get(ProfileId=profile_id)

                    horo_data=ProfileHoroscope.objects.get(profile_id=profile_id)


                    my_star_id=horo_data.birthstar_name
                    my_rasi_id=horo_data.birth_rasi_name
            
                    my_gender=profile_data.Gender

                    
                    # mutual_condition = Q(status=2) & (Q(profile_from=profile_id) | Q(profile_to=profile_id))
                    # mutual_int_count = count_records_forQ(models.Express_interests, mutual_condition)
                    
                    restricted_profile_details = [
                        {
                            "mutint_profileid": detail.get("ProfileId"),
                            "mutint_profile_name": detail.get("Profile_name"),
                            "mutint_Profile_img":  Get_profile_image(detail.get("ProfileId"),my_gender,1,detail.get("Photo_protection")),                           
                            "mutint_profile_age": calculate_age(detail.get("Profile_dob")),
                            "mutint_verified":detail.get("Profile_verified"),
                            "mutint_height":detail.get("Profile_height"),
                            "mutint_star":detail.get("star_name"),
                            "mutint_profession":getprofession(detail.get("profession")),
                            "mutint_city":detail.get("Profile_city"),
                            "mutint_degree":get_degree(detail.get("ug_degeree")),
                            "mutint_match_score":Get_matching_score(my_star_id,my_rasi_id,detail.get("birthstar_name"),detail.get("birth_rasi_name"),my_gender),
                            "mutint_views":count_records(Profile_visitors, {'status': 1,'viewed_profile':detail.get("ProfileId")}),
                            "mutint_lastvisit": get_user_statusandlastvisit(detail.get("Last_login_date"))[0],
                            "mutint_userstatus": get_user_statusandlastvisit(detail.get("Last_login_date"))[1],
                            "mutint_horoscope": "Horoscope Available" if detail.get("horoscope_file") else "Horoscope Not Available",
                            "mutint_profile_wishlist":Get_wishlist(profile_id,detail.get("ProfileId")),
                        }
                        for detail in profile_details
                    ]
                    
                    #serialized_fetch_data = serializers.ExpressintrSerializer(fetch_data, many=True).data
                    #serialized_profile_details = serializers.ProfileDetailsSerializer(profile_details, many=True).data

                    combined_data = {
                        #"interests": serialized_fetch_data,
                        "profiles": restricted_profile_details,
                        "page": page,
                        "per_page": per_page,
                        "total_pages": (total_records + per_page - 1) // per_page,  # Calculate total pages
                        "total_records": total_records,
                        "all_profile_ids":all_profile_ids_1,
                        "page_id":4
                    }

                    return JsonResponse({"Status": 1, "message": "Fetched interests and profile details successfully", "data": combined_data,"mut_int_count":total_records}, status=status.HTTP_200_OK)
                else:
                    return JsonResponse({"Status": 0, "message": "No interests found for the given profile ID"}, status=status.HTTP_200_OK)
            except Express_interests.DoesNotExist:
                return JsonResponse({"Status": 0, "message": "No interests found for the given profile ID"}, status=status.HTTP_200_OK)
        else:
            return JsonResponse(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


        
#Fetch Login Details
@api_view(['GET'])
def fetch_login_details(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    # If both dates are provided, parse them, else fetch all records
    if from_date and to_date:
        try:
            from_date = datetime.strptime(from_date, "%Y-%m-%d")
            to_date = datetime.strptime(to_date, "%Y-%m-%d")
            login_logs = ProfileLoginLogs.objects.filter(login_datetime__range=[from_date, to_date])
        except ValueError:
            return JsonResponse({'error': 'Invalid date format. Use YYYY-MM-DD'}, status=400)
    else:
        # Fetch all records if no date range is provided
        login_logs = ProfileLoginLogs.objects.all()

    paginator = StandardResultsPaging()
    paginated_logs = paginator.paginate_queryset(login_logs, request)

    # Prepare response data
    response_data = []
    for log in paginated_logs:
        profile = get_object_or_404(LoginDetails, ProfileId=log.profile_id)

        # Check if the city is an ID or a name
        city_value = profile.Profile_city
        
        if city_value.isdigit():  # Check if it's a numeric ID
            # Fetch city by ID
            city = get_object_or_404(City, id=int(city_value), is_deleted=False)
            city_name = city.city_name
        else:
            # Otherwise, it's a city name, so use it directly
            city_name = city_value

      
        

        response_data.append({
            'ContentId': profile.ContentId,
            'ProfileId': profile.ProfileId,
            'Name': profile.Profile_name,
            'City': city_name,  # Use the fetched city name
            'Email': profile.EmailId,
            'LastLoginDate': log.login_datetime.strftime('%Y-%m-%d %H:%M:%S'),
            'Status': profile.status
        })

    return paginator.get_paginated_response(response_data)

@api_view(['GET'])
def fetch_login_details_profile(request):
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    profile_id = request.GET.get('profile_id')
    
    # If both dates are provided, parse them, else fetch all records
    if from_date and to_date:
        try:
            from_date = datetime.strptime(from_date, "%Y-%m-%d")
            to_date = datetime.strptime(to_date, "%Y-%m-%d")
            login_logs = ProfileLoginLogs.objects.filter(login_datetime__range=[from_date, to_date],profile_id=profile_id)
        except ValueError:
            return JsonResponse({'error': 'Invalid date format. Use YYYY-MM-DD'}, status=400)
    else:
        # Fetch all records if no date range is provided
        login_logs = ProfileLoginLogs.objects.filter(profile_id=profile_id)

    paginator = StandardResultsPaging()
    paginated_logs = paginator.paginate_queryset(login_logs, request)

    # Prepare response data
    response_data = []
    for log in paginated_logs:
        profile = get_object_or_404(LoginDetails, ProfileId=log.profile_id)

        # Check if the city is an ID or a name
        city_value = profile.Profile_city
        
        if city_value.isdigit():  # Check if it's a numeric ID
            # Fetch city by ID
            city = get_object_or_404(City, id=int(city_value), is_deleted=False)
            city_name = city.city_name
        else:
            # Otherwise, it's a city name, so use it directly
            city_name = city_value

      
        

        response_data.append({
            'ContentId': profile.ContentId,
            'ProfileId': profile.ProfileId,
            'Name': profile.Profile_name,
            'City': city_name,  # Use the fetched city name
            'Email': profile.EmailId,
            'LastLoginDate': log.login_datetime.strftime('%Y-%m-%d %H:%M:%S'),
            'Status': profile.status
        })

    return paginator.get_paginated_response(response_data)



class ProfileSendTo(APIView):
    def post(self, request, profile_from_id):
        try:
            # Fetch all records from ProfileSendFromAdmin for profile_from_id
            send_from_admin_list = ProfileSendFromAdmin.objects.filter(profile_from=profile_from_id)
            if not send_from_admin_list.exists():
                return JsonResponse({"status": "error", "message": "Profile not found in ProfileSendFromAdmin"}, status=404)

            profiles_data = []

            for send_from_admin in send_from_admin_list:
                profile_to_id = send_from_admin.profile_to  

                # Get details from LoginDetails
                login_details = LoginDetails.objects.get(ProfileId=profile_to_id)
                age = calculate_age(login_details.Profile_dob)

                # Determine the city name
                if login_details.Profile_city.isdigit():
                    try:
                        city = City.objects.get(id=int(login_details.Profile_city))
                        profile_city = city.city_name
                    except City.DoesNotExist:
                        profile_city = None
                else:
                    profile_city = login_details.Profile_city

                # Get details from ProfileEduDetails
                edu_details = ProfileEduDetails.objects.get(profile_id=profile_to_id)

                # Determine the highest education level
                if edu_details.highest_education.isdigit():
                    try:
                        education_level = EducationLevel.objects.get(row_id=int(edu_details.highest_education))
                        highest_education = education_level.EducationLevel
                    except EducationLevel.DoesNotExist:
                        highest_education = None
                else:
                    highest_education = edu_details.highest_education

                # Determine profession
                if edu_details.profession.isdigit():
                    try:
                        profession_obj = Profession.objects.get(row_id=int(edu_details.profession))
                        profession = profession_obj.profession
                    except Profession.DoesNotExist:
                        profession = None
                else:
                    profession = edu_details.profession

                # Determine annual income
                if edu_details.anual_income.isdigit():
                    try:
                        income_obj = AnnualIncome.objects.get(pk=int(edu_details.anual_income))
                        anual_income = income_obj.income
                    except AnnualIncome.DoesNotExist:
                        anual_income = None
                else:
                    anual_income = edu_details.anual_income

                # Get details from ProfileFamilyDetails
                family_details = ProfileFamilyDetails.objects.get(profile_id=profile_to_id)

                # Get details from ProfileHoroscope
                horoscope_details = ProfileHoroscope.objects.get(profile_id=profile_to_id)

                if horoscope_details.birthstar_name.isdigit():
                    try:
                        birthstar_obj = BirthStar.objects.get(pk=int(horoscope_details.birthstar_name))
                        birthstar_name = birthstar_obj.star
                    except BirthStar.DoesNotExist:
                        birthstar_name = None
                else:
                    birthstar_name = horoscope_details.birthstar_name

                # Get status from the current ProfileSendFromAdmin record
                status = send_from_admin.status

                # Structure the data for this profile
                profile_data = {
                    'ContentId': login_details.ContentId,
                    'ProfileId': login_details.ProfileId,
                    'Profile_name': login_details.Profile_name,
                    'Age': age,
                    'Profile_city': profile_city,
                    'highest_education': highest_education,
                    'profession': profession,
                    'anual_income': anual_income,
                    'suya_gothram': family_details.suya_gothram,
                    'birthstar_name': birthstar_name,
                    'status': status,
                }

                # Append the profile data to the list
                profiles_data.append(profile_data)

            # Return the response as JSON with all profiles data
            return JsonResponse({
                'status': 'success',
                'message': 'Profiles sent to fetch successfully',
                'data': profiles_data
            })

        except LoginDetails.DoesNotExist:
            return JsonResponse({"status": "error", "message": "Profile not found in LoginDetails"}, status=404)
        except ProfileEduDetails.DoesNotExist:
            return JsonResponse({"status": "error", "message": "Profile not found in ProfileEduDetails"}, status=404)
        except ProfileFamilyDetails.DoesNotExist:
            return JsonResponse({"status": "error", "message": "Profile not found in ProfileFamilyDetails"}, status=404)
        except ProfileHoroscope.DoesNotExist:
            return JsonResponse({"status": "error", "message": "Profile not found in ProfileHoroscope"}, status=404)
        except (AnnualIncome.DoesNotExist, BirthStar.DoesNotExist):
            return JsonResponse({"status": "error", "message": "Data not found in master table"}, status=404)
        


class ProfileVysAssistFollowupSerializer(serializers.ModelSerializer):
    class Meta:
        model = ProfileVysAssistFollowup
        fields = '__all__'

    def validate(self, data):
        required_fields = ['assist_id', 'owner_id', 'comments']
        for field in required_fields:
            if field not in data or data[field] in [None, '']:
                raise serializers.ValidationError({field: f"{field} is required."})
        return data
    
# List and Create API
class ProfileVysAssistFollowupListCreateView(generics.ListCreateAPIView):
    serializer_class = ProfileVysAssistFollowupSerializer

    def get_queryset(self):
        """
        Fetch data by assist_id in descending order of update_at.
        """
        assist_id = self.request.query_params.get('assist_id', None)
        queryset = ProfileVysAssistFollowup.objects.all().order_by('-update_at')

        if assist_id:
            queryset = queryset.filter(assist_id=assist_id)

        return queryset

    def create(self, request, *args, **kwargs):
        """
        Ensure required fields are present.
        """
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            self.perform_create(serializer)
            return JsonResponse(serializer.data, status=status.HTTP_201_CREATED)
        return JsonResponse(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# Retrieve, Update, Delete API
class ProfileVysAssistFollowupRetrieveUpdateDeleteView(generics.RetrieveUpdateDestroyAPIView):
    queryset = ProfileVysAssistFollowup.objects.all()
    serializer_class = ProfileVysAssistFollowupSerializer


#Call action sent by profile
class CallactionSent(APIView):
    pagination_class = StandardResultsPaging

    def get(self, request):
        profile_id = request.query_params.get('profile_id')

        if not profile_id :
            return Response({"error": "Profile_id is required"}, status=400)

        # Fetch profile visitors within the given date range
        profile_call_sent = Profile_callogs.objects.filter(
           profile_from=profile_id
        )

        if not profile_call_sent.exists():
            return Response({"message": "No call action found for the profile id"}, status=404)

        # Create a result list to include profile information
        result = []

        for callsent in profile_call_sent:
            # Fetch profile_id (the user who viewed the profile)
            profile_call = LoginDetails.objects.filter(ProfileId=callsent.profile_to).first()


            # Only add to the result if both profile_call and viewed_profile exist
            if profile_call:
                # Fetch plan_name from PlanDetails based on Plan_id
                profile_call_plan = PlanDetails.objects.filter(id=profile_call.Plan_id).first()

                # Fetch mode_name from Mode table based on Profile_for
                profile_call_mode = Mode.objects.filter(mode=profile_call.Profile_for).first()

                profile_call_city_name=get_city_name(profile_call.Profile_city)
                profile_call_state_name=get_state_name(profile_call.Profile_state)

                # Fetch profile_call city name
                result.append({
                    'profile_id': profile_call.ProfileId,
                    'profile_name': profile_call.Profile_name,
                    'profile_dob': profile_call.Profile_dob.isoformat() if profile_call.Profile_dob else None,
                    'profile_state': profile_call_state_name,
                    'profile_city': profile_call_city_name,  # City name resolved
                    'profile_mobile': profile_call.Mobile_no,
                    'profile_gender': profile_call.Gender,
                    'profile_planname': profile_call_plan.plan_name if profile_call_plan else None,  # Plan name
                    'profile_created_by': profile_call_mode.mode_name if profile_call_mode else None,  # Mode name
                      # State name

                })

        # Implement pagination if necessary
        paginator = self.pagination_class()
        paginated_result = paginator.paginate_queryset(result, request)

        # If there are paginated results, return the paginated response
        if paginated_result is not None:
            return paginator.get_paginated_response(paginated_result)

        # If no pagination is needed, return the full result set
        return Response(result, status=200)
    

#Call action sent by profile
class CallactionReceived(APIView):
    
    pagination_class = StandardResultsPaging

    def get(self, request):
        profile_id = request.query_params.get('profile_id')

        if not profile_id :
            return Response({"error": "Profile_id is required"}, status=400)

        # Fetch profile visitors within the given date range
        profile_call_sent = Profile_callogs.objects.filter(
           profile_to=profile_id
        )

        if not profile_call_sent.exists():
            return Response({"message": "No call action found for the provider id"}, status=404)

        # Create a result list to include profile information
        result = []

        for callsent in profile_call_sent:
            # Fetch profile_id (the user who viewed the profile)
            profile_call = LoginDetails.objects.filter(ProfileId=callsent.profile_from).first()


            # Only add to the result if both profile_call and viewed_profile exist
            if profile_call:
                # Fetch plan_name from PlanDetails based on Plan_id
                profile_call_plan = PlanDetails.objects.filter(id=profile_call.Plan_id).first()

                # Fetch mode_name from Mode table based on Profile_for
                profile_call_mode = Mode.objects.filter(mode=profile_call.Profile_for).first()

                profile_call_city_name=get_city_name(profile_call.Profile_city)
                profile_call_state_name=get_state_name(profile_call.Profile_state)

                # Fetch profile_call city name
                result.append({
                    'profile_id': profile_call.ProfileId,
                    'profile_name': profile_call.Profile_name,
                    'profile_dob': profile_call.Profile_dob.isoformat() if profile_call.Profile_dob else None,
                    'profile_state': profile_call_state_name,
                    'profile_city': profile_call_city_name,  # City name resolved
                    'profile_mobile': profile_call.Mobile_no,
                    'profile_gender': profile_call.Gender,
                    'profile_planname': profile_call_plan.plan_name if profile_call_plan else None,  # Plan name
                    'profile_created_by': profile_call_mode.mode_name if profile_call_mode else None,  # Mode name
                      # State name
                })

        # Implement pagination if necessary
        paginator = self.pagination_class()
        paginated_result = paginator.paginate_queryset(result, request)

        # If there are paginated results, return the paginated response
        if paginated_result is not None:
            return paginator.get_paginated_response(paginated_result)

        # If no pagination is needed, return the full result set
        return Response(result, status=200)
    

# Matchingprint profile
class Matchingprintprofile(APIView):
    def post(self, request, *args, **kwargs):
        format = request.data.get('format')
        profile_ids = request.data.get('profile_id')
        to_profile_id = request.data.get('to_profile_id')  # Single recipient profile ID


        if not format:
            return JsonResponse({"status": "error", "message": "format is required"}, status=400)
        if not profile_ids:
            return JsonResponse({"status": "error", "message": "profile_id is required"}, status=400)
        if not to_profile_id:
            return JsonResponse({"status": "error", "message": "to_profile_id is required"}, status=400)

        try:
            if format == "fullprofile":
                pdf_response = SendFullProfilePrintPDF().post(request)

            elif format == "withoutaddress":
                pdf_response = WithoutAddressPrintPDF().post(request)

            elif format == "shortprofile":
                pdf_response = SendShortProfilePrintPDF().post(request)

            else:
                return JsonResponse({"status": "error", "message": "Invalid format type"}, status=400)

            # Check if the response contains PDF content
            if pdf_response.status_code == 200 and pdf_response.get('Content-Type') == 'application/pdf':
                return pdf_response  # Return the PDF file directly

            return JsonResponse({"status": "error", "message": "PDF generation failed"}, status=500)

        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=500)



class Partnersettings(APIView):
    def post(self, request):
        profile_id=request.data.get('profile_id')

        # Validate profile_id
        if not profile_id:
            return JsonResponse({"status": "error", "message": "Profile ID is required"}, status=400)
        
        try:
            partner_pref = ProfilePartnerPref.objects.get(profile_id=profile_id)

            # Convert model instance to dictionary with all fields
            partner_data = model_to_dict(partner_pref)

            return JsonResponse({"status": "success", "data": partner_data}, status=200)

        except ObjectDoesNotExist:
            return JsonResponse({"status": "error", "message": "No partner settings found for this profile ID"}, status=404)
        
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=500)
        
class Suggestsettings(APIView):
    def post(self, request):
        profile_id=request.data.get('profile_id')

        # Validate profile_id
        if not profile_id:
            return JsonResponse({"status": "error", "message": "Profile ID is required"}, status=400)
        
        try:
            partner_pref = ProfilePartnerPref.objects.get(profile_id=profile_id)

            # Convert model instance to dictionary with all fields
            partner_data = model_to_dict(partner_pref)

            return JsonResponse({"status": "success", "data": partner_data}, status=200)

        except ObjectDoesNotExist:
            return JsonResponse({"status": "error", "message": "No partner settings found for this profile ID"}, status=404)
        
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=500)

#Matchingwhatsapp profile
class Matchingwhatsapp(APIView):
    def post(self, request, *args, **kwargs):
        format = request.data.get('format')
        profile_ids = request.data.get('profile_id')
        to_profile_id = request.data.get('to_profile_id')  # Single recipient profile ID
        action_type=request.data.get('action_type')

        if not format:
            return JsonResponse({"status": "error", "message": "format is required"}, status=400)
        if not profile_ids:
            return JsonResponse({"status": "error", "message": "profile_id is required"}, status=400)
        if not to_profile_id:
            return JsonResponse({"status": "error", "message": "to_profile_id is required"}, status=400)
    
        if not action_type:
            return JsonResponse({"status": "error", "message": "action_type is required"}, status=400)   

        try:
            if format == "fullprofile":
                pdf_response = SendFullProfilePrintPDF().post(request)

            elif format == "withoutaddress":
                pdf_response = WithoutAddressPrintPDF().post(request)

            elif format == "shortprofile":
                pdf_response = SendShortProfilePrintPDF().post(request)

            else:
                return JsonResponse({"status": "error", "message": "Invalid format type"}, status=400)

            # Check if the response contains PDF content
            if pdf_response.status_code == 200 and pdf_response.get('Content-Type') == 'application/pdf':
                return pdf_response  # Return the PDF file directly

            return JsonResponse({"status": "error", "message": "PDF generation failed"}, status=500)

        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=500)



# class Matchingsendemail(APIView):
#     def post(self, request):
#         format=request.data.get('format')
#         profile_ids=request.data.get('profile_ids')
#         if not format :
#              return JsonResponse({"status": "error", "message": "format is required"}, status=404)
#         if not profile_ids :
#              return JsonResponse({"status": "error", "message": "Profile_id is required"}, status=404)


class Matchingsendemail(APIView):
    def post(self, request, *args, **kwargs):
        format = request.data.get('format')
        profile_ids = request.data.get('profile_id')
        to_profile_id = request.data.get('to_profile_id')  # Single recipient profile ID

        if not format:
            return JsonResponse({"status": "error", "message": "format is required"}, status=400)
        if not profile_ids:
            return JsonResponse({"status": "error", "message": "profile_id is required"}, status=400)
        if not to_profile_id:
            return JsonResponse({"status": "error", "message": "to_profile_id is required"}, status=400)

        # Fetch recipient email for to_profile_id
        recipient_email = LoginDetails.objects.filter(ProfileId__iexact=to_profile_id).values_list('EmailId', flat=True).first()

        if not recipient_email:
            return JsonResponse({"status": "error", "message": "No email found for to_profile_id"}, status=400)

        try:
            # Create a mutable copy of request data
            mutable_data = request.data.copy()  
            mutable_data["email"] = recipient_email  

            # Wrap the modified request data into a DRF Request object
            new_request = Request(request._request)  
            new_request._full_data = mutable_data  # Override request data with the modified one

            # Call the appropriate API class
            if format == "fullprofile":
                response = SendFullProfilePDFEmail().post(new_request)

            elif format == "withoutaddress":
                response = WithoutAddressSendEmailAPI().post(new_request)

            elif format == "shortprofile":
                response = SendShortProfilePDFEmail().post(new_request)

            else:
                return JsonResponse({"status": "error", "message": "Invalid format type"}, status=400)

            return response  # Return the response from the called API

        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=500)



        



def get_city_name(city_id):
    try:
        # Attempt to retrieve the city object using the string city_id
        city = City.objects.get(id=city_id)
        return city.city_name  # Return the city name if found
    except City.DoesNotExist:
        return city_id  # Return city_id if the city does not exist
    except Exception as e:
        return city_id 

def get_state_name(state_id):
    try:
        # Attempt to retrieve the city object using the string city_id
        state = State.objects.get(id=state_id)
        return state.name  # Return the city name if found
    except State.DoesNotExist:
        return state_id  # Return city_id if the city does not exist
    except Exception as e:
        return state_id 
    

    #Print short profiles code
def get_country_name(country_id):
    try:
        # Attempt to retrieve the city object using the string city_id
        country = models.Country.objects.get(id=country_id)
        return country.name  # Return the city name if found
    except models.Country.DoesNotExist:
        return country_id  # Return city_id if the city does not exist
    except Exception as e:
        return country_id 


from django.shortcuts import get_object_or_404
from rest_framework.views import APIView
from django.http import HttpResponse
from xhtml2pdf import pisa
from django.template.loader import render_to_string

# Planet mapping dictionary
# planet_mapping = {
#     "1": "Sun",
#     "2": "Moo",
#     "3": "Mar",
#     "4": "Mer",
#     "5": "Jup",
#     "6": "Ven",
#     "7": "Sat",
#     "8": "Rahu",
#     "9": "Kethu",
#     "10": "Lagnam",
# }

planet_mapping = {
    "1": "Sun",
    "2": "Moo",
    "3": "Rahu",
    "4": "Kethu",
    "5": "Mar",
    "6": "Ven",
    "7": "Jup",
    "8": "Mer",
    "9": "Sat",
    "10": "Lagnam",
}


# Define a default placeholder for empty values
default_placeholder = '-'

def GetMarsRahuKethuDoshamDetails(raw_input):

        rasi_grid_data = {}
        pattern = r"Grid (\d+):\s*([\d,]*|empty)"
        matches = re.findall(pattern, raw_input)

        for match in matches:
            grid_number = int(match[0])
            if match[1].lower() == "empty" or match[1].strip() == "":
                rasi_grid_data[f'Grid {grid_number}'] = []
            else:
                rasi_grid_data[f'Grid {grid_number}'] = [
                    int(x) for x in match[1].split(',') if x.strip()
                ]

        planet_mapping = {
                    "1": "Sun",
                    "2": "Moon",
                    "3": "Rahu",
                    "4": "Kethu",
                    "5": "Mars",
                    "6": "Venus",
                    "7": "Jupiter",
                    "8": "Mercury",
                    "9": "Saturn",
                    "10": "Lagnam",
                }

        # Create a grid of 12 cells with mapped planet names
        grid = []
        for i in range(1, 13):
            if f'Grid {i}' in rasi_grid_data:
                planets = [planet_mapping.get(x, '') for x in rasi_grid_data[f'Grid {i}']]
                grid.append(", ".join(planets))
            else:
                grid.append("")

        # Calculation for identifying the positions
        mars_position = None
        rahu_positions = []
        kethu_positions = []
        lagnam_position = None

        for grid_num, planets in rasi_grid_data.items():
            if 5 in planets:  # Mars
                mars_position = int(grid_num.split()[1])
            if 3 in planets:  # Rahu
                rahu_positions.append(int(grid_num.split()[1]))
            if 4 in planets:  # Kethu
                kethu_positions.append(int(grid_num.split()[1]))
            if 10 in planets:  # Lagnam
                lagnam_position = int(grid_num.split()[1])

        def calculate_position(from_position, to_position):
            if from_position is None or to_position is None:
                return None
            if to_position >= from_position:
                return to_position - from_position + 1
            else:
                return 12 - from_position + to_position + 1

        # Calculate positions relative to Lagnam
        rahu_positions_from_lagnam = [
            calculate_position(lagnam_position, pos) for pos in rahu_positions
        ]
        kethu_positions_from_lagnam = [
            calculate_position(lagnam_position, pos) for pos in kethu_positions
        ]

        print('rahu_positions_from_lagnam',rahu_positions_from_lagnam)
        print('kethu_positions_from_lagnam',kethu_positions_from_lagnam)


        # Calculate mars position from lagnam
        mars_position_from_lagnam = calculate_position(lagnam_position, mars_position)


        print('mars_position_from_lagnam',mars_position_from_lagnam)

        # Determine if there is Mars dosham
        mars_dosham = False
        # if mars_position_from_lagnam in {1, 2, 4, 7, 8, 12}:
        if mars_position_from_lagnam in {2, 4, 7, 8, 12}:
            mars_dosham = True

        # Determine if there is Rahu-Kethu dosham
        critical_positions = {1, 2, 7, 8}
        rahu_kethu_dosham = False

        # Check if any Rahu or Kethu position falls within the critical positions
        if any(pos in critical_positions for pos in rahu_positions_from_lagnam) or \
           any(pos in critical_positions for pos in kethu_positions_from_lagnam):
            rahu_kethu_dosham = True
        
        return mars_dosham, rahu_kethu_dosham



def parse_data(data):
    # Clean up and split data
    items = data.strip('{}').split(', ')
    parsed_items = []
    for item in items:
        parts = item.split(':')
        if len(parts) > 1:
            values = parts[-1].strip()
            # Handle multiple values separated by comma
            if ',' in values:
                values = '/'.join(planet_mapping.get(v.strip(), default_placeholder) for v in values.split(','))
            else:
                values = planet_mapping.get(values, default_placeholder)
        else:
            values = default_placeholder
        parsed_items.append(values)
    return parsed_items


class ShortProfilePDFView(APIView):
    def post(self, request):
        profile_id = request.data.get('profile_id')
        format_type = request.data.get('format') or "shortprofile"

        if not format_type:
            return JsonResponse({"status": "error", "message": "format is required"}, status=400)

        if not profile_id:
            return JsonResponse({"status": "error", "message": "profile_id is required"}, status=400)

        try:
            if format_type == "shortprofile":
                return self.generate_short_profile_pdf(profile_id)

            elif format_type == "fullprofile":
                return My_horoscope_generate(request, profile_id, filename=f"fullprofile_{profile_id}.pdf")

            elif format_type == "withoutaddress":
                return generate_pdf_without_address(request, profile_id,filename=f"profile_withoutaddress_{profile_id}.pdf")

            else:
                return JsonResponse({"status": "error", "message": "Invalid format"}, status=400)

        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=500)

    def generate_short_profile_pdf(self, profile_id):
        login = get_object_or_404(LoginDetails, ProfileId=profile_id)
        family = get_object_or_404(ProfileFamilyDetails, profile_id=profile_id)
        edu = get_object_or_404(ProfileEduDetails, profile_id=profile_id)
        horoscope = get_object_or_404(ProfileHoroscope, profile_id=profile_id)

        def get_safe_value(model, lookup_field, id_value, return_field, default="N/A"):
            if not id_value:
                return default
            return model.objects.filter(**{lookup_field: id_value}).values_list(return_field, flat=True).first() or default

        complexion = get_safe_value(Complexion, 'complexion_id', login.Profile_complexion, 'complexion_desc')
        birthstar = get_safe_value(BirthStar, 'id', horoscope.birthstar_name, 'star')

        rasi_kattam = parse_data(horoscope.rasi_kattam or '') + ['N/A'] * 12
        rasi_kattam = rasi_kattam[:12]  # ensure exactly 12 elements

        html_content = f"""
        <html>
        <head>
          <style>
            body {{ font-family: Arial, sans-serif; }}
            .details p {{ font-size: 14px; margin: 5px 0; }}
            .rasi-kattam td {{
                width: 60px; height: 50px; text-align: center; border: 1px solid #000;
            }}
            table.rasi-kattam {{
                border-collapse: collapse;
            }}
          </style>
        </head>
        <body>
          <h2>Short Profile PDF</h2>
          <p><strong>Profile ID:</strong> {profile_id}</p>
          <div class="details">
            <p><strong>Name:</strong> {login.Profile_name}</p>
            <p><strong>DOB:</strong> {login.Profile_dob}</p>
            <p><strong>Father's Name:</strong> {family.father_name}</p>
            <p><strong>Complexion:</strong> {complexion}</p>
            <p><strong>Birth Star:</strong> {birthstar}</p>
          </div>
          <h4>Rasi Kattam</h4>
          <table class="rasi-kattam">
            <tr><td>{rasi_kattam[0]}</td><td>{rasi_kattam[1]}</td><td>{rasi_kattam[2]}</td><td>{rasi_kattam[3]}</td></tr>
            <tr><td>{rasi_kattam[11]}</td><td colspan="2" rowspan="2">Rasi</td><td>{rasi_kattam[4]}</td></tr>
            <tr><td>{rasi_kattam[10]}</td><td>{rasi_kattam[5]}</td></tr>
            <tr><td>{rasi_kattam[9]}</td><td>{rasi_kattam[8]}</td><td>{rasi_kattam[7]}</td><td>{rasi_kattam[6]}</td></tr>
          </table>
        </body>
        </html>
        """

        return self.render_pdf(html_content, f"short_profile_{profile_id}.pdf")

    def generate_full_profile_pdf(self, profile_id):
        # Reuse short profile template or add more fields for full profile
        html_content = f"""
        <html><body><h2>Full Profile</h2><p>Profile ID: {profile_id}</p>
        <p>This is a placeholder for full profile generation logic.</p></body></html>"""
        return self.render_pdf(html_content, f"full_profile_{profile_id}.pdf")

    def generate_profile_without_address(self, profile_id):
        html_content = f"""
        <html><body><h2>Profile Without Address</h2><p>Profile ID: {profile_id}</p>
        <p>This PDF is generated without sensitive address information.</p></body></html>"""
        return self.render_pdf(html_content, f"without_address_profile_{profile_id}.pdf")

    def render_pdf(self, html, filename):
        pdf_file = io.BytesIO()
        pisa_status = pisa.CreatePDF(io.StringIO(html), dest=pdf_file)

        if pisa_status.err:
            return JsonResponse({"status": "error", "message": "Error generating PDF."}, status=500)

        pdf_file.seek(0)
        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = 'inline'
        return response

class SendShortProfilePDFEmail(APIView):
    def post(self, request):
        profile_ids = request.data.get('profile_id')
        to_profile_id = request.data.get('to_profile_id')

        if not profile_ids or not to_profile_id:
            return Response({"error": "profile_id and to_profile_id are required"}, status=status.HTTP_400_BAD_REQUEST)

        profile_ids_list = [pid.strip() for pid in profile_ids.split(',') if pid.strip()]

        if not profile_ids_list:
            return Response({"error": "Invalid profile_id format"}, status=status.HTTP_400_BAD_REQUEST)

        recipient_email = LoginDetails.objects.filter(ProfileId__iexact=to_profile_id).values_list('EmailId', flat=True).first()
        
        if not recipient_email:
            return Response({"error": "No email found for to_profile_id"}, status=status.HTTP_400_BAD_REQUEST)

        profiles_html = ""
        for profile_id in profile_ids_list:
            login_details = get_object_or_404(LoginDetails, ProfileId=profile_id)
            family_details = get_object_or_404(ProfileFamilyDetails, profile_id=profile_id)
            edu_details = get_object_or_404(ProfileEduDetails, profile_id=profile_id)
            horoscope_details = get_object_or_404(ProfileHoroscope, profile_id=profile_id)

            complexion_desc = (
                get_object_or_404(Complexion, complexion_id=login_details.Profile_complexion).complexion_desc
                if login_details.Profile_complexion else "N/A"
            )
            birthstar_name = (
                get_object_or_404(BirthStar, id=horoscope_details.birthstar_name).star
                if horoscope_details.birthstar_name else "N/A"
            )
            highest_education = (
                get_object_or_404(EducationLevel, row_id=edu_details.highest_education).EducationLevel
                if edu_details.highest_education else "N/A"
            )
            profession = (
                get_object_or_404(Profession, row_id=edu_details.profession).profession
                if edu_details.profession else "N/A"
            )
            annual_income = (
                get_object_or_404(AnnualIncome, id=edu_details.anual_income).income
                if edu_details.anual_income else "N/A"
            )
            state_name = (
                get_object_or_404(State, id=login_details.Profile_state).name
                if login_details.Profile_state else "N/A"
            )

            profiles_html += f"""
            <div class="profile">
                <table class="vysyamala-flex">
                    <tr>
                        <td>Vysyamala  https://www.vysyamala.com  |  {login_details.Mobile_no}  |  UserId: {login_details.ProfileId}</td>
                    </tr>
                </table>
                <div class="details">
                    <p><strong>Name:</strong> {login_details.Profile_name} <strong>S/o</strong> {family_details.father_name} DOB: {login_details.Profile_dob}</p>
                    <p><strong>Height:</strong> {login_details.Profile_height} cm | <strong>Complexion:</strong> {complexion_desc} | <strong>Birth Star:</strong> {birthstar_name} | <strong>Gothram:</strong> {family_details.suya_gothram}</p>
                    <p><strong>Education:</strong> {highest_education} | <strong>State:</strong> {state_name}</p>
                    <p><strong>Profession:</strong> {profession}</p>
                    <p><strong>Annual Income:</strong> {annual_income}</p>
                </div>
                <hr/>
            </div>
            """

        html_content = f"""
        <html>
        <head>
            <style>
                .print-heading {{ font-size: 28px; font-weight: bold; text-align: center; }}
                .details p {{ font-size: 14px; margin: 5px 0; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1 class="print-heading"><strong>Print Short Profiles</strong></h1>
                <br>
            </div>
            {profiles_html}
        </body>
        </html>
        """

        temp_dir = tempfile.gettempdir()
        pdf_filename = f"Profiles_{'_'.join(profile_ids_list)}.pdf"
        pdf_file_path = os.path.join(temp_dir, pdf_filename)

        with open(pdf_file_path, "wb") as pdf_file:
            pisa_status = pisa.CreatePDF(html_content, dest=pdf_file)

        if not os.path.exists(pdf_file_path) or pisa_status.err:
            return Response({"error": "Error generating PDF"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        email_subject = "Short Profile PDF"
        email_body = "Please find the attached PDF containing the requested short profiles."
        email = EmailMessage(subject=email_subject, body=email_body, to=[recipient_email])
        email.attach_file(pdf_file_path)

        try:
            email.send()
            email_status = "sent"
            message = f"Email sent successfully to {recipient_email}"
        except Exception as e:
            email_status = "failed"
            message = f"Error sending email: {str(e)}"

        # Log Email Sending
        SentShortProfileEmailLog.objects.create(
            profile_id=profile_ids,
            to_ids=to_profile_id,
            profile_owner=profile_ids_list[0],  # Assuming first profile ID as owner
            status=email_status,
            sent_datetime=datetime.now()
        )

        os.remove(pdf_file_path)

        return Response({"message": message}, status=status.HTTP_200_OK if email_status == "sent" else status.HTTP_500_INTERNAL_SERVER_ERROR)




class SendFullProfilePDFEmail(APIView):
    def post(self, request):
        """API to generate multiple horoscope PDFs and send them to a single recipient, while logging the details."""

        profile_ids = request.data.get('profile_id')  # Expecting a comma-separated string
        to_profile_id = request.data.get('to_profile_id')  # Single recipient profile ID

        if not profile_ids or not to_profile_id:
            return Response({"error": "profile_id and to_profile_id are required"}, status=status.HTTP_400_BAD_REQUEST)

        # Convert comma-separated IDs into a list
        profile_ids_list = [pid.strip() for pid in profile_ids.split(',') if pid.strip()]

        if not profile_ids_list:
            return Response({"error": "Invalid profile_id format"}, status=status.HTTP_400_BAD_REQUEST)

        # Fetch recipient email for to_profile_id
        recipient_email = LoginDetails.objects.filter(ProfileId__iexact=to_profile_id).values_list('EmailId', flat=True).first()

        if not recipient_email:
            return Response({"error": "No email found for to_profile_id"}, status=status.HTTP_400_BAD_REQUEST)

        pdf_attachments = []  # Store generated PDFs

        for profile_id in profile_ids_list:
            try:
                # Get user details
                user = get_object_or_404(Registration1, ProfileId=profile_id)

                # Generate PDF
                response = My_horoscope_generate(request, profile_id)
                if response.status_code != 200:
                    return Response({"error": f"Error generating PDF for {profile_id}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

                pdf_content = response.getvalue()  # Extract PDF content
                pdf_attachments.append((f"Horoscope_{profile_id}.pdf", pdf_content, "application/pdf"))

            except Exception as e:
                return Response({"error": f"Error processing profile {profile_id}: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # Email setup
        subject = "Your Horoscope Profile Details"
        message = f"Dear User,\n\nPlease find your full horoscope details attached.\n\nBest Regards,\nYour Astrology Team"
        email = EmailMessage(subject, message, settings.DEFAULT_FROM_EMAIL, [recipient_email])

        # Attach all PDFs
        for filename, content, mimetype in pdf_attachments:
            email.attach(filename, content, mimetype)

        # Send email and log the details
        try:
            email.send()

            # Log the email sending in `SentProfileEmailLog`**
            SentFullProfileEmailLog.objects.create(
                profile_id=profile_ids,  # Store all profile_ids as a comma-separated string
                to_ids=to_profile_id,
                profile_owner=profile_ids_list[0],  # Assuming the first profile in the list is the owner
                status="sent",
                sent_datetime=datetime.now()
            )

            return Response({"message": f"Email sent successfully to {recipient_email}"}, status=status.HTTP_200_OK)

        except Exception as e:
            # Log failure
            SentFullProfileEmailLog.objects.create(
                profile_id=profile_ids,
                to_ids=to_profile_id,
                profile_owner=profile_ids_list[0],
                status="failed",
                sent_datetime=datetime.datetime.now()
            )

            return Response({"error": f"Error sending email: {e}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



class SendFullProfilePrintPDF(APIView):
    def post(self, request):
        """API to generate and merge multiple horoscope PDFs into a single response."""

        profile_ids = request.data.get('profile_id')  # Expecting a comma-separated string
        action_type = request.data.get('action_type')  # 'print' or 'whatsapp'
        to_profile_id = request.data.get('to_profile_id') 

        # Check if profile_ids and action_type are provided
        if not profile_ids:
            return JsonResponse({"error": "profile_id is required"}, status=400)

        # Convert comma-separated IDs into a list
        profile_ids_list = [pid.strip() for pid in profile_ids.split(',') if pid.strip()]
        missing_profiles = []
        pdf_merger = PdfMerger()
        successful_profiles = []
        log_status = "failed"
        profile_owner = request.data.get('profile_owner')

        for profile_id in profile_ids_list:
            try:
                # Generate PDF for each profile
                pdf_response = My_horoscope_generate(request, profile_id)

                if not pdf_response or pdf_response.status_code != 200:
                    missing_profiles.append(profile_id)
                    continue  # Skip this profile

                pdf_content = pdf_response.getvalue()  # Extract PDF content

                # Store the PDF content in memory
                pdf_file = io.BytesIO(pdf_content)
                pdf_merger.append(pdf_file)  # Merge PDF into one file
                successful_profiles.append(profile_id)

                # Dynamically fetch profile_owner (assuming 'Profile' is your model)
                if profile_owner is None:
                    try:
                        profile = Profile.objects.get(id=profile_id)  # Replace Profile with your actual model
                        profile_owner = profile.owner  # Assuming 'owner' is the field that holds the owner information
                    except ObjectDoesNotExist:
                        missing_profiles.append(profile_id)
                        continue  # Skip if the profile owner is not found
                        
            except Exception as e:
                missing_profiles.append(profile_id)
                continue  # Skip this profile if an error occurs

        if successful_profiles:
            # Create a final merged PDF file in memory
            merged_pdf = io.BytesIO()
            pdf_merger.write(merged_pdf)
            pdf_merger.close()
            merged_pdf.seek(0)

            # Log success
            log_status = "sent"

            # Return the merged PDF file
            response = HttpResponse(merged_pdf.read(), content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="Merged_Horoscope_Profiles.pdf"'
        else:
            # Log failure
            response = JsonResponse({"error": f"Failed to generate PDF for profiles: {', '.join(missing_profiles)}"},
                                    status=500)

        # Log based on action_type (either print or whatsapp)
        if action_type == 'whatsapp':
            SentFullProfilePrintwpLog.objects.create(
                profile_id=profile_ids,
                to_ids=to_profile_id,  # Since it's a file download, the recipient is self
                profile_owner=profile_owner if profile_owner else "Unknown",  # Dynamic profile owner
                status=log_status,
                sent_datetime=datetime.now()
            )
        else:
            SentFullProfilePrintPDFLog.objects.create(
                profile_id=profile_ids,
                to_ids=to_profile_id,  # Since it's a file download, the recipient is self
                profile_owner=profile_owner if profile_owner else "Unknown",  # Dynamic profile owner
                status=log_status,
                sent_datetime=datetime.now()
            )

        return response




class SendShortProfilePrintPDF(APIView):
    def post(self, request):
        profile_ids = request.data.get('profile_id')
        action_type = request.data.get('action_type')  # 'print' or 'whatsapp'
        to_profile_id = request.data.get('to_profile_id')
       
        if not profile_ids:
            return JsonResponse({"error": "profile_id is required"}, status=400)

        profile_ids_list = [pid.strip() for pid in profile_ids.split(',')]
        profiles_html = ""
        log_status = "failed"  # Default status for logging

        for profile_id in profile_ids_list:
            login_details = get_object_or_404(LoginDetails, ProfileId=profile_id)
            family_details = get_object_or_404(ProfileFamilyDetails, profile_id=profile_id)
            edu_details = get_object_or_404(ProfileEduDetails, profile_id=profile_id)
            horoscope_details = get_object_or_404(ProfileHoroscope, profile_id=profile_id)

            complexion_desc = "N/A"
            if login_details.Profile_complexion:
                complexion_instance = get_object_or_404(Complexion, complexion_id=login_details.Profile_complexion)
                complexion_desc = complexion_instance.complexion_desc

            birthstar_name = "N/A"
            if horoscope_details.birthstar_name:
                birthstar_instance = get_object_or_404(BirthStar, id=horoscope_details.birthstar_name)
                birthstar_name = birthstar_instance.star

            highest_education = "N/A"
            if edu_details.highest_education:
                education_instance = get_object_or_404(EducationLevel, row_id=edu_details.highest_education)
                highest_education = education_instance.EducationLevel

            profession = "N/A"
            if edu_details.profession:
                profession_instance = get_object_or_404(Profession, row_id=edu_details.profession)
                profession = profession_instance.profession

            annual_income = "N/A"
            if edu_details.anual_income:
                income_instance = get_object_or_404(AnnualIncome, id=edu_details.anual_income)
                annual_income = income_instance.income

            state_name = "N/A"
            if login_details.Profile_state:
                state_instance = get_object_or_404(State, id=login_details.Profile_state)
                state_name = state_instance.name

            profiles_html += f"""
            <div class="profile">
                <table class="vysyamala-flex">
                    <tr>
                        <td>Vysyamala  https://www.vysyamala.com  |  {login_details.Mobile_no}  |  UserId: {login_details.ProfileId}</td>
                    </tr>
                </table>
                <div class="details">
                    <p><strong>Name:</strong> {login_details.Profile_name} <strong>S/o</strong> {family_details.father_name} DOB: {login_details.Profile_dob}</p>
                    <p><strong>Height:</strong> {login_details.Profile_height} cm | <strong>Complexion:</strong> {complexion_desc} | <strong>Birth Star:</strong> {birthstar_name} | <strong>Gothram:</strong> {family_details.suya_gothram}</p>
                    <p><strong>Education:</strong> {highest_education} | <strong>State:</strong> {state_name}</p>
                    <p><strong>Profession:</strong> {profession}</p>
                    <p><strong>Annual Income:</strong> {annual_income}</p>
                </div>
                <hr/>
            </div>
            """

        html_content = f"""
        <html>
        <head>
            <style>
                .print-heading {{ font-size: 28px; font-weight: bold; text-align: center; }}
                .details p {{ font-size: 14px; margin: 5px 0; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1 class="print-heading"><strong>Print Short Profiles</strong></h1>
                <br>
            </div>
            {profiles_html}
        </body>
        </html>
        """

        # Generate PDF in memory
        pdf_response = HttpResponse(content_type='application/pdf')
        pdf_response['Content-Disposition'] = f'attachment; filename="Short_Profiles.pdf"'
        profile_owner = request.data.get('profile_owner')


        pisa_status = pisa.CreatePDF(html_content, dest=pdf_response)

        if pisa_status.err:
           # Log failure based on action_type
           log_model = SentShortProfilePrintwpLog if action_type == 'whatsapp' else SentShortProfilePrintPDFLog
           log_model.objects.create(
               profile_id=profile_ids,
               to_ids=to_profile_id,  # Since it's a file download or WhatsApp share, recipient is "self"
               profile_owner=profile_owner if profile_owner else "Unknown",  # Dynamic profile owner
               status="failed",
               sent_datetime=datetime.now()
           )
           return JsonResponse({"error": "Error generating PDF"}, status=500)

        # Log success based on action_type
        log_model = SentShortProfilePrintwpLog if action_type == 'whatsapp' else SentShortProfilePrintPDFLog
        log_model.objects.create(
            profile_id=profile_ids,
            to_ids=to_profile_id,
            profile_owner=profile_owner if profile_owner else "Unknown",  # Dynamic profile owner
            status="sent",
            sent_datetime=datetime.now()
        )
        
        return pdf_response  # Return the generated PDF file

    
class CallTypeListCreateView(generics.ListCreateAPIView):
    queryset = CallType.objects.all()
    serializer_class = CallTypeSerializer

class CallStatusListCreateView(generics.ListCreateAPIView):
    queryset = CallStatus.objects.all()
    serializer_class = CallStatusSerializer


class CallActionListCreateView(generics.ListCreateAPIView):
    queryset = CallAction.objects.all()
    serializer_class = CallActionSerializer

# Insert (Create) API
class ProfileCallManagementCreateView(generics.CreateAPIView):
    queryset = ProfileCallManagement.objects.all()
    serializer_class = ProfileCallManagementSerializer

# Get (List) API
class ProfileCallManagementListView(generics.ListAPIView):
    serializer_class = ProfileCallManagementSerializer

    def get_queryset(self):
        profile_id = self.request.query_params.get('profile_id')
        if profile_id:
            return ProfileCallManagement.objects.filter(profile_id=profile_id)
        return ProfileCallManagement.objects.none()


# Create view
class MarriageSettleDetailsCreateView(generics.CreateAPIView):
   def post(self, request):
        profile_id = request.data.get('profile_id')
        if not profile_id:
            return Response({'status': 'error', 'message': 'profile_id is required'}, status=status.HTTP_400_BAD_REQUEST)

        existing_records = MarriageSettleDetails.objects.filter(profile_id=profile_id)

        if existing_records.exists():
            instance = existing_records.first()  # Use first matching instance
            serializer = MarriageSettleDetailsSerializer(instance, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response({'status': 'success',"message": "Marriage Settle Details updated successfully",'data': serializer.data}, status=status.HTTP_200_OK)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        # If no existing records, create new one
        serializer = MarriageSettleDetailsSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'status': 'message',"message": "Marriage Settle Details inserted successfully", 'data': serializer.data}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
   
# List view
class MarriageSettleDetailsListView(generics.ListAPIView):
    serializer_class = MarriageSettleDetailsSerializer

    def get_queryset(self):
        profile_id = self.request.query_params.get('profile_id')
        if profile_id:
            return MarriageSettleDetails.objects.filter(profile_id=profile_id)
        return MarriageSettleDetails.objects.none()

class PaymentTransactionCreateView(generics.CreateAPIView):
    def post(self, request):
        profile_id = request.data.get('profile_id')
        payment_type = request.data.get('payment_type')
        payment_status = request.data.get('status')

        # Validate required fields
        if not profile_id:
            return Response({'status': 'error', 'message': 'profile_id is required'}, status=status.HTTP_400_BAD_REQUEST)
        if not payment_type:
            return Response({'status': 'error', 'message': 'payment_type is required'}, status=status.HTTP_400_BAD_REQUEST)
        if payment_status is None:
            return Response({'status': 'error', 'message': 'status is required'}, status=status.HTTP_400_BAD_REQUEST)

        existing_records = PaymentTransaction.objects.filter(profile_id=profile_id)

        if existing_records.exists():
            instance = existing_records.first()
            serializer = PaymentTransactionSerializer(instance, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response({'status': 'success', 'message': 'Payment Transaction updated successfully', 'data': serializer.data}, status=status.HTTP_200_OK)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        # Create new record
        serializer = PaymentTransactionSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'status': 'success', 'message': 'Payment Transaction inserted successfully', 'data': serializer.data}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
class PaymentTransactionListView(generics.ListAPIView):
    serializer_class = PaymentTransactionSerializer

    def get_queryset(self):
        profile_id = self.request.query_params.get('profile_id')
        if profile_id:
            return PaymentTransaction.objects.filter(profile_id=profile_id)
        return PaymentTransaction.objects.none()

    def list(self, request, *args, **kwargs):
        profile_id = request.query_params.get('profile_id')
        if not profile_id:
            return Response({
                'status': 'error',
                'message': 'profile_id is required'
            }, status=status.HTTP_400_BAD_REQUEST)

        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)

        # Add extra fields to each serialized item
        enriched_data = []
        for item in serializer.data:
            item_copy = dict(item)
            item_copy['balance_amount'] = "0"
            item_copy['payment_details'] = "null"
            enriched_data.append(item_copy)

        return Response({
            'status': 'success',
            'message': 'Payment transactions fetched successfully',
            'data': enriched_data
        })
    
# class GenerateInvoicePDF(APIView):
#     def post(self, request):
#         serializer = InvoiceSerializer(data=request.data)
#         if serializer.is_valid():
#             data = serializer.validated_data
#             # Load and encode the logo image
#             logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'newlogo.png')
#             print(logo_path)
#             try:
#                 with open(logo_path, "rb") as image_file:
#                     encoded_logo = base64.b64encode(image_file.read()).decode()
#             except FileNotFoundError:
#                 encoded_logo = ""

#             html_string = f"""
#             <!DOCTYPE html>
#             <html>
#             <head>
#                 <style>
#                     body {{
#                         font-family: Arial, sans-serif;
#                         padding: 25px;
#                         font-size: 13px;
#                     }}
#                     .invoice-meta {{
#                         text-align: right;
#                         font-size: 13px;
#                     }}
#                     .invoice-meta h2 {{
#                         color: #9c9c9c;
#                         margin: 0 0 10px 0;
#                     }}
#                     .table {{
#                         width: 100%;
#                         border-collapse: collapse;
#                         margin-top: 10px;
#                     }}
#                     .table th, .table td {{
#                         border: 1px solid #000;
#                         padding: 10px;
#                         text-align: left;
#                     }}
#                     .table th {{
#                         background-color: #f2f2f2;
#                     }}
#                     .footer {{
#                         margin-top: 20px;
#                         font-size: 12px;
#                     }}
#                     .note {{
#                         margin-top: 30px;
#                         font-style: italic;
#                         font-weight: bold;
#                         text-align: center;
#                     }}
#                     .bottom-contact {{
#                         margin-top: 10px;  /* reduced from 30px */
#                         font-size: 11px;
#                         border-top: 1px solid #ccc;
#                         padding-top: 5px;  /* optional: tighter padding */
#                     }}
                    
#                 </style>
#             </head>
#             <body>
            
#             <table style="width: 100%;">
#                 <tr>
#                     <td style="vertical-align: top; width: 60%; font-size: 13px;">
#                         {'<img src="data:image/png;base64,' + encoded_logo + '" style="height: 70px;"><br>' if encoded_logo else ''}
#                         <strong>To</strong><br>
#                         {data.get('customer_name', '')}<br>
#                         {data['address'].replace(chr(10), '<br>') if data.get("address") else ""}
#                     </td>
#                     <td style="vertical-align: top; text-align: right;">
#                         <h2 style="margin: 0; color: #9c9c9c;">Invoice</h2>
#                         Date: {data['date']}<br>
#                         Invoice #: {data['invoice_number']}<br>
#                         Vysyamala ID: {data['vysyamala_id']}<br>
#                     </td>
#                 </tr>
#             </table>

#             <table class="table">
#                 <tr>
#                     <th>Service Description</th>
#                     <th>Price</th>
#                     <th>Net Price</th>
#                 </tr>
#                 <tr>
#                     <td>
#                         {data['service_description']}<br>
#                         <small>Valid till: {data['valid_till']} or engagement date whichever is earlier</small>
#                     </td>
#                     <td>{data['price']}</td>
#                     <td>{data['price']}</td>
#                 </tr>
#                 <tr>
#                     <td colspan="2"><strong>Total</strong></td>
#                     <td><strong>{data['price']}</strong></td>
#                 </tr>
#             </table>

#             <p><strong>In words:</strong></p>

#             <table class="table" style="width: 50%; margin-top: 20px;">
#                 <tr>
#                     <th>Payment Mode</th>
#                     <td>Online Transfer</td>
#                 </tr>
#             </table>

#             <p class="note">Thank you for your opportunity to serve you!</p>

#             <div class="bottom-contact">
#                 <table style="width: 100%;">
#                     <tr>
#                         <td style="vertical-align: top; width: 50%;">
#                             Vysyamala<br>
#                             C/o. YK Lavanya<br>
#                             No.2, Krishnaswamy Street (Lane)<br>
#                             A6-2nd Floor, Sri Vinayaga Flats<br>
#                             Ganapathipuram, Chrompet<br>
#                             Chennai – 600 044
#                         </td>
#                         <td style="vertical-align: top; text-align: right; font-size: 11px;">
#                             Web: <a href="http://www.vysyamala.com">www.vysyamala.com</a> |
#                             Email: vysyamala@gmail.com |
#                             Facebook: www.fb.com/vysyamala<br>
#                             Whatsapp: 9043085524 |
#                             Customer Support: 9944851550 (8 a.m. to 8 p.m.)
#                         </td>
#                     </tr>
#                     <tr>
#                         <td colspan="2" style="text-align: center; font-style: italic; font-size: 12px; padding-top: 10px;">
#                             May Goddess Sri Vasavi Kanyaka Parameswari bless you and your family with peace & prosperity!
#                         </td>
#                     </tr>
#                 </table>
#             </div>
        

#             </body>
#             </html>
#             """

#             result = BytesIO()
#             pisa_status = pisa.CreatePDF(src=html_string, dest=result)

#             if pisa_status.err:
#                 return Response({"error": "Error generating PDF"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

#             response = HttpResponse(result.getvalue(), content_type='application/pdf')
#             response['Content-Disposition'] = f'inline; filename="invoice_{data["invoice_number"]}.pdf"'
#             return response

#         return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

def number_to_words(n):
    units = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine"]
    teens = ["Ten", "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen", "Seventeen", "Eighteen", "Nineteen"]
    tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]

    def two_digit(num):
        if num < 10:
            return units[num]
        elif num < 20:
            return teens[num - 10]
        else:
            return tens[num // 10] + (" " + units[num % 10] if num % 10 != 0 else "")

    def three_digit(num):
        if num < 100:
            return two_digit(num)
        else:
            return units[num // 100] + " Hundred" + (" and " + two_digit(num % 100) if num % 100 != 0 else "")

    if n == 0:
        return "Zero"

    result = ""
    if n >= 1000:
        result += two_digit(n // 1000) + " Thousand"
        if n % 1000 != 0:
            result += " " + three_digit(n % 1000)
    else:
        result += three_digit(n)

    return result + " only"

class GenerateInvoicePDF(APIView):
    def get(self, request):
        subscription_id = request.query_params.get('subscription_id')
        if not subscription_id:
            return Response({"error": "subscription_id is required"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            subscription = PlanSubscription.objects.get(id=subscription_id)
        except PlanSubscription.DoesNotExist:
            return Response({"error": "Subscription not found"}, status=status.HTTP_404_NOT_FOUND)

        try:
            profile = Registration1.objects.get(ProfileId=subscription.profile_id)
            state = get_state_name(profile.Profile_state)
        except Exception:
            profile = None
            state = None
            
        if profile:
            customer_name = profile.Profile_name
            address = f"{profile.Profile_address or ''}<br/>{profile.Profile_city or ''}, {state or ''} - {profile.Profile_pincode or ''}<br/>Phone: {profile.Mobile_no or ''}<br/>Email: {profile.EmailId or ''}"
        else:
            customer_name = None
            address = "Address not available"
        # Load and encode the logo image from URL
        image_url = "https://vysyamat.blob.core.windows.net/vysyamala/newvysyamalalogo2.png"
        try:
            response = requests.get(image_url)
            encoded_logo = base64.b64encode(response.content).decode() if response.status_code == 200 else ""
        except Exception:
            encoded_logo = ""

        if subscription.payment_mode:
            mode = subscription.payment_mode.strip().lower()
            print(mode)
            if mode in ["razorpay","razor pay","onlinegpay","manualgpay","accounttransfer","online","payu online payment","by account transfer","online payment","by google pay"]:
                payment_mode = "Online Transfer"
            else:
                payment_mode = "Cash/Cheque/DD"
        else:
            payment_mode = "Cash/Cheque/DD"
        
        try:
            addon_ids = [int(pk.strip()) for pk in subscription.addon_package.split(",") if pk.strip().isdigit()]
            addon_qs = Addonpackages.objects.filter(package_id__in=addon_ids)

            addon_items = []
            addon_total = 0
            for addon in addon_qs:
                addon_items.append({
                    "name": addon.name,
                    "description": addon.description,
                    "amount": addon.amount or 0
                })
                addon_total += addon.amount or 0
        except Exception:
            addon_items = []
            addon_total = 0
        try:
            if subscription.package_amount:
                plan_amount = subscription.package_amount
            else:
                if subscription.plan_id:
                    plan=PlanDetails.objects.get(id=subscription.plan_id)
                    plan_amount = plan.plan_price
                else:
                    plan_amount=0
        except Exception:
            plan_amount=0
              
        base_price = plan_amount
        discount = subscription.discount or 0
        total_price = base_price + addon_total- discount
        net_price = base_price + addon_total
        if subscription.plan_id and subscription.plan_id > 0:
            plan_name = PlanDetails.objects.filter(id=subscription.plan_id).values_list('plan_name', flat=True).first()   
        else:
            plan_name = ""
        data = {
            'encoded_logo': encoded_logo,
            'customer_name': customer_name if customer_name else "Valued Customer",
            'address': address,
            'date': subscription.payment_date.strftime("%d/%m/%Y") if subscription.payment_date else "",
            'invoice_number': subscription.id,
            'vysyamala_id': subscription.profile_id or "",
            'service_description':plan_name or "" ,
            'offer': subscription.offer or "",
            'price': f"{base_price:.0f}",
            'valid_till': subscription.validity_enddate.strftime("%d-%m-%Y") if subscription.validity_enddate else "",
            'payment_mode': payment_mode or "N/A",
            'addon_items': addon_items,
            'addon_total': f"{addon_total:.0f}",
            'discount': f"{discount:.0f}",
            'total_price': f"{total_price:.0f}",
            'num_to_words': number_to_words(int(total_price)),
            'net_price': f"{net_price:.0f}",
        }

        # Render HTML template with data
        html_string = render_to_string("invoice.html", data)

        # Generate PDF
        result = BytesIO()
        pisa_status = pisa.CreatePDF(src=html_string, dest=result)

        if pisa_status.err:
            return Response({"error": "Error generating PDF"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="invoice_{data["invoice_number"]}.pdf"'
        return response
       
def process_and_blur_image(image_bytes):
    """Process image bytes and return blurred image bytes"""
    try:
        with Image.open(BytesIO(image_bytes)) as img:
            # Convert to RGB if needed (for PNGs with alpha channel)
            if img.mode != 'RGB':
                img = img.convert('RGB')
            
            # Resize (optional, improves performance)
            img = img.resize((img.width // 2, img.height // 2))
            
            # Apply blur effect
            blurred_img = img.filter(ImageFilter.GaussianBlur(radius=10))
            
            # Save to bytes
            output = BytesIO()
            blurred_img.save(output, format='JPEG', quality=85)
            return output.getvalue()
    
    except Exception as e:
        logger.error(f"Image processing failed: {e}")
        raise


@csrf_exempt
@require_http_methods(["GET", "POST"])
def GetPhotoProofDetails(request):
    if request.method == "GET":
        profile_id = request.GET.get('profile_id')

        if not profile_id:
            return JsonResponse({'status': 'error', 'message': 'profile_id is required'}, status=400)

        try:
            login = LoginDetails.objects.get(ProfileId=profile_id)
            horoscope = Horoscope.objects.get(profile_id=profile_id)
            profile_images = Image_Upload.objects.filter(
                    profile_id=profile_id
                ).filter(
                    Q(is_deleted=False) | Q(is_deleted__isnull=True)
                )

            image_list = [
                {
                    'id': image.id,
                    'image_url': image.image.url if image.image else None,
                    'image_approved': image.image_approved,
                    'uploaded_at': image.uploaded_at,
                    'is_deleted': image.is_deleted
                }
                for image in profile_images
            ]

            return JsonResponse({
                'status': 'success',
                'message': "Photo proof details fetched successfully.",
                'data': {
                    'photo_password': login.Photo_password,
                    'id_proof': login.Profile_idproof.url if login.Profile_idproof else None,
                    'divorce_certificate': login.Profile_divorceproof.url if login.Profile_divorceproof else None,
                    'horoscope_file': horoscope.horoscope_file.url if horoscope.horoscope_file else None,
                    'horoscope_file_admin': horoscope.horoscope_file_admin.url if horoscope.horoscope_file_admin else None,
                    'profile_images': image_list,
                    'profile_martial_status': login.Profile_marital_status,
                    'Profile_name':login.Profile_name,
                    'photo_protection':login.Photo_protection
                }
            })

        except LoginDetails.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Login details not found'}, status=404)
        except Horoscope.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Horoscope not found'}, status=404)

    elif request.method == "POST":
        profile_id = request.POST.get('profile_id')
        image_ids_csv = request.POST.get('image_id')
        is_deleted_csv = request.POST.get('is_deleted')
        image_approved_csv = request.POST.get('image_approved')
        photo_password = request.POST.get('photo_password')
        photo_protection = request.POST.get('photo_protection')
        owner_id = request.POST.get('admin_user_id')
        try:
            owner_id = int(owner_id)
            user = User.objects.get(id=owner_id)
        except Exception:
            user = None
             
        if user:
            role = user.role
            permissions = RolePermission.objects.filter(role=role).select_related('action')
            data = permissions.values('action__code', 'value')
            edit_permission = data.filter(action__code='edit_horo_photo').first()
            edit=edit_permission['value'] if edit_permission else None
        else:
            edit =None
           
        if user:  
            if edit ==1:
                pass
            else:
                return JsonResponse({
                    "status": "error",
                    "message": "Permission Error"
                }, status=403)

        if not profile_id:
            return JsonResponse({'status': 'error', 'message': 'profile_id is required'}, status=400)
    
        update_summary = {}
    
        try:
            # IMAGE UPDATES WITH INDIVIDUAL VALUES
            if image_ids_csv and is_deleted_csv and image_approved_csv:
                image_ids = [i.strip() for i in image_ids_csv.split(',')]
                is_deleted_vals = [int(i.strip()) for i in is_deleted_csv.split(',')]
                image_approved_vals = [int(i.strip()) for i in image_approved_csv.split(',')]
    
                if len(image_ids) != len(is_deleted_vals) or len(image_ids) != len(image_approved_vals):
                    return JsonResponse({'status': 'error', 'message': 'Length of image_id, is_deleted, and image_approved must match'}, status=400)
    
                updated_images = []
                deleted_images = []

                for idx, image_id in enumerate(image_ids):
                    try:
                        image = Image_Upload.objects.get(id=image_id, profile_id=profile_id)
                        file_name = os.path.basename(image.image.name) if image.image else None

                        if is_deleted_vals[idx] == 1:
                            try:
                                container_name = 'vysyamala'
                                connection_string = settings.AZURE_CONNECTION_STRING
                                blob_service = BlobServiceClient.from_connection_string(connection_string)
                                container_client = blob_service.get_container_client(container_name)

                                if file_name:
                                    for folder in ["profile_images/", "blurred_images/"]:
                                        blob_name = f"{folder}{file_name}"
                                        blob_client = container_client.get_blob_client(blob_name)
                                        if blob_client.exists():
                                            blob_client.delete_blob()
                                            update_summary.setdefault('azure_blobs_deleted', []).append(blob_name)

                            except Exception as azure_delete_error:
                                update_summary.setdefault('azure_delete_errors', []).append({
                                    'image_id': image.id,
                                    'error': str(azure_delete_error)
                                })

                            image.delete()
                            deleted_images.append(image_id)

                        else:
                            image.image_approved = bool(image_approved_vals[idx])
                            image.save()
                            updated_images.append(image.id)

                            if image.image_approved and image.image:
                                try:
                                    container_name = 'vysyamala'
                                    connection_string = settings.AZURE_CONNECTION_STRING
                                    source_folder = "profile_images/"
                                    dest_folder = "blurred_images/"
                                    file_name = os.path.basename(image.image.name)
                                    source_blob_name = f"{source_folder}{file_name}"
                                    dest_blob_name = f"{dest_folder}{file_name}"

                                    blob_service = BlobServiceClient.from_connection_string(connection_string)
                                    container_client = blob_service.get_container_client(container_name)
                                    blob_client = container_client.get_blob_client(source_blob_name)

                                    if blob_client.exists():
                                        image_data = blob_client.download_blob().readall()
                                        blurred_img = process_and_blur_image(image_data)
                                        container_client.get_blob_client(dest_blob_name).upload_blob(
                                            blurred_img,
                                            overwrite=True,
                                            content_settings=ContentSettings(content_type="image/jpeg")
                                        )
                                        update_summary.setdefault('blurred_images_uploaded', []).append(file_name)
                                    else:
                                        update_summary.setdefault('blurred_images_skipped', []).append(file_name)

                                except Exception as blur_e:
                                    update_summary.setdefault('blur_errors', []).append({
                                        'image_id': image.id,
                                        'error': str(blur_e)
                                    })

                    except Image_Upload.DoesNotExist:
                        continue

                update_summary['images_updated'] = updated_images
                update_summary['images_deleted'] = deleted_images

            # PHOTO PASSWORD UPDATE
            if photo_password is not None:
                try:
                    login = LoginDetails.objects.get(ProfileId=profile_id)
                    login.Photo_password = photo_password
                    login.save()
                    update_summary['photo_password'] = "updated"
                except LoginDetails.DoesNotExist:
                    update_summary['photo_password'] = "login details not found"

            try:
                login = LoginDetails.objects.get(ProfileId=profile_id)
                login.Photo_protection = photo_protection
                login.save()
            except Exception:
                pass
    
            if not update_summary:
                return JsonResponse({'status': 'error', 'message': 'No update data provided'}, status=400)
    
            return JsonResponse({
                'status': 'success',
                'message': 'Updates applied successfully',
                'update_summary': update_summary
            })
    
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)



class SendmobileOtp(APIView):
    
    def generate_otp(self):
        # Implement your OTP generation logic here
        import random
        return str(random.randint(100000, 999999))
    
    def post(self, request):
        profile_id=request.data.get('profile_id')
        if not profile_id:
            return JsonResponse({"status": "error", "message": "Profile ID is required"}, status=400)
        
        # mobile_number='91'+mobile_number

        # print('mobile_number',mobile_number)
        # Check if the mobile number exists in Registration table
        try:
            profile = models.Registration1.objects.get(ProfileId=profile_id)
        except models.Registration1.DoesNotExist:
            return JsonResponse({"status": 0, "message": "Invalid profile id."}, status=status.HTTP_200_OK)
        
        mobile_number=profile.Mobile_no
        # Generate OTP
        otp = self.generate_otp()

        # Send OTP via SMS (implement SendSMS() appropriately)
        
        #Below code commented on 30th jully 2024 harcode value set as 1234

        sms_sender = SendSMS()  # Ensure SendSMS class is implemented and imported correctly
        message_id = sms_sender.send_sms(otp, mobile_number)
        dlr_status = sms_sender.check_dlr(message_id)
        available_credit = sms_sender.available_credit()

        # Save OTP to UserProfile
        profile.Otp = otp
        #profile.Otp = 123456 #otp
        profile.save()

        # Prepare response data
        response_data = {
            "message": "OTP sent successfully.",
            "Send Message Response": message_id,
            "Delivery Report Status": dlr_status,
            "Available Credit": available_credit
        }

        return JsonResponse({"status": 1, "response_data": response_data, "message": "OTP sent successfully."}, status=status.HTTP_201_CREATED)
    

class SendSMS:
    def __init__(self):
        self.url = 'http://pay4sms.in'
        self.token = '76111ad0d3c72d750e36ec22c6e5105d'
        self.credit = '2'
        self.sender = 'VYSYLA'
        # self.message_template = 'Dear Customer, {} is the OTP for Edit profile. Please enter it in the space provided in the Website. Thank you for using Vysyamala.com'
        self.message_template = 'Dear Customer,{} is the OTP for mobile verification. Please enter it in the space provided in the Website. Thank you for using Vysyamala.com'

    def send_sms(self, otp, numbers):
        message = self.message_template.format(otp)
        message = requests.utils.quote(message)
        sms_url = f"{self.url}/sendsms/?token={self.token}&credit={self.credit}&sender={self.sender}&number={numbers}&message={message}"
        response = requests.get(sms_url)
        return response.text

    def check_dlr(self, message_id):
        dlr_url = f"{self.url}/Dlrcheck/?token={self.token}&msgid={message_id}"
        response = requests.get(dlr_url)
        return response.text

    def available_credit(self):
        credit_url = f"{self.url}/Credit-Balance/?token={self.token}"
        response = requests.get(credit_url)
        return response.text
    

class VerifymobileOtp(APIView):
    def post(self, request):
        profile_id = request.data.get('profile_id')
        otp = request.data.get('otp')

        if not profile_id or not otp:
            return JsonResponse({"status": "error", "message": "profile_id and otp are required"}, status=400)       

        try:
            profile = models.Registration1.objects.get(ProfileId=profile_id, Otp=otp)
            
            # Update the Otp_verify column
            profile.Otp_verify = 1
            profile.save()

            return JsonResponse({"status": "success", "message": "Otp verified successfully"}, status=status.HTTP_200_OK)

        except models.Registration1.DoesNotExist:
            return JsonResponse({"status": "error", "message": "Invalid otp for the profile id"}, status=status.HTTP_200_OK)

class AdminUserDropdownAPIView(APIView):
    def get(self, request):
        users = AdminUser.objects.filter(deleted=False)
        serializer = AdminUserDropdownSerializer(users, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
def get_star_lookup():
    with connection.cursor() as cursor:
        cursor.execute("SELECT id, star FROM masterbirthstar WHERE is_deleted = 0")
        return {str(row[0]): row[1] for row in cursor.fetchall()}
  
  
def build_profession_map(profile_details):
    profession_ids = {
        int(d["profession"])
        for d in profile_details
        if d.get("profession") not in [None, "0", "N/A", "~", " ", ""]
        and str(d.get("profession")).isdigit()
    }

    return {
        p.RowId: p.profession
        for p in Profespref.objects.filter(RowId__in=profession_ids)
    }
    
def build_degree_map(profile_details):
    degree_ids = set()

    for d in profile_details:
        if d.get("degree"):
            ids = [
                int(x) for x in str(d["degree"]).split(",")
                if x.strip().isdigit() and int(x) != 86
            ]
            degree_ids.update(ids)

    return {
        d.id: d.degeree_name
        for d in models.MasterhighestEducation.objects.filter(id__in=degree_ids)
    }
    
def getprofession_fast(profession, profession_map):
    if profession in [None, "0", "N/A", "~", " ", ""]:
        return "N/A"
    return profession_map.get(int(profession), "N/A")
  
  
def degree_fast(degree_ids, other_degree, degree_map):
    if not degree_ids:
        return other_degree or "N/A"

    ids = [
        int(x) for x in str(degree_ids).split(",")
        if x.strip().isdigit() and int(x) != 86
    ]

    names = [degree_map.get(i) for i in ids if degree_map.get(i)]

    if other_degree:
        names.append(other_degree)

    return ", ".join(names) if names else "N/A"

class CommonProfileSearchAPIView(APIView):

    def post(self, request):
        owner_id = request.data.get('admin_user_id')
        try:
            owner_id = int(owner_id)
            user = User.objects.get(id=owner_id)
        except Exception:
            user = None
             
        if user:
            role = user.role
            permissions = RolePermission.objects.filter(role=role).select_related('action')
            data = permissions.values('action__code', 'value')
            edit_permission = data.filter(action__code='search_profile').first()
            edit=edit_permission['value'] if edit_permission else None
        else:
            edit =None
          
        if user:   
            if edit ==1:
                pass
            else:
                return Response({
                    "status": "error",
                    "message": "Permission Error"
                }, status=status.HTTP_403_FORBIDDEN)
        serializer = CommonProfileSearchSerializer(data=request.data)
        if not serializer.is_valid():
            return JsonResponse(serializer.errors, status=400)
        
        filters = serializer.validated_data
        export_type = filters.pop('export_type', None)
        
        per_page = filters.pop('per_page', 10)
        page_number = filters.pop('page_number', 1)
        if export_type:
            start = 0
            per_page = 100000
        else:
            start = (page_number - 1) * per_page
        print("wrerw")
        profile_details, total_count, profile_with_indices = Get_profiledata_Matching.get_common_profile_list(
            start=start,
            per_page=per_page,
            is_export=bool(export_type),
            **filters
        )
        print("fsdf")
        if not profile_details:
            return JsonResponse({
                "Status": 0,
                "message": "No profiles found for the given criteria.",
                "search_result": "0"
            }, status=200)
        star_lookup = get_star_lookup()
        profiles = []
        print("gdgdfg")
        today = date.today()
        def fast_age(dob):
            if not dob:
                return None
            return today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))
        if not export_type:
            for detail in profile_details: 
                star_id = str(detail.get("birthstar_name", "")).strip()
                star_name = star_lookup.get(star_id, "")
                profiles.append({
                    "profile_id": detail["ProfileId"],
                    "profile_name": detail["Profile_name"],
                    "profile_img": Get_profile_image(detail["ProfileId"], detail["Gender"], 1, 0, is_admin=True),
                    "profile_age": calculate_age(detail["Profile_dob"]),
                    "profile_gender": detail["Gender"],
                    "height": detail["Profile_height"],
                    "degree": degree(detail.get("degree"),detail.get("other_degree")),
                    "profession": getprofession(detail.get("profession")),
                    "location": detail["Profile_city"],
                    "photo_protection": detail["Photo_protection"],
                    "verified": detail.get('Profile_verified'),
                    "star": star_name
                })
        print("fdsgvdfg")
        if export_type:
            export_rows = []
            print("itha")
            profession_map = build_profession_map(profile_details)
            degree_map = build_degree_map(profile_details)


            export_rows = [
                {
                "Profile ID": d.get("ProfileId"),
                "Name": d.get("Profile_name"),
                "Gender": d.get("Gender"),
                "Age": calculate_age(d.get("Profile_dob")),
                "Height": d.get("Profile_height"),
                "City": d.get("Profile_city"),
                "Profession": getprofession_fast(d.get("profession"), profession_map),
                "Education": degree_fast(d.get("degree"), d.get("other_degree"), degree_map),
                "Mobile": d.get("mobile_no"),
                "Email": d.get("email_id")
               }
                for d in profile_details
            ]
            print("itha")
            if export_type == "csv":
                response = HttpResponse(content_type='text/csv')
                response['Content-Disposition'] = 'attachment; filename="profiles.csv"'

                writer = csv.DictWriter(response, fieldnames=export_rows[0].keys())
                writer.writeheader()
                writer.writerows(export_rows)

                return response
            if export_type == "excel":
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Profiles"

                headers = export_rows[0].keys()
                ws.append(list(headers))

                for row in export_rows:
                    ws.append(list(row.values()))

                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = 'attachment; filename="profiles.xlsx"'

                wb.save(response)
                return response

        return JsonResponse({
            "Status": 1,
            "message": "Profiles fetched successfully.",
            "profiles": profiles,
            "total_count": total_count,
            "received_per_page": per_page,
            "received_page_number": page_number,
            "all_profile_ids": profile_with_indices,
            "search_result": "1"
        }, status=200)


def generate_pdf_from_template(template_name, context, filename):
    html_string = render_to_string(template_name, context)  # Removed "templates/"
    pdf_file = io.BytesIO()
    pisa_status = pisa.CreatePDF(io.StringIO(html_string), dest=pdf_file)

    if pisa_status.err:
        return JsonResponse({"status": "error", "message": "Error generating PDF."}, status=500)

    pdf_file.seek(0)
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response




def get_district_name(district_id):
    try:
        # Attempt to retrieve the city object using the string city_id
        district = models.District.objects.get(id=district_id)
        return district.name  # Return the city name if found
    except models.District.DoesNotExist:
        return district_id  # Return city_id if the city does not exist
    except Exception as e:
        return district_id 

def get_degree_name(degree_ids, other_degree,highest_edu,field_ofstudy_id,about_edu):
        if not degree_ids:
            
            return other_degree if other_degree else None

        try:
            id_list = [int(x) for x in str(degree_ids).split(',') if x.strip().isdigit()]
            id_list = [x for x in id_list if x != 86]
            degree_names = list(
                models.MasterhighestEducation.objects.filter(id__in=id_list)
                .values_list("degeree_name", flat=True)
            )
            print(" degree")
            if other_degree:
                degree_names.append(other_degree)
            final_names = ", ".join(degree_names) if degree_names else None
            if final_names:
                return final_names
            else:
                if highest_edu:
                    highest_education = models.EducationLevel.objects.filter(RowId=highest_edu).values_list('EducationLevel', flat=True).first() or "Unknown"
                    if field_ofstudy_id:
                        fieldof_study = models.Profilefieldstudy.objects.filter(id=field_ofstudy_id).values_list('field_of_study', flat=True).first() or "Unknown"
                    
                        about_edu=about_edu
                    
                        final_education = (highest_education + ' ' + fieldof_study).strip() or about_edu
                        return final_education
                    return highest_education
                return "N/A"
        except Exception:
            return "N/A"
        
def degree(degree_ids,other_degree):
    if not degree_ids:      
            return other_degree if other_degree else None
    try:
        id_list = [int(x) for x in str(degree_ids).split(',') if x.strip().isdigit()]
        id_list = [x for x in id_list if x != 86]
        degree_names = list(
            models.MasterhighestEducation.objects.filter(id__in=id_list)
            .values_list("degeree_name", flat=True)
        )
        if other_degree:
            degree_names.append(other_degree)
        final_names = ", ".join(degree_names) if degree_names else None
        
        if final_names==None:
            return "N/A"
        else:
            return final_names
    except Exception:
        final_names = "N/A"
        return final_names

def get_primary_sign(value):
    if not value:
        return "N/A"
    return value.split('/')[0]

def cm_to_feet_inches(cm):
    if cm is None or cm == "":
        return "N/A"

    try:
        cm = float(cm)
    except ValueError:
        return "N/A"

    total_inches = cm / 2.54
    feet = int(total_inches // 12)
    inches = round(total_inches % 12)

    return f"{feet} ft {inches} in"

def get_work_address(city, district, state, country):
    try:
        parts = []
        if city:
                parts.append(city)
        else:
            if district:
                parts.append(get_district_name(district))
        if state:
            parts.append(get_state_name(state))
        if country:
            parts.append(get_country_name(country))

        return "-".join(parts) if parts else "N/A"
    except Exception:
        return " "
    
def is_valid_file(url):
    try:
        response = requests.head(url, timeout=3)
        return response.status_code == 200
    except requests.RequestException:
        return False

class AdminProfilePDFView(APIView):
    def get(self, request, profile_id=None, pdf_format=None):
        profile_id = profile_id or request.query_params.get('profile_id')
        format_type = pdf_format or request.query_params.get('pdf_format')
        
        # get details
        login = get_object_or_404(models.Registration1, ProfileId=profile_id)
        family = models.ProfileFamilyDetails.objects.filter(profile_id=profile_id).first()
        horoscope_data = get_object_or_404(models.ProfileHoroscope, profile_id=profile_id)
        education_details = get_object_or_404(models.ProfileEduDetails, profile_id=profile_id)
        family_details = models.ProfileFamilyDetails.objects.filter(profile_id=profile_id)
        if family_details.exists():
                family_detail = family_details.first()  

                father_name = family_detail.father_name  
                father_occupation = family_detail.father_occupation
                family_status = family_detail.family_status
                mother_name = family_detail.mother_name
                mother_occupation = family_detail.mother_occupation
                no_of_sis_married = family_detail.no_of_sis_married
                no_of_bro_married = family_detail.no_of_bro_married
                suya_gothram = family_detail.suya_gothram
                madulamn = family_detail.madulamn if family_detail.madulamn != None else "N/A" 
                no_of_sister = family_detail.no_of_sister or "No"
                no_of_brother = family_detail.no_of_brother or "No"
        else:
            # Handle case where no family details are found
            father_name = father_occupation = family_status = ""
            mother_name = mother_occupation = ""
            no_of_sis_married = no_of_bro_married = 0

        try:
            num_sisters_married = int(no_of_sis_married)
        except ValueError:
            num_sisters_married = 0     
    
        try:
            num_brothers_married = int(no_of_bro_married)
        except ValueError:
            num_brothers_married = 0   
        if int(num_sisters_married) == 0:
            no_of_sis_married = "No"

        if  int(num_brothers_married) == 0:
            no_of_bro_married="No"
        if no_of_sister=="0" or no_of_sister =='':
            no_of_sis_married="No"
            no_of_sister ='No'

        if no_of_brother=="0" or no_of_brother =='':
            no_of_bro_married="No"
            no_of_brother ='No'
            
        complexion_id = login.Profile_complexion
        complexion = "Unknown"
        if complexion_id:
            complexion = models.Complexion.objects.filter(complexion_id=complexion_id).values_list('complexion_desc', flat=True).first() or "Unknown"

        # Safely handle education level
        highest_education_id = education_details.highest_education
        highest_education = "Unknown"
        if highest_education_id:
            highest_education = models.EducationLevel.objects.filter(row_id=highest_education_id).values_list('EducationLevel', flat=True).first() or "Unknown"
        
        field_ofstudy_id = education_details.field_ofstudy
        fieldof_study=" "
        if field_ofstudy_id:
            fieldof_study = models.Profilefieldstudy.objects.filter(id=field_ofstudy_id).values_list('field_of_study', flat=True).first() or "Unknown"
        
        about_edu=education_details.about_edu
        
        final_education = (highest_education + ' ' + fieldof_study).strip() or about_edu
        try:
            degree= get_degree_name(education_details.degree,education_details.other_degree,education_details.highest_education,education_details.field_ofstudy,education_details.about_edu)
        except Exception:
            degree=None

        annual_income = "Unknown"
        actual_income = str(education_details.actual_income).strip()
        annual_income_id = education_details.anual_income

        if not actual_income or actual_income in ["", "~"]:
            if annual_income_id and str(annual_income_id).isdigit():
                annual_income = models.AnnualIncome.objects.filter(id=int(annual_income_id)).values_list('income', flat=True).first() or "Unknown"
        else:
            annual_income = actual_income


        profession_id = education_details.profession
        profession = "Unknown"
        if profession_id:
            profession = models.Profespref.objects.filter(RowId=profession_id).values_list('profession', flat=True).first() or "Unknown"

        work_place =get_work_address(city=education_details.work_city,state=education_details.work_state,district=education_details.work_district,country=education_details.work_country)
        occupation_title=''
        occupation=''

        try:
            prof_id_int = int(profession_id)
            if prof_id_int == 1:
                occupation_title = 'Employment Details'
                occupation = f"{education_details.company_name or ''} / {education_details.designation or ''}"
            elif prof_id_int == 2:
                occupation_title = 'Business Details'
                occupation = f"{education_details.business_name or ''} / {education_details.nature_of_business or ''}"
        except (ValueError, TypeError):
            occupation_title = 'Other'
            occupation = ''
        
    

        #father_occupation_id = family_detail.father_occupation
        father_occupation = family_detail.father_occupation or "N/A"

            #mother_occupation_id = family_detail.mother_occupation
        mother_occupation = family_detail.mother_occupation or "N/A"
        father_name = family_detail.father_name or "N/A"
        mother_name = family_detail.mother_name or "N/A"
        family_status = "Unknown"
        family_status_id = family_detail.family_status

        if family_status_id:
            family_status = models.FamilyStatus.objects.filter(id=family_status_id).values_list('status', flat=True).first() or "Unknown"

        def safe_get_value(model, pk_field, value, name_field='name', default='N/A'):
                    try:
                        if value and str(value).isdigit():
                            return model.objects.filter(**{pk_field: value}).values_list(name_field, flat=True).first() or default
                    except Exception:
                        pass
                    return default

        if horoscope_data.horoscope_file:
                    horoscope_image_url = horoscope_data.horoscope_file.url
                    if is_valid_file(horoscope_image_url):
                        if horoscope_image_url.lower().endswith(('.png', '.jpg', '.jpeg', '.gif')):
                            horoscope_content = f'<img src="{horoscope_image_url}" alt="Horoscope Image" style="max-width: 200%; height: auto;">'
                        else:
                            horoscope_content = f'<a href="{horoscope_image_url}" download>Download Horoscope File</a>'
                    else:
                        horoscope_content = "empty"
        else:
            horoscope_content = "empty"
            
        if horoscope_data.horoscope_file_admin:
                    horoscope_image_url = horoscope_data.horoscope_file_admin.url
                    if is_valid_file(horoscope_image_url):
                        if horoscope_image_url.lower().endswith(('.png', '.jpg', '.jpeg', '.gif')):
                            horoscope_content_admin = f'<img src="{horoscope_image_url}" alt="Horoscope Image" style="max-width: 200%; height: auto;">'
                        else:
                            horoscope_content_admin = f'<a href="{horoscope_image_url}" download>Download Horoscope File</a>'
                    else:
                        horoscope_content_admin = "empty"
        else:
            horoscope_content_admin = "empty"
                # Get matching stars data
        birthstar = safe_get_value(models.BirthStar, 'id', horoscope_data.birthstar_name, 'star')
        birth_rasi = get_primary_sign(safe_get_value(models.Rasi, 'id', horoscope_data.birth_rasi_name, 'name'))

        complexion_id = login.Profile_complexion
        complexion = safe_get_value(models.Complexion, 'complexion_id', complexion_id, 'complexion_desc')
        father_name = family.father_name if family else "N/A"
        if not format_type:
            return JsonResponse({"status": "error", "message": "format is required"}, status=400)

        if not profile_id:
            return JsonResponse({"status": "error", "message": "profile_id is required"}, status=400)
        birth_star_id = horoscope_data.birthstar_name
        birth_rasi_id = horoscope_data.birth_rasi_name
        gender = login.Gender
        porutham_data = models.MatchingStarPartner.get_matching_stars_pdf(birth_rasi_id, birth_star_id, gender)
        didi = horoscope_data.didi or "Not specified"
        nalikai = horoscope_data.nalikai or "Not specified"
        lagnam="Unknown"
        try:
            if horoscope_data.lagnam_didi and str(horoscope_data.lagnam_didi).isdigit() and int(horoscope_data.lagnam_didi) > 0:
                lagnam = models.Rasi.objects.filter(pk=int(horoscope_data.lagnam_didi)).first()
                lagnam= get_primary_sign(lagnam.name)
        except models.Rasi.DoesNotExist:
            lagnam = "Unknown"
        def format_time_am_pm(time_str):
            if not time_str:  # Handles None or empty strings
                return "N/A"
            try:
                time_obj = datetime.strptime(str(time_str), "%H:%M:%S")
                return time_obj.strftime("%I:%M %p")  # 12-hour format with AM/PM
            except ValueError:
                return str(time_str)

        dob = login.Profile_dob
        age = calculate_age(dob) if dob else "N/A"

        birth_time=format_time_am_pm(horoscope_data.time_of_birth)
        image_status = models.Image_Upload.get_image_status(profile_id=profile_id)
        horo_hint = horoscope_data.horoscope_hints or "N/A"
        # Prepare the Porutham sections for the PDF
        def format_star_names(poruthams):
            return ', '.join([item['matching_starname'] for item in poruthams])
        if horoscope_data.rasi_kattam or  horoscope_data.amsa_kattam:
            rasi_kattam_data = parse_data(horoscope_data.rasi_kattam)
            amsa_kattam_data = parse_data(horoscope_data.amsa_kattam)

        else:
            rasi_kattam_data=parse_data('{Grid 1: empty, Grid 2: empty, Grid 3: empty, Grid 4: empty, Grid 5: empty, Grid 6: empty, Grid 7: empty, Grid 8: empty, Grid 9: empty, Grid 10: empty, Grid 11: empty, Grid 12: empty}')
            amsa_kattam_data=parse_data('{Grid 1: empty, Grid 2: empty, Grid 3: empty, Grid 4: empty, Grid 5: empty, Grid 6: empty, Grid 7: empty, Grid 8: empty, Grid 9: empty, Grid 10: empty, Grid 11: empty, Grid 12: empty}')

        if all(not str(val).strip() for val in [
            login.Profile_address,
            get_district_name(login.Profile_district),
            get_city_name(login.Profile_city),
            login.Profile_pincode
        ]):
            address_content = f"""
                <p>Not Specified</p>"""
        else:
            address_content = f"""
                <p>{login.Profile_address}</p>
                <p>{get_district_name(login.Profile_district)}, {get_city_name(login.Profile_city)}</p>
                <p>{login.Profile_pincode}.</p>
            """
        mobile_email_content = f"""
                        <p>Mobile: {login.Mobile_no or ''}</p>
                        <p>WhatsApp: {login.Profile_whatsapp or ''}</p>
                        <p>Email: {login.EmailId or ''}</p>
                """
        # Ensure that we have exactly 12 values for the grid
        rasi_kattam_data.extend([default_placeholder] * (12 - len(rasi_kattam_data)))
        amsa_kattam_data.extend([default_placeholder] * (12 - len(amsa_kattam_data)))   
        dasa_day = dasa_month = dasa_year = 0
        dasa_balance_str=dasa_format_date(horoscope_data.dasa_balance)
        match = re.match(
                        r"(?:(\d{2})/(\d{2})/(\d{2}))|(?:(\d+)\s+Year[s]?,\s+(\d+)\s+Month[s]?,\s+(\d+)\s+Day[s]?)",
                        dasa_balance_str or ""
                    )
        if match:
            if match.group(1):
                dasa_year, dasa_month, dasa_day = match.group(1), match.group(2), match.group(3)
            else:
                dasa_year, dasa_month, dasa_day = match.group(4), match.group(5), match.group(6) 
            
        date =  format_date_of_birth(login.Profile_dob)
        context_data = {
            "profile_id": login.ProfileId,
            "name": login.Profile_name,
            "height":cm_to_feet_inches(login.Profile_height),
            "image_status":image_status,
            "dob": date,
            "age":age,
            "didi":didi,
            "nalikai":nalikai,
            "father_name": father_name if father_name not in [None, ""] else "N/A" ,
            "suya_gothram":suya_gothram if suya_gothram not in [None, ""] else "N/A",
            "madulamn":madulamn if madulamn not in [None, ""] else "N/A",
            "work_place":work_place if work_place not in [None, ""] else "N/A",
            "highest_education":final_education if final_education not in [None, ""] else "N/A",
            'degree':degree if degree not in [None,""] else "N/A",
            "annual_income":annual_income if annual_income not in [None, ""] else "N/A",
            "father_occupation":father_occupation if father_occupation not in [None, ""] else "N/A",
            "family_status":family_status if family_status not in [None, ""] else "N/A",
            "no_of_brother_married":no_of_bro_married if no_of_bro_married not in [None, ""] else "N/A",
            "no_of_sister": no_of_sister if no_of_sister not in [None, ""] else "No",
            "no_of_brother": no_of_brother if no_of_brother not in [None, ""] else "No ",
            "mother_name":mother_name if mother_name not in [None, ""] else "N/A",
            "mother_occupation":mother_occupation if mother_occupation not in [None, ""] else "N/A",
            "no_of_sister_married":no_of_sis_married if no_of_sis_married not in [None, ""] else "N/A",
            "contact": login.Mobile_no if login.Mobile_no not in [None, ""] else "N/A",
            "alternate_number":login.Profile_alternate_mobile if login.Profile_alternate_mobile not in [None, ""] else "N/A",
            "whatsapp": login.Profile_whatsapp,
            "email":login.EmailId,
            "complexion": complexion if complexion not in [None, ""] else "N/A",
            "birth_star": birthstar if birthstar not in [None, ""] else "N/A",
            "birth_rasi": birth_rasi if birth_rasi not in [None, ""] else "N/A",
            "birth_place": horoscope_data.place_of_birth if horoscope_data.place_of_birth not in [None, ""] else "N/A",
            "address": address_content,
            "padham":horoscope_data.padham if horoscope_data.padham not in [None,""] else None,
            "lagnam":lagnam,
            "dasa_year":dasa_year,
            "dasa_month":dasa_month,
            "dasa_day":dasa_day,
            "dasa_name":get_dasa_name(horoscope_data.dasa_name),
            "occupation":occupation,
            "birth_start":birth_time,
            "occupation_title":occupation_title,
            "profession":profession,
            "horoscope_content": horoscope_content,
            "horoscope_content_admin":horoscope_content_admin,
            "horo_hint":horo_hint,
            "rasi_kattam_data": rasi_kattam_data,
            "amsa_kattam_data": amsa_kattam_data,
            "mobile_content":mobile_email_content,
            "porutham_stars": OrderedDict([
                ("9", format_star_names(porutham_data.get("9 Poruthams"))),
                ("8", format_star_names(porutham_data.get("8 Poruthams"))),
                ("7", format_star_names(porutham_data.get("7 Poruthams"))),
                ("6", format_star_names(porutham_data.get("6 Poruthams"))),
                ("5", format_star_names(porutham_data.get("5 Poruthams"))),
            ]),
            "view_profile_url": f"https://www.vysyamala.com/ProfileDetails?id={login.ProfileId}"
        }


        try:
            if format_type == "withoutcontact":
                return generate_pdf_from_template("without_contact.html", context_data, f"profile_with_contact_{profile_id}.pdf")
            elif format_type == "withoutaddress":
                return generate_pdf_from_template("without_address.html", context_data, f"profile_with_contact_{profile_id}.pdf")
            elif format_type == "withaddress":
                return generate_pdf_from_template("with_address.html", context_data, f"profile_with_contact_{profile_id}.pdf")
            elif format_type == "withonlystar":
                return generate_pdf_from_template("with_star_list.html", context_data, f"profile_with_contact_{profile_id}.pdf")
            # elif format_type == "withintimationlist":
            #     return generate_pdf_from_template("with_intimation_list.html", context_data, f"profile_with_contact_{profile_id}.pdf")
            elif format_type == "withcontactonly":
                return generate_pdf_from_template("with_contact_only.html", context_data, f"profile_with_contact_{profile_id}.pdf")
            elif format_type == "withoutcontactonly":
                return generate_pdf_from_template("without_contact_only.html", context_data, f"profile_with_contact_{profile_id}.pdf")
            
            else:
                return JsonResponse({"status": "error", "message": "Invalid format"}, status=400)

        except Exception as e:
            print(f"error{str(e)}")
            return JsonResponse({"status": "error", "message": str(e)}, status=500)

class AdminMatchProfilePDFView(APIView):
    def get(self, request, profile_ids=None, pdf_format=None, profile_to=None,action_type=None):
        # Prefer path params; fall back to query params for flexibility
        profile_ids = profile_ids or request.query_params.get('profile_ids')
        format_type = pdf_format or request.query_params.get('pdf_format')  
        profile_to = profile_to or request.query_params.get('profile_to')
        action = action_type or request.query_params.get('action_type')
        if format_type == "whatsapp_link_profile" or format_type == "whatsapp_link_profile_img":
            
            print('whatapp link profiles')
            whatsapp_profiles = []
            
        if not profile_ids:
            return JsonResponse({"status": "error", "message": "profile_id is required"}, status=400)
        elif not format_type:
            return JsonResponse({"status": "error", "message": "format is required"}, status=400)
        profile_ids_list = [pid.strip() for pid in profile_ids.split(',') if pid.strip()]
        pdf_merger = PdfMerger()
        errors = []

        for profile_id in profile_ids_list:
            try:
                login = get_object_or_404(models.Registration1, ProfileId=profile_id)
                family = models.ProfileFamilyDetails.objects.filter(profile_id=profile_id).first()
                horoscope_data = get_object_or_404(models.ProfileHoroscope, profile_id=profile_id)
                education_details = get_object_or_404(models.ProfileEduDetails, profile_id=profile_id)
                family_details = models.ProfileFamilyDetails.objects.filter(profile_id=profile_id)
                login_my  = get_object_or_404(models.Registration1, ProfileId=profile_to)
                horoscope_my = get_object_or_404(models.ProfileHoroscope, profile_id=profile_to)
                education_my = get_object_or_404(models.ProfileEduDetails, profile_id=profile_to)
                if family_details.exists():
                        family_detail = family_details.first()  

                        father_name = family_detail.father_name  
                        father_occupation = family_detail.father_occupation
                        family_status = family_detail.family_status
                        mother_name = family_detail.mother_name
                        mother_occupation = family_detail.mother_occupation
                        no_of_sis_married = family_detail.no_of_sis_married
                        no_of_bro_married = family_detail.no_of_bro_married
                        suya_gothram = family_detail.suya_gothram
                        madulamn = family_detail.madulamn if family_detail.madulamn != None else "N/A" 
                        no_of_sister = family_detail.no_of_sister or "No"
                        no_of_brother = family_detail.no_of_brother or "No"
                else:
                    # Handle case where no family details are found
                    father_name = father_occupation = family_status = ""
                    mother_name = mother_occupation = ""
                    no_of_sis_married = no_of_bro_married = 0

                try:
                    num_sisters_married = int(no_of_sis_married)
                except ValueError:
                    num_sisters_married = 0     
            
                try:
                    num_brothers_married = int(no_of_bro_married)
                except ValueError:
                    num_brothers_married = 0   
                if int(num_sisters_married) == 0:
                    no_of_sis_married = "No"

                if  int(num_brothers_married) == 0:
                    no_of_bro_married="No"
                if no_of_sister=="0" or no_of_sister =='':
                    no_of_sis_married="No"
                    no_of_sister ='No'

                if no_of_brother=="0" or no_of_brother =='':
                    no_of_bro_married="No"
                    no_of_brother ='No'

                try:
                    degree= get_degree_name(education_details.degree,education_details.other_degree,education_details.highest_education,education_details.field_ofstudy,education_details.about_edu)
                except Exception:
                    degree=None
                complexion_id = login.Profile_complexion
                complexion = "Unknown"
                if complexion_id:
                    complexion = models.Complexion.objects.filter(complexion_id=complexion_id).values_list('complexion_desc', flat=True).first() or "Unknown"

                # Safely handle education level
                highest_education_id = education_details.highest_education
                highest_education = "Unknown"
                if highest_education_id:
                    highest_education = models.EducationLevel.objects.filter(row_id=highest_education_id).values_list('EducationLevel', flat=True).first() or "Unknown"
                
                field_ofstudy_id = education_details.field_ofstudy
                fieldof_study=" "
                if field_ofstudy_id:
                    fieldof_study = models.Profilefieldstudy.objects.filter(id=field_ofstudy_id).values_list('field_of_study', flat=True).first() or "Unknown"
                
                about_edu=education_details.about_edu
                
                final_education = (highest_education + ' ' + fieldof_study).strip() or about_edu
                
                
                highest_education_id_my = education_my.highest_education
                highest_education_my="N/A"
                if highest_education_id_my:
                    highest_education_my = models.EducationLevel.objects.filter(row_id=highest_education_id_my).values_list('EducationLevel', flat=True).first() or "N/A"

                field_ofstudy_id_my = education_my.field_ofstudy
                fieldof_study_my=" "
                if field_ofstudy_id_my:
                    fieldof_study_my = models.Profilefieldstudy.objects.filter(id=field_ofstudy_id_my).values_list('field_of_study', flat=True).first() or "N/A"
                
                about_edu=education_my.about_edu
                
                final_education_my = (highest_education_my + ' ' + fieldof_study_my).strip() or about_edu

                annual_income = "Unknown"
                actual_income = str(education_details.actual_income).strip()
                annual_income_id = education_details.anual_income

                if not actual_income or actual_income in ["", "~"]:
                    if annual_income_id and str(annual_income_id).isdigit():
                        annual_income = models.AnnualIncome.objects.filter(id=int(annual_income_id)).values_list('income', flat=True).first() or "Unknown"
                else:
                    annual_income = actual_income


                profession_id = education_details.profession
                profession = "Unknown"
                if profession_id:
                    profession = models.Profespref.objects.filter(RowId=profession_id).values_list('profession', flat=True).first() or "Unknown"

                work_place =get_work_address(city=education_details.work_city,state=education_details.work_state,district=education_details.work_district,country=education_details.work_country)
                occupation_title=''
                occupation=''

                try:
                    prof_id_int = int(profession_id)
                    if prof_id_int == 1:
                        occupation_title = 'Employment Details'
                        occupation = f"{education_details.company_name or ''} / {education_details.designation or ''}"
                    elif prof_id_int == 2:
                        occupation_title = 'Business Details'
                        occupation = f"{education_details.business_name or ''} / {education_details.nature_of_business or ''}"
                except (ValueError, TypeError):
                    occupation_title = 'Other'
                    occupation = ''

                #father_occupation_id = family_detail.father_occupation
                father_occupation = family_detail.father_occupation or "N/A"

                    #mother_occupation_id = family_detail.mother_occupation
                mother_occupation = family_detail.mother_occupation or "N/A"
                father_name = family_detail.father_name or "N/A"
                mother_name = family_detail.mother_name or "N/A"
                family_status = "Unknown"
                family_status_id = family_detail.family_status

                if family_status_id:
                    family_status = models.FamilyStatus.objects.filter(id=family_status_id).values_list('status', flat=True).first() or "Unknown"

                def safe_get_value(model, pk_field, value, name_field='name', default='N/A'):
                    try:
                        if value and str(value).isdigit():
                            return model.objects.filter(**{pk_field: value}).values_list(name_field, flat=True).first() or default
                    except Exception:
                        pass
                    return default

                if horoscope_data.horoscope_file:
                    horoscope_image_url = horoscope_data.horoscope_file.url
                    if is_valid_file(horoscope_image_url):
                        if horoscope_image_url.lower().endswith(('.png', '.jpg', '.jpeg', '.gif')):
                            horoscope_content = f'<img src="{horoscope_image_url}" alt="Horoscope Image" style="width: 100%; height: auto;">'
                        else:
                            horoscope_content = f'<a href="{horoscope_image_url}" download>Download Horoscope File</a>'
                    else:
                        horoscope_content = "empty"
                else:
                    horoscope_content = "empty"
                if horoscope_data.horoscope_file_admin:
                    horoscope_image_url = horoscope_data.horoscope_file_admin.url
                    if is_valid_file(horoscope_image_url):
                        if horoscope_image_url.lower().endswith(('.png', '.jpg', '.jpeg', '.gif')):
                            horoscope_content_admin = f'<img src="{horoscope_image_url}" alt="Horoscope Image" style="width: 100%; height: auto;">'
                        else:
                            horoscope_content_admin = f'<a href="{horoscope_image_url}" download>Download Horoscope File</a>'
                    else:
                        horoscope_content_admin = "empty"
                else:
                    horoscope_content_admin = "empty"
                birthstar = safe_get_value(models.BirthStar, 'id', horoscope_data.birthstar_name, 'star')
                birth_rasi = get_primary_sign(safe_get_value(models.Rasi, 'id', horoscope_data.birth_rasi_name, 'name'))

                birthstar_my = safe_get_value(models.BirthStar, 'id', horoscope_my.birthstar_name, 'star')
                birth_rasi_my = get_primary_sign(safe_get_value(models.Rasi, 'id', horoscope_my.birth_rasi_name, 'name'))

                complexion_id = login.Profile_complexion
                complexion = safe_get_value(models.Complexion, 'complexion_id', complexion_id, 'complexion_desc')
                father_name = family.father_name if family else "N/A"

                birth_star_id = horoscope_data.birthstar_name
                birth_rasi_id = horoscope_data.birth_rasi_name
                gender = login.Gender

                porutham_data1 = models.MatchingStarPartner.get_matching_stars_pdf(birth_rasi_id, birth_star_id, gender)
                porutham_data = fetch_porutham_details(profile_id, profile_to)
                didi = horoscope_data.didi or "Not specified"
                nalikai = horoscope_data.nalikai or "Not specified"
                def format_time_am_pm(time_str):
                    if not time_str:  # Handles None or empty strings
                        return "N/A"
                    try:
                        time_obj = datetime.strptime(str(time_str), "%H:%M:%S")
                        return time_obj.strftime("%I:%M %p")  # 12-hour format with AM/PM
                    except ValueError:
                        return str(time_str)
                    
                birth_time=format_time_am_pm(horoscope_data.time_of_birth)
                my_birth_time=format_time_am_pm(horoscope_my.time_of_birth)
                horo_hint = horoscope_data.horoscope_hints or "N/A"
                valid_rows = []
                # Define the HTML content with custom styles
                porutham_rows = ""
                for porutham in porutham_data['porutham_results']:
                    if 'porutham_name' not in porutham or 'status' not in porutham:
                        continue  # skip invalid rows
                    valid_rows.append(porutham)

                # Step 2: Use actual count for rowspan
                rowspan_count = len(valid_rows)

                # Step 3: Build HTML rows
                for idx, porutham in enumerate(valid_rows):
                    extra_td = ""
                    if idx == 0:
                        extra_td = (
                            f"<td rowspan='{rowspan_count}'>"
                            f"<p class='matching-score'>{porutham_data['matching_score']}</p>"
                            f"<p style='font-weight:500; font-size:13px;'>Please check with your astrologer for detailed compatibility.</p>"
                            f"<p style='margin-top:10px;'>Jai Vasavi</p>"
                            f"</td>"
                        )

                    # Color logic
                    span_color = 'green' if porutham['status'].startswith('YES') else 'red'
                    if format_type not in ["match_compatability_color", "match_compatability_without_horo"]:
                        span_color = 'black'

                    porutham_rows += (
                        f"<tr>"
                        f"<td>{porutham['porutham_name']}</td>"
                        f"<td><span style='color: {span_color};'>{porutham['status']}</span></td>"
                        f"{extra_td}"
                        f"</tr>"
                    )

                porutham_show=True
                if porutham_data['matching_score']=='0/10' or porutham_data['matching_score']=='0' or porutham_data['matching_score']=='0.0' or porutham_data['matching_score']=='10/10' or porutham_data['matching_score']==0.0:
                    porutham_show= False
                # print("porutham show:",porutham_show) 
                def format_star_names(poruthams):
                    if not poruthams:
                        return "N/A"
                    return ', '.join([item['matching_starname'] for item in poruthams])

                if horoscope_data.rasi_kattam or horoscope_data.amsa_kattam:
                    rasi_kattam_data = parse_data(horoscope_data.rasi_kattam)
                    amsa_kattam_data = parse_data(horoscope_data.amsa_kattam)
                else:
                    rasi_kattam_data = parse_data('{Grid 1: empty, Grid 2: empty, Grid 3: empty, Grid 4: empty, Grid 5: empty, Grid 6: empty, Grid 7: empty, Grid 8: empty, Grid 9: empty, Grid 10: empty, Grid 11: empty, Grid 12: empty}')
                    amsa_kattam_data = parse_data('{Grid 1: empty, Grid 2: empty, Grid 3: empty, Grid 4: empty, Grid 5: empty, Grid 6: empty, Grid 7: empty, Grid 8: empty, Grid 9: empty, Grid 10: empty, Grid 11: empty, Grid 12: empty}')

                if all(not str(val).strip() for val in [
                    login.Profile_address,
                    get_district_name(login.Profile_district),
                    get_city_name(login.Profile_city),
                    login.Profile_pincode
                ]):
                    address_content = f"""
                        <p>Not Specified</p>"""
                else:
                    address_content = f"""
                        <p>{login.Profile_address}</p>
                        <p>{get_district_name(login.Profile_district)}, {get_city_name(login.Profile_city)}</p>
                        <p>{login.Profile_pincode}.</p>
                    """
                try:
                    rasi = models.Rasi.objects.get(pk=horoscope_data.birth_rasi_name)
                    rasi_name = rasi.name  # Or use rasi.tamil_series, telugu_series, etc. as per your requirement
                except Exception:
                    rasi_name = "Unknown"
                lagnam="Unknown"
                try:
                    if horoscope_data.lagnam_didi and str(horoscope_data.lagnam_didi).isdigit() and int(horoscope_data.lagnam_didi) > 0:
                        lagnam = models.Rasi.objects.filter(pk=int(horoscope_data.lagnam_didi)).first()
                        lagnam= get_primary_sign(lagnam.name) or "N/A"
                except Exception:
                    lagnam = "Unknown"
                    
                dob = login.Profile_dob
                age = calculate_age(dob) if dob else "N/A"
                image_status = models.Image_Upload.get_image_status(profile_id=profile_id)
                rasi_kattam_data.extend([default_placeholder] * (12 - len(rasi_kattam_data)))
                amsa_kattam_data.extend([default_placeholder] * (12 - len(amsa_kattam_data)))
                dasa_day, dasa_month, dasa_year = 0, 0, 0
                dasa_balance_str=dasa_format_date(horoscope_data.dasa_balance)
                # match = re.match(r"(\d+)\s+Years,\s+(\d+)\s+Months,\s+(\d+)\s+Days", dasa_balance_str or "")
                match = re.match(
                        r"(?:(\d{2})/(\d{2})/(\d{2}))|(?:(\d+)\s+Year[s]?,\s+(\d+)\s+Month[s]?,\s+(\d+)\s+Day[s]?)",
                        dasa_balance_str or ""
                    )
                if match:
                    if match.group(1):
                        dasa_year, dasa_month, dasa_day = match.group(1), match.group(2), match.group(3)
                    else:
                        dasa_year, dasa_month, dasa_day = match.group(4), match.group(5), match.group(6)
                # print("porutham",porutham_data)
                
                if format_type == "whatsapp_link_profile_img":
                    profile_link = f"https://app.vysyamala.com/auth/profile/{signing.dumps(profile_id)}/"
                    profile_data = {
                        "profile_link": profile_link,
                        "profile_id": login.ProfileId,
                        "profile_name": login.Profile_name or "N/A",
                        "highest_education_my":final_education_my if final_education_my not in [None, ""] else "N/A",
                        "highest_education":final_education if final_education not in [None, ""] else "N/A",
                        "annual_income":annual_income if annual_income not in [None, ""] else "N/A",
                        "occupation":occupation,
                        "work_place":work_place if work_place not in [None, ""] else "N/A",
                        "age": calculate_age(login.Profile_dob) if login.Profile_dob else "Unknown",
                        "star_name": birthstar if birthstar not in [None, ""] else "N/A",
                        "padham":horoscope_data.padham if horoscope_data.padham not in [None,""] else None
                    }
                    whatsapp_profiles.append(profile_data)

                if format_type == "whatsapp_link_profile":
                    profile_link = f"https://app.vysyamala.com/auth/profile_view/{signing.dumps(profile_id)}/"
                    profile_data = {
                        "profile_link": profile_link,
                        "profile_id": login.ProfileId,
                        "profile_name": login.Profile_name or "N/A",
                        "highest_education_my":final_education_my if final_education_my not in [None, ""] else "N/A",
                        "highest_education":final_education if final_education not in [None, ""] else "N/A",
                        "annual_income":annual_income if annual_income not in [None, ""] else "N/A",
                        "occupation":occupation,
                        "work_place":work_place if work_place not in [None, ""] else "N/A",
                        "age": calculate_age(login.Profile_dob) if login.Profile_dob else "Unknown",
                        "star_name": birthstar if birthstar not in [None, ""] else "N/A",
                        "padham":horoscope_data.padham if horoscope_data.padham not in [None,""] else None
                    }
                    whatsapp_profiles.append(profile_data)
                
                date =  format_date_of_birth(login.Profile_dob)
                my_date =  format_date_of_birth(login_my.Profile_dob)
                context_data = {
                    "profile_id": login.ProfileId,
                    "my_profile_id":login_my.ProfileId,
                    "name": login.Profile_name,
                    "my_name":login_my.Profile_name,
                    "dob": date,
                    "my_dob":my_date,
                    "age":age,
                    "image_status":image_status,
                    "height":cm_to_feet_inches(login.Profile_height),
                    "my_height":cm_to_feet_inches(login_my.Profile_height),
                    "didi":didi,
                    "nalikai":nalikai,
                    "degree":degree if degree not in [None,""] else "N/a",
                    "father_name": father_name if father_name not in [None, ""] else "N/A" ,
                    "suya_gothram":suya_gothram if suya_gothram not in [None, ""] else "N/A",
                    "madulamn":madulamn if madulamn not in [None, ""] else "N/A",
                    "work_place":work_place if work_place not in [None, ""] else "N/A",
                    "occupation_title":occupation_title,
                    "occupation":occupation,
                    "highest_education":final_education if final_education not in [None, ""] else "N/A",
                    "highest_education_my":final_education_my if final_education_my not in [None, ""] else "N/A",
                    "annual_income":annual_income if annual_income not in [None, ""] else "N/A",
                    "father_occupation":father_occupation if father_occupation not in [None, ""] else "N/A",
                    "family_status":family_status if family_status not in [None, ""] else "N/A",
                    "no_of_brother_married":no_of_bro_married if no_of_bro_married not in [None, ""] else "N/A",
                    "no_of_brother":no_of_brother if no_of_brother not in [None, ""] else "No",
                    "mother_name":mother_name if mother_name not in [None, ""] else "N/A",
                    "mother_occupation":mother_occupation if mother_occupation not in [None, ""] else "N/A",
                    "no_of_sister_married":no_of_sis_married if no_of_sis_married not in [None, ""] else "N/A",
                    "no_of_sister":no_of_sister if no_of_sister not in [None, ""] else "No",
                    "contact": login.Mobile_no if login.Mobile_no not in [None, ""] else "N/A",
                    "whatsapp": login.Profile_whatsapp,
                    "alternate_number":login.Profile_alternate_mobile if login.Profile_alternate_mobile not in [None, ""] else "N/A",
                    "email":login.EmailId,
                    "complexion": complexion if complexion not in [None, ""] else "N/A",
                    "birth_star": birthstar if birthstar not in [None, ""] else "N/A",
                    "birth_star_my": birthstar_my if birthstar_my not in [None, ""] else "N/A",
                    "birth_rasi": birth_rasi if birth_rasi not in [None, ""] else "N/A",
                    "birth_rasi_my": birth_rasi_my if birth_rasi_my not in [None, ""] else "N/A",
                    "birth_place": horoscope_data.place_of_birth if horoscope_data.place_of_birth not in [None, ""] else "N/A",
                    "my_birth_place": horoscope_my.place_of_birth if horoscope_my.place_of_birth not in [None, ""] else "N/A",
                    "address": address_content,
                    "lagnam":lagnam,
                    "dasa_year":dasa_year,
                    "dasa_month":dasa_month,
                    "dasa_day":dasa_day,
                    "dasa_name":get_dasa_name(horoscope_data.dasa_name),
                    "padham":horoscope_data.padham if horoscope_data.padham not in [None,""] else None,
                    "birth_start":birth_time,
                    "my_time_of_birth":my_birth_time,
                    "profession":profession,
                    "horoscope_content": horoscope_content,
                    "horoscope_content_admin":horoscope_content_admin,
                    "horo_hint":horo_hint,
                    "rasi_kattam_data": rasi_kattam_data,
                    "amsa_kattam_data": amsa_kattam_data,
                    "porutham_stars": OrderedDict([
                        ("9", format_star_names(porutham_data1.get("9 Poruthams")or [])),
                        ("8", format_star_names(porutham_data1.get("8 Poruthams")or [])),
                        ("7", format_star_names(porutham_data1.get("7 Poruthams")or [])),
                        ("6", format_star_names(porutham_data1.get("6 Poruthams")or [])),
                        ("5", format_star_names(porutham_data1.get("5 Poruthams")or [])),
                    ]),
                    "porutham_rows":porutham_rows ,
                    "porutham_show":porutham_show,
                    "view_profile_url": f"https://www.vysyamala.com/ProfileDetails?id={login.ProfileId}"
                }

                template_map = {
                    "match_full_profile": "full_profile.html",
                    "match_full_profile_black": "full_profile_black.html",
                    "match_compatability_color": "compatability_color.html",
                    "match_compatability_black": "compatability_black.html",
                    "match_compatability_without_horo": "compatability_only.html",
                    "match_compatability_without_horo_black": "compatability_only_black.html",
                    "whatsapp_link_profile": "whatsapp_profile.html",
                    "whatsapp_link_profile_img": "whatsapp_profile.html"
                }

                if format_type not in template_map:
                    return JsonResponse({"status": "error", "message": "Invalid format"}, status=400)

                html_string = render_to_string(template_map[format_type], context_data)
                pdf_buffer = io.BytesIO()
                pisa_status = pisa.CreatePDF(html_string, dest=pdf_buffer)

                if pisa_status.err:
                    print(pisa_status.log)  
                    errors.append(f"{profile_id} (pisa error: {pisa_status.err})")

                    continue

                pdf_buffer.seek(0)
                pdf_merger.append(pdf_buffer)
                try:
                    exitsting_logs = AdminPrintLogs.objects.filter(
                        profile_id=profile_to,
                        sentprofile_id=profile_id,
                        action_type=action,
                        format_type=format_type,
                        status=1
                    )
                    if exitsting_logs.exists():
                        exitsting_logs.update(
                            updated_at= timezone.now(),
                        )
                    else:
                        AdminPrintLogs.objects.create(
                            profile_id=profile_to,
                            sentprofile_id = profile_id,
                            action_type = action,
                            format_type = format_type,
                            sent_date= timezone.now(),
                            updated_at= timezone.now(), 
                            status=1
                        )
                except Exception as e:
                    print(f"Error for profile {profile_id}: {str(e)}")
                    pass
                    
                    
            except Exception as e:
                import traceback
                print(f"Error for profile {profile_id}: {str(e)}")
                traceback.print_exc()
                errors.append(profile_id)
                continue


        merged_pdf = io.BytesIO()
        pdf_merger.write(merged_pdf)
        pdf_merger.close()
        merged_pdf.seek(0)
        if format_type == "whatsapp_link_profile" or format_type == "whatsapp_link_profile_img" :

            return render(request, "whatsapp_profile.html", {"profiles": whatsapp_profiles})



            # html_string = render_to_string("whatsapp_profile.html", { "profiles": whatsapp_profiles })
            # pdf_buffer = io.BytesIO()
            # pisa_status = pisa.CreatePDF(html_string, dest=pdf_buffer)
            # if pisa_status.err:
            #     return JsonResponse({"status": "error", "message": "PDF generation failed."}, status=500)
            
            # pdf_buffer.seek(0)
            # response = HttpResponse(pdf_buffer.read(), content_type='application/pdf')
            # response['Content-Disposition'] = 'inline; filename="WhatsAppProfiles.pdf"'
            # return response  
        if not errors:
            if action == 'email':
                try:
                    if login_my.EmailId:
                        subject = "Matched Profiles PDF"
                        body = f"Dear User,\nPlease find the attached PDF for the matched profiles.\n\nRegards,\nVysyamala Team"
                        from_email = settings.DEFAULT_FROM_EMAIL
                        to_email = [login_my.EmailId]
                        email = EmailMessage(subject, body, from_email, to_email)
                        email.attach('MatchedProfiles.pdf', merged_pdf.read(), 'application/pdf')
                        email.send()

                        return JsonResponse({"status": "success", "message": "PDF emailed successfully."})
                    else:
                        return JsonResponse({"status": "Failed", "message": "No email address is associated with this profile. Please add one." })
                except Exception as e:
                    print(f"Email Error :{str(e)}")
                    pass
            else:
                response = HttpResponse(merged_pdf.read(), content_type='application/pdf')
                response['Content-Disposition'] = 'inline; filename="MatchedProfiles.pdf"'
                return response
        else:
            return JsonResponse({"status": "error", "message": f"PDF generated with errors for: {', '.join(errors)}"}, status=206)

class RenewalProfilesView(generics.ListAPIView):
    serializer_class = Renewalprofiledata
    pagination_class = StandardResultsPaging
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['ProfileId', 'Gender', 'EmailId', 'Profile_dob', 'Profile_city']


    def get_queryset(self):
        search_query = self.request.query_params.get('search', None)
        from_date = self.request.query_params.get('from_date', None)
        to_date = self.request.query_params.get('to_date', None)

        plan_ids="4,6,7,8,9"
        status_id = 1
            
        sql = """
                SELECT ld.ContentId, ld.ProfileId, ld.Profile_name, ld.Gender, ld.Mobile_no, ld.EmailId, 
                    ld.Profile_dob,  ld.Profile_whatsapp, ld.Profile_alternate_mobile, ld.Plan_id, ld.status, 
                    ld.DateOfJoin, ld.Last_login_date, ld.Profile_for, ms.MaritalStatus, cm.complexion_desc, s.name AS state_name, 
                    cy.city_name AS Profile_city, cy.city_name , c.name AS country_name, d.name AS district_name,pl.plan_name,
                    pfd.family_status, ped.highest_education,ped.degree,ped.other_degree, ped.anual_income, ph.birthstar_name , mp.profession AS profession,ld.membership_startdate,ld.membership_enddate
                FROM logindetails ld
                LEFT JOIN maritalstatusmaster ms ON ld.Profile_marital_status = ms.StatusId
                LEFT JOIN complexionmaster cm ON ld.Profile_complexion = cm.complexion_id
                LEFT JOIN masterstate s ON ld.Profile_state = s.id
                LEFT JOIN mastercity cy ON ld.Profile_city = cy.id
                LEFT JOIN mastercountry c ON ld.Profile_country = c.id
                LEFT JOIN masterdistrict d ON ld.Profile_district = d.name
                LEFT JOIN profile_familydetails pfd ON ld.ProfileId = pfd.profile_id
                LEFT JOIN profile_edudetails ped ON ld.ProfileId = ped.profile_id
                LEFT JOIN profile_horoscope ph ON ld.ProfileId = ph.profile_id 
                LEFT JOIN masterprofession mp ON ped.profession = mp.RowId
                LEFT JOIN plan_master pl ON ld.Plan_id = pl.id
                """
            
        where_clauses = ["ld.status = %s"]
        params = [status_id]

        if search_query:
            search_pattern = f'%{search_query}%'
            where_clauses.insert(0, """(
                ld.ProfileId LIKE %s OR
                ld.temp_profileid LIKE %s OR
                ld.Gender LIKE %s OR
                ld.Mobile_no LIKE %s OR
                ld.EmailId LIKE %s OR
                ms.MaritalStatus LIKE %s OR
                ld.Profile_dob LIKE %s OR
                cm.complexion_desc LIKE %s OR
                ld.Profile_address LIKE %s OR
                ld.Profile_country LIKE %s OR
                s.name LIKE %s OR
                cy.city_name LIKE %s OR
                ld.Profile_pincode LIKE %s
            )""")
            params = [search_pattern] * 13 + params


        # Add plan_id filter
        if plan_ids:
            plan_id_list = [int(pid.strip()) for pid in plan_ids.split(',') if pid.strip().isdigit()]
            if plan_id_list:
                placeholders = ','.join(['%s'] * len(plan_id_list))
                where_clauses.append(f"ld.Plan_id NOT IN ({placeholders})")
                params.extend(plan_id_list)

        valid_from = from_date and from_date.strip()
        valid_to = to_date and to_date.strip()

        if valid_from and valid_to:
            try:
                start_date = datetime.strptime(valid_from, "%Y-%m-%d").date()
                end_date = datetime.strptime(valid_to, "%Y-%m-%d").date()
                
                where_clauses.append("DATE(ld.membership_enddate) BETWEEN %s AND %s")
                params.extend([start_date, end_date])
            except ValueError:
                pass
        else:
            where_clauses.append("""
                (
                    ld.membership_enddate < CURDATE()
                    OR (
                        MONTH(ld.membership_enddate) = MONTH(CURDATE()) 
                        AND YEAR(ld.membership_enddate) = YEAR(CURDATE())
                    )
                )
            """)
            
        # Combine all WHERE clauses
        sql += " WHERE " + " AND ".join(where_clauses)
        sql += """
            ORDER BY ld.membership_enddate DESC
        """
        

        with connection.cursor() as cursor:
            cursor.execute(sql, params)
            rows = dictfetchall(cursor)  # Fetch rows as a dictionary
        # print("Final SQL:", sql)
        # print("Params:", params)

        return rows
    def list(self, request, *args, **kwargs):
        export_type = request.query_params.get('export')

        queryset = self.get_queryset()  

        if export_type == 'csv':
            return export_renew_csv("renewal_profiles", queryset)

        if export_type == 'xlsx':
            return export_renew_xlsx("renewal_profiles", queryset)

        return super().list(request, *args, **kwargs)
 
def export_renew_csv(filename, data):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'

    writer = csv.writer(response)

    if not data:
        return response

    headers = data[0].keys()
    writer.writerow(headers)

    for row in data:
        writer.writerow([normalize_value(row[h]) for h in headers])

    return response


def export_renew_xlsx(filename, data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Renewal Profiles"

    if not data:
        return wb

    headers = list(data[0].keys())
    ws.append(headers)

    for row in data:
        ws.append([normalize_value(row[h]) for h in headers])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    wb.save(response)

    return response
 
def normalize_value(val):
    if val in ("0000-00-00", "0000-00-00 00:00:00", None):
        return ""
    if isinstance(val, (date, datetime)):
        return val.strftime("%Y-%m-%d")
    return val

   
class CustomPageNumberPagination(PageNumberPagination):
    page_size = 10  # Default
    page_size_query_param = 'per_page'
    page_query_param = 'page_number'
    max_page_size = 100    

class LoginLogView(generics.ListAPIView):
    serializer_class = LoginLogSerializer
    pagination_class = CustomPageNumberPagination

    def get_queryset(self):
        qs = Registration1.objects.filter(Last_login_date__isnull=False)

        date_str = self.request.GET.get('date')
        start_date_str = self.request.GET.get('from_date')
        end_date_str = self.request.GET.get('to_date')
        plan = self.request.GET.get('plan')

        if date_str:
            try:
                date = datetime.strptime(date_str, '%Y-%m-%d').date()
                qs = qs.filter(
                    Last_login_date__startswith=date.strftime('%Y-%m-%d')
                )
            except ValueError:
                pass

        if start_date_str and end_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d') + timedelta(days=1)
                qs = qs.filter(Last_login_date__gte=start_date, Last_login_date__lt=end_date)
            except ValueError:
                pass
        if plan:
            try:
                qs = qs.filter(Plan_id=plan)
            except Exception:
                pass

        return qs.order_by('-Last_login_date')[:100]
    
    def get(self, request, *args, **kwargs):
        export_type = request.query_params.get("export")

        queryset = self.get_queryset()

        if export_type in ("csv", "xlsx"):
            data = queryset.values(
                "ProfileId",
                "Profile_name",
                "Gender",
                "EmailId",
                "Mobile_no",
                "Plan_id",
                "Last_login_date"
            )

            filename = f"login_logs_{datetime.today().date()}.{export_type}"

            if export_type == "csv":
                return export_login_csv(filename, data)

            return export_login_xlsx(filename, data)

        return super().get(request, *args, **kwargs)



class PlanSubscriptionCreateView(generics.CreateAPIView):
    queryset = PlanSubscription.objects.all()
    serializer_class = PlanSubscriptionSerializer

    def create(self, request, *args, **kwargs):

        # Step A: Admin user permission check
        owner_id = request.data.get('admin_user_id')

        try:
            owner_id = int(owner_id)
            user = User.objects.get(id=owner_id)
        except Exception:
            user = None

        if user:
            role = user.role
            perm = RolePermission.objects.filter(
                role=role,
                action__code="membership_activation"
            ).first()
            can_add = perm.value if perm else 0
        else:
            can_add = 0

        print("USER =", user)
        print("CAN_ADD =", can_add)

        # Stop API here (this works in CREATE, unlike perform_create)
        if not user or can_add != 1:
            return Response(
                {"status": "error", "message": "Permission denied"},
                status=status.HTTP_403_FORBIDDEN
            )

        # If allowed → continue normal DRF flow
        return super().create(request, *args, **kwargs)

    def perform_create(self, serializer):
        subscription = serializer.save()

        transaction = PaymentTransaction.objects.create(
            profile_id=subscription.profile_id,
            plan_id=subscription.plan_id,
            order_id=subscription.order_id,
            payment_id=subscription.payment_id,
            amount=subscription.paid_amount,
            status='2',
            created_at=timezone.now(),
            payment_type=subscription.payment_mode,
            discount_amont=subscription.discount,
            payment_refno=subscription.gpay_no,
            addon_package=subscription.addon_package
        )

        subscription.trans_id = transaction.id
        subscription.save(update_fields=["trans_id"])



# List subscriptions (only status=1)
class PlanSubscriptionListView(generics.ListAPIView):
    serializer_class = PlanSubscriptionListSerializer
 
    def get_queryset(self):
        profile_id = self.request.query_params.get("profile_id")  # get ?profile_id=123 from URL
        queryset = PlanSubscription.objects.filter(status=1)
 
        if profile_id:
            queryset = queryset.filter(profile_id=profile_id)
 
        return queryset
 
 
# Update subscription (PATCH/PUT)
class PlanSubscriptionUpdateView(generics.UpdateAPIView):
    queryset = PlanSubscription.objects.all()
    serializer_class = PlanSubscriptionSerializer
    lookup_field = "id"
    
    def update(self, request, *args, **kwargs):
        owner_id = request.data.get('admin_user_id')
        try:
            owner_id = int(owner_id)
            user = User.objects.get(id=owner_id)
        except Exception:
            user = None

        if user:
            role = user.role
            permissions = RolePermission.objects.filter(role=role).select_related('action')
            data = permissions.values('action__code', 'value')
            edit_permission = data.filter(action__code='membership_activation').first()
            edit = edit_permission['value'] if edit_permission else None
        else:
            edit = None

        if user:
            if edit == 1:
                return super().update(request, *args, **kwargs)
            else:
                return Response({
                    "status": "error",
                    "message": "Permission Error"
                }, status=status.HTTP_403_FORBIDDEN)
        else:
            return super().update(request, *args, **kwargs)


@api_view(["POST"])
def process_transaction(request):
    """
    Accept or Reject a payment transaction.
    """
    transaction_id = request.data.get("transaction_id")
    action = request.data.get("action")  # accept / reject
    payment_for = request.data.get("payment_for")
    admin_user = request.data.get("admin_user")
    admin_user_id = request.data.get("admin_user_id")

    if not transaction_id or not action:
        return Response(
            {"status": "error", "message": "transaction_id and action are required"},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        admin_user_id = int(admin_user_id)
        user = User.objects.get(id=admin_user_id)
    except Exception:
        user = None

    if user:
        role = user.role
        permissions = RolePermission.objects.filter(role=role).select_related('action')
        data = permissions.values('action__code', 'value')
        edit_permission = data.filter(action__code='membership_activation').first()
        edit = edit_permission['value'] if edit_permission else None
    else:
        edit = None

    if user:
        if edit == 1:
            pass
        else:
            return Response({
                "status": "error",
                "message": "Permission Error"
            }, status=status.HTTP_403_FORBIDDEN)

    try:
        transaction = PaymentTransaction.objects.get(id=transaction_id)
    except PaymentTransaction.DoesNotExist:
        return Response(
            {"status": "error", "message": "Transaction not found"},
            status=status.HTTP_404_NOT_FOUND
        )
   
 
    # Build payment_for string
    plan_name = None
    addon_name = None
 
    if transaction.plan_id:
        try:
            plan_obj = PlanDetails.objects.get(id=transaction.plan_id)
            plan_name = plan_obj.plan_name
        except PlanDetails.DoesNotExist:
            pass
 
    if transaction.addon_package:
        try:
            addon_ids = str(transaction.addon_package).split(",")  # list of IDs: ['5','4','2','1']
            addon_ids = [id.strip() for id in addon_ids if id.strip().isdigit()]

            addon_objects = Addonpackages.objects.filter(package_id__in=addon_ids)

            if addon_objects.exists():
                addon_name = ", ".join([a.name for a in addon_objects])
            # addon_name = addon_obj.name
        except Addonpackages.DoesNotExist:
            pass
 
    if plan_name and addon_name:
        payment_for = f"{plan_name} + {addon_name}"
    elif plan_name:
        payment_for = plan_name
    elif addon_name:
        payment_for = addon_name
    else:
        payment_for = "N/A"
 
 
    if action.lower() == "accept":
        # Check if PlanSubscription for this transaction already exists
        existing_plan = PlanSubscription.objects.filter(trans_id=transaction.id).first()

        if not existing_plan:
            # Create plan subscription only if not already inserted
            plan_sub = PlanSubscription.objects.create(
                profile_id=transaction.profile_id,
                plan_id=transaction.plan_id,
                addon_package=transaction.addon_package,
                paid_amount=transaction.amount,
                discount=transaction.discount_amont or 0,
                payment_mode=transaction.payment_type,
                payment_for=payment_for,
                status=1,  # assuming 1 = active
                payment_by=transaction.profile_id,  # adjust if different
                admin_user=admin_user,
                order_id=transaction.order_id,
                payment_id=transaction.payment_id,
                gpay_no="",  # set if needed
                trans_id=transaction.id,
                payment_date=datetime.now()
            )

            transaction.status = "2"  # or 2 if int based
            transaction.owner_id = admin_user
            transaction.save()
            subscription = plan_sub
            message = "Transaction accepted and subscription created successfully"
        else:
            subscription = existing_plan
            message = f"Subscription already exists for transaction ID {transaction.id}"

        return Response(
            {"status": "success", "message": message, "subscription_id": subscription.id},
            status=status.HTTP_200_OK
        )
 
    elif action.lower() == "reject":
        transaction.status = "3"  # or 3 if int based
        transaction.owner_id = admin_user
        transaction.save()
 
        return Response(
            {"status": "success", "message": "Transaction rejected"},
            status=status.HTTP_200_OK
        )
 
    else:
        return Response(
            {"status": "error", "message": "Invalid action (must be 'accept' or 'reject')"},
            status=status.HTTP_400_BAD_REQUEST
        )
 

class PaymentTransactionListView(APIView):
    def get(self, request):
        profile_id = request.query_params.get('profile_id')
        if not profile_id:
            return Response({"error": "profile_id is required"}, status=status.HTTP_400_BAD_REQUEST)

        transactions = PaymentTransaction.objects.filter(profile_id=profile_id).order_by('-created_at')
        serializer = PaymentTransactionSerializer(transactions, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
# class TransactionHistoryView(generics.ListAPIView):
#     serializer_class = PaymentTransactionListSerializer
#     pagination_class = StandardResultsPaging

#     def get_queryset(self):
#         from_date = self.request.query_params.get('from_date')
#         to_date = self.request.query_params.get('to_date')
#         filter_type = self.request.query_params.get('filter_type')
#         search = self.request.query_params.get('search')
#         status_ids = [1, 2, 3]
#         placeholders = ', '.join(['%s'] * len(status_ids))

#         sql_pt = f"""
#             SELECT ld.ContentId, ld.ProfileId, ld.Profile_name, ld.Gender, ld.Mobile_no, ld.EmailId,
#                    ld.status as profile_status, ld.Profile_dob, ld.Profile_whatsapp, pt.Plan_id,
#                    ld.Profile_city, ld.Profile_state, pt.status, ld.DateOfJoin, pl.plan_name,
#                    ld.membership_startdate, ld.membership_enddate, pt.id AS transaction_id,
#                    pt.order_id, pt.created_at, pt.payment_id, pt.amount, pt.discount_amont,
#                    pt.payment_type, pt.admin_status, pt.payment_refno,
#                    'payment_transaction' AS source
#             FROM payment_transaction pt
#             LEFT JOIN plan_master pl ON pt.Plan_id = pl.id
#             LEFT JOIN logindetails ld ON ld.ProfileId = pt.profile_id
#             WHERE pt.status IN ({placeholders})
#         """

#         sql_ps = f"""
#             SELECT ld.ContentId, ld.ProfileId, ld.Profile_name, ld.Gender, ld.Mobile_no, ld.EmailId,
#                    ld.status as profile_status, ld.Profile_dob, ld.Profile_whatsapp, ps.plan_id,
#                    ld.Profile_city, ld.Profile_state, ps.status, ld.DateOfJoin, pl.plan_name,
#                    ps.validity_startdate AS membership_startdate, ps.validity_enddate AS membership_enddate,
#                    ps.id AS transaction_id, ps.order_id, ps.payment_date AS created_at,
#                    ps.payment_id, ps.paid_amount AS amount, ps.discount AS discount_amont,
#                    ps.payment_mode AS payment_type, ps.admin_user AS admin_status,
#                    ps.gpay_no AS payment_refno,
#                    'plan_subscription' AS source
#             FROM plan_subscription ps
#             LEFT JOIN plan_master pl ON ps.plan_id = pl.id
#             LEFT JOIN logindetails ld ON ld.ProfileId = ps.profile_id
#             WHERE ps.profile_id NOT IN (SELECT profile_id FROM payment_transaction)
#               AND ps.status IN ({placeholders})
#         """

#         sql = f"""
#             SELECT * FROM (
#                 SELECT *,
#                     ROW_NUMBER() OVER (PARTITION BY combined.ProfileId ORDER BY combined.created_at DESC) AS rn
#                 FROM (
#                     {sql_pt}
#                     UNION ALL
#                     {sql_ps}
#                 ) AS combined
#                 WHERE 1=1
#             """
#         params = status_ids.copy() + status_ids.copy()

#         try:
#             if filter_type == "today":
#                 today = datetime.today().date()
#                 sql += " AND DATE(combined.created_at) = %s"
#                 params.append(today)

#             elif filter_type == "last_week":
#                 today = datetime.today().date()
#                 last_week_start = today - timedelta(days=today.weekday() + 7)
#                 last_week_end = last_week_start + timedelta(days=6)
#                 sql += " AND DATE(combined.created_at) BETWEEN %s AND %s"
#                 params += [last_week_start, last_week_end]

#             elif filter_type == "new_approved":
#                 sql += " AND combined.profile_status IN (%s, %s)"
#                 params += [0, 1]

#             elif filter_type == "delete_others":
#                 sql += " AND combined.profile_status NOT IN (%s, %s)"
#                 params += [0, 1]

#             elif from_date and to_date:
#                 start_date = datetime.strptime(from_date, "%Y-%m-%d").date()
#                 end_date = datetime.strptime(to_date, "%Y-%m-%d").date()
#                 sql += " AND DATE(combined.created_at) BETWEEN %s AND %s"
#                 params += [start_date, end_date]
#         except Exception:
#             pass

#         if search:
#             sql += """
#                 AND (
#                     CAST(combined.ProfileId AS CHAR) LIKE %s OR
#                     combined.Profile_name LIKE %s OR
#                     combined.Mobile_no LIKE %s
#                 )
#             """
#             search_term = f"%{search}%"
#             params += [search_term, search_term, search_term]

#         sql += """
#             ) AS ranked
#             WHERE rn = 1
#             ORDER BY ranked.created_at DESC
#         """

#         # print("Final SQL:", sql)
#         # print("Params:", params)
#         self.sql = sql
#         self.params = params
#         with connection.cursor() as cursor:
#             cursor.execute(sql, params)
#             rows = dictfetchall(cursor)

#         return rows

def dictfetchall(cursor):
    "Return all rows from a cursor as a dict"
    desc = cursor.description
    return [
        dict(zip([col[0] for col in desc], row))
        for row in cursor.fetchall()
    ]

class TransactionHistoryView(generics.ListAPIView):
    serializer_class = PaymentTransactionListSerializer
    pagination_class = StandardResultsPaging

    def get_queryset(self):
        request = self.request
        from_date = request.query_params.get('from_date')
        to_date = request.query_params.get('to_date')
        filter_type = request.query_params.get('filter_type')
        search = request.query_params.get('search')
        t_status = request.query_params.get('t_status')
        a_status = request.query_params.get('a_status')

        status_ids = [1, 2, 3]
        placeholders = ', '.join(['%s'] * len(status_ids))

        def build_transaction_sql(source, table_alias, date_field, start_date_field, end_date_field, amount_field, discount_field, payment_type_field, admin_field, ref_field):
            return f"""
                SELECT ld.ContentId, ld.ProfileId, ld.Profile_name, ld.Gender, ld.Mobile_no, ld.EmailId,
                    ld.status AS profile_status, ld.Profile_dob, ld.Profile_whatsapp,
                    ld.Plan_id ,ld.Plan_status AS profile_plan_id, {table_alias}.plan_id AS transaction_plan_id,
                    ld.Profile_city, ld.Profile_state, {table_alias}.status, ld.DateOfJoin, pl.plan_name,cpl.plan_name AS current_plan_name,
                    {start_date_field} AS membership_startdate, {end_date_field} AS membership_enddate,
                    {table_alias}.id AS transaction_id, {table_alias}.order_id, {table_alias}.{date_field} AS created_at,
                    {table_alias}.payment_id, {table_alias}.{amount_field} AS amount, {table_alias}.{discount_field} AS discount_amont,
                    {table_alias}.{payment_type_field} AS payment_type, {table_alias}.{admin_field} AS admin_status,
                    {table_alias}.{ref_field} AS payment_refno,
                    '{source}' AS source
                FROM {source} {table_alias}
                LEFT JOIN plan_master pl ON {table_alias}.plan_id = pl.id
                LEFT JOIN logindetails ld ON ld.ProfileId = {table_alias}.profile_id
                LEFT JOIN plan_master cpl ON ld.plan_id = cpl.id
            """


        sql_pt = build_transaction_sql(
            source="payment_transaction", table_alias="pt",
            date_field="created_at", amount_field="amount", discount_field="discount_amont",
            payment_type_field="payment_type", admin_field="admin_status", ref_field="payment_refno",
            start_date_field="ld.membership_startdate", end_date_field="ld.membership_enddate"
        ) + f" WHERE pt.status IN ({placeholders}) "

        sql_ps = build_transaction_sql(
            source="plan_subscription", table_alias="ps",
            date_field="payment_date", amount_field="paid_amount", discount_field="discount",
            payment_type_field="payment_mode", admin_field="admin_user", ref_field="gpay_no",
            start_date_field="ps.validity_startdate", end_date_field="ps.validity_enddate"
        ) + f"""
            WHERE ps.profile_id NOT IN (SELECT profile_id FROM payment_transaction)
            AND ps.status IN ({placeholders}) 
        """

        sql = f"""
            SELECT * FROM (
                SELECT *,
                    ROW_NUMBER() OVER (PARTITION BY combined.ProfileId ORDER BY combined.created_at DESC) AS rn
                FROM (
                    {sql_pt}
                    UNION ALL
                    {sql_ps}
                ) AS combined
                WHERE 1=1
        """
        params = status_ids + status_ids

        try:
            if filter_type == "today":
                today = datetime.today().date()
                sql += " AND DATE(combined.created_at) = %s"
                params.append(today)
            elif filter_type == "last_week":
                today = datetime.today().date()
                last_week_start = today - timedelta(days=today.weekday() + 7)
                last_week_end = last_week_start + timedelta(days=6)
                sql += " AND DATE(combined.created_at) BETWEEN %s AND %s"
                params.extend([last_week_start, last_week_end])
            elif filter_type == "new_approved":
                sql += " AND combined.profile_status IN (%s, %s)"
                params.extend([0, 1])
            elif filter_type == "delete_others":
                sql += " AND combined.profile_status NOT IN (%s, %s)"
                params.extend([0, 1])
            elif from_date and to_date:
                start_date = datetime.strptime(from_date, "%Y-%m-%d").date()
                end_date = datetime.strptime(to_date, "%Y-%m-%d").date()
                sql += " AND DATE(combined.created_at) BETWEEN %s AND %s"
                params.extend([start_date, end_date])
        except Exception as e:
            print(f"Filter error: {e}")

        if t_status:
            try:
                t_status = int(t_status)
                if t_status == 1:
                    sql += " AND ((combined.source = 'payment_transaction' AND combined.status = 1))"
                elif t_status == 2:
                    sql += " AND ((combined.source = 'payment_transaction' AND combined.status = 2) OR (combined.source = 'plan_subscription' AND combined.status = 1))"
                elif t_status == 3:
                    sql += " AND ((combined.source = 'payment_transaction' AND combined.status = 3) OR (combined.source = 'plan_subscription' AND combined.status = 0))"
            except Exception as e:
                print(f"Invalid t_status value:{str(e)}")
                pass

        if search:
            sql += """
                AND (
                    CAST(combined.ProfileId AS CHAR) LIKE %s OR
                    combined.Profile_name LIKE %s OR
                    combined.Mobile_no LIKE %s
                )
            """
            search_term = f"%{search}%"
            params.extend([search_term] * 3)

        sql += """
            ) AS ranked
            WHERE rn = 1
            ORDER BY ranked.created_at DESC
        """
        self.sql = sql
        self.params = params

        # print("Sql",sql,params)
        with connection.cursor() as cursor:
            cursor.execute(sql, params)
            main_rows = dictfetchall(cursor)

        if not main_rows:
            return []

        profile_ids = [row["ProfileId"] for row in main_rows]
        transaction_ids = [row["transaction_id"] for row in main_rows]

        placeholders_profiles = ', '.join(['%s'] * len(profile_ids))
        placeholders_txns = ', '.join(['%s'] * len(transaction_ids))

        other_txn_sql = f"""
            SELECT pt.profile_id AS ProfileId,
                pt.status,
                pt.created_at,
                pt.id AS transaction_id,
                pl.plan_name,
                (
                    SELECT GROUP_CONCAT(mp.name SEPARATOR ', ')
                    FROM masteradonpackages mp
                    WHERE FIND_IN_SET(mp.package_id, pt.addon_package)
                ) AS addon_packages
            FROM payment_transaction pt
            LEFT JOIN plan_master pl ON pt.Plan_id = pl.id
            WHERE pt.profile_id IN ({placeholders_profiles})
            AND pt.id NOT IN ({placeholders_txns})
            ORDER BY pt.created_at DESC
        """

        with connection.cursor() as cursor:
            cursor.execute(other_txn_sql, profile_ids + transaction_ids)
            action_logs = dictfetchall(cursor)

        logs_by_profile = defaultdict(list)
        for log in action_logs:
            logs_by_profile[log["ProfileId"]].append({
                "status": self.map_status_code(log["status"]),
                "plan_name": log.get("plan_name"),
                "addon_packages": log.get("addon_packages"),
                "created_at": log["created_at"].date() if isinstance(log["created_at"], datetime) else log["created_at"],
            })
        for row in main_rows:
            row["action_log"] = logs_by_profile.get(row["ProfileId"], [])
            transaction_plan_id = row.get("transaction_plan_id")
            profile_plan_id = row.get("profile_plan_id")
            # print("tran_id:",transaction_plan_id, profile_plan_id)
            row["a_status"] = "active" if transaction_plan_id == profile_plan_id else "inactive" 

        if a_status in ["0", "1"]:
            status_label = "active" if a_status == "1" else "inactive"
            main_rows = [row for row in main_rows if row["a_status"] == status_label]
            
        return main_rows    

    def map_status_code(self, code):
        return {
            1: "Initialized",
            2: "Paid",
            3: "Failed"
        }.get(code, "Unknown")

def iter_cursor(cursor, arraysize=1000):
    cols = [col[0] for col in cursor.description]
    while True:
        batch = cursor.fetchmany(arraysize)
        if not batch:
            break
        for row in batch:
            yield dict(zip(cols, row))

class Echo:
    def write(self, value):
        return value

class TransactionHistoryExportView(generics.GenericAPIView):
    serializer_class = PaymentTransactionListSerializer

    def get(self, request, *args, **kwargs):
        view = TransactionHistoryView()
        view.request = request
        main_rows = view.get_queryset()  # use returned rows directly

        serializer = self.get_serializer(main_rows, many=True)
        field_names = list(serializer.child.fields.keys())

        pseudo_buffer = Echo()
        writer = csv.writer(pseudo_buffer)

        def row_stream():
            yield writer.writerow(field_names)
            for item in serializer.data:
                yield writer.writerow([smart_str(item.get(f)) for f in field_names])

        response = StreamingHttpResponse(row_stream(), content_type="text/csv; charset=utf-8")
        response['Content-Disposition'] = 'attachment; filename="transaction_history.csv"'
        return response

class DataHistoryListView(generics.GenericAPIView):
         
        def get(self, request, *args, **kwargs):
            profile_id = request.query_params.get('profile_id')
            
            sql = f"""
                SELECT dh.profile_id,dh.date_time,mp.status_name,dh.others,pm.plan_name
                FROM datahistory dh LEFT JOIN masterprofilestatus mp ON mp.status_code = dh.profile_status
                LEFT JOIN plan_master pm ON pm.id = dh.plan_id
                WHERE dh.profile_id = %s
            """
            params= (profile_id)
            
            sql += " ORDER BY dh.date_time DESC"
            
            with connection.cursor() as cursor:
                cursor.execute(sql, params)
                rows = dictfetchall(cursor)

            return Response(rows)

class FeaturedProfilesView(APIView):
    pagination_class = StandardResultsPaging

    def get(self, request):
        from_date = request.query_params.get('from_date')
        to_date = request.query_params.get('to_date')
        search = request.query_params.get('search')
        export_type = request.query_params.get('export')
        sql = """
            SELECT pf.profile_id,ld.Profile_name,ld.Gender,masterstate.name,pl.plan_name,pf.boosted_date,pf.boosted_enddate,ms.status_name
            FROM profile_plan_feature_limits pf
            LEFT JOIN logindetails ld ON pf.profile_id = ld.ProfileId
            LEFT JOIN plan_master pl ON pf.Plan_id = pl.id
            LEFT JOIN masterstate ON ld.Profile_state = masterstate.id
            LEFT JOIN masterprofilestatus ms ON ld.status = ms.status_code
            Where pf.featured_profile = 1 AND ld.status = 1 
            
        """
        params = []

        if from_date and to_date:
            try:
                start_date = datetime.strptime(from_date, "%Y-%m-%d").date()
                end_date = datetime.strptime(to_date, "%Y-%m-%d").date()
                sql += " AND DATE(pf.boosted_date) BETWEEN %s AND %s"
                params += [start_date, end_date]
            except ValueError:
                pass
            
        if search:
            search_pattern = f"%{search}%"
            sql += """ AND (ld.Profile_name LIKE %s OR ld.ProfileId LIKE %s )"""
            params += [search_pattern, search_pattern]
            
        sql += "ORDER BY pf.boosted_date DESC"
        with connection.cursor() as cursor:
            cursor.execute(sql, params)
            rows = dictfetchall(cursor)
            
        three_months_ago = datetime.today().date() - timedelta(days=90)
        for row in rows:
            boosted_date = row.get("boosted_date")
            if boosted_date and boosted_date >= three_months_ago:
                row["active"] = "Yes"
            else:
                row["active"] = "No"
                
        if export_type == "csv":
            return export_renew_csv("featured_profiles", rows)

        if export_type == "xlsx":
            return export_renew_xlsx("featured_profiles", rows)
                
        paginator = self.pagination_class()
        paginated_data = paginator.paginate_queryset(rows, request)

        return paginator.get_paginated_response(paginated_data)
    
class FeaturedProfileAddView(APIView):
    def post(self, request):
        profile_id = request.data.get('profile_id')
        boosted_date = request.data.get('boosted_startdate')
        boosted_enddate = request.data.get('boosted_enddate')
        owner_id = request.data.get('admin_user_id')

        try:
            owner_id = int(owner_id)
            user = User.objects.get(id=owner_id)
        except Exception:
            user = None
             
        if user:
            role = user.role
            permissions = RolePermission.objects.filter(role=role).select_related('action')
            data = permissions.values('action__code', 'value')
            edit_permission = data.filter(action__code='featured_profile_add').first()
            edit=edit_permission['value'] if edit_permission else None
        else:
            edit =None
        
        if user:
            if edit ==1:
                pass
            else:
                return Response({
                    "status": "error",
                    "message": "Permission Error"
                }, status=status.HTTP_403_FORBIDDEN)
        if not profile_id or not boosted_date or not boosted_enddate:
            return Response({
                "status": "error",
                "message": "profile_id, boosted_startdate, and boosted_enddate are required"
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            profile = Registration1.objects.filter(ProfileId=profile_id).first()
            if not profile:
                return Response({
                    "status": "error",
                    "message": "Profile not found"
                }, status=status.HTTP_404_NOT_FOUND)

            pf = Profile_PlanFeatureLimit.objects.get(profile_id=profile_id)
            pf.featured_profile = 1
            pf.boosted_date = boosted_date
            pf.boosted_enddate = boosted_enddate

            pf.save()

            return Response({
                "status": "success",
                "message": "Updated to featured profile successfully"
            }, status=status.HTTP_200_OK)

        except Profile_PlanFeatureLimit.DoesNotExist:
            return Response({
                "status": "error",
                "message": "Feature record not found for this profile"
            }, status=status.HTTP_404_NOT_FOUND)

        except Exception as e:
            print(f"Error adding featured profile: {str(e)}")
            return Response({
                "status": "error",
                "message": f"An unexpected error occurred: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



class Renewalplans(generics.GenericAPIView):
    def post(self, request):
        plan_type = request.data.get('type')

        if plan_type == 'new':
            statuses = PlanDetails.objects.filter(master_substatus__in=[5, 6],plan_status=1)
        elif plan_type == 'renewal':
            statuses = PlanDetails.objects.filter(master_substatus=7).exclude(id=4)
        else:
            statuses = PlanDetails.objects.filter(master_substatus__in=[5, 6, 7],plan_status=1)

        serializer = PlandetailsSerializer(statuses, many=True)
        return Response({
            'status': 'success',
            'data': serializer.data
        })
        
class SendInvoicePDF(APIView):
    def get(self, request):
        subscription_id = request.query_params.get('subscription_id')
        owner_id = request.query_params.get('admin_user_id')
        try:
            owner_id = int(owner_id)
            user = User.objects.get(id=owner_id)
        except Exception:
            user = None

        if user:
            role = user.role
            permissions = RolePermission.objects.filter(role=role).select_related('action')
            data = permissions.values('action__code', 'value')
            edit_permission = data.filter(action__code='membership_activation').first()
            edit = edit_permission['value'] if edit_permission else None
        else:
            edit = None

        if user:
            if edit == 1:
                pass
            else:
                return Response({
                    "status": "error",
                    "message": "Permission Error"
                }, status=status.HTTP_403_FORBIDDEN)
            
        if not subscription_id:
            return Response({"error": "subscription_id is required"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            subscription = PlanSubscription.objects.get(id=subscription_id)
        except PlanSubscription.DoesNotExist:
            return Response({"error": "Subscription not found"}, status=status.HTTP_404_NOT_FOUND)

        try:
            profile = Registration1.objects.get(ProfileId=subscription.profile_id)
            state = get_state_name(profile.Profile_state)
        except Exception:
            profile = None
            state = None

        if profile:
            customer_name = profile.Profile_name
            address = f"{profile.Profile_address or ''}\n{profile.Profile_city or ''}, {state or ''} - {profile.Profile_pincode or ''}\nPhone: {profile.Mobile_no or ''}\nEmail: {profile.EmailId or ''}"
            recipient_email = profile.EmailId
        else:
            customer_name = None
            address = "Address not available"
            recipient_email = None

        image_url = "https://vysyamat.blob.core.windows.net/vysyamala/newvysyamalalogo2.png"
        try:
            response = requests.get(image_url)
            encoded_logo = base64.b64encode(response.content).decode() if response.status_code == 200 else ""
        except Exception:
            encoded_logo = ""

        if subscription.payment_mode:
            mode = subscription.payment_mode.strip().lower()
            if mode in ["razorpay","razor pay","onlinegpay","manualgpay","accounttransfer","online","payu online payment","by account transfer","online payment","by google pay"]:
                payment_mode = "Online Transfer"
            else:
                payment_mode = "Cash/Cheque/DD"
        else:
            payment_mode = "Cash/Cheque/DD"

        try:
            addon_ids = [int(pk.strip()) for pk in subscription.addon_package.split(",") if pk.strip().isdigit()]
            addon_qs = Addonpackages.objects.filter(package_id__in=addon_ids)

            addon_items = []
            addon_total = 0
            for addon in addon_qs:
                addon_items.append({
                    "name": addon.name,
                    "description": addon.description,
                    "amount": addon.amount or 0
                })
                addon_total += addon.amount or 0
        except Exception:
            addon_items = []
            addon_total = 0
        try:
            if subscription.package_amount:
                plan_amount = subscription.package_amount
            else:
                if subscription.plan_id:
                    plan=PlanDetails.objects.get(id=subscription.plan_id)
                    plan_amount = plan.plan_price
                else:
                    plan_amount=0
        except Exception:
            plan_amount=0
              
        base_price = plan_amount
        discount = subscription.discount or 0
        total_price = base_price + addon_total - discount
        net_price = base_price + addon_total
        
        if subscription.plan_id and subscription.plan_id > 0:
            plan_name = PlanDetails.objects.filter(id=subscription.plan_id).values_list('plan_name', flat=True).first()   
        else:
            plan_name = ""
        data = {
            'encoded_logo': encoded_logo,
            'customer_name': customer_name if customer_name else "Valued Customer",
            'address': address,
            'date': subscription.payment_date.strftime("%d/%m/%Y") if subscription.payment_date else "",
            'invoice_number': subscription.id,
            'vysyamala_id': subscription.profile_id or "",
            'service_description':plan_name or "" ,
            'offer': subscription.offer or "",
            'price': f"{base_price:.0f}",
            'valid_till': subscription.validity_enddate.strftime("%d-%m-%Y") if subscription.validity_enddate else "",
            'payment_mode': payment_mode or "N/A",
            'addon_items': addon_items,
            'addon_total': f"{addon_total:.0f}",
            'discount': f"{discount:.0f}",
            'total_price': f"{total_price:.0f}",
            'num_to_words': number_to_words(int(total_price)),
            'net_price': f"{net_price:.0f}",
        }

        html_string = render_to_string("invoice.html", data)
        pdf_buffer = BytesIO()
        pisa_status = pisa.CreatePDF(src=html_string, dest=pdf_buffer)

        if pisa_status.err:
            return Response({"error": "Error generating PDF"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        email_body = """
Dear Mrs. {CustomerName}, 🙏

Thank you for choosing Vysyamala and opting for the {MembershipPlan} membership.

Your invoice is attached for your reference ✅  
For any assistance, kindly contact our support: 99448 51550.

With the blessings of Goddess Sri Kannika Parameswari,  
we wish you a happy and blessed journey towards finding your life partner 🌸

Team Vysyamala  
www.vysyamala.com
        """.format(
            CustomerName=customer_name or "Customer",
            MembershipPlan=plan_name or ""
        )
        if recipient_email:
            email = EmailMessage(
                subject=f"Invoice V{subscription.id} from Vysyamala",
                body=email_body,
                to=[recipient_email],
                cc=["vysyamala@gmail.com"]
            )
            email.attach(f"invoice_{subscription.id}.pdf", pdf_buffer.getvalue(), "application/pdf")
            try:
                email.send()
                subscription.is_sent_email = True
                subscription.save()
                return Response({"success": "Email sent successfully!"}, status=status.HTTP_200_OK)
            except Exception as e:
                return Response({"error": f"Failed to send email: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        else:
            return Response({"error": "Recipient email not found"}, status=status.HTTP_400_BAD_REQUEST)
        





class UserViewSet(viewsets.ModelViewSet):
    """
    CRUD for users:
    - list()    → GET /api/users/
    - retrieve()→ GET /api/users/{id}/
    - create()  → POST /api/users/
    - update()  → PUT /api/users/{id}/
    - partial_update() → PATCH /api/users/{id}/
    - destroy() → DELETE /api/users/{id}/
    """
    queryset = User.objects.filter(is_deleted=0).order_by('id')
    serializer_class = UserSerializer
    #permission_classes = [IsAuthenticated]
    permission_classes = [permissions.AllowAny]

    # Optional: current user profile
    @action(detail=False, methods=['get'])
    def me(self, request):
        serializer = self.get_serializer(request.user)
        return Response(serializer.data)
    

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.status = 0
        instance.is_deleted = 1
        instance.save(update_fields=['status', 'is_deleted'])
        return Response(
            {"message": "User soft deleted successfully"},
            status=status.HTTP_200_OK
        )
    


        # Override create to handle Excel file with user creation
    def create(self, request, *args, **kwargs):

        # print('76565vghvchg')
        excel_file = request.FILES.get('excel_file')
        
        # First, create the user
        response = super().create(request, *args, **kwargs)
        
        # If user creation successful and Excel file provided, update profiles
        if response.status_code == status.HTTP_201_CREATED and excel_file:
            user_id = response.data.get('id')
        
            # print('user_id',user_id)
            
            # Use the utility function
            update_results = ProfileUpdater.update_profiles_from_excel(
                excel_file=excel_file, 
                owner_id=user_id
            )
            
            # Include results in response
            response.data['profile_update_results'] = update_results
        
        return response

    
    # Override partial_update for PATCH requests
    def partial_update(self, request, *args, **kwargs):
        # print('PATCH method - updating user with Excel')
        excel_file = request.FILES.get('excel_file')
        
        # First, update the user using partial_update
        response = super().partial_update(request, *args, **kwargs)
        
        # If user update successful and Excel file provided, update profiles
        if response.status_code == status.HTTP_200_OK and excel_file:
            user_id = kwargs.get('pk')  # Get user ID from URL parameters
            # print('user_id', user_id)
            
            # Use the utility function
            update_results = ProfileUpdater.update_profiles_from_excel(
                excel_file=excel_file, 
                owner_id=user_id
            )
            
            # Include results in response
            response.data['profile_update_results'] = update_results
        
        return response

    # # Override update for PUT requests (if you want Excel support for PUT as well)
    # def update(self, request, *args, **kwargs):
    #     print('PUT method - updating user with Excel')
    #     excel_file = request.FILES.get('excel_file')
        
    #     # First, update the user
    #     response = super().update(request, *args, **kwargs)
        
    #     # If user update successful and Excel file provided, update profiles
    #     if response.status_code == status.HTTP_200_OK and excel_file:
    #         user_id = kwargs.get('pk')
    #         print('user_id', user_id)
            
    #         # Use the utility function
    #         update_results = ProfileUpdater.update_profiles_from_excel(
    #             excel_file=excel_file, 
    #             owner_id=user_id
    #         )
            
    #         # Include results in response
    #         response.data['profile_update_results'] = update_results
        
    #     return response

    # Assign profiles to existing user
    @action(detail=True, methods=['post'], url_path='assign-profiles')
    def assign_profiles(self, request, pk=None):
        """
        Assign profiles to an existing user using Excel file
        """
        excel_file = request.FILES.get('excel_file')
        
        if not excel_file:
            return Response(
                {"error": "No Excel file provided"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        user = self.get_object()
        
        # Use the utility function
        update_results = ProfileUpdater.update_profiles_from_excel(
            excel_file=excel_file, 
            owner_id=user.id
        )
        
        response_data = {
            "message": "Profile assignment completed",
            "user_id": user.id,
            "profile_update_results": update_results
        }
        
        return Response(response_data, status=status.HTTP_200_OK)

    # Bulk update with Excel (multiple owners)
    @action(detail=False, methods=['post'], url_path='bulk-assign-profiles')
    def bulk_assign_profiles(self, request):
        """
        Bulk assign profiles to multiple owners using Excel file
        Excel should contain both profile_id and owner_id columns
        """
        excel_file = request.FILES.get('excel_file')
        
        if not excel_file:
            return Response(
                {"error": "No Excel file provided"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Use the utility function without owner_id (will read from Excel)
        update_results = ProfileUpdater.update_profiles_from_excel(excel_file=excel_file)
        
        response_data = {
            "message": "Bulk profile assignment completed",
            "results": update_results
        }
        
        return Response(response_data, status=status.HTTP_200_OK)

    # Download Excel template
    @action(detail=False, methods=['get'], url_path='download-template')
    def download_template(self, request):
        """
        Download Excel template for profile assignment
        """
        return ProfileUpdater.get_excel_template()


class ProfileUpdater:
    """
    Utility class to handle Excel file operations for profile updates
    """

    @staticmethod
    def update_profiles_from_excel(excel_file, owner_id=None):
        
        # print('owner_id',owner_id)
        """
        Read Excel file and update profiles in LoginDetails table
        
        Args:
            excel_file: The uploaded Excel file
            owner_id: The owner ID to assign to profiles (if None, uses owner_id from Excel)
        
        Returns:
            dict: Results with success/failure counts and errors
        """
        # Validate file
        if not excel_file:
            return {
                "success": False,
                "error": "No Excel file provided",
                "successful_updates": 0,
                "failed_updates": 0,
                "errors": []
            }
        
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            return {
                "success": False,
                "error": "File must be an Excel file (.xlsx or .xls)",
                "successful_updates": 0,
                "failed_updates": 0,
                "errors": []
            }
        
        try:
            # Read Excel file
            df = pd.read_excel(excel_file)
            
            # Determine the mode (with or without owner_id in Excel)
            if owner_id:
                # Mode 1: Single owner_id for all profiles (Excel has only profile_id column)
                return ProfileUpdater._update_with_single_owner(df, owner_id)

                
        except Exception as e:
            return {
                "success": False,
                "error": f"Error processing Excel file: {str(e)}",
                "successful_updates": 0,
                "failed_updates": 0,
                "errors": []
            }


    @staticmethod
    def _update_with_single_owner(df, owner_id):
        """
        Update profiles with a single owner ID (Excel has only profile_id column)
        """
        # Validate required columns
        if 'profile_id' not in df.columns:
            return {
                "success": False,
                "error": "Missing required column: profile_id",
                "successful_updates": 0,
                "failed_updates": 0,
                "errors": []
            }
        
        # Remove rows with empty profile_id
        df = df.dropna(subset=['profile_id'])
        
        # Validate owner exists
        try:
            owner = User.objects.get(id=owner_id, is_deleted=0)
        except User.DoesNotExist:
            return {
                "success": False,
                "error": f"Owner with ID {owner_id} not found or is deleted",
                "successful_updates": 0,
                "failed_updates": 0,
                "errors": []
            }
        
        results = {
            "success": True,
            "successful_updates": 0,
            "failed_updates": 0,
            "errors": []
        }
        
        # Process each profile
        for index, row in df.iterrows():
            try:
                # Handle different data types safely
                raw_value = row['profile_id']
                
                # Convert to string and clean
                if pd.isna(raw_value):
                    results['errors'].append(f"Row {index+2}: Empty profile_id")
                    results['failed_updates'] += 1
                    continue
                    
                profile_id = str(raw_value).strip()
                
                if not profile_id:
                    results['errors'].append(f"Row {index+2}: Empty profile_id after cleaning")
                    results['failed_updates'] += 1
                    continue
                
                # Update LoginDetails - using exact string match
                updated_count = LoginDetails.objects.filter(ProfileId=profile_id).update(Owner_id=owner_id)
                
                if updated_count > 0:
                    results['successful_updates'] += 1
                else:
                    results['errors'].append(f"Row {index+2}: Profile with ID '{profile_id}' not found")
                    results['failed_updates'] += 1

                # print('Updated sucessfully')
                    
            except Exception as e:
                error_msg = f"Row {index+2}: {str(e)}"
                results['errors'].append(error_msg)
                results['failed_updates'] += 1
        
        return results

class LoginView(APIView):
    def post(self, request):
        serializer = LoginSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        username = serializer.validated_data['username']
        password = serializer.validated_data['password']

        # try:
        #     user = User.objects.select_related('role').prefetch_related('role__permissions__action').get(
        #         username=username, is_deleted=False
        #     )

        try:
            
            user = (
                    User.objects
                    .select_related('role')
                    .prefetch_related('role__permissions__action')
                    .get(
                        Q(username=username) | Q(email=username),
                        is_deleted=False
                    )
                )
            if not user.check_password(password):
                return Response({"error": "Invalid password"}, status=status.HTTP_401_UNAUTHORIZED)

            user_data = UserSerializer(user).data
            role_data = RoleSerializers(user.role).data if user.role else None

            return Response({
                "message": "Login successful",
                "user": user_data,
                "role": role_data
            }, status=status.HTTP_200_OK)

        except User.DoesNotExist:
            return Response({"error": "Invalid username"}, status=status.HTTP_404_NOT_FOUND)




class RoleDropdownView(APIView):
    def get(self, request):
        roles = Roles.objects.all().order_by('id')
        serializer = RoleDropdownSerializer(roles, many=True)
        return Response({
            "message": "Roles fetched successfully",
            "roles": serializer.data
        }, status=status.HTTP_200_OK)
        
class OwnerUpdateView(APIView):
    def post(self,request):
        try:
            profile_id = request.POST.get('profile_id')
            owner_id = request.POST.get('owner_id')
            assign_by = request.POST.get('assign_by')
            if not profile_id or not owner_id or not assign_by:
                return Response({"status":"error","message":"profile_id,owner_id and assign_by are required"},status=status.HTTP_400_BAD_REQUEST)
            
            profile = LoginDetails.objects.filter(ProfileId=profile_id).first()
            if not profile:
                return Response({"status":"error","message":"Profile not found"},status=status.HTTP_404)
            
            profile.Owner_id = owner_id
            profile.save()
            return Response({"status":"success","message":"Owner updated successfully"},status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"status":"error","message":"server error:"},status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
class EditProfileWithPermissionAPIView(APIView):
    """
    This API view will allow users to edit and update their profile details based on ProfileId.
    """

    @transaction.atomic
    def put(self, request, profile_id, *args, **kwargs):
        # Extract the data from the request payload
        login_data = request.data.get('login_details', {})
        family_data = request.data.get('family_details', {})
        edu_data = request.data.get('education_details', {})
        horoscope_data = request.data.get('horoscope_details', {})
        partner_pref_data = request.data.get('partner_pref_details', {})
        suggested_pref_data = request.data.get('suggested_pref_details', {})
        profile_common_data = request.data.get('profile_common_details', {})
        profile_visibility_data=request.data.get('profile_visibility_details', {})
        owner_id = request.data.get('admin_user_id')

        profile_owner_id = request.data.get('profile_owner_id')
        

        # print(profile_visibility_data,'123456')

        # Initialize error tracking
        errors = {}

        # Step 1: Retrieve and update LoginDetails based on ProfileId
        try:
            owner_id = int(owner_id)
            user = User.objects.get(id=owner_id)
        except Exception:
            user = None
             
        if user:
            role = user.role
            permissions = RolePermission.objects.filter(role=role).select_related('action')
            data = permissions.values('action__code', 'value')
            edit_permission = data.filter(action__code='edit_profile_admin').first()
            edit=edit_permission['value'] if edit_permission else None
            membership_permission = data.filter(action__code='membership_activation').first()
            edit_mem=membership_permission['value'] if membership_permission else None
            print("edit",edit_mem)
        else:
            edit =None
            
        if edit == 1:
            try:
                login_detail = LoginDetails.objects.get(ProfileId=profile_id)
            except LoginDetails.DoesNotExist:
                return Response({'error': 'Profile not found.'}, status=status.HTTP_404_NOT_FOUND)
            if login_data:
                login_serializer = LoginEditSerializer(instance=login_detail, data=login_data, partial=True)
                if login_serializer.is_valid():
                    login_serializer.save()
                else:
                    errors['login_details'] = login_serializer.errors 

            if family_data:
                try:
                    family_detail = ProfileFamilyDetails.objects.get(profile_id=profile_id)
                except ProfileFamilyDetails.DoesNotExist:
                    return Response({'error': 'Family details not found.'}, status=status.HTTP_404_NOT_FOUND)

                family_serializer = ProfileFamilyDetailsSerializer(instance=family_detail, data=family_data, partial=True)
                if family_serializer.is_valid():
                    updated_instance = family_serializer.save()
                    uncle_gothram = family_data.get('uncle_gothram')
                    if uncle_gothram:
                        updated_instance.madulamn = uncle_gothram
                        updated_instance.save()
                else:
                    errors['family_details'] = family_serializer.errors

            # Step 3: Retrieve and update ProfileEduDetails
            if edu_data:
                try:
                    edu_detail = ProfileEduDetails.objects.get(profile_id=profile_id)
                except ProfileEduDetails.DoesNotExist:
                    return Response({'error': 'Education details not found.'}, status=status.HTTP_404_NOT_FOUND)

                edu_serializer = ProfileEduDetailsSerializer(instance=edu_detail, data=edu_data, partial=True)
                if edu_serializer.is_valid():
                    edu_serializer.save()
                else:
                    errors['education_details'] = edu_serializer.errors

            # Step 4: Retrieve and update ProfileHoroscope
            if horoscope_data:
                # print('1234567890')
                try:
                    horoscope_detail = ProfileHoroscope.objects.get(profile_id=profile_id)
                except ProfileHoroscope.DoesNotExist:
                    return Response({'error': 'Horoscope details not found.'}, status=status.HTTP_404_NOT_FOUND)
                
                # Get input text
                rasi_input_text = horoscope_data.get("rasi_kattam")
                # print(rasi_input_text,'123456')
                if rasi_input_text:
                    # Update input field
                    horoscope_detail.rasi_kattam = rasi_input_text
            
                    # Run dosham logic
                    mars_dosham, rahu_kethu_dosham = GetMarsRahuKethuDoshamDetails(rasi_input_text)
                    # print(mars_dosham)
                    # print(rahu_kethu_dosham)
                    # Save dosham results directly to model fields
                    horoscope_detail.calc_chevvai_dhosham = "True" if mars_dosham else "False"
                    horoscope_detail.calc_raguketu_dhosham = "True" if rahu_kethu_dosham else "False"
            
                # Update other fields in horoscope_data using serializer (excluding the calculated fields)
                # horoscope_data.pop("calc_chevvai_dhosham", None)
                # horoscope_data.pop("calc_raguketu_dhosham", None)
            
                horoscope_serializer = ProfileHoroscopeSerializer(
                    instance=horoscope_detail,
                    data=horoscope_data,
                    partial=True
                )

                horoscope_serializer = ProfileHoroscopeSerializer(instance=horoscope_detail, data=horoscope_data, partial=True)
                if horoscope_serializer.is_valid():
                    horoscope_serializer.save()
                else:
                    errors['horoscope_details'] = horoscope_serializer.errors

                # Step 5: Retrieve and update ProfilePartnerPref
            if partner_pref_data:
                    try:
                        partner_pref_detail = ProfilePartnerPref.objects.get(profile_id=profile_id)
                    except ProfilePartnerPref.DoesNotExist:
                        return Response({'error': 'Partner preference details not found.'}, status=status.HTTP_404_NOT_FOUND)
            
                    #prefered porutham rasi-stat value storing in the database mythili code 25-06-25

                    # Make a proper mutable copy of the input dict
                    if isinstance(partner_pref_data, dict):
                        partner_pref_payload = partner_pref_data.copy()
                    else:
                        # If it's a QueryDict (e.g., from request.data), convert to normal dict
                        partner_pref_payload = dict(partner_pref_data.lists())
                        # flatten single-item lists: {'key': ['value']} -> {'key': 'value'}
                        for key in partner_pref_payload:
                            if isinstance(partner_pref_payload[key], list) and len(partner_pref_payload[key]) == 1:
                                partner_pref_payload[key] = partner_pref_payload[key][0]
            
                    # Extract and process 'pref_porutham_star'
                    pref_star_ids = partner_pref_payload.get('pref_porutham_star')
                    if pref_star_ids:
                        try:
                            id_list = [int(i.strip()) for i in str(pref_star_ids).split(',') if i.strip().isdigit()]
                            matches = MatchingStarPartner.objects.filter(id__in=id_list)
            
                            star_rasi_pairs = [f"{m.dest_star_id}-{m.dest_rasi_id}" for m in matches]
            
                            # Save both cleaned values
                            partner_pref_payload['pref_porutham_star'] = ",".join(map(str, id_list))
                            partner_pref_payload['pref_porutham_star_rasi'] = ",".join(star_rasi_pairs)
            
                        except Exception as e:
                            errors['partner_pref_details'] = {
                                'pref_porutham_star': [f"Invalid input or failed to process star IDs: {str(e)}"]
                            }

                    partner_pref_serializer = ProfilePartnerPrefSerializer(instance=partner_pref_detail, data=partner_pref_data, partial=True)
                    if partner_pref_serializer.is_valid():
                        partner_pref_serializer.save()
                    else:
                        errors['partner_pref_details'] = partner_pref_serializer.errors
        

            # Step 6: RetriSuggestedeve and update ProfilePartnerPref
            if suggested_pref_data:
                try:
                    suggested_pref_detail = ProfileSuggestedPref.objects.get(profile_id=profile_id)
                except ProfileSuggestedPref.DoesNotExist:
                    #return Response({'error': 'suggested pref not found.'}, status=status.HTTP_404_NOT_FOUND)
                    suggested_pref_detail = ProfileSuggestedPref.objects.create(
                        profile_id=profile_id
                    )

                suggested_pref_serializer = ProfileSuggestedPrefSerializer(instance=suggested_pref_detail, data=suggested_pref_data, partial=True)
                if suggested_pref_serializer.is_valid():
                    suggested_pref_serializer.save()
                else:
                    errors['suggested_pref_details'] = suggested_pref_serializer.errors
         
         

            # Step 7: Retrieve and update ProfileEduDetails
            if profile_visibility_data:
                # print('inside profile visibility')
                try:
                    # print('update the existing record')
                    profvis_detail = ProfileVisibility.objects.get(profile_id=profile_id)
                    provis_serializer = ProfileVisibilitySerializer(instance=profvis_detail, data=profile_visibility_data, partial=True)

                except ProfileVisibility.DoesNotExist:
                    # print('insert the new record')
                    # return Response({'error': 'Profile Visibility details not found.'}, status=status.HTTP_404_NOT_FOUND)
                    profile_visibility_data['profile_id'] = profile_id
                    provis_serializer = ProfileVisibilitySerializer(data=profile_visibility_data)
                    
                    #Insert if data not exists
            
                if provis_serializer.is_valid():
                    provis_serializer.save()
                else:
                    errors['profile_visibility_details'] = provis_serializer.errors


        #common data to be update code is below

        
        if profile_common_data:
            if edit_mem ==3 or edit_mem ==1 or edit_mem ==2:
                owner = profile_common_data.get("owner_id")
                # print('inside profile common data update',profile_common_data.get("primary_status"))
                # Only include the common data keys that are available in the request
                if edit_mem ==3:
                    login_detail = LoginDetails.objects.get(ProfileId=profile_id)
                    get_plan_status = profile_common_data.get("secondary_status")
                    get_profile_status = profile_common_data.get("status")
                    old_profile_status = getattr(login_detail, 'status', None)

                    if old_profile_status == 1:

                        return Response({'error': 'You do not have permission to edit this profile.'}, status=status.HTTP_403_FORBIDDEN)
                    
                    if (old_profile_status == 0 or old_profile_status == 1) and (get_plan_status is not None) and (get_plan_status not in [6, 7, 8, 9]):

                        return Response({'error': 'You do not have permission to edit this profile.'}, status=status.HTTP_403_FORBIDDEN)

                    if old_profile_status != 0 and old_profile_status !=1 :
                        return Response({'error': 'You do not have permission to edit this profile.'}, status=status.HTTP_403_FORBIDDEN)
                    # print('old_profile_status',old_profile_status)
                    # print('get_profile_status',get_profile_status)
                    
                    if (old_profile_status in [0, 1]) and (get_profile_status not in [1, 2, 3, 4]):
                        return Response(
                            {'error': 'You do not have permission to edit this profile.'}, 
                            status=status.HTTP_403_FORBIDDEN
                        )

                    if (old_profile_status in [4,3,2]):
                        return Response(
                            {'error': 'You do not have permission to edit this profile.'}, 
                            status=status.HTTP_403_FORBIDDEN
                        )
                    
                    if ((old_profile_status ==1) and (get_profile_status in [2,3,4])):
                        return Response(
                            {'error': 'You do not have permission to edit this profile.'}, 
                            status=status.HTTP_403_FORBIDDEN
                        )


                if edit_mem ==2:
                    login_detail = LoginDetails.objects.get(ProfileId=profile_id)
                    get_plan_status = profile_common_data.get("secondary_status")
                    get_profile_status = profile_common_data.get("status")
                    old_profile_status = getattr(login_detail, 'status', None)
                    old_plan_status = getattr(login_detail, 'plan_status', None)
                    # print('old_profile_status',old_profile_status)
                    # print('get_plan_status',get_plan_status)
                    # print('get_profile_status',get_profile_status)
                    
                
                    if (old_profile_status == 0 or old_profile_status == 1) and (get_plan_status is not None) and (get_plan_status not in [6, 7, 8, 9]):
                        return Response({'error': 'You do not have permission to edit this profile.'}, status=status.HTTP_403_FORBIDDEN)
                    if old_profile_status != 0 and old_profile_status !=1 :
                        return Response({'error': 'You do not have permission to edit this profile.'}, status=status.HTTP_403_FORBIDDEN)

                    if (old_profile_status in [4,3,2]):
                        return Response(
                            {'error': 'You do not have permission to edit this profile.'}, 
                            status=status.HTTP_403_FORBIDDEN
                        )

                    if ((old_profile_status ==1) and (old_plan_status not in [6, 7, 8, 9])):
                        return Response(
                            {'error': 'You do not have permission to edit this profile.'}, 
                            status=status.HTTP_403_FORBIDDEN
                        )

                login_common_data = clean_none_fields({
                    "Addon_package": profile_common_data.get("Addon_package"),
                    "Notifcation_enabled": profile_common_data.get("Notifcation_enabled"),
                    "PaymentExpire": profile_common_data.get("PaymentExpire"),
                    "Package_name": profile_common_data.get("Package_name"),
                    "status": profile_common_data.get("status"),
                    "DateOfJoin": profile_common_data.get("DateOfJoin"),
                    "Profile_name": profile_common_data.get("Profile_name"),
                    "Gender": profile_common_data.get("Gender"),
                    "Mobile_no": profile_common_data.get("Mobile_no"),
                    "membership_startdate": parse_membership_date(profile_common_data.get("membership_fromdate")),
                    "membership_enddate": parse_membership_date(profile_common_data.get("membership_todate")),
                    "Profile_for": profile_common_data.get("Profile_for"),
                    "primary_status":profile_common_data.get("status"),
                    "secondary_status":profile_common_data.get("primary_status"),
                    "plan_status":profile_common_data.get("secondary_status"),
                    "Plan_id": str(profile_common_data.get("secondary_status")),
                    "Otp_verify":profile_common_data.get("mobile_otp_verify"),
                    "Owner_id":profile_common_data.get("profile_owner_id"),
                })
                family_common_data=clean_none_fields({
                    "family_status":profile_common_data.get("family_status")
                })
                horos_common_data=clean_none_fields({
                    "calc_chevvai_dhosham":profile_common_data.get("calc_chevvai_dhosham"),
                    "calc_raguketu_dhosham":profile_common_data.get("calc_raguketu_dhosham"),
                    "horoscope_hints":profile_common_data.get("horoscope_hints")
                })
                profileplan_common_data=clean_none_fields({
                    "exp_int_lock":profile_common_data.get("exp_int_lock"),
                    "express_int_count":profile_common_data.get("exp_int_count"),
                    "profile_permision_toview":profile_common_data.get("visit_count"),
                    "plan_id":profile_common_data.get("secondary_status"),
                    # "membership_fromdate":profile_common_data.get("membership_fromdate"),
                    # "membership_todate":profile_common_data.get("membership_todate")
                    "membership_fromdate": parse_membership_date(profile_common_data.get("membership_fromdate")),
                    "membership_todate": parse_membership_date(profile_common_data.get("membership_todate")),

                })
            else:
                return Response({'error': 'You do not have permission to edit this profile.'}, status=status.HTTP_403_FORBIDDEN)

            
            # try:
            #     if owner:
            #         owner_id =int(owner)
            #     else:
            #         owner_id = None
            #     old_status = getattr(login_detail, 'status', None)
            #     new_status = profile_common_data.get("status") or old_status

            #     old_plan_id = getattr(login_detail, 'Plan_id', None)
            #     new_plan_id = profile_common_data.get("secondary_status") or old_plan_id
            #     others_id = profile_common_data.get("primary_status")
            #     try:
            #         status_sub = ProfileSubStatus.objects.get(id=others_id)
            #         others = status_sub.sub_status_name
            #     except:    
            #         others = None
                
            #     if old_status is not None and int(old_status) != int(new_status) and int(old_plan_id) != int(new_plan_id):
            #         try:
            #             DataHistory.objects.create(
            #                 profile_id=profile_id,
            #                 profile_status=new_status,
            #                 plan_id=new_plan_id,
            #                 owner_id = owner_id
            #             )
            #         except Exception as e:
            #             pass
            #     elif old_status is not None and int(old_status) != int(new_status):
            #         try:
            #             DataHistory.objects.create(
            #                 profile_id=profile_id,
            #                 profile_status=new_status,
            #                 owner_id = owner_id,
            #                 others=others
            #             )
            #         except Exception as e:
            #             pass
                    
            #     elif int(old_plan_id) != int(new_plan_id):
            #         try:
            #             DataHistory.objects.create(
            #                 profile_id=profile_id,
            #                 profile_status=new_status, 
            #                 plan_id=new_plan_id,
            #                 owner_id = owner_id
            #             )
            #         except Exception as e:
            #             pass
            
            try:
                owner_id = int(owner) if owner else None

                old_status = getattr(login_detail, 'status', None)
                new_status = profile_common_data.get("status") or old_status
                old_sub_status = getattr(login_detail, 'secondary_status', None)
                new_sub_status = profile_common_data.get("primary_status") or old_sub_status

                old_plan_id = getattr(login_detail, 'Plan_id', None)
                new_plan_id = profile_common_data.get("secondary_status") or old_plan_id
                other = profile_common_data.get("others")
                status_changed = (
                    old_status is not None and
                    new_status is not None and
                    int(old_status) != int(new_status)
                )
                sub_status_changed = (
                    old_sub_status is not None and
                    new_sub_status is not None and
                    int(old_sub_status) != int(new_sub_status)
                )
                plan_changed = (
                    old_plan_id is not None and
                    new_plan_id is not None and
                    int(old_plan_id) != int(new_plan_id)
                )

                if not status_changed and not plan_changed and not sub_status_changed:
                    if old_status == 4:
                        others_field = 'delete_others'
                    elif old_status == 3:
                        others_field = 'hide_others'
                    elif old_status == 2:
                        others_field = 'pending_others'
                    else:
                        others_field = None
                    latest_log = (
                        DataHistory.objects
                        .filter(profile_id=profile_id)
                        .order_by('-date_time')
                        .first()
                    )

                    if latest_log:
                        update_data = {
                            others_field: other,
                        }

                        DataHistory.objects.filter(id=latest_log.id).update(**update_data)
                else:
                    others_id = profile_common_data.get("primary_status")

                    try:
                        status_sub = ProfileSubStatus.objects.get(id=others_id)
                        others_text = status_sub.sub_status_name
                    except ProfileSubStatus.DoesNotExist:
                        others_text = None
                        
                    others_field = None
                    if status_changed:
                        if new_status == 4:
                            others_field = 'delete_others'
                        elif new_status == 3:
                            others_field = 'hide_others'
                        elif new_status == 2:
                            others_field = 'pending_others'
                        else:
                            others_field = None

                    data_history_payload = {
                        'profile_id': profile_id,
                        'owner_id': owner_id,
                        'profile_status': new_status,
                    }

                    if plan_changed:
                        data_history_payload['plan_id'] = new_plan_id

                    if status_changed and others_field:
                        data_history_payload['others'] = others_text
                        if other:
                            data_history_payload[others_field] = other

                    DataHistory.objects.create(**data_history_payload)
                    
            except Exception as e:
                # print('Error logging data history:', str(e))
                pass
            # print('login_common_data', login_common_data)
            # Update Login Details
            login_detail = LoginDetails.objects.get(ProfileId=profile_id)
            login_serializer = LoginEditSerializer(instance=login_detail, data=login_common_data, partial=True)
            if login_serializer.is_valid():
                login_serializer.save()
            else:
                return Response({'error': login_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

            # Update Family Details
            family_detail = ProfileFamilyDetails.objects.get(profile_id=profile_id)
            family_serializer = ProfileFamilyDetailsSerializer(instance=family_detail, data=family_common_data, partial=True)
            if family_serializer.is_valid():
                family_serializer.save()
            else:
                return Response({'error': family_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

            # Update Horoscope Details
            horoscope_detail = ProfileHoroscope.objects.get(profile_id=profile_id)
            horoscope_serializer = ProfileHoroscopeSerializer(instance=horoscope_detail, data=horos_common_data, partial=True)
            if horoscope_serializer.is_valid():
                horoscope_serializer.save()
            else:
                return Response({'error': horoscope_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

            plan_id = profile_common_data.get("secondary_status")
            plan_features = models.PlanFeatureLimit.objects.filter(plan_id=plan_id).values().first()
            

            if plan_features:
                # Remove the 'id' field if present
                plan_features.pop('id', None)
                plan_features.pop('plan_id', None)  # optional, if you don't want to override plan_id

                # Add membership dates
                plan_features.update({
                    'plan_id': plan_id,
                    'membership_fromdate': parse_membership_date(profile_common_data.get("membership_fromdate")),
                    'membership_todate': parse_membership_date(profile_common_data.get("membership_todate")),
                    'status':1
                })
                # print(plan_features,'plan features updated')
                # Update the profile_plan_features row for profile_id
                models.Profile_PlanFeatureLimit.objects.filter(profile_id=profile_id).update(**plan_features)
                # print(pro_plan,'profile plan feature updated')
                
            # # Update profileplan Details
            profileplan_detail = Profile_PlanFeatureLimit.objects.get(profile_id=profile_id,status=1)
            profileplan_serializer = ProfileplanSerializer(instance=profileplan_detail, data=profileplan_common_data, partial=True)
            if profileplan_serializer.is_valid():
                # print('profile plan serializer is valid', profileplan_serializer.validated_data)
                profileplan_serializer.save()
            else:
                return Response({'error': profileplan_serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

            addon_package_ids = profile_common_data.get("Addon_package", "")

            if addon_package_ids:
                # Split comma-separated string into list of ints
                addon_package_id_list = [int(pk.strip()) for pk in addon_package_ids.split(",") if pk.strip().isdigit()]

                # Check if ID 1 is in the list
                if 1 in addon_package_id_list:
                    # print("Addon Package ID 1 found. Updating Profile_plan_feature...")

                    # Example: update all rows (or filter if needed)
                    Profile_PlanFeatureLimit.objects.filter(profile_id=profile_id).update(vys_assist=1,vys_assist_count=5)
     
        # If there are any validation errors, return them
        if errors:
            return Response(errors, status=status.HTTP_400_BAD_REQUEST)

        # Success response
        return Response({"status": "success", "message": "Profile updated successfully."}, status=status.HTTP_200_OK)

class DashboardAPIView(APIView):
    def get(self, request, *args, **kwargs):
        owner_id = request.query_params.get('owner_id')
        if not owner_id:
            return Response({"error": "owner_id is required"}, status=400)

        serializer = DashboardSerializer(instance={"owner_id": int(owner_id)})
        return Response(serializer.data)

class Files_upload(APIView):

    def post(self, request):
        profile_id = request.data.get('profile_id')
        admin_user_id = request.data.get('admin_user_id')

        if not profile_id:
            return JsonResponse({"error": "profile_id is required"}, status=status.HTTP_400_BAD_REQUEST)

        # Get user & permission
        try:
            user = User.objects.get(id=admin_user_id)
            role = user.role

            perm = RolePermission.objects.filter(role=role, action__code='edit_horo_photo').first()
            can_edit = perm.value if perm else 0

        except Exception:
            user = None
            can_edit = 0   # no permission

        horoscope_file = request.FILES.get('horoscope_file')
        horoscope_file_admin = request.FILES.get('horoscope_file_admin')
        idproof_file = request.FILES.get('idproof_file')
        divorcepf_file = request.FILES.get('divorcepf_file')

        if not any([horoscope_file, horoscope_file_admin, idproof_file, divorcepf_file]):
            return JsonResponse({"error": "Upload at least one file."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            horoscope_instance, _ = models.ProfileHoroscope.objects.get_or_create(profile_id=profile_id)
            registration_instance = models.LoginDetails.objects.get(ProfileId=profile_id)

            max_file_size = 10 * 1024 * 1024
            valid_extensions = ['doc', 'docx', 'pdf', 'png', 'jpeg', 'jpg']

            # --------------------------
            # 1️ Horoscope File (User Upload)
            # --------------------------
            if horoscope_file:

                # Editing existing? → check permission
                if horoscope_instance.horoscope_file and can_edit != 1:
                    return JsonResponse({"error": "Permission denied for editing horoscope file."},
                                        status=status.HTTP_403_FORBIDDEN)

                # Validate file
                if horoscope_file.size > max_file_size:
                    return JsonResponse({"error": "Horoscope file too large"}, status=status.HTTP_400_BAD_REQUEST)

                ext = horoscope_file.name.split(".")[-1].lower()
                if ext not in valid_extensions:
                    return JsonResponse({"error": "Invalid horoscope file format"}, status=status.HTTP_400_BAD_REQUEST)

                horoscope_instance.horoscope_file.save(horoscope_file.name, ContentFile(horoscope_file.read()), save=True)
                horoscope_instance.horo_file_updated = timezone.now()
                horoscope_instance.save()

            # --------------------------
            # 2️ Horoscope File (Admin Upload)
            # --------------------------
            if horoscope_file_admin:

                if horoscope_instance.horoscope_file_admin and can_edit != 1:
                    return JsonResponse({"error": "Permission denied for editing admin horoscope file."},
                                        status=status.HTTP_403_FORBIDDEN)

                if horoscope_file_admin.size > max_file_size:
                    return JsonResponse({"error": "Admin horoscope file too large"}, status=status.HTTP_400_BAD_REQUEST)

                ext = horoscope_file_admin.name.split(".")[-1].lower()
                if ext not in valid_extensions:
                    return JsonResponse({"error": "Invalid horoscope admin file format"}, status=status.HTTP_400_BAD_REQUEST)

                horoscope_instance.horoscope_file_admin.save(horoscope_file_admin.name, ContentFile(horoscope_file_admin.read()), save=True)
                horoscope_instance.horo_file_updated = timezone.now()
                horoscope_instance.save()

            # --------------------------
            # 3️ ID Proof
            # --------------------------
            if idproof_file:

                if registration_instance.Profile_idproof and can_edit != 1:
                    return JsonResponse({"error": "Permission denied for editing ID proof."},
                                        status=status.HTTP_403_FORBIDDEN)

                ext = idproof_file.name.split(".")[-1].lower()
                if ext not in valid_extensions:
                    return JsonResponse({"error": "Invalid ID proof file format"}, status=status.HTTP_400_BAD_REQUEST)

                registration_instance.Profile_idproof.save(idproof_file.name, ContentFile(idproof_file.read()), save=True)
                registration_instance.save()

            # --------------------------
            # 4️ Divorce Proof
            # --------------------------
            if divorcepf_file:

                if registration_instance.Profile_divorceproof and can_edit != 1:
                    return JsonResponse({"error": "Permission denied for editing divorce proof."},
                                        status=status.HTTP_403_FORBIDDEN)

                ext = divorcepf_file.name.split(".")[-1].lower()
                if ext not in valid_extensions:
                    return JsonResponse({"error": "Invalid divorce proof file format"}, status=status.HTTP_400_BAD_REQUEST)

                registration_instance.Profile_divorceproof.save(divorcepf_file.name, ContentFile(divorcepf_file.read()), save=True)
                registration_instance.save()

            return JsonResponse({"status": "success"}, status=status.HTTP_200_OK)

        except models.Registration1.DoesNotExist:
            return JsonResponse({"error": "Profile does not exist"}, status=status.HTTP_404_NOT_FOUND)


#Call management New Api


class GetDropdownMasters(APIView):
    def get(self, request):
        try:
            call_types = list(CallTypeMaster.objects.values("id", "call_type"))
            particulars = list(ParticularsMaster.objects.values("id", "particulars"))
            call_status = list(CallStatusMaster.objects.values("id", "status"))
            action_points = list(ActionPointMaster.objects.values("id", "action_point"))

            return Response({
                "call_types": call_types,
                "particulars": particulars,
                "call_status": call_status,
                "action_points": action_points
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



@api_view(['POST'])
def save_call_management(request):

    data = request.data
    call_management_id = data.get("call_management_id", None)
    profile_id = data["profile_id"]

    # ------------------------------
    # 1. Create or get call_management entry
    # ------------------------------
    if call_management_id:
        #cm = CallManagement.objects.get(id=call_management_id)

        try:
            cm = CallManagement.objects.get(id=call_management_id)

            # Check if created more than 24 hours ago
            time_diff = timezone.now() - cm.created_at

            if time_diff > timedelta(hours=24):
                return Response({
                    "status": "failed",
                    "message": "You cannot update this record. Edit time limit (24 hours) has expired."
                }, status=403)

        except CallManagement.DoesNotExist:
            return Response({"status": "failed", "message": "Invalid call_management_id"}, status=400)
    else:
        cm = CallManagement.objects.create(profile_id=profile_id)

    # ------------------------------
    # 2. Save Call Logs (Insert + Update)
    # ------------------------------
    submitted_call_ids = []

    for item in data.get("call_logs", []):
        log_id = item.get("id", None)

        if log_id:
            log = CallLog.objects.get(id=log_id)
        else:
            log = CallLog(call_management=cm)

        log.call_date = item["call_date"]
        log.call_owner = item.get("call_owner")
        log.next_call_date= item.get("next_call_date")
        log.call_type_id = item.get("call_type_id")
        log.particulars_id = item.get("particulars_id")
        log.call_status_id = item.get("call_status_id")
        log.comments = item.get("comments")

        log.save()
        submitted_call_ids.append(log.id)

    # Delete removed call logs
    CallLog.objects.filter(call_management=cm).exclude(id__in=submitted_call_ids).delete()

    # ------------------------------
    # 3. Save Action Logs
    # ------------------------------
    submitted_action_ids = []

    for item in data.get("action_logs", []):
        log_id = item.get("id", None)

        if log_id:
            log = ActionLog.objects.get(id=log_id)
        else:
            log = ActionLog(call_management=cm)

        log.action_date = item["action_date"]
        log.next_action_date=item["next_action_date"]
        log.action_owner = item.get("action_owner")
        log.action_point_id = item.get("action_point_id")
        log.next_action_id = item.get("next_action_id")
        log.comments = item.get("comments")

        log.save()
        submitted_action_ids.append(log.id)

    ActionLog.objects.filter(call_management=cm).exclude(id__in=submitted_action_ids).delete()

    # ------------------------------
    # 4. Save Assign Logs
    # ------------------------------
    submitted_assign_ids = []

    for item in data.get("assign_logs", []):
        log_id = item.get("id", None)

        if log_id:
            log = AssignLog.objects.get(id=log_id)
        else:
            log = AssignLog(call_management=cm)

        log.assigned_date = item["assigned_date"]
        log.assigned_to = item.get("assigned_to")
        log.assigned_by = item.get("assigned_by")
        log.notes = item.get("notes")

        log.save()
        submitted_assign_ids.append(log.id)

    AssignLog.objects.filter(call_management=cm).exclude(id__in=submitted_assign_ids).delete()

    return Response({
        "status": "success",
        "call_management_id": cm.id
    })



@api_view(['GET'])
def get_all_call_logs_by_profile(request, profile_id):

    call_ids = CallManagement.objects.filter(profile_id=profile_id).values_list('id', flat=True)

    call_logs = (
        CallLog.objects
        .filter(call_management_id__in=call_ids,is_deleted=0)
        .select_related('call_type', 'particulars', 'call_status')
        .annotate(
            call_type_name=F('call_type__call_type'),
            particulars_name=F('particulars__particulars'),
            call_status_name=F('call_status__status'),
        )
        .order_by('-call_date')
        .values(
            'id',
            'call_management_id',
            'call_date',
            'comments',
            'next_call_date',
            'call_owner',
            'created_at',

            'call_type_id',
            'call_type_name',

            'particulars_id',
            'particulars_name',

            'call_status_id',
            'call_status_name',
        )
    )


    # Collect all call_owner user IDs
    user_ids = set()
    for log in call_logs:
        if log["call_owner"]:
            try:
                user_ids.add(int(log["call_owner"]))   # convert charfield to int
            except:
                pass

    # Fetch all users in ONE query
    users = User.objects.in_bulk(user_ids)

    # Add action_owner_name to each log
    for log in call_logs:
        owner_id = log["call_owner"]
        try:
            owner_id = int(owner_id)
        except:
            owner_id = None

        log["call_owner_name"] = (
            users[owner_id].username if owner_id in users else None
        )

    return Response({
        "profile_id": profile_id,
        "call_logs": list(call_logs)
    })

from django.db.models import F

@api_view(['GET'])
def get_all_action_logs_by_profile(request, profile_id):

    call_ids = CallManagement.objects.filter(profile_id=profile_id).values_list('id', flat=True)

    # Fetch main action logs
    action_logs = list(
        ActionLog.objects
        .filter(call_management_id__in=call_ids,is_deleted=0)
        .select_related('action_point', 'next_action')
        .annotate(
            action_point_name=F('action_point__action_point'),
            next_action_name=F('next_action__action_point'),
        )
        .order_by('-action_date')
        .values(
            'id',
            'call_management_id',
            'action_date',
            'comments',
            'created_at',
            
            'action_point_id',
            'action_point_name',
            'next_action_date',
            'next_action_id',
            'next_action_name',

            'action_owner'
        )
    )

    # Collect all action_owner user IDs
    user_ids = set()
    for log in action_logs:
        if log["action_owner"]:
            try:
                user_ids.add(int(log["action_owner"]))   # convert charfield to int
            except:
                pass

    # Fetch all users in ONE query
    users = User.objects.in_bulk(user_ids)

    # Add action_owner_name to each log
    for log in action_logs:
        owner_id = log["action_owner"]
        try:
            owner_id = int(owner_id)
        except:
            owner_id = None

        log["action_owner_name"] = (
            users[owner_id].username if owner_id in users else None
        )

    return Response({
        "profile_id": profile_id,
        "action_logs": action_logs
    })







# @api_view(['GET'])
# def get_all_assign_logs_by_profile(request, profile_id):

#     # get all call management ids for the profile
#     call_ids = CallManagement.objects.filter(profile_id=profile_id).values_list('id', flat=True)

#     # get all assign logs for these call ids
#     assign_logs = AssignLog.objects.filter(call_management_id__in=call_ids).values()

#     return Response({
#         "profile_id": profile_id,
#         "assign_logs": list(assign_logs)
#     })

@api_view(['GET'])
def get_all_assign_logs_by_profile(request, profile_id):

    call_ids = CallManagement.objects.filter(profile_id=profile_id).values_list('id', flat=True)
    assign_logs = list(AssignLog.objects.filter(call_management_id__in=call_ids,is_deleted=0).order_by('-assigned_date').values())

    # Collect all unique user IDs
    user_ids = set()
    for log in assign_logs:
        if log["assigned_by"]:
            user_ids.add(log["assigned_by"])
        if log["assigned_to"]:
            user_ids.add(log["assigned_to"])

    # Fetch user objects in a single query
    users = User.objects.in_bulk(user_ids)

    # Append only the user names
    for log in assign_logs:
        log["assigned_by_name"] = (
            users[log["assigned_by"]].username if log["assigned_by"] in users else None
        )
        log["assigned_to_name"] = (
            users[log["assigned_to"]].username if log["assigned_to"] in users else None
        )

    return Response({
        "profile_id": profile_id,
        "assign_logs": assign_logs
    })


@api_view(['GET'])
def get_logs_by_profile(request, profile_id):

    call_ids = CallManagement.objects.filter(profile_id=profile_id).values_list('id', flat=True)


    try:
        login_detail = LoginDetails.objects.get(ProfileId=profile_id)

        # Get Owner details from Users table
        owner_user = None
        owner_name = None
        if login_detail.Owner_id:
            try:
                owner_user = User.objects.get(id=login_detail.Owner_id)
                owner_name = owner_user.username   # or first_name / full_name
            except User.DoesNotExist:
                owner_name = None

        # Get Status name from ProfileStatus table
        status_name = None
        if login_detail.status:
            try:
                status = ProfileStatus.objects.get(status_code=login_detail.status)
                status_name = status.status_name
            except ProfileStatus.DoesNotExist:
                status_name = None

    except LoginDetails.DoesNotExist:
            return Response({'error': 'Profile not found.'}, status=status.HTTP_404_NOT_FOUND)

    # -----------------------
    # CALL LOGS
    # -----------------------
    call_logs = CallLog.objects.filter(
        call_management_id__in=call_ids
    ).order_by('-call_date').select_related("call_type", "particulars", "call_status")
    

    call_log_list = []
    for c in call_logs:
        call_log_list.append({
            "id": c.id,
            "call_date": c.call_date,
            "call_type_id": c.call_type_id,
            "call_type_name": c.call_type.call_type if c.call_type else None,
            "particulars_id": c.particulars_id,
            "particulars_name": c.particulars.particulars if c.particulars else None,
            "call_status_id": c.call_status_id,
            "call_status_name": c.call_status.status if c.call_status else None,
            "comments": c.comments,
            "call_management_id": c.call_management_id,
            "created_at": c.created_at
        })

    # -----------------------
    # ACTION LOGS
    # -----------------------
    action_logs = ActionLog.objects.filter(
        call_management_id__in=call_ids
    ).order_by('-action_date').select_related("action_point", "next_action")

    action_log_list = []
    for a in action_logs:
        action_log_list.append({
            "id": a.id,
            "action_date": a.action_date,
            "next_action_date":a.next_action_date,
            "action_point_id": a.action_point_id,
            "action_point_name": a.action_point.action_point if a.action_point else None,
            "next_action_id": a.next_action_id,
            "next_action_name": a.next_action.action_point if a.next_action else None,
            "comments": a.comments,
            "call_management_id": a.call_management_id,
            "created_at": a.created_at
        })

    # -----------------------
    # ASSIGN LOGS (simple)
    # -----------------------
    assign_logs = list(
        AssignLog.objects.filter(call_management_id__in=call_ids).values(
            "id", "assigned_date", "assigned_to", "assigned_by",
            "notes", "call_management_id", "created_at"
        ).order_by('-assigned_date')
    )

    return Response({
        "profile_id": profile_id,
        "owner_name": owner_name,
        "status_name": status_name,
        "call_logs": call_log_list,
        "action_logs": action_log_list,
        "assign_logs": assign_logs
    })





@api_view(['GET'])
def get_all_logs_by_call_id(request, call_id):

    return Response({
        "call_management_id": call_id,

        "call_logs": list(CallLog.objects.filter(call_management_id=call_id).values()),
        "action_logs": list(ActionLog.objects.filter(call_management_id=call_id).values()),
        "assign_logs": list(AssignLog.objects.filter(call_management_id=call_id).values())
    })



class CallManageDeleteView(APIView):

    # Map table names → model classes
    TABLE_MODEL_MAP = {
        "call_logs": CallLog,
        "action_logs": ActionLog,
        "assign_logs": AssignLog,
    }

    def post(self, request):

        table_name = request.data.get("delete_module")
        record_id = request.data.get("call_id")   # corrected key
        deleted_by = request.data.get("deleted_by")

        
        # Validate table
        if table_name not in self.TABLE_MODEL_MAP:
            return Response({
                "status": "failed",
                "message": "Invalid table name. Allowed: call_logs, action_logs, assign_logs"
            }, status=400)

        model = self.TABLE_MODEL_MAP[table_name]

        # Fetch the record
        try:
            record = model.objects.get(id=record_id)
        except model.DoesNotExist:
            return Response({
                "status": "failed",
                "message": "Record not found"
            }, status=404)

        # ------------------------------------------------
        # 24-HOUR VALIDATION (Do NOT allow delete)
        # ------------------------------------------------
        created_time = record.created_at
        time_diff = timezone.now() - created_time

        if time_diff > timedelta(hours=24):
            return Response({
                "status": "failed",
                "message": "You cannot delete this record. Delete time limit (24 hours) has expired."
            }, status=403)

        # Soft delete logic
        record.is_deleted = 1
        record.deleted_at = timezone.now()

        if hasattr(record, "deleted_by"):
            record.deleted_by = deleted_by

        record.save()

        return Response({
            "status": "success",
            "message": "Call Record deleted successfully",
            "table": table_name,
            "record_id": record_id
        })
        
class DeleteAttachFile(APIView):

    def post(self, request, *args, **kwargs):
        try:
            model_type = request.data.get('model_type')
            profile_id = request.data.get('profile_id')
            field_name = request.data.get('field_name')
            print("model_type:",model_type,profile_id,field_name)
            if not model_type or not profile_id or not field_name:
                return JsonResponse({
                    'success': 0,
                    'message': 'model_type, profile_id and field_name are required.'
                }, status=400)

            if model_type == 'registration':
                instance = get_object_or_404(LoginDetails, ProfileId=profile_id)
                allowed_fields = ['Profile_idproof', 'Profile_divorceproof']
            elif model_type == 'horoscope':
                instance = get_object_or_404(Horoscope, profile_id=profile_id)
                allowed_fields = ['horoscope_file', 'horoscope_file_admin']
            else:
                return JsonResponse({
                    'success': 0,
                    'message': 'Invalid model_type. Must be "registration" or "horoscope".'
                }, status=400)

            if field_name not in allowed_fields:
                return JsonResponse({
                    'success': 0,
                    'message': f'Invalid field_name for {model_type}. Allowed: {allowed_fields}'
                }, status=400)

            file_field = getattr(instance, field_name, None)
            if file_field:
                try:
                    file_field.delete(save=False) 
                except Exception as e:
                    return JsonResponse({
                        'success': 0,
                        'message': f'Failed to delete file from Azure: {str(e)}'
                    }, status=500)

                setattr(instance, field_name, None)
                instance.save()

                return JsonResponse({
                    'success': 1,
                    'message': f'{field_name} deleted successfully.'
                }, status=200)
            else:
                return JsonResponse({
                    'success': 0,
                    'message': f'{field_name} is already empty or does not exist.'
                }, status=404)

        except Exception as e:
            return JsonResponse({
                'success': 0,
                'message': str(e)
            }, status=500)
        





@api_view(['POST'])
def new_save_call_management(request):

    data = request.data
    call_management_id = data.get("call_management_id", None)
    profile_id = data.get("profile_id")
    mobile_no = data.get("mobile_no")


    # ------------------------------
    # 1. Create or get call_management entry
    # ------------------------------
    if call_management_id:
        #cm = CallManagement.objects.get(id=call_management_id)

        try:
            cm = CallManagement_New.objects.get(id=call_management_id)

            # Check if created more than 24 hours ago
            time_diff = timezone.now() - cm.created_at

            if time_diff > timedelta(hours=24):
                return Response({
                    "status": "failed",
                    "message": "You cannot update this record. Edit time limit (24 hours) has expired."
                }, status=403)

        except CallManagement_New.DoesNotExist:
            return Response({"status": "failed", "message": "Invalid call_management_id"}, status=400)
    else:
        cm = CallManagement_New.objects.create()


    cm.profile_id = profile_id
    cm.mobile_no = mobile_no
    cm.save()
    # ------------------------------
    # 2. Save Call Logs (Insert + Update)
    # ------------------------------
    submitted_call_ids = []

    for item in data.get("call_logs", []):
        log_id = item.get("id", None)

        if log_id:
            log = CallLog_New.objects.get(id=log_id)
        else:
            log = CallLog_New(call_management=cm)

        log.call_date = item["call_date"]
        log.call_owner = item.get("call_owner")
        log.next_call_date= item.get("next_call_date")
        log.call_type_id = item.get("call_type_id")
        log.particulars_id = item.get("particulars_id")
        log.call_status_id = item.get("call_status_id")
        log.comments = item.get("comments")

        log.save()
        submitted_call_ids.append(log.id)

    # Delete removed call logs
    CallLog_New.objects.filter(call_management=cm).exclude(id__in=submitted_call_ids).delete()

    # ------------------------------
    # 3. Save Action Logs
    # ------------------------------
    submitted_action_ids = []

    for item in data.get("action_logs", []):
        log_id = item.get("id", None)

        if log_id:
            log = ActionLog_New.objects.get(id=log_id)
        else:
            log = ActionLog_New(call_management=cm)

        log.action_date = item["action_date"]
        log.next_action_date=item["next_action_date"]
        log.action_owner = item.get("action_owner")
        log.action_point_id = item.get("action_point_id")
        log.next_action_id = item.get("next_action_id")
        log.comments = item.get("comments")

        log.save()
        submitted_action_ids.append(log.id)

    ActionLog_New.objects.filter(call_management=cm).exclude(id__in=submitted_action_ids).delete()

    # ------------------------------
    # 4. Save Assign Logs
    # ------------------------------
    submitted_assign_ids = []

    for item in data.get("assign_logs", []):
        log_id = item.get("id", None)

        if log_id:
            log = AssignLog_New.objects.get(id=log_id)
        else:
            log = AssignLog_New(call_management=cm)

        log.assigned_date = item["assigned_date"]
        log.assigned_to = item.get("assigned_to")
        log.assigned_by = item.get("assigned_by")
        log.notes = item.get("notes")

        log.save()
        submitted_assign_ids.append(log.id)

    AssignLog_New.objects.filter(call_management=cm).exclude(id__in=submitted_assign_ids).delete()

    return Response({
        "status": "success",
        "call_management_id": cm.id
    })



@api_view(['GET'])
def new_get_all_call_logs(request):

    call_ids = CallManagement_New.objects.filter().values_list('id', flat=True)

    call_logs = (
        CallLog_New.objects
        .filter(call_management_id__in=call_ids,is_deleted=0)
        .select_related('call_type', 'particulars', 'call_status')
        .annotate(
            call_type_name=F('call_type__call_type'),
            particulars_name=F('particulars__particulars'),
            call_status_name=F('call_status__status'),
        )
        .order_by('-call_date')
        .values(
            'id',
            'call_management_id',
            'call_date',
            'comments',
            'next_call_date',
            'call_owner',
            'created_at',

            'call_type_id',
            'call_type_name',

            'particulars_id',
            'particulars_name',

            'call_status_id',
            'call_status_name',
            'call_management__profile_id',
            'call_management__mobile_no',

        )
    )


    # Collect all call_owner user IDs
    user_ids = set()
    for log in call_logs:
        if log["call_owner"]:
            try:
                user_ids.add(int(log["call_owner"]))   # convert charfield to int
            except:
                pass

    # Fetch all users in ONE query
    users = User.objects.in_bulk(user_ids)

    # Add action_owner_name to each log
    for log in call_logs:
        owner_id = log["call_owner"]
        try:
            owner_id = int(owner_id)
        except:
            owner_id = None

        log["call_owner_name"] = (
            users[owner_id].username if owner_id in users else None
        )

    return Response({
        "call_logs": list(call_logs)
    })

@api_view(['GET'])
def new_get_all_action_logs(request):

    call_ids = CallManagement_New.objects.filter().values_list('id', flat=True)

    # Fetch main action logs
    action_logs = list(
        ActionLog_New.objects
        .filter(call_management_id__in=call_ids,is_deleted=0)
        .select_related('action_point', 'next_action')
        .annotate(
            action_point_name=F('action_point__action_point'),
            next_action_name=F('next_action__action_point'),
        )
        .order_by('-action_date')
        .values(
            'id',
            'call_management_id',
            'action_date',
            'comments',
            'created_at',
            'action_point_id',
            'action_point_name',
            'next_action_date',
            'next_action_id',
            'next_action_name',

            'action_owner',
            'call_management__profile_id',
            'call_management__mobile_no',
        )
    )

    # Collect all action_owner user IDs
    user_ids = set()
    for log in action_logs:
        if log["action_owner"]:
            try:
                user_ids.add(int(log["action_owner"]))   # convert charfield to int
            except:
                pass

    # Fetch all users in ONE query
    users = User.objects.in_bulk(user_ids)

    # Add action_owner_name to each log
    for log in action_logs:
        owner_id = log["action_owner"]
        try:
            owner_id = int(owner_id)
        except:
            owner_id = None

        log["action_owner_name"] = (
            users[owner_id].username if owner_id in users else None
        )

    return Response({
        "action_logs": action_logs
    })

@api_view(['GET'])
def new_get_all_assign_logs(request):

    call_ids = CallManagement_New.objects.filter().values_list('id', flat=True)

    # Fetch assign logs WITH profile + mobile using JOIN
    assign_logs = list(
        AssignLog_New.objects
        .filter(call_management_id__in=call_ids, is_deleted=0)
        .select_related("call_management")  # JOIN
        .order_by('-assigned_date')
        .values(
            'id',
            'call_management_id',
            'assigned_date',
            'assigned_to',
            'assigned_by',
            'notes',
            'created_at',
            'is_deleted',
            'deleted_by',

            #Add profile & mobile from CallManagement_New
            'call_management__profile_id',
            'call_management__mobile_no',
        )
    )

    # Collect unique user IDs
    user_ids = set()
    for log in assign_logs:
        if log["assigned_by"]:
            user_ids.add(log["assigned_by"])
        if log["assigned_to"]:
            user_ids.add(log["assigned_to"])

    # Fetch user objects in a single query
    users = User.objects.in_bulk(user_ids)

    # Add assigned_by_name & assigned_to_name
    for log in assign_logs:
        log["assigned_by_name"] = (
            users[log["assigned_by"]].username if log["assigned_by"] in users else None
        )
        log["assigned_to_name"] = (
            users[log["assigned_to"]].username if log["assigned_to"] in users else None
        )

        #Rename keys for clean response:
        log["profile_id"] = log.pop('call_management__profile_id')
        log["mobile_no"] = log.pop('call_management__mobile_no')

    return Response({
        "assign_logs": assign_logs
    })



class CallManageDeleteView_New(APIView):

    # Map table names → model classes
    TABLE_MODEL_MAP = {
        "call_logs": CallLog_New,
        "action_logs": ActionLog_New,
        "assign_logs": AssignLog_New,
    }

    def post(self, request):

        table_name = request.data.get("delete_module")
        record_id = request.data.get("call_id")   # corrected key
        deleted_by = request.data.get("deleted_by")

        
        # Validate table
        if table_name not in self.TABLE_MODEL_MAP:
            return Response({
                "status": "failed",
                "message": "Invalid table name. Allowed: call_logs, action_logs, assign_logs"
            }, status=400)

        model = self.TABLE_MODEL_MAP[table_name]

        # Fetch the record
        try:
            record = model.objects.get(id=record_id)
        except model.DoesNotExist:
            return Response({
                "status": "failed",
                "message": "Record not found"
            }, status=404)

        # ------------------------------------------------
        # 24-HOUR VALIDATION (Do NOT allow delete)
        # ------------------------------------------------
        created_time = record.created_at
        time_diff = timezone.now() - created_time

        if time_diff > timedelta(hours=24):
            return Response({
                "status": "failed",
                "message": "You cannot delete this record. Delete time limit (24 hours) has expired."
            }, status=403)

        # Soft delete logic
        record.is_deleted = 1
        record.deleted_at = timezone.now()

        if hasattr(record, "deleted_by"):
            record.deleted_by = deleted_by

        record.save()

        return Response({
            "status": "success",
            "message": "Call Record deleted successfully",
            "table": table_name,
            "record_id": record_id
        })



@api_view(['GET'])
def new_get_all_logs_by_call_id(request, call_id):



    try:
        cm = CallManagement_New.objects.get(id=call_id)
    except CallManagement_New.DoesNotExist:
        return Response({
            "status": "failed",
            "message": "Invalid call_id"
        }, status=400)

    profile_id = cm.profile_id
    mobile_no = cm.mobile_no

    return Response({
        "call_management_id": call_id,
        "profile_id":profile_id,
        "mobile_no":mobile_no,
        "call_logs": list(CallLog_New.objects.filter(call_management_id=call_id).values()),
        "action_logs": list(ActionLog_New.objects.filter(call_management_id=call_id).values()),
        "assign_logs": list(AssignLog_New.objects.filter(call_management_id=call_id).values())
    })
    
def safe_get_value(model, pk_field, value, name_field='name', default='N/A'):
    try:
        if value and str(value).isdigit():
            return model.objects.filter(**{pk_field: value}).values_list(name_field, flat=True).first() or default
    except Exception:
        pass
    return default

class WhatsappShareView(APIView):
    def get(self, request, profile_id=None, *args, **kwargs):
        if not profile_id:
            return Response(
                {'status': 'error', 'message': 'profile_id is required'},
                status=400
            )
        format_type = request.GET.get("format_type")
        try:
            profile = LoginDetails.objects.get(ProfileId=profile_id)
        except Exception:
            return Response(
                {'status': 'error', 'message': 'profile not found'},
                status=404
            )
        try:
            horoscope_data = get_object_or_404(models.ProfileHoroscope, profile_id=profile_id)
            birthstar = safe_get_value(models.BirthStar, 'id', horoscope_data.birthstar_name, 'star')
            patham = horoscope_data.padham
        except Exception:
            birthstar=None
            patham = None
        try:
            education_details = get_object_or_404(models.ProfileEduDetails, profile_id=profile_id)
            annual_income = "Unknown"
            actual_income = str(education_details.actual_income).strip()
            annual_income_id = education_details.anual_income
            try:
                if not actual_income or actual_income in ["", "~"]:
                    if annual_income_id and str(annual_income_id).isdigit():
                        annual_income = models.AnnualIncome.objects.filter(id=int(annual_income_id)).values_list('income', flat=True).first() or "Unknown"
                else:
                    annual_income = actual_income
            except Exception:
                annual_income=None
            try:
                highest_education_id = education_details.highest_education
                highest_education = "Unknown"
                if highest_education_id:
                    highest_education = models.EducationLevel.objects.filter(row_id=highest_education_id).values_list('EducationLevel', flat=True).first() or "Unknown"
            except Exception:
                highest_education = None 
            try:
                work_place =get_work_address(city=education_details.work_city,state=education_details.work_state,district=education_details.work_district,country=education_details.work_country)
            except Exception:
                work_place = None
            
            profession_id = education_details.profession
            profession = "Unknown"
            if profession_id:
                profession = models.Profespref.objects.filter(RowId=profession_id).values_list('profession', flat=True).first() or "Unknown"

            occupation=''

            try:
                prof_id_int = int(profession_id)
                if prof_id_int == 1:
                    occupation = f"{education_details.company_name or ''} / {education_details.designation or ''}"
                elif prof_id_int == 2:
                    occupation = f"{education_details.business_name or ''} / {education_details.nature_of_business or ''}"
            except Exception:
                occupation = None
        except Exception:
            annual_income=None
            highest_education = None
            occupation = None
            work_place = None
        print("format",format_type)
        if format_type == "with_image":
            profile_link = f"https://app.vysyamala.com/auth/profile/{signing.dumps(profile.ProfileId)}/"
        else:  # default to without image
            profile_link = f"https://app.vysyamala.com/auth/profile_view/{signing.dumps(profile.ProfileId)}/"
               
        # Build context for template
        profile_data = {
            "profile_id": profile.ProfileId,
            "profile_name": profile.Profile_name or "N/A",
            "age": calculate_age(profile.Profile_dob) if profile.Profile_dob else "Unknown",
            "star_name": birthstar if birthstar else "N/A",
            "annual_income": annual_income if annual_income else "N/A",
            "highest_education":highest_education if highest_education else "N/A",
            "occupation": occupation if occupation else "N/A",
            "work_place": work_place if work_place else "N/A",
            "padham":patham if patham else None,
            "profile_link": profile_link
        }

        context = {
            "profiles": [profile_data]  # template loops over profiles
        }

        return render(request, "whatsapp_profile.html", context)

    

EXPORT_COLUMNS = [
    ("ProfileId", "Profile ID"),
    ("Profile_name", "Name"),
    ("age", "Age"),
    ("family_status_name", "Family Status"),
    ("education", "Education Details"),
    ("income", "Annual Income"),
    ("Profile_city", "City"),
    ("plan_name", "Mode"),
    ("owner_name", "Owner"),
    ("membership_startdate", "From Date"),
    ("membership_enddate", "To Date"),
    ("Last_login_date", "Last Login"),
    ("idle_days", "Idle Days"),
    ("call_status", "Status"),
]


class ExpiredMembersReport(APIView):

    def get(self, request, *args, **kwargs):

        status_param      = request.GET.get("status", "1")
        owner_param       = request.GET.get("owner", "26")
        age_filter        = request.GET.get("ageFilter", "")
        gender_filter     = request.GET.get("genderFilter", "")
        family_filter     = request.GET.get("familyFilter", "")
        login_filter      = request.GET.get("loginFilter", "")
        expiring_filter   = request.GET.get("expiringFilter", "")
        call_status_filter= request.GET.get("callStatusFilter", "")
        idle_days_filter  = request.GET.get("idleDaysFilter", "")
        action_filter     = request.GET.get("actionFilter", "")
        photo_filter      = request.GET.get("photoFilter", "")   
        horo_filter       = request.GET.get("horoFilter", "") 
        from_date         = request.GET.get("from_date", "")
        to_date           = request.GET.get("to_date", "")
        age_from          = request.GET.get("age_from", "")
        age_to            = request.GET.get("age_to", "")
        plan_id           = request.GET.get("plan_id", "")
        search            = request.GET.get("search", "").strip().lower()
        export_type       = request.GET.get("export", "").lower()
        today = date.today()
        yesterday = today - timedelta(days=1)

        def calculate_age(dob):
            if not dob:
                return None
            try:
                if isinstance(dob, datetime):
                    dob = dob.date()
                return today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))
            except:
                return None

        def safe_date(d):
            try:
                return datetime.strptime(d, "%Y-%m-%d").date()
            except:
                return None

        from_date_obj = safe_date(from_date)
        to_date_obj   = safe_date(to_date)

        final_filtered = []
        with connection.cursor() as cursor:

            cursor.callproc("GetExpiredMembersReport", [
                status_param,
                owner_param,
                age_filter,
                gender_filter,
                family_filter,
                login_filter,
                expiring_filter,
                call_status_filter,
                idle_days_filter,
                from_date,
                to_date,
                age_from,
                age_to,
                plan_id
            ])

            base_rows = cursor.fetchall()
            base_columns = [col[0] for col in cursor.description]
            base_data = [dict(zip(base_columns, row)) for row in base_rows]

            cursor.nextset()
            overall_row = cursor.fetchone()
            overall_count = overall_row[0] if overall_row else len(base_data)

            cursor.nextset()
            filtered_rows = cursor.fetchall() if cursor.description else []
            filtered_columns = [col[0] for col in cursor.description] if cursor.description else []
            filtered_data = [dict(zip(filtered_columns, row)) for row in filtered_rows]

        for row in filtered_data:

            row["age"] = calculate_age(row.get("Profile_dob"))
            age = row["age"]

            next_action_raw = row.get("next_action_date")
            next_action = next_action_raw.date() if isinstance(next_action_raw, datetime) else next_action_raw

            if next_action:
                diff = (today - next_action).days
                row["idle_days"] = diff if diff > 0 else None
            else:
                row["idle_days"] = None

            include = True
            if action_filter:
                today = date.today()

                next_call = row.get("next_call_date")
                if isinstance(next_call, datetime):
                    next_call = next_call.date()

                next_action = row.get("next_action_date")
                if isinstance(next_action, datetime):
                    next_action = next_action.date()

                if action_filter == "today_work":
                    if not next_call or next_call != today:
                        continue

                elif action_filter == "pending_work":
                    if not next_call or next_call >= today:
                        continue

                elif action_filter == "today_task":
                    if not next_action or next_action != today:
                        continue

                elif action_filter == "pending_task":
                    if not next_action or next_action >= today:
                        continue
                
            if include and photo_filter:
                has_photo = bool(row.get("has_photo"))

                if photo_filter == "has_photo" and not has_photo:
                    include = False
                elif photo_filter == "no_photo" and has_photo:
                    include = False

            if include and horo_filter:
                has_horo = bool(row.get("has_horo"))

                if horo_filter == "has_horo" and not has_horo:
                    include = False
                elif horo_filter == "no_horo" and has_horo:
                    include = False

            if search:
                if search not in str(row.get("ProfileId", "")).lower() and \
                   search not in str(row.get("Profile_name", "")).lower():
                    include = False

            if include:
                final_filtered.append(row)
                
                
        if export_type in ["csv", "excel"]:

            export_data = []
            for row in final_filtered:
                education = ""
                if row.get("degree_name") and row.get("other_degree"):
                    education = f"{row.get('degree_name')} / {row.get('other_degree')}"
                else:
                    education = row.get("degree_name") or row.get("other_degree") or ""

                export_row = {
                    "Profile ID": row.get("ProfileId"),
                    "Name": row.get("Profile_name"),
                    "Age": row.get("age"),
                    "Family Status": row.get("family_status_name"),
                    "Education Details": education,
                    "Annual Income": row.get("income"),
                    "City": row.get("Profile_city"),
                    "Mode": row.get("plan_name"),
                    "Owner": row.get("owner_name"),
                    "From Date": row.get("membership_startdate"),
                    "To Date": row.get("membership_enddate"),
                    "Last Login": row.get("Last_login_date"),
                    "Idle Days": row.get("idle_days"),
                    "Status": row.get("call_status"),
                }

                export_data.append(export_row)

            df = pd.DataFrame(export_data)

            if export_type == "csv":
                output = StringIO()
                df.to_csv(output, index=False)
                response = HttpResponse(output.getvalue(), content_type="text/csv")
                response["Content-Disposition"] = 'attachment; filename="expired_members_report.csv"'
                return response

            elif export_type == "excel":
                output = BytesIO()
                with pd.ExcelWriter(output, engine="openpyxl") as writer:
                    df.to_excel(writer, index=False, sheet_name="Report")
                response = HttpResponse(
                    output.getvalue(),
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                response["Content-Disposition"] = 'attachment; filename="expired_members_report.xlsx"'
                return response

        under_30 = above_30 = 0
        male_count = female_count = 0
        today_login = yesterday_login = 0
        expired_count = expiring_count = 0
        hot = warm = cold = not_interested = 0
        idle_45_count = idle_90_count = 0
        today_work_count = pending_work_count = 0
        today_task_count = pending_task_count = 0
        no_photo_count = 0
        no_horo_count = 0
        has_photo_count = 0
        has_horo_count = 0


        family_status_counts = {3: 0, 4: 0, 6: 0}
        family_status_labels = {3: "Upper Middle Class", 4: "Rich", 6: "Affluent"}

        for item in base_data:

            age = calculate_age(item.get("Profile_dob"))
            if age is not None:
                if age < 30:
                    under_30 += 1
                else:
                    above_30 += 1

            gender = str(item.get("Gender", "")).lower()
            if gender in ["male", "m", "1"]:
                male_count += 1
            elif gender in ["female", "f", "2"]:
                female_count += 1

            try:
                fs = int(item.get("family_status"))
                if fs in family_status_counts:
                    family_status_counts[fs] += 1
            except:
                pass

            last_login = item.get("Last_login_date")
            last_login = last_login.date() if isinstance(last_login, datetime) else last_login
            if last_login == today:
                today_login += 1
            elif last_login == yesterday:
                yesterday_login += 1

            mem_end = item.get("membership_enddate")
            mem_end = mem_end.date() if isinstance(mem_end, datetime) else mem_end
            if mem_end:
                if mem_end.month == today.month and mem_end.year == today.year:
                    expiring_count += 1
                elif mem_end < today:
                    expired_count += 1

            call_st = item.get("last_call_status")
            try:
                call_st = int(call_st)
                if call_st == 1: hot += 1
                elif call_st == 2: warm += 1
                elif call_st == 3: cold += 1
                elif call_st == 4: not_interested += 1
            except:
                pass
            last_action = item.get("last_action_date")
            last_action = last_action.date() if isinstance(last_action, datetime) else last_action
            if last_action:
                gap = (today - last_action).days
                if gap > 45: idle_45_count += 1
                if gap > 90: idle_90_count += 1

            next_call = item.get("next_call_date")
            next_call = next_call.date() if isinstance(next_call, datetime) else next_call
            if next_call:
                if next_call == today:
                    today_work_count += 1
                elif next_call < today:
                    pending_work_count += 1
                    
            next_action = item.get("next_action_date")
            next_action = next_action.date() if isinstance(next_action, datetime) else next_action
            if next_action:
                if next_action == today:
                    today_task_count += 1
                elif next_action < today:
                    pending_task_count += 1
                    
            if item.get("has_photo"):
                has_photo_count += 1
            else:
                no_photo_count += 1

            if item.get("has_horo"):
                has_horo_count += 1
            else:
                no_horo_count += 1

        family_status_counts_named = {
            family_status_labels[k]: v for k, v in family_status_counts.items()
        }

        return Response({
            "status": True,
            "overall_count": overall_count,
            "filtered_count": len(final_filtered),
            "under_30": under_30,
            "above_30": above_30,
            "male_count": male_count,
            "female_count": female_count,
            "family_status_counts": family_status_counts_named,
            "today_login_count": today_login,
            "yesterday_login_count": yesterday_login,
            "expired_this_month_count": expiring_count,
            "call_status_counts": {
                "hot": hot,
                "warm": warm,
                "cold": cold,
                "not_interested": not_interested
            },
            "last_action_counts": {
                "over_45_days": idle_45_count,
                "over_90_days": idle_90_count
            },
            "action_counts": {
                "today_work": today_work_count,
                "pending_work": pending_work_count,
                "today_task": today_task_count,
                "pending_task": pending_task_count
            },
            "no_photo": no_photo_count,
            "no_horo": no_horo_count,
            "data": final_filtered
        })

class GetPlans(APIView):
    def get(self,request):
        plans = PlanDetails.objects.exclude(id__in=[4,6,7,8,9])
        return Response({
            "status":True,
            "plans":plans.values('id', 'plan_name')
        })  





class AutoAssignProfile(APIView):
    def post(self, request, profile_id):
        profile = get_object_or_404(LoginDetails, ProfileId=profile_id)

        assigned_admin = assign_user_for_state(profile.Profile_state)

        if assigned_admin:
            profile.Owner_id = str(assigned_admin.id)
            profile.save()

        return Response({
            "status": True,
            "profile_id": profile_id,
            "assigned_admin": assigned_admin.username if assigned_admin else None
        })
        

@api_view(['GET'])
def get_call_log_by_id(request, id):
    call_logs = (
        CallLog.objects
        .filter(id=id, is_deleted=0)
        .select_related('call_type', 'particulars', 'call_status')
        .annotate(
            call_type_name=F('call_type__call_type'),
            particulars_name=F('particulars__particulars'),
            call_status_name=F('call_status__status'),
        )
        .order_by('-call_date')
        .values(
            'id',
            'call_management_id',
            'call_date',
            'comments',
            'next_call_date',
            'call_owner',
            'created_at',

            'call_type_id',
            'call_type_name',

            'particulars_id',
            'particulars_name',

            'call_status_id',
            'call_status_name',
        )
    )


    user_ids = set()
    for log in call_logs:
        if log["call_owner"]:
            try:
                user_ids.add(int(log["call_owner"]))  
            except:
                pass

    users = User.objects.in_bulk(user_ids)

    for log in call_logs:
        owner_id = log["call_owner"]
        try:
            owner_id = int(owner_id)
        except:
            owner_id = None

        log["call_owner_name"] = (
            users[owner_id].username if owner_id in users else None
        )

    return Response({
        "call_logs": list(call_logs)
    })


@api_view(['GET'])
def get_action_log_by_id(request, id):

    action_logs = list(
        ActionLog.objects
        .filter(id=id, is_deleted=0)
        .select_related('action_point', 'next_action')
        .annotate(
            action_point_name=F('action_point__action_point'),
            next_action_name=F('next_action__action_point'),
        )
        .order_by('-action_date')
        .values(
            'id',
            'call_management_id',
            'action_date',
            'comments',
            'created_at',
            
            'action_point_id',
            'action_point_name',
            'next_action_date',
            'next_action_id',
            'next_action_name',

            'action_owner'
        )
    )

    user_ids = set()
    for log in action_logs:
        if log["action_owner"]:
            try:
                user_ids.add(int(log["action_owner"]))   
            except:
                pass

    users = User.objects.in_bulk(user_ids)

    for log in action_logs:
        owner_id = log["action_owner"]
        try:
            owner_id = int(owner_id)
        except:
            owner_id = None

        log["action_owner_name"] = (
            users[owner_id].username if owner_id in users else None
        )

    return Response({
        "action_logs": action_logs
    })

@api_view(['GET'])
def get_action_summary(request, profile_id):
    summary = { 
               "interest_sent": 0, 
               "interest_received": 0, 
               "interest_accepted": 0, 
               "interest_rejected": 0, 
               "bookmarked": 0, 
               "bookmark_received": 0, 
               "photo_request_sent": 0, 
               "photo_request_received": 0, 
               "visited_count": 0, 
               "viewed_count": 0, 
               }
    interests = Express_interests.objects.filter( Q(profile_from=profile_id) | Q(profile_to=profile_id), status__in=[1, 2, 3] ) 
    for ei in interests: 
        if ei.profile_from == profile_id: 
            if ei.status == 1: 
                summary["interest_sent"] += 1 
            elif ei.status == 2: 
                summary["interest_accepted"] += 1 
            elif ei.status == 3: 
                summary["interest_rejected"] += 1 
            else: 
                if ei.status == 1: summary["interest_received"] += 1 
                
    wishlists = Profile_wishlists.objects.filter( Q(profile_from=profile_id, status=1) | Q(profile_to=profile_id, status=1) ) 
    for wl in wishlists: 
        if wl.profile_from == profile_id: 
            summary["bookmarked"] += 1 
        else: 
            summary["bookmark_received"] += 1 
     
    visitors = Profile_visitors.objects.filter( Q(profile_id=profile_id, status=1) | Q(viewed_profile=profile_id, status=1) ) 
    for v in visitors: 
        if v.profile_id == profile_id: 
            summary["visited_count"] += 1 
        else: 
            summary["viewed_count"] += 1 
            
    return Response(summary)

def safe_int(val, default=0):
    try:
        return int(val)
    except (TypeError, ValueError):
        return default


def safe_str(val):
    return val if val not in (None, "null", "undefined") else ""


def is_missing_file(val):
    return (
        val is None
        or val == ""
        or val == 0
        or str(val).lower() in ("null", "undefined", "false")
    )


def to_date_safe(val):
    try:
        return val.date() if hasattr(val, "date") else val
    except Exception:
        return None


def dictfetchall(cursor):
    columns = [col[0] for col in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]


# ---------------- API VIEW ----------------

class NewRegistrationsDashboard(APIView):

    def get(self, request):

        # ---------------- GET PARAMS ----------------
        owner = safe_str(request.GET.get("owner", "26"))
        gender = safe_str(request.GET.get("genderFilter", ""))
        login = safe_str(request.GET.get("loginFilter", ""))
        call_status = safe_str(request.GET.get("callStatusFilter", ""))
        idle = safe_str(request.GET.get("idleDaysFilter", ""))
        from_date = safe_str(request.GET.get("from_date", ""))
        to_date = safe_str(request.GET.get("to_date", ""))
        age_from = safe_str(request.GET.get("age_from", ""))
        age_to = safe_str(request.GET.get("age_to", ""))
        plan_id = safe_str(request.GET.get("plan_id", ""))
        count_filter = safe_str(request.GET.get("countFilter", ""))
        search = safe_str(request.GET.get("search", ""))
        export_type = safe_str(request.GET.get("export", "")).lower()


        today = date.today()

        # STATE IDS
        TN = 2
        KAT = 4

        # ---------------- CALL STORED PROCEDURE ----------------
        with connection.cursor() as cursor:
            cursor.callproc(
                "GetNewRegistrationsDashboard",
                [
                    owner, "", gender, "", login,
                    call_status, idle,
                    from_date, to_date,
                    age_from, age_to,
                    plan_id,
                    count_filter,
                    search
                ]
            )

            base = dictfetchall(cursor) or []

            cursor.nextset()
            overall_row = cursor.fetchone()
            overall = overall_row[0] if overall_row else 0

            cursor.nextset()
            filtered = dictfetchall(cursor) or []

        base = base or []

        # ---------------- BASIC COUNTS ----------------
        total = len(base)

        approved = sum(1 for x in base if safe_int(x.get("status")) == 1)
        unapproved = total - approved

        non_login = sum(1 for x in base if not x.get("Last_login_date"))

        premium = 0
        for x in base:
            pid = safe_int(x.get("Plan_id"), None)
            if pid and pid not in [4, 6, 7, 8, 9]:
                premium += 1

        # ---------------- WORK / TASK COUNTS ----------------
        today_work = pending_work = 0
        today_task = pending_task = 0

        # ---------------- ONLINE / ADMIN COUNTS ----------------
        online_approved = admin_approved = 0
        online_unapproved = admin_unapproved = 0

        online_approved_tn = online_approved_kat = 0
        online_unapproved_tn = online_unapproved_kat = 0
        admin_approved_tn = admin_approved_kat = 0
        admin_unapproved_tn = admin_unapproved_kat = 0

        # ---------------- INTEREST COUNTS ----------------
        hot = warm = cold = not_interested = 0

        # ---------------- MAIN LOOP ----------------
        
        # ---------------- EXPORT ----------------
        if export_type in ["csv", "excel"]:

            export_rows = []

            for row in filtered:
                education = ""
                if row.get("degree_name") and row.get("other_degree"):
                    education = f"{row.get('degree_name')} / {row.get('other_degree')}"
                else:
                    education = row.get("degree_name") or row.get("other_degree") or ""

                export_rows.append({
                    "Profile ID": row.get("ProfileId"),
                    "Name": row.get("Profile_name"),
                    "Gender": row.get("Gender"),
                    "Age": row.get("age"),
                    "City": row.get("Profile_city"),
                    "State": row.get("state"),
                    "Education Details": education,
                    "Annual Income": row.get("income"),
                    "Family Status": row.get("family_status_name"),
                    "Plan": row.get("plan_name"),
                    "Status": row.get("status_name"),
                    "Owner": row.get("owner_name"),
                    "Registered Date": row.get("DateOfJoin"),
                    "Last Login": row.get("Last_login_date"),
                    "Call Status": row.get("call_status"),
                })

            df = pd.DataFrame(export_rows)

            if export_type == "csv":
                output = StringIO()
                df.to_csv(output, index=False)

                response = HttpResponse(
                    output.getvalue(),
                    content_type="text/csv"
                )
                response["Content-Disposition"] = (
                    'attachment; filename="new_registrations_report.csv"'
                )
                return response

            elif export_type == "excel":
                output = BytesIO()
                with pd.ExcelWriter(output, engine="openpyxl") as writer:
                    df.to_excel(writer, index=False, sheet_name="New Registrations")

                response = HttpResponse(
                    output.getvalue(),
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                response["Content-Disposition"] = (
                    'attachment; filename="new_registrations_report.xlsx"'
                )
                return response

        
        for x in base:

            status = safe_int(x.get("status"))
            state = safe_int(x.get("Profile_state"))
            profile_for = safe_int(x.get("Profile_for"))

            next_call = to_date_safe(x.get("next_call_date"))
            next_action = to_date_safe(x.get("next_action_date"))

            # Work
            if next_call:
                if next_call == today:
                    today_work += 1
                elif next_call < today:
                    pending_work += 1

            # Task
            if next_action:
                if next_action == today:
                    today_task += 1
                elif next_action < today:
                    pending_task += 1

            is_online = profile_for != 8
            is_admin = profile_for == 8
            is_approved = status == 1

            # Online
            if is_online:
                if is_approved:
                    online_approved += 1
                    if state == TN:
                        online_approved_tn += 1
                    else:
                        online_approved_kat += 1
                else:
                    online_unapproved += 1
                    if state == TN:
                        online_unapproved_tn += 1
                    else:
                        online_unapproved_kat += 1

            # Admin
            if is_admin:
                if is_approved:
                    admin_approved += 1
                    if state == TN:
                        admin_approved_tn += 1
                    else:
                        admin_approved_kat += 1
                else:
                    admin_unapproved += 1
                    if state == TN:
                        admin_unapproved_tn += 1
                    else:
                        admin_unapproved_kat += 1

            # Interest
            interest = safe_int(x.get("last_call_status"))
            if interest == 1:
                hot += 1
            elif interest == 2:
                warm += 1
            elif interest == 3:
                cold += 1
            elif interest == 4:
                not_interested += 1

        # ---------------- EXTRA COUNTS ----------------
        today_login = sum(
            1 for x in base
            if x.get("Last_login_date")
            and to_date_safe(x.get("Last_login_date")) == today
        )

        today_birthday = sum(
            1 for x in base
            if x.get("Profile_dob")
            and to_date_safe(x.get("Profile_dob")).day == today.day
            and to_date_safe(x.get("Profile_dob")).month == today.month
        )

        # ---------------- SAFE FILE COUNTS ----------------
        no_id = no_photo = no_horo = 0

        for x in base:
            if is_missing_file(x.get("Profile_idproof")):
                no_id += 1
            if is_missing_file(x.get("has_photo")):
                no_photo += 1
            if is_missing_file(x.get("has_horo")):
                no_horo += 1

        # ---------------- RESPONSE ----------------
        return Response({
            "status": True,
            "overall_count": overall,
            "filtered_count": len(filtered),

            "total_registration": total,
            "approved": approved,
            "unapproved": unapproved,
            "non_logged_in": non_login,
            "premium": premium,

            "today_login": today_login,
            "today_birthday": today_birthday,

            "work_counts": {
                "today_work": today_work,
                "pending_work": pending_work
            },

            "task_counts": {
                "today_task": today_task,
                "pending_task": pending_task
            },

            "online": {
                "total_approved": online_approved,
                "total_unapproved": online_unapproved,
                "approved": {"TN": online_approved_tn, "KAT": online_approved_kat},
                "unapproved": {"TN": online_unapproved_tn, "KAT": online_unapproved_kat}
            },

            "admin": {
                "total_approved": admin_approved,
                "total_unapproved": admin_unapproved,
                "approved": {"TN": admin_approved_tn, "KAT": admin_approved_kat},
                "unapproved": {"TN": admin_unapproved_tn, "KAT": admin_unapproved_kat}
            },

            "interest": {
                "hot": hot,
                "warm": warm,
                "cold": cold,
                "not_interested": not_interested
            },

            "no_photo": no_photo,
            "no_horo": no_horo,
            "no_id": no_id,

            "data": filtered
        })
        
PAYMENT_STATUS_MAP = {
    1: "created",
    2: "success",
    3: "failed"
}

class ProspectDashboard(APIView):

    def get(self, request):
        from_date = safe_str(request.GET.get("from_date", ""))
        to_date = safe_str(request.GET.get("to_date", ""))
        age_from = safe_str(request.GET.get("age_from", ""))
        age_to = safe_str(request.GET.get("age_to", ""))
        plan_id = safe_str(request.GET.get("plan_id", ""))
        count_filter = safe_str(request.GET.get("countFilter", ""))
        search = safe_str(request.GET.get("search", ""))
        export_type = safe_str(request.GET.get("export", "")).lower()

        owner = safe_str(request.GET.get("owner", "26"))
        gender = safe_str(request.GET.get("genderFilter", ""))
        login = safe_str(request.GET.get("loginFilter", ""))
        call_status = safe_str(request.GET.get("callStatusFilter", ""))
        idle = safe_str(request.GET.get("idleDaysFilter", ""))

        today = date.today()
        yesterday = today - timedelta(days=1)

        with connection.cursor() as cursor:
            cursor.callproc(
                "GetProspectDashboard",
                [
                    owner, "", gender, "", login,
                    call_status, idle,
                    from_date, to_date,
                    age_from, age_to,
                    plan_id,
                    count_filter,
                    search
                ]
            )

            base = dictfetchall(cursor) or []
            cursor.nextset()
            overall_row = cursor.fetchone()
            overall = overall_row[0] if overall_row else 0
            cursor.nextset()
            filtered = dictfetchall(cursor) or []

        if export_type in ["csv", "excel"]:
            return self.export_to_file(filtered, export_type)

        # Counter variables
        stats = {
            "under_30": 0,
            "above_30": 0,
            "today_work": 0,
            "pending_work": 0,
            "today_task": 0,
            "pending_task": 0,
            "prospect": 0,
            "prospect_tn":0,
            "prospect_non_tn":0,
            "free": 0,
            "free_tn":0,
            "free_non_tn":0,
            "basic": 0,
            "basic_tn":0,
            "basic_non_tn":0,
            "offer": 0,
            "offer_tn":0,
            "offer_non_tn":0,
            "hot": 0,
            "warm": 0,
            "cold": 0,
            "not_interested": 0,
            "no_id": 0,
            "no_photo": 0,
            "no_horo": 0,
            "payment_success": 0,
            "payment_failed": 0,
            "express_interest_20": 0,
            "idle_90_count": 0,
            "today_login": 0,
            "yesterday_login": 0,
            "today_birthday": 0,
            "cur_month_registrations": 0,
            "assigned_to_me": 0,
        }

        for row in base:
            dob = to_date_safe(row.get("Profile_dob"))
            login_date = to_date_safe(row.get("Last_login_date"))
            join_date = to_date_safe(row.get("DateOfJoin"))
            plan = safe_int(row.get("Plan_id"))
            age = calculate_age(dob)
            last_call_date = to_date_safe(row.get("last_call_date"))
            next_call = to_date_safe(row.get("next_call_date"))
            next_action = to_date_safe(row.get("next_action_date"))
            ps = safe_int(row.get("payment_status"))
            express_interest = safe_int(row.get("express_interest_count"))
            call_status = safe_int(row.get("last_call_status"))
            state = safe_int(row.get("Profile_state"))

            # Age
            if age is not None:
                if age < 30:
                    stats["under_30"] += 1
                else:
                    stats["above_30"] += 1

            # Login counts
            if login_date:
                if login_date == today:
                    stats["today_login"] += 1
                elif login_date == yesterday:
                    stats["yesterday_login"] += 1

            # Birthday count
            if dob and dob.day == today.day and dob.month == today.month:
                stats["today_birthday"] += 1

            # Registrations
            if join_date and join_date.month == today.month and join_date.year == today.year:
                stats["cur_month_registrations"] += 1

            # Work & tasks
            if next_call:
                if next_call == today:
                    stats["today_work"] += 1
                elif next_call < today:
                    stats["pending_work"] += 1
            if next_action:
                if next_action == today:
                    stats["today_task"] += 1
                elif next_action < today:
                    stats["pending_task"] += 1

            # Plan count
            plan_map = {6: "basic", 7: "free", 8: "prospect", 9: "offer"}
            if plan in plan_map:
                stats[plan_map[plan]] += 1
                if plan ==6:
                    if state == 2:
                        stats["basic_tn"] +=1
                    else:
                        stats["basic_non_tn"] +=1
                elif plan ==7:
                    if state == 2:
                        stats["free_tn"] +=1
                    else:
                        stats["free_non_tn"] +=1
                elif plan ==8:
                    if state == 2:
                        stats["prospect_tn"] +=1
                    else:
                        stats["prospect_non_tn"] +=1
                elif plan ==9:
                    if state == 2:
                        stats["offer_tn"] +=1
                    else:
                        stats["offer_non_tn"] +=1       

            # Interest status
            interest_map = {1: "hot", 2: "warm", 3: "cold", 4: "not_interested"}
            if call_status in interest_map:
                stats[interest_map[call_status]] += 1

            # Missing data
            if is_missing_file(row.get("Profile_idproof")):
                stats["no_id"] += 1
            if is_missing_file(row.get("has_photo")):
                stats["no_photo"] += 1
            if is_missing_file(row.get("has_horo")):
                stats["no_horo"] += 1

            # Payment
            if ps == 2:
                stats["payment_success"] += 1
            elif ps in (1, 3):
                stats["payment_failed"] += 1

            # Express Interest
            if express_interest > 20:
                stats["express_interest_20"] += 1

            # Idle count
            if last_call_date and (today - last_call_date).days > 90:
                stats["idle_90_count"] += 1
            
            if safe_int(row.get("assigned_to_owner")) == 1:
                stats["assigned_to_me"] += 1

        # Add payment status label to response
        for row in filtered:
            ps_filter = safe_int(row.get("payment_status"))
            row["payment_status_name"] = PAYMENT_STATUS_MAP.get(ps_filter, None)

        return Response({
            "status": True,
            "overall_count": overall,
            "filtered_count": len(filtered),
            "age_under_30": stats["under_30"],
            "age_above_30": stats["above_30"],
            "assigned_to_me": stats["assigned_to_me"],
            "total_profiles": len(base),
            "today_login": stats["today_login"],
            "yesterday_login": stats["yesterday_login"],
            "today_birthday": stats["today_birthday"],
            "idle_90_count": stats["idle_90_count"],
            "payment_counts": {
                "success": stats["payment_success"],
                "failed": stats["payment_failed"]
            },
            "work_counts": {
                "today_work": stats["today_work"],
                "pending_work": stats["pending_work"]
            },
            "task_counts": {
                "today_task": stats["today_task"],
                "pending_task": stats["pending_task"]
            },
            "express_interest": {
                "above_20": stats["express_interest_20"]
            },
            "plan_counts": {
                "prospect": stats["prospect"],
                "prospect_tn": stats["prospect_tn"],
                "prospect_non_tn": stats["prospect_non_tn"],
                "free": stats["free"],
                "free_tn": stats["free_tn"],
                "free_non_tn": stats["free_non_tn"],
                "basic": stats["basic"],
                "basic_tn": stats["basic_tn"],
                "basic_non_tn": stats["basic_non_tn"],
                "offer": stats["offer"],
                "offer_tn": stats["offer_tn"],
                "offer_non_tn": stats["offer_non_tn"],
            },
            "interest": {
                "hot": stats["hot"],
                "warm": stats["warm"],
                "cold": stats["cold"],
                "not_interested": stats["not_interested"]
            },
            "no_photo": stats["no_photo"],
            "no_horo": stats["no_horo"],
            "no_id": stats["no_id"],
            "cur_month_registrations": stats["cur_month_registrations"],
            "data": filtered
        })

    def export_to_file(self, rows, export_type):
        data = []

        for row in rows:
            education = " / ".join(filter(None, [row.get("degree_name"), row.get("other_degree")]))
            data.append({
                "Profile ID": row.get("ProfileId"),
                "Name": row.get("Profile_name"),
                "Gender": row.get("Gender"),
                "Age": row.get("age"),
                "City": row.get("Profile_city"),
                "State": row.get("state"),
                "Education Details": education,
                "Annual Income": row.get("income"),
                "Family Status": row.get("family_status_name"),
                "Plan": row.get("plan_name"),
                "Status": row.get("status_name"),
                "Owner": row.get("owner_name"),
                "Registered Date": row.get("DateOfJoin"),
                "Last Login": row.get("Last_login_date"),
                "Call Status": row.get("call_status"),
            })

        df = pd.DataFrame(data)

        if export_type == "csv":
            output = StringIO()
            df.to_csv(output, index=False)
            return HttpResponse(
                output.getvalue(),
                content_type="text/csv",
                headers={"Content-Disposition": 'attachment; filename="prospects.csv"'}
            )

        elif export_type == "excel":
            output = BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="Prospects")
            return HttpResponse(
                output.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": 'attachment; filename="prospects.xlsx"'}
            )  
            
                 
def month_diff(from_date, to_date):
    return (to_date.year - from_date.year) * 12 + (to_date.month - from_date.month)


class PremiumDashboard(APIView):

    def get(self, request):

        owner = safe_str(request.GET.get("owner", "26"))
        gender = safe_str(request.GET.get("genderFilter", ""))
        login = safe_str(request.GET.get("loginFilter", ""))
        call_status = safe_str(request.GET.get("callStatusFilter", ""))
        idle = safe_str(request.GET.get("idleDaysFilter", ""))
        from_date = safe_str(request.GET.get("from_date", ""))
        to_date = safe_str(request.GET.get("to_date", ""))
        age_from = safe_str(request.GET.get("age_from", ""))
        age_to = safe_str(request.GET.get("age_to", ""))
        plan_id = safe_str(request.GET.get("plan_id", ""))
        count_filter = safe_str(request.GET.get("countFilter", ""))
        search = safe_str(request.GET.get("search", ""))
        export_type = safe_str(request.GET.get("export", "")).lower()


        today = date.today()
        yesterday = today - timedelta(days=1)

        TN = 2
        KAT = 4

        with connection.cursor() as cursor:
            cursor.callproc(
                "GetPremiumDashboard",
                [
                    owner, "", gender, "", login,
                    call_status, idle,
                    from_date, to_date,
                    age_from, age_to,
                    plan_id,
                    count_filter,
                    search
                ]
            )

            base = dictfetchall(cursor) or []

            cursor.nextset()
            overall_row = cursor.fetchone()
            overall = overall_row[0] if overall_row else 0

            cursor.nextset()
            filtered = dictfetchall(cursor) or []

        base = base or []

        total = len(base)
        under_30 = above_30 = 0

        today_work = pending_work = 0
        today_task = pending_task = 0
        hot = warm = cold = not_interested = 0
        gold = platinum = platinum_private = vys_delight = 0
        g_call_count =g_action_count=p_call_count=p_action_count=0
        pp_call_count=pp_action_count =vysd_call_count=vysd_action_count=0
        no_id = no_photo = no_horo = 0
        current_month_registration = 0
        male_count = female_count = 0
        express_interest_20 = 0
        idle_90_count=0
        tn=not_tn =0
        exp_four_month = first_three_month=0
        selected_month = 3 
        selected_formonth = 4
        yesterday_vys_assist = 0
        yesterday_express_sent = 0
        yesterday_express_received = 0
        yesterday_bookmark = 0



        
        if export_type in ["csv", "excel"]:

            export_rows = []

            for row in filtered:
                education = ""
                if row.get("degree_name") and row.get("other_degree"):
                    education = f"{row.get('degree_name')} / {row.get('other_degree')}"
                else:
                    education = row.get("degree_name") or row.get("other_degree") or ""

                export_rows.append({
                    "Profile ID": row.get("ProfileId"),
                    "Name": row.get("Profile_name"),
                    "Gender": row.get("Gender"),
                    "Age": row.get("age"),
                    "City": row.get("Profile_city"),
                    "State": row.get("state"),
                    "Education Details": education,
                    "Annual Income": row.get("income"),
                    "Family Status": row.get("family_status_name"),
                    "Plan": row.get("plan_name"),
                    "Owner": row.get("owner_name"),
                    "Activation Date": row.get("latest_plan_date"),
                    "Last Login": row.get("Last_login_date"),
                    "Call Status": row.get("call_status"),
                })

            df = pd.DataFrame(export_rows)

            if export_type == "csv":
                output = StringIO()
                df.to_csv(output, index=False)

                response = HttpResponse(
                    output.getvalue(),
                    content_type="text/csv"
                )
                response["Content-Disposition"] = (
                    'attachment; filename="new_registrations_report.csv"'
                )
                return response

            elif export_type == "excel":
                output = BytesIO()
                with pd.ExcelWriter(output, engine="openpyxl") as writer:
                    df.to_excel(writer, index=False, sheet_name="New Registrations")

                response = HttpResponse(
                    output.getvalue(),
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                response["Content-Disposition"] = (
                    'attachment; filename="new_registrations_report.xlsx"'
                )
                return response

        
        for x in base:
            state = safe_int(x.get("Profile_state"))
            next_call = to_date_safe(x.get("next_call_date"))
            next_action = to_date_safe(x.get("next_action_date"))
            plan = safe_int(x.get("Plan_id"))
            age = calculate_age(x.get("Profile_dob"))
            particular = safe_int(x.get("particulars_id"))
            today_action = safe_int(x.get("action_point_id"))
            call_days = safe_int(x.get("call_gap_days"))
            action_days = safe_int(x.get("action_gap_days"))
            mem_end = to_date_safe(x.get("membership_enddate"))
            mem_start = to_date_safe(x.get("membership_startdate"))
            if age is not None:
                if age < 30:
                    under_30 += 1
                else:
                    above_30 += 1
                    
            gender = str(x.get("Gender", "")).lower()
            if gender in ["male", "m", "1"]:
                male_count += 1
            elif gender in ["female", "f", "2"]:
                female_count += 1

            if state == 2:
                tn+=1
            else:
                not_tn+=1
            # Work
            if next_call:
                if next_call == today:
                    today_work += 1
                elif next_call < today:
                    pending_work += 1

            # Task
            if next_action:
                if next_action == today:
                    today_task += 1
                elif next_action < today:
                    pending_task += 1

            # Interest
            interest = safe_int(x.get("last_call_status"))
            if interest == 1:
                hot += 1
            elif interest == 2:
                warm += 1
            elif interest == 3:
                cold += 1
            elif interest == 4:
                not_interested += 1
                
            if is_missing_file(x.get("Profile_idproof")):
                no_id += 1
            if is_missing_file(x.get("has_photo")):
                no_photo += 1
            if is_missing_file(x.get("has_horo")):
                no_horo += 1
                
            if safe_int(x.get("express_interest_count")) > 20:
                express_interest_20 += 1
                
            last_action = x.get("last_call_date")
            last_action = last_action.date() if isinstance(last_action, datetime) else last_action
            if last_action:
                gap = (today - last_action).days
                if gap > 90: idle_90_count += 1


            if mem_start:
                mem_start = mem_start.date() if isinstance(mem_start, datetime) else mem_start
                first_three = (today - mem_start).days
                if first_three <= 90:
                    first_three_month += 1

            if mem_end:
                mem_end = mem_end.date() if isinstance(mem_end, datetime) else mem_end
                last_four = (mem_end-today).days
                if 0 <= last_four <= 120:
                    exp_four_month += 1
 
            if plan == 1:
                gold += 1
                if call_days > 60 and particular == 4:
                    g_call_count += 1
                if action_days > 60 and today_action == 11:
                    g_action_count += 1
            elif plan == 2:
                platinum += 1
                if call_days > 45 and particular == 4:
                    p_call_count += 1
                if action_days > 45 and today_action == 11:
                    p_action_count += 1
            elif plan == 3:
                platinum_private +=1
                if call_days > 30 and particular == 4:
                    pp_call_count += 1
                if action_days > 30 and today_action == 11:
                    pp_action_count += 1
            elif plan == 16:
                vys_delight +=1
                if call_days > 30 and particular == 4:
                    vysd_call_count += 1
                if action_days > 30 and today_action == 11:
                    vysd_action_count += 1
                    
            if safe_int(x.get("yesterday_vys_assist")) == 1:
                yesterday_vys_assist += 1

            if safe_int(x.get("yesterday_express_sent")) == 1:
                yesterday_express_sent += 1

            if safe_int(x.get("yesterday_express_received")) == 1:
                yesterday_express_received += 1

            if safe_int(x.get("yesterday_bookmark")) == 1:
                yesterday_bookmark += 1



        # ---------------- EXTRA COUNTS ----------------
        today_login = sum(
            1 for x in base
            if x.get("Last_login_date")
            and to_date_safe(x.get("Last_login_date")) == today
        )

        today_birthday = sum(
            1 for x in base
            if x.get("Profile_dob")
            and to_date_safe(x.get("Profile_dob")).day == today.day
            and to_date_safe(x.get("Profile_dob")).month == today.month
        )
        yesterday_login = sum(
            1 for x in base
            if x.get("Last_login_date")
            and to_date_safe(x.get("Last_login_date")) == yesterday
        )
        current_month_registration = sum(
            1 for x in base
            if x.get("membership_startdate")
            and to_date_safe(x.get("membership_startdate")).month == today.month
            and to_date_safe(x.get("membership_startdate")).year == today.year
        )
            

        # ---------------- RESPONSE ----------------
        return Response({
            "status": True,
            "overall_count": overall,
            "filtered_count": len(filtered),
            "age_under_30":under_30,
            "age_above_30":above_30,
            "male_count":male_count,
            "female":female_count,
            "total_profiles": total,
            "cur_month_registrations":current_month_registration,

            "today_login": today_login,
            "yesterday_login": yesterday_login,
            "today_birthday": today_birthday,
            "idle_90_count":idle_90_count,
            "yesterday_activity": {
                "vys_assist": yesterday_vys_assist,
                "express_interest_sent": yesterday_express_sent,
                "express_interest_received": yesterday_express_received,
                "bookmark": yesterday_bookmark
            },

            "state_count":{
                "tn":tn,
                "non-tn":not_tn
                },
            "fisrt_three_month":first_three_month,
            "last_four_month":exp_four_month,
            "plan_counts":{
                "gold":{
                    "total":gold,
                    "call":g_call_count,
                    "action":g_action_count
                    },
                "platinum":{
                    "total":platinum,
                    "call":p_call_count,
                    "action":p_action_count
                },
                "platinum_private":{
                    "total":platinum_private,
                    "call":pp_call_count,
                    "action":pp_action_count
                },
                "vysyamala_delight":{
                    "total":vys_delight,
                    "call":vysd_call_count,
                    "action":vysd_action_count
                }},
            "work_counts": {
                "today_work": today_work,
                "pending_work": pending_work
            },
            "express_interest": {
                "above_20": express_interest_20
            },

            "task_counts": {
                "today_task": today_task,
                "pending_task": pending_task
            },

            "interest": {
                "hot": hot,
                "warm": warm,
                "cold": cold,
                "not_interested": not_interested
            },

            "no_photo": no_photo,
            "no_horo": no_horo,
            "no_id": no_id,

            "data": filtered
        })
        
def dictfetchall(cursor):
    "Return all rows from a cursor as a dict"
    columns = [col[0] for col in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]


class DailyWorkDashboard(APIView):

    def get(self, request):
        owner_id = safe_str(request.GET.get("owner", "26"))
        count_filter = safe_str(request.GET.get("countFilter", ""))
        export_type = safe_str(request.GET.get("export", "")).lower()

        with connection.cursor() as cursor:
            cursor.callproc("GetDailyWorkDashboard", [owner_id, count_filter])

            filtered_profiles = dictfetchall(cursor)  # resultset 1: filtered profiles
            cursor.nextset()
            counts_by_type_list = dictfetchall(cursor)  # resultset 2.1
            cursor.nextset()
            global_counts = dictfetchall(cursor) 

        counts_by_type = {row["dashboard_type"]: row for row in counts_by_type_list}
        counts_by_type["all"] = global_counts[0]
        # Export if needed
        if export_type in ["csv", "excel"]:
            return self.export_to_file(filtered_profiles, export_type)

        return Response({
            "status": True,
            "counts_by_type": counts_by_type,
            "filtered_count": len(filtered_profiles),
            "data": filtered_profiles
        })

    def export_to_file(self, rows, export_type):
        data = []
        for row in rows:
            data.append({
                "Profile ID": row.get("ProfileId"),
                "Gender": row.get("Gender"),
                "Plan": row.get("plan_name"),
                "Membership End Date": row.get("membership_enddate"),
                "Next Call Date": row.get("next_call_date"),
                "Next Action Date": row.get("next_action_date"),
                "Owner": row.get("owner_name"),
            })

        df = pd.DataFrame(data)

        if export_type == "csv":
            output = StringIO()
            df.to_csv(output, index=False)
            return HttpResponse(
                output.getvalue(),
                content_type="text/csv",
                headers={"Content-Disposition": 'attachment; filename="daily_work_dashboard.csv"'}
            )

        elif export_type == "excel":
            output = BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="DailyWork")
            return HttpResponse(
                output.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": 'attachment; filename="daily_work_dashboard.xlsx"'}
            )     
            
class MarriageDashboard(APIView):

    def get(self, request):

        owner = safe_str(request.GET.get("owner", "26"))
        profile_id = safe_str(request.GET.get("profile_id", ""))
        particular_id = safe_str(request.GET.get("particular_id", ""))
        gender = safe_str(request.GET.get("genderFilter", ""))
        family = safe_str(request.GET.get("familyFilter", ""))
        login = safe_str(request.GET.get("loginFilter", ""))
        call_status = safe_str(request.GET.get("callStatusFilter", ""))
        idle = safe_str(request.GET.get("idleDaysFilter", ""))
        from_date = safe_str(request.GET.get("from_date", ""))
        to_date = safe_str(request.GET.get("to_date", ""))
        age_from = safe_str(request.GET.get("age_from", ""))
        age_to = safe_str(request.GET.get("age_to", ""))
        plan_id = safe_str(request.GET.get("plan_id", ""))
        count_filter = safe_str(request.GET.get("countFilter", ""))
        search = safe_str(request.GET.get("search", ""))
        export_type = safe_str(request.GET.get("export", "")).lower()
        order_by = safe_str(request.GET.get("order_by", "asc")).lower()

        today = date.today()
        yesterday = today - timedelta(days=1)

        with connection.cursor() as cursor:
            cursor.callproc(
                "GetMarriageDashboard",
                [
                    owner,
                    profile_id,
                    particular_id,
                    gender,
                    family,
                    login,
                    call_status,
                    idle,
                    from_date,
                    to_date,
                    age_from,
                    age_to,
                    plan_id,
                    count_filter,
                    search,
                    order_by
                ]
            )

            base = dictfetchall(cursor) or []

            cursor.nextset()
            overall_row = cursor.fetchone()
            overall = overall_row[0] if overall_row else 0

            cursor.nextset()
            filtered = dictfetchall(cursor) or []

        total = len(base)

        if export_type in ["csv", "excel"]:

            export_rows = []
            for row in filtered:
                education = ""
                if row.get("degree_name") and row.get("other_degree"):
                    education = f"{row.get('degree_name')} / {row.get('other_degree')}"
                else:
                    education = row.get("degree_name") or row.get("other_degree") or ""

                export_rows.append({
                    "Profile ID": row.get("ProfileId"),
                    "Name": row.get("Profile_name"),
                    "Gender": row.get("Gender"),
                    "Age": row.get("age"),
                    "City": row.get("Profile_city"),
                    "State": row.get("state"),
                    "Education": education,
                    "Annual Income": row.get("income"),
                    "Family Status": row.get("family_status_name"),
                    "Plan": row.get("plan_name"),
                    "Owner": row.get("owner_name"),
                    "Last Login": row.get("Last_login_date"),
                    "Call Status": row.get("call_status"),
                })

            df = pd.DataFrame(export_rows)

            if export_type == "csv":
                output = StringIO()
                df.to_csv(output, index=False)
                return HttpResponse(
                    output.getvalue(),
                    content_type="text/csv",
                    headers={
                        "Content-Disposition":
                        'attachment; filename="marriage_dashboard.csv"'
                    }
                )

            output = BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df.to_excel(writer, index=False)
            return HttpResponse(
                output.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition":
                    'attachment; filename="marriage_dashboard.xlsx"'
                }
            )
            
        premium = free_offer = propect = 0
        premium_tn = premium_not_tn = free_offer_tn = free_offer_not_tn = propect_tn = propect_not_tn =0
        today_work = pending_work = 0
        today_task = pending_task = 0
        pre_settle_vys = pre_settle_oth = pre_both = pre_single =0
        fop_settle_vys = fop_settle_oth = fop_both = fop_single =0
        nm_date = ne_date = nm_photo = ne_photo = 0
        wish_card_accept = wish_card_rejected = instagram_accept = 0
        upcoming_marriage = current_month_marriage =0
        assigned_to=0

        for x in base:
            plan = safe_int(x.get("Plan_id"))
            state = safe_int(x.get("Profile_state"))
            marriage_date = to_date_safe(x.get("marriagedate"))
            settle = to_date_safe(x.get("settledthru"))
            mr_id = to_date_safe(x.get("groombridevysysaid"))
            eng_date = to_date_safe(x.get("engagementdate"))
            marriage_photo = to_date_safe(x.get("marriagephotodetails"))
            eng_photo = to_date_safe(x.get("engagementphotodetails"))
            wish_card_accept_val = to_date_safe(x.get("wish_card_accept"))
            instagram_accept_val = to_date_safe(x.get("instagram_accept"))
            if wish_card_accept_val ==1:
                wish_card_accept +=1
            else:
                wish_card_rejected +=1
            if instagram_accept_val ==1:
                instagram_accept +=1
                
            if eng_date:
                pass
            else:
                ne_date +=1
            if marriage_date:
                pass
            else:
                nm_date +=1
            if eng_photo:
                pass
            else:
                ne_photo +=1
            if marriage_photo:
                pass
            else:
                nm_photo +=1
                            
            if settle:
                if settle.lower()=="vysyamala":
                    if plan in [1,2,3,16]:
                        pre_settle_vys += 1
                    elif plan in [7,8,9]:
                        fop_settle_vys += 1
                else:
                    if plan in [1,2,3,16]:
                        pre_settle_oth += 1
                    elif plan in [7,8,9]:
                        fop_settle_oth += 1        
            if mr_id:
                if plan in [1,2,3,16]:
                    pre_both += 1
                elif plan in [7,8,9]:
                    fop_both += 1
            else:
                if plan in [1,2,3,16]:
                    pre_single += 1
                elif plan in [7,8,9]:
                    fop_single += 1
                    
            if safe_int(x.get("assigned_to_owner")) == 1:
                assigned_to += 1
            if marriage_date:
                if marriage_date.month == today.month and marriage_date.year == today.year:
                    current_month_marriage +=1
                if marriage_date >= today:
                    upcoming_marriage +=1
                    
            if plan in [1, 2, 3, 16]:
                premium += 1
                if state == 2:
                    premium_tn += 1
                else:
                    premium_not_tn += 1
            elif plan in [7, 9]:
                free_offer += 1
                if state == 2:
                    free_offer_tn += 1
                else:
                    free_offer_not_tn += 1
            elif plan == 8:
                propect += 1
                if state == 2:
                    propect_tn += 1
                else:
                    propect_not_tn += 1


            next_call = to_date_safe(x.get("next_call_date"))
            next_action = to_date_safe(x.get("next_action_date"))

            if next_call:
                today_work += next_call == today
                pending_work += next_call < today

            if next_action:
                today_task += next_action == today
                pending_task += next_action < today


        # ---------------- RESPONSE ----------------
        return Response({
            "status": True,
            "overall_count": overall,
            "filtered_count": len(filtered),
            "total_profiles": total,
            "assigned_to_me": assigned_to,
            "pre_settlement": {
                "vysyamala": pre_settle_vys,
                "others": pre_settle_oth,
                "both": pre_both,
                "single": pre_single
            },
            "free_offer_settlement": {
                "vysyamala": fop_settle_vys,
                "others": fop_settle_oth,   
                "both": fop_both,
                "single": fop_single
            },
            "missing_details": {
                "no_marriage_date": nm_date,
                "no_engagement_date": ne_date,
                "no_marriage_photo": nm_photo,
                "no_engagement_photo": ne_photo,
                "wish_card_accepted": wish_card_accept,
                "wish_card_rejected": wish_card_rejected,
                "instagram_accepted": instagram_accept
            },
            "marriage_counts": {
                "upcoming_marriage": upcoming_marriage,
                "current_month_marriage": current_month_marriage
            },
            "plan_counts":{
                "premium":{
                    "total":premium,
                    "tn":premium_tn,
                    "non-tn":premium_not_tn
                    },
                "free_offer":{
                    "total":free_offer,
                    "tn":free_offer_tn,
                    "non-tn":free_offer_not_tn
                },
                "propect":{
                    "total":propect,
                    "tn":propect_tn,
                    "non-tn":propect_not_tn
                }
                },
            
            "work_counts": {
                "today_work": today_work,
                "pending_work": pending_work
            },

            "task_counts": {
                "today_task": today_task,
                "pending_task": pending_task
            },


            "data": filtered
        })

class DeleteDashboard(APIView):

    def get(self, request):

        owner = safe_str(request.GET.get("owner", "26"))
        profile_id = safe_str(request.GET.get("profile_id", ""))
        gender = safe_str(request.GET.get("genderFilter", ""))
        family = safe_str(request.GET.get("familyFilter", ""))
        login = safe_str(request.GET.get("loginFilter", ""))
        call_status = safe_str(request.GET.get("callStatusFilter", ""))
        idle = safe_str(request.GET.get("idleDaysFilter", ""))
        from_date = safe_str(request.GET.get("from_date", ""))
        to_date = safe_str(request.GET.get("to_date", ""))
        age_from = safe_str(request.GET.get("age_from", ""))
        age_to = safe_str(request.GET.get("age_to", ""))
        plan_id = safe_str(request.GET.get("plan_id", ""))
        count_filter = safe_str(request.GET.get("countFilter", ""))
        search = safe_str(request.GET.get("search", ""))
        order_dir = safe_str(request.GET.get("order", "desc")).lower()
        export_type = safe_str(request.GET.get("export", "")).lower()
        pending_param = safe_str(request.GET.get("pending", ""))
        hidden_param = safe_str(request.GET.get("hidden", ""))


        with connection.cursor() as cursor:
            cursor.callproc(
                "GetDeleteDashboard",
                [
                    owner,
                    profile_id,
                    gender,
                    family,
                    login,
                    call_status,
                    idle,
                    from_date,
                    to_date,
                    age_from,
                    age_to,
                    plan_id,
                    count_filter,
                    search,
                    order_dir,
                    int(pending_param or 0),
                    int(hidden_param or 0)
                ]
            )

            base = dictfetchall(cursor) or []

            cursor.nextset()
            overall_row = cursor.fetchone()
            overall_count = overall_row[0] if overall_row else 0

            cursor.nextset()
            filtered = dictfetchall(cursor) or []
            cursor.nextset()
            row = cursor.fetchone() or (0, 0)
            pending = row[0]
            hidden = row[1]
            pending_current_month = row[2]
            hidden_current_month = row[3]
            
            if pending_param == "1" or hidden_param == "1":
                cursor.nextset()
                filtered = dictfetchall(cursor) or []


        total_profiles = len(base)

        if export_type in ["csv", "excel"]:

            export_rows = []
            for row in filtered:
                education = ""
                if row.get("degree_name") and row.get("other_degree"):
                    education = f"{row['degree_name']} / {row['other_degree']}"
                else:
                    education = row.get("degree_name") or row.get("other_degree") or ""

                export_rows.append({
                    "Profile ID": row.get("ProfileId"),
                    "Name": row.get("Profile_name"),
                    "Gender": row.get("Gender"),
                    "Age": row.get("age"),
                    "City": row.get("Profile_city"),
                    "State": row.get("state"),
                    "Education": education,
                    "Annual Income": row.get("income"),
                    "Family Status": row.get("family_status_name"),
                    "Plan": row.get("plan_name"),
                    "Owner": row.get("owner_name"),
                    "Last Login": row.get("Last_login_date"),
                    "Deleted On": row.get("dh_date_time"),
                })

            df = pd.DataFrame(export_rows)

            if export_type == "csv":
                output = StringIO()
                df.to_csv(output, index=False)
                return HttpResponse(
                    output.getvalue(),
                    content_type="text/csv",
                    headers={
                        "Content-Disposition":
                        'attachment; filename="delete_dashboard.csv"'
                    }
                )

            output = BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df.to_excel(writer, index=False)

            return HttpResponse(
                output.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition":
                    'attachment; filename="delete_dashboard.xlsx"'
                }
            )
            
        tn=non_tn=0
        premium = pre_tn = pre_non_tn =0
        free =free_tn =free_non_tn =0
        offer =offer_tn =offer_non_tn =0
        prospect =prospect_tn =prospect_non_tn =0
        cur_month_delete = 0
        duplicate = fake = marriage = others = 0
        
        for x in base:
            state = safe_int(x.get("Profile_state"))
            plan = safe_int(x.get("Plan_id"))
            delete_date = to_date_safe(x.get("dh_date_time"))
            sub_id = safe_int(x.get("sub_status_id"))
            if delete_date:
                if delete_date.month == date.today().month and delete_date.year == date.today().year:
                    cur_month_delete += 1
            if sub_id == 18:
                duplicate += 1
            elif sub_id == 19:
                fake += 1
            elif sub_id in [20,21]:
                marriage += 1
            elif sub_id == 22:
                others += 1
                
            if state == 2:
                tn += 1
            else:
                non_tn += 1
            if plan in [1,2,3,16]:
                premium +=1
                if state == 2:
                    pre_tn += 1
                else:
                    pre_non_tn += 1
            elif plan ==7:
                free += 1
                if state == 2:
                    free_tn += 1
                else:
                    free_non_tn += 1
            elif plan == 9:
                offer += 1
                if state == 2:
                    offer_tn += 1
                else:
                    offer_non_tn += 1
            elif plan == 8:
                prospect += 1
                if state == 2:
                    prospect_tn += 1
                else:
                    prospect_non_tn += 1

        return Response({
            "status": True,
            "overall_count": overall_count,
            "filtered_count": len(filtered),
            "current_month_deletions": cur_month_delete,
            "state_counts":{
              "tn":tn,
              "non_tn":non_tn
            },
            "plan_counts":{
                "premium":{
                    "total":premium,
                    "tn":pre_tn,
                    "non_tn":pre_non_tn
                    },
                "free": {
                    "total":free,
                    "tn":free_tn,
                    "non_tn":free_non_tn
                    },
                "offer":{
                    "total":offer,
                    "tn":offer_tn,
                    "non_tn":offer_non_tn
                    },
                "prospect": {
                    "total":prospect,
                    "tn":prospect_tn,
                    "non_tn":prospect_non_tn
                    }
            },
            
            "other_status_counts": {
                "pending": pending,
                "hidden": hidden,
                "pending_current_month": pending_current_month,
                "hidden_current_month": hidden_current_month
            },
            "status_counts":{
                "duplicate":duplicate,
                "fake":fake,
                "marriage":marriage,
                "others":others
                },
            "total_profiles": total_profiles,
            "data": filtered
        })
        
        
def detect_input_type(value: str):
    if not value:
        return None

    value = value.strip()

    if re.match(r'^(VM|VF)\d+$', value):
        return "PROFILE_ID"

    if len(value) > 4:
        return "MOBILE"

    return None

def as_list(value):
    if value is None:
        return None
    if isinstance(value, list):
        return value
    return [value]

from django.db.models import Q, F, Window
from django.db.models.functions import RowNumber
from rest_framework.views import APIView
from rest_framework.response import Response
from datetime import datetime, timedelta
from django.db.models import Max



def get_status(status_id):
    if status_id or status_id==0:
        try:
            status = ProfileStatus.objects.get(status_code=status_id)
            return status.status_name
        except Exception as e:
            return None
    return None

def aware_date(date_str, end=False):
    if not date_str:
        return None
    dt = datetime.strptime(date_str, "%Y-%m-%d")
    if end:
        dt += timedelta(days=1)
    return timezone.make_aware(dt)

def excel_safe_datetime(dt):
    if not dt:
        return None
    if isinstance(dt, datetime) and timezone.is_aware(dt):
        return dt.replace(tzinfo=None)
    return dt


class CallManagementSearchAPI(APIView):
    def post(self, request):
        try:
            data = request.data
            search_value = data.get("search_value")
            export_type = data.get("export_type")
            input_type = detect_input_type(search_value)

            if input_type == "MOBILE":
                latest_ids = (
                        CallManagement_New.objects
                        .filter(mobile_no__icontains=search_value)
                        .values('mobile_no')
                        .annotate(latest_created=Max('created_at'))
                    )
                
                cms = CallManagement_New.objects.filter(
                    mobile_no__icontains=search_value,
                    created_at__in=[x['latest_created'] for x in latest_ids]
                ).order_by('-created_at')

                if not cms.exists():
                    return Response({"status": True, "count": 0, "profiles": []})

                profiles = []
                for cm in cms:
                    call_filters = Q(call_management=cm, is_deleted=0)
                    action_filters = Q(call_management=cm, is_deleted=0)
                    assign_filters = Q(call_management=cm, is_deleted=0)

                    if cl_from := data.get("call_from_date"):
                        call_filters &= Q(call_date__gte=aware_date(cl_from))
                    if cl_to := data.get("call_to_date"):
                        call_filters &= Q(call_date__lt=aware_date(cl_to,end=True))

                    if nc_from := data.get("next_call_from_date"):
                        call_filters &= Q(next_call_date__gte=aware_date(nc_from))
                    if nc_to := data.get("next_call_to_date"):
                        call_filters &= Q(next_call_date__lte=aware_date(nc_to,end=True))

                    if ct := data.get("call_type"):
                        call_filters &= Q(call_type_id__in=ct)
                    if cs := data.get("call_status"):
                        call_filters &= Q(call_status_id__in=cs)
                    if pt := data.get("particulars"):
                        call_filters &= Q(particulars_id__in=pt)
                    if co := data.get("call_owner"):
                        call_filters &= Q(call_owner__in=co)
                    if cc := data.get("call_comments"):
                        call_filters &= Q(comments__icontains=cc)

                    if af := data.get("action_from_date"):
                        action_filters &= Q(action_date__gte=aware_date(af))
                    if at := data.get("action_to_date"):
                        action_filters &= Q(action_date__lt=aware_date(at,end=True))

                    if naf := data.get("next_action_from_date"):
                        action_filters &= Q(next_action_date__gte=aware_date(naf))
                    if nat := data.get("next_action_to_date"):
                        action_filters &= Q(next_action_date__lte=aware_date(nat,end=True))

                    if ap := data.get("action_point"):
                        action_filters &= Q(action_point_id__in=ap)
                    if na := data.get("next_action"):
                        action_filters &= Q(next_action_id__in=na)
                    if ao := data.get("action_owner"):
                        action_filters &= Q(action_owner__in=ao)
                    if ac := data.get("action_comments"):
                        action_filters &= Q(comments__icontains=ac)
                    if nac := data.get("next_action_comments"):
                        action_filters &= Q(next_action_comments__icontains=nac)

                    if adf := data.get("assign_from_date"):
                        assign_filters &= Q(assigned_date__gte=aware_date(adf))
                    if adt := data.get("assign_to_date"):
                        assign_filters &= Q(assigned_date__lt=aware_date(adt,end=True))
                    if ab := data.get("assigned_by"):
                        assign_filters &= Q(assigned_by__in=ab)
                    if at := data.get("assigned_to"):
                        assign_filters &= Q(assigned_to__in=at)
                    if an := data.get("assign_notes"):
                        assign_filters &= Q(notes__icontains=an)

                    latest_call = CallLog_New.objects.filter(call_filters).order_by('-call_date', '-created_at').first()
                    latest_action = ActionLog_New.objects.filter(action_filters).order_by('-action_date', '-created_at').first()
                    latest_assign = AssignLog_New.objects.filter(assign_filters).order_by('-assigned_date', '-created_at').first()

                    profiles.append({
                        "mobile_no": cm.mobile_no,
                        "profile_id": cm.profile_id,
                        "particulars": getattr(latest_call.particulars, 'particulars', None) if latest_call else None,
                        "call_type": getattr(latest_call.call_type, 'call_type', None) if latest_call else None,
                        "call_comments": getattr(latest_call, 'comments', None),
                        "call_status": getattr(latest_call.call_status, 'status', None) if latest_call else None,
                        "call_date": excel_safe_datetime(getattr(latest_call, 'call_date', None)),
                        "next_call_date": excel_safe_datetime(getattr(latest_call, 'next_call_date', None)),
                        "action_point": getattr(latest_action.action_point, 'action_point', None) if latest_action else None,
                        "next_action_point": getattr(latest_action.next_action, 'action_point', None) if latest_action else None,
                        "next_action_date": excel_safe_datetime(getattr(latest_action, 'next_action_date', None)),
                        "owner": get_owner_name(
                            latest_call.call_owner if latest_call
                            else latest_action.action_owner if latest_action else None
                        ),
                        "work_assign": getattr(latest_assign, 'assigned_to', None),
                    })
                if export_type == "csv":
                    return export_csv(profiles, "general_call_management.csv")
                if export_type == "excel":
                    return export_excel_call(profiles, "general_call_management.xlsx")
                
                return Response({
                    "status": True,
                    "count": len(profiles),
                    "profiles": profiles
                })


            profiles = LoginDetails.objects.only("ProfileId", "Owner_id", "status", "Plan_id")
            latest_call_from = data.get("latest_call_date_from")
            latest_call_to   = data.get("latest_call_date_to")

            latest_action_from = data.get("latest_action_date_from")
            latest_action_to   = data.get("latest_action_date_to")

            filters = Q()
            if input_type == "PROFILE_ID":
                filters &= Q(ProfileId=search_value)

            if owner := data.get("owner"):
                filters &= Q(Owner_id__in=as_list(owner))
            if plan := data.get("plan"):
                filters &= Q(Plan_id__in=as_list(plan))
            if status := data.get("status"):
                filters &= Q(status__in=as_list(status))

            if filters:
                profiles = profiles.filter(filters)

            from_date, to_date = data.get("from_date"), data.get("to_date")
            call_filters = Q(is_deleted=0)
            action_filters = Q(is_deleted=0)
            assign_filters = Q(is_deleted=0)
            if cl_from := data.get("call_from_date"):
                call_filters &= Q(call_date__gte=aware_date(cl_from))
            if cl_to := data.get("call_to_date"):
                call_filters &= Q(call_date__lt=aware_date(cl_to,end=True))

            if nc_from := data.get("next_call_from_date"):
                call_filters &= Q(next_call_date__gte=aware_date(nc_from))
            if nc_to := data.get("next_call_to_date"):
                call_filters &= Q(next_call_date__lte=aware_date(nc_to,end=True))

            if ct := data.get("call_type"):
                call_filters &= Q(call_type_id__in=ct)
            if cs := data.get("call_status"):
                call_filters &= Q(call_status_id__in=cs)
            if pt := data.get("particulars"):
                call_filters &= Q(particulars_id__in=pt)
            if co := data.get("call_owner"):
                call_filters &= Q(call_owner__in=co)
            if cc := data.get("call_comments"):
                call_filters &= Q(comments__icontains=cc)

            if af := data.get("action_from_date"):
                action_filters &= Q(action_date__gte=aware_date(af))
            if at := data.get("action_to_date"):
                action_filters &= Q(action_date__lt=aware_date(at,end=True))

            if naf := data.get("next_action_from_date"):
                action_filters &= Q(next_action_date__gte=aware_date(naf))
            if nat := data.get("next_action_to_date"):
                action_filters &= Q(next_action_date__lte=aware_date(nat,end=True))

            if ap := data.get("action_point"):
                action_filters &= Q(action_point_id__in=ap)
            if na := data.get("next_action"):
                action_filters &= Q(next_action_id__in=na)
            if ao := data.get("action_owner"):
                action_filters &= Q(action_owner__in=ao)
            if ac := data.get("action_comments"):
                action_filters &= Q(comments__icontains=ac)
            if nac := data.get("next_action_comments"):
                action_filters &= Q(next_action_comments__icontains=nac)

            if adf := data.get("assign_from_date"):
                assign_filters &= Q(assigned_date__gte=aware_date(adf))
            if adt := data.get("assign_to_date"):
                assign_filters &= Q(assigned_date__lt=aware_date(adt,end=True))
            if ab := data.get("assigned_by"):
                assign_filters &= Q(assigned_by__in=ab)
            if at := data.get("assigned_to"):
                assign_filters &= Q(assigned_to__in=at)
            if an := data.get("assign_notes"):
                assign_filters &= Q(notes__icontains=an)

            if from_date and to_date:
                fd = aware_date(from_date)
                td = aware_date(to_date, end=True)
                call_filters &= Q(call_date__gte=fd, call_date__lt=td)
                action_filters &= Q(action_date__gte=fd, action_date__lt=td)
                assign_filters &= Q(assigned_date__gte=fd, assigned_date__lt=td)

            call_qs = CallLog.objects.filter(call_filters).annotate(
                rn=Window(
                    expression=RowNumber(),
                    partition_by=[F('call_management__profile_id')],
                    order_by=[F('call_date').desc(), F('created_at').desc()]
                )
            ).filter(rn=1).values(
                'call_management__profile_id',
                'particulars__particulars',
                'call_type__call_type',
                'comments',
                'call_status__status',
                'call_date',
                'next_call_date',
            )

            action_qs = ActionLog.objects.filter(action_filters).annotate(
                rn=Window(
                    expression=RowNumber(),
                    partition_by=[F('call_management__profile_id')],
                    order_by=[F('action_date').desc(), F('created_at').desc()]
                )
            ).filter(rn=1).values(
                'call_management__profile_id',
                'action_date',
                'action_point__action_point',
                'next_action__action_point',
                'next_action_date',
            )

            assign_qs = AssignLog.objects.filter(assign_filters).annotate(
                rn=Window(
                    expression=RowNumber(),
                    partition_by=[F('call_management__profile_id')],
                    order_by=[F('assigned_date').desc(), F('created_at').desc()]
                )
            ).filter(rn=1).values(
                'call_management__profile_id',
                'assigned_to',
            )

            call_dict = {c['call_management__profile_id']: c for c in call_qs}
            action_dict = {a['call_management__profile_id']: a for a in action_qs}
            assign_dict = {a['call_management__profile_id']: a for a in assign_qs}
            
            def in_range(dt, start, end):
                if not dt:
                    return False

                if timezone.is_naive(dt):
                    dt = timezone.make_aware(dt)

                if start and dt < start:
                    return False
                if end and dt > end:
                    return False
                return True
            
            latest_call_from_dt = aware_date(latest_call_from)
            latest_call_to_dt   = aware_date(latest_call_to, end=True)

            latest_action_from_dt = aware_date(latest_action_from)
            latest_action_to_dt   = aware_date(latest_action_to, end=True)

            call_filters_applied = bool(
                data.get("call_from_date") or
                data.get("call_to_date") or
                data.get("call_type") or
                data.get("call_status") or
                data.get("particulars") or
                data.get("call_owner") or
                data.get("call_comments") or
                data.get("next_call_from_date") or
                data.get("next_call_to_date")
            )
            action_filters_applied = bool(
                data.get("action_from_date") or
                data.get("action_to_date") or
                data.get("next_action_from_date") or
                data.get("next_action_to_date") or
                data.get("action_point") or
                data.get("next_action") or
                data.get("action_owner") or
                data.get("action_comments") or
                data.get("next_action_comments")
            )
            assign_filters_applied = bool(
                data.get("assign_from_date") or
                data.get("assign_to_date") or
                data.get("assigned_by") or
                data.get("assigned_to") or
                data.get("assign_notes") 
            )

            valid_ids = None

            if call_filters_applied:
                valid_ids = set(call_dict.keys())

            if action_filters_applied:
                valid_ids = valid_ids & set(action_dict.keys()) if valid_ids else set(action_dict.keys())

            if assign_filters_applied:
                valid_ids = valid_ids & set(assign_dict.keys()) if valid_ids else set(assign_dict.keys())

            if valid_ids is not None:
                profiles = profiles.filter(ProfileId__in=valid_ids)
                
            if from_date or to_date:
                valid_ids = (
                    set(call_dict.keys()) |
                    set(action_dict.keys()) |
                    set(assign_dict.keys())
                )
                profiles = profiles.filter(ProfileId__in=valid_ids)

            owner_map = {
                u.id: u.username
                for u in User.objects.filter(id__in={p.Owner_id for p in profiles if p.Owner_id})
            }

            status_map = {
                s.status_code: s.status_name
                for s in ProfileStatus.objects.filter(
                    status_code__in={p.status for p in profiles if p.status is not None}
                )
            }
            if latest_call_from_dt or latest_call_to_dt or latest_action_from_dt or latest_action_to_dt:
                valid_ids = set()

                for pid in profiles.values_list('ProfileId', flat=True):
                    latest_call = call_dict.get(pid)
                    latest_action = action_dict.get(pid)

                    call_ok = True
                    action_ok = True

                    if latest_call_from_dt or latest_call_to_dt:
                        call_ok = in_range(
                            latest_call.get('call_date') if latest_call else None,
                            latest_call_from_dt,
                            latest_call_to_dt
                        )

                    if latest_action_from_dt or latest_action_to_dt:
                        action_ok = in_range(
                            latest_action.get('action_date') if latest_action else None,
                            latest_action_from_dt,
                            latest_action_to_dt
                        )


                    if call_ok and action_ok:
                        valid_ids.add(pid)

                profiles = profiles.filter(ProfileId__in=valid_ids)


            profiles_data_dict = {}
            for p in profiles:
                pid = p.ProfileId
                if pid in profiles_data_dict:
                    continue 
                c = call_dict.get(pid)
                a = action_dict.get(pid)
                s = assign_dict.get(pid)

                profiles_data_dict[pid]={
                    "profile_id": pid,
                    "particulars": c.get('particulars__particulars') if c else None,
                    "call_type": c.get('call_type__call_type') if c else None,
                    "call_comments": c.get('comments') if c else None,
                    "call_status": c.get('call_status__status') if c else None,
                    "call_date": excel_safe_datetime(c.get('call_date')) if c else None,
                    "next_call_date": excel_safe_datetime(c.get('next_call_date')) if c else None,
                    "next_action_date": excel_safe_datetime(a.get('next_action_date')) if a else None,
                    "action_point": a.get('action_point__action_point') if a else None,
                    "next_action_point": a.get('next_action__action_point') if a else None,
                    "work_assign": s.get('assigned_to') if s else None,
                    "owner": owner_map.get(int(p.Owner_id)) if p.Owner_id else None,
                    "profile_status": status_map.get(p.status),
                    "lad_call_date": excel_safe_datetime(c.get('call_date')) if c else None,
                }
                
            profiles_data = list(profiles_data_dict.values())
            if export_type == "csv":
                return export_csv(profiles_data, "profile_call_management.csv")
            if export_type == "excel":
                return export_excel_call(profiles_data, "profile_call_management.xlsx")
                            
            return Response({
                "status": True,
                "count": len(profiles_data),
                "profiles": profiles_data
            })

        except Exception as e:
            return Response({"status": False, "error": str(e)}, status=500)


class ClickToCallAPI(APIView):
    pagination_class = StandardResultsPaging
    def get(self,request):
        
        from_date = request.query_params.get("from_date")
        to_date = request.query_params.get("to_date")
        profile_id = request.query_params.get("profile_id")
        export_type = request.query_params.get("export")
        
        calls = Profile_callogs.objects.filter(status=1).order_by('-req_datetime')
        
        if from_date and to_date:
            try:
                from_date = datetime.strptime(from_date, "%Y-%m-%d")
                to_date = datetime.strptime(to_date, "%Y-%m-%d")
                to_date = to_date.replace(hour=23, minute=59, second=59)


                calls = calls.filter(
                req_datetime__range=(from_date, to_date)
                )
            except ValueError:
                return Response(
                {"status": False, "message": "Invalid date format. Use YYYY-MM-DD"},
                status=400
                )
        
        if profile_id:
            calls = calls.filter(
            Q(profile_from=profile_id) |
            Q(profile_to=profile_id)
            )
        
        profile_ids = set()
        for c in calls:
            profile_ids.add(c.profile_from)
            profile_ids.add(c.profile_to)

        profiles = LoginDetails.objects.filter(ProfileId__in=profile_ids)
        profile_map = {p.ProfileId: p for p in profiles}
        plan_map = {
            p.id: p.plan_name
            for p in PlanDetails.objects.all()
        }
        
        profile_status_map = {
            ps.status_code: ps.status_name
            for ps in ProfileStatus.objects.all()
        }
        
        state_map = {
            s.id: s.name
            for s in State.objects.filter(is_deleted=False)
        }
        
        
        def resolve_state(profile):
            state = profile.Profile_state
            if isinstance(state, str) and not state.isdigit():
                return state
            return state_map.get(int(state)) if state else None
        
        result = []

        for c in calls:
            profile_from = profile_map.get(c.profile_from)
            profile_to = profile_map.get(c.profile_to)

            if not profile_from or not profile_to:
                continue

            result.append({
                    'profile_from_id': profile_from.ProfileId,
                    'profile_from_name': profile_from.Profile_name,
                    'profile_from_gender': profile_from.Gender,
                    'profile_from_city': profile_from.Profile_city,
                    'profile_from_state': resolve_state(profile_from),
                    'profile_from_plan': plan_map.get(safe_int(profile_from.Plan_id)) if profile_from.Plan_id else None,
                    'profile_from_status': profile_status_map.get(safe_int(profile_from.status)) if profile_from.status else None,

                    'profile_to_id': profile_to.ProfileId,
                    'profile_to_name': profile_to.Profile_name,
                    'profile_to_gender': profile_to.Gender,
                    'profile_to_city': profile_to.Profile_city,
                    'profile_to_state': resolve_state(profile_to),
                    'profile_to_plan': plan_map.get(safe_int(profile_to.Plan_id)) if profile_to.Plan_id else None,
                    'profile_to_status': profile_status_map.get(safe_int(profile_to.status)) if profile_to.status else None,

                    'click_to_call_datetime': c.req_datetime.isoformat(),
                })
        filename = f"click_to_call_logs_.{export_type}"
        if export_type == "csv":
            return export_login_csv(filename, result)
        if export_type == "xlsx":
            return export_login_xlsx(filename, result)

        paginator = self.pagination_class()
        page = paginator.paginate_queryset(result, request)

        if page is not None:
            return paginator.get_paginated_response(page)

        return Response(result, status=200)
        
    
